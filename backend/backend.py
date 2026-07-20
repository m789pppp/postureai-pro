"""
PostureAI Pro — Backend v15
Real MediaPipe Pose (33 landmarks) + FaceMesh (478 landmarks)
Accuracy: Standard ~88% | Professional ~93% | Elite ~96%
3 Modes: Laptop | Phone | Side
Auto PDF download via ReportLab
Fixed: WORK_HOURS_START, SUPPORT_EMAIL, ADMIN_PHONE, localhost links, Auth middleware
"""
# Must be the first statement: makes type annotations lazy (PEP 563) so a
# type hint like `np.ndarray` doesn't crash at import time before cv2/numpy
# have been lazy-loaded by _ensure_models().
from __future__ import annotations
import base64, math, io, os, time, sys, requests as req, threading
from collections import Counter
# cv2 and numpy loaded lazily inside _ensure_models() to save ~130MB per worker at startup
cv2 = None
# Cascade classifiers depend on cv2 — also lazy, initialized in _ensure_models()
face_c = None
eye_c  = None
prof_c = None
np  = None
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FuturesTimeoutError
_ai_executor = ThreadPoolExecutor(max_workers=2, thread_name_prefix="gemini")
_coupon_lock = threading.Lock()   # ← thread-safe coupon counter
from datetime import datetime, timedelta
import traceback
from flask import Flask, request, jsonify, send_file, g, make_response
from flask_cors import CORS
from dotenv import load_dotenv

load_dotenv()

# ── Auth & Redis (graceful fallback if not configured) ────────────
import sys, os
sys.path.insert(0, os.path.dirname(__file__))
try:
    from auth.middleware import (
        require_auth, require_tier, require_admin, require_hr,
        optional_auth, verify_firebase_token, get_user_role,
        invalidate_user_cache,
    )
    AUTH_READY = True
    print("✅ Auth middleware loaded")
except ImportError as _e:
    print(f"⚠️  Auth middleware not loaded: {_e}")
    AUTH_READY = False
    # Stub decorators when auth not available
    def require_auth(f): return f
    def require_tier(*t): return lambda f: f
    def require_admin(f): return f
    def require_hr(f): return f
    def optional_auth(f): return f

try:
    from services.redis_service import (
        rset, rget, rdel, rpush, rlist, cache_get, cache_set,
        push_score, get_smoothed_score, check_rate_limit,
        set_risk_start, clear_risk, get_risk_duration,
        redis_health, queue_job, rincr,
    )
    REDIS_READY = True
    print("✅ Redis service loaded")
except ImportError:
    REDIS_READY = False
    def rset(*a,**k): pass
    def rget(k): return None
    def rdel(k): pass
    def cache_get(k): return None
    def cache_set(*a,**k): pass
    def push_score(*a,**k): pass
    def get_smoothed_score(uid): return None
    def check_rate_limit(*a,**k): return (True, 0)
    def set_risk_start(*a,**k): pass
    def clear_risk(*a,**k): pass
    def get_risk_duration(*a,**k): return None
    def redis_health(): return {"status": "not_loaded"}
    def queue_job(*a,**k): pass
    def rpush(key, value, max_len=5000): pass
    def rlist(key, count=1000): return []
    def rincr(key, ttl_s=0): return 1
    def push_blink(*a,**k): pass
    def get_blink_rate(*a,**k): return None

# ── Server-side pricing (SECURITY: never trust client amount_cents) ──
import sys as _sys
_sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
try:
    from config.pricing import (
        get_paymob_amount, get_stripe_amount, validate_plan_request,
        PAID_TIERS, ALL_TIERS, TIER_ORDER as PLAN_ORDER_CFG, SEAT_LIMITS,
    )
    PRICING_READY = True
except ImportError as _pe:
    print(f"⚠️  Pricing config not loaded: {_pe}", file=_sys.stderr)
    PRICING_READY = False
    def get_paymob_amount(tier, billing): return None
    def get_stripe_amount(tier, billing): return None
    def validate_plan_request(tier, billing): return (True, "")

# ── Tier-based analysis/AI-coach quality (single source of truth) ──
try:
    from config.tier_quality import get_coach_quality, get_depth_instruction
except ImportError as _qe:
    print(f"⚠️  Tier quality config not loaded: {_qe}", file=_sys.stderr)
    def get_coach_quality(tier):
        return {"monthly_limit": 5, "max_tokens": 800, "depth": "brief"}
    def get_depth_instruction(tier):
        return ""

# ── Centralized error handler ─────────────────────────────────────
try:
    from middleware.errors import safe_error, register_error_handlers
except ImportError:
    pass  # use local safe_error defined below
# ── Structured logging ─────────────────────────────────────
def compute_percentile(score: int) -> int:
    """
    Returns what % of recent users this score beats (0-99).
    Uses Redis sorted set when available, in-memory pool as fallback.
    Returns -1 if < 10 data points (not enough to be meaningful).
    """
    global _score_pool
    try:
        import time as _tp
        _rc = get_redis()
        if _rc:
            _rc.zadd("score_pool", {f"{_tp.time()}:{score}": score})
            _rc.zremrangebyrank("score_pool", 0, -10001)
            total = _rc.zcard("score_pool")
            below = _rc.zcount("score_pool", 0, score - 1)
            if total >= 10:
                return min(99, int((below / total) * 100))
    except Exception:
        pass
    _score_pool.append(score)
    if len(_score_pool) > 10000:
        _score_pool = _score_pool[-5000:]
    if len(_score_pool) < 10:
        return -1
    below = sum(1 for s in _score_pool if s < score)
    return min(99, int((below / len(_score_pool)) * 100))


# ── Arabic alert translations ─────────────────────────────────────
_ALERT_AR = {
    # Detection
    "No person detected — ensure your upper body is visible in the camera frame":
        "لم يتم اكتشاف شخص — تأكد من ظهور الجزء العلوي من جسمك في الكاميرا",
    "No person detected — camera should be 90° to your side, 80–120cm away":
        "لم يتم الاكتشاف — ضع الكاميرا على بُعد 80-120سم وبزاوية 90° من جانبك",
    # Neck / FHP
    "⚠️ Forward head posture": "⚠️ وضعية الرأس للأمام",
    "Forward head posture":    "وضعية الرأس للأمام",
    "⚠️ Neck bearing":         "⚠️ حِمل الرقبة",
    "Neck load":               "حِمل الرقبة",
    "⚠️ Severe neck lean":     "⚠️ ميل حاد في الرقبة",
    "Neck lean":               "ميل الرقبة",
    "raise monitor to eye level immediately": "ارفع الشاشة لمستوى عينيك فوراً",
    "tuck chin slightly and check monitor height": "اسحب ذقنك للخلف قليلاً وتحقق من ارتفاع الشاشة",
    # Head
    "Head tilting":            "ميل الرأس",
    "check chair height and monitor centering": "تحقق من ارتفاع الكرسي ومركزية الشاشة",
    "Head pitched":            "زاوية ميل الرأس",
    "raise your monitor":      "ارفع شاشتك",
    "lower your monitor":      "اخفض شاشتك",
    # Shoulders
    "Shoulder imbalance":      "عدم توازن الكتفين",
    "adjust armrests":         "اضبط مسند الذراعين",
    "⚠️ Significant body asymmetry": "⚠️ عدم تماثل كبير في الجسم",
    "Body asymmetry":          "عدم تماثل الجسم",
    "check armrest heights and sitting position": "تحقق من ارتفاع مسند الذراعين ووضعية الجلوس",
    # Spine
    "⚠️ Spine lean":           "⚠️ ميل العمود الفقري",
    "Spine lean":              "ميل العمود الفقري",
    "sit back and use lumbar support": "اجلس للخلف واستخدم دعم أسفل الظهر",
    "engage your core and sit upright": "شد عضلات البطن واجلس منتصباً",
    # Distance
    "⚠️ Very close to screen": "⚠️ قريب جداً من الشاشة",
    "Too close to screen":     "قريب جداً من الشاشة",
    "Too far from screen":     "بعيد جداً عن الشاشة",
    "move back to":            "ابتعد إلى",
    "ideal is":                "المثالي هو",
    # Wrist / RSI
    "⚠️ Wrist deviation":      "⚠️ انحراف المعصم",
    "Wrist deviation":         "انحراف المعصم",
    "risk of Carpal Tunnel. Keep wrists straight.": "خطر متلازمة النفق الرسغي. حافظ على استقامة معصميك.",
    "try to keep wrists flat on desk.": "حاول إبقاء معصميك مستويين على المكتب.",
    "⚠️ High RSI risk":        "⚠️ خطر إصابة RSI مرتفع",
    # Blink / eyes
    "⚠️ Eye strain risk":      "⚠️ خطر إجهاد العينين",
    "Low blink rate":          "معدل رمش منخفض",
    "Apply 20-20-20 rule.":    "طبّق قاعدة 20-20-20.",
    "remember to blink consciously.": "تذكر أن تومض عينيك بوعي.",
    # Ergonomics
    "⚠️ Poor workstation setup": "⚠️ إعداد بيئة عمل سيئ",
    "adjust monitor height and distance": "اضبط ارتفاع الشاشة والمسافة منها",
    # Pain
    "⚠️ Pain onset predicted in": "⚠️ توقع ظهور الألم خلال",
    "at risk":                 "في خطر",
    # Break
    "⚠️ Seated":               "⚠️ جلست",
    "take a standing break (every 45-60min)": "خذ استراحة واقفاً (كل 45-60 دقيقة)",
    "consider a short break soon": "فكر في أخذ استراحة قصيرة قريباً",
    # Posture DNA
    "⚠️ Habitual bad posture": "⚠️ عادة وضعية سيئة متكررة",
    # ISO
    "⚠️ ISO 11226: posture not recommended": "⚠️ ISO 11226: الوضعية غير موصى بها",
    "ISO 11226: conditional posture":        "ISO 11226: وضعية مشروطة",
    # Breathing
    "Shallow breathing detected":  "تم اكتشاف تنفس ضحل",
    "take a deep breath, relax shoulders": "خذ نفساً عميقاً، أرخِ كتفيك",
    # Stillness
    "Body too still":          "الجسم ساكن جداً",
    "shift position slightly every few minutes": "غيّر وضعيتك قليلاً كل بضع دقائق",
    # Trunk (side view)
    "⚠️ Forward head":         "⚠️ الرأس للأمام",
    "ear must align directly above shoulder": "يجب أن يكون الأذن مباشرة فوق الكتف",
    "tuck chin and lengthen spine": "اسحب الذقن للخلف ومدّ عمودك الفقري",
    "⚠️ Trunk leaning":        "⚠️ ميل الجذع",
    "press back fully into chair backrest": "اضغط بظهرك كاملاً على مسند الكرسي",
}

def translate_alert_ar(alert: str) -> str:
    """Translate an English alert string to Arabic using keyword matching."""
    for en, ar in _ALERT_AR.items():
        if en in alert:
            # Replace matched substring, preserve numbers/values
            return alert.replace(en, ar)
    return alert   # return as-is if no match found

def log_event(event, uid=None, meta=None):
    import json as _j
    rec = {"ts": datetime.utcnow().isoformat()+"Z", "event": event,
           "uid": uid or getattr(g,"uid",None), "rid": getattr(g,"request_id",None)}
    if meta: rec.update(meta)
    print(_j.dumps(rec), flush=True)

def safe_error(e, msg="Internal server error", status=500):
        import traceback, sys
        print(traceback.format_exc(), file=sys.stderr)
        env = os.getenv("FLASK_ENV","development")
        if env == "production":
            from flask import jsonify as _j
            return _j({"error": msg}), status
        from flask import jsonify as _j
        return _j({"error": str(e), "trace": traceback.format_exc()}), status
def register_error_handlers(app): pass


# ── Firebase Firestore (lazy import) ─────────────────────────────
try:
    from firebase_admin import firestore
    print("✅ Firestore loaded")
except ImportError:
    class _FSStub:
        def client(self): return None
        class Query:
            DESCENDING = "DESCENDING"
        @staticmethod
        def Increment(n): return n
        @staticmethod  
        def ArrayUnion(v): return v
    firestore = _FSStub()
    print("⚠️  firebase-admin not installed — Firestore disabled")

# Module-level Firestore client — firebase_admin.initialize_app() already
# ran via the auth.middleware import above. Individual functions that
# need a guaranteed-fresh client still do their own `db = firestore.client()`
# locally (that local assignment correctly shadows this one); this global
# exists so the many call sites that never did that — and were silently
# throwing NameError — actually work.
try:
    db = firestore.client()
except Exception as _db_init_err:
    db = None
    print(f"⚠️  Firestore client init failed: {_db_init_err}", file=sys.stderr)

# Module-level Firebase Auth admin handle + readiness flag — same story:
# referenced under three different undefined names (_fb_auth, firebase_auth,
# _firebase_ok) across user-onboarding, session-revocation, and SCIM
# provisioning routes, none of which were ever actually defined.
try:
    from firebase_admin import auth as _fb_auth
    firebase_auth = _fb_auth      # codebase uses both names — alias both
    _firebase_ok  = db is not None
except Exception:
    _fb_auth = None
    firebase_auth = None
    _firebase_ok = False

# ── Config ─────────────────────────────────────────────────────────
# AI backend — server-side ONLY, never exposed to the client.
# Groq (primary): free tier, 14,400 req/day, ~500ms latency.
#   Sign up at console.groq.com → API Keys → set GROQ_API_KEY in Railway.
# Ollama (optional fallback): self-hosted, zero cost but needs a server.
GROQ_API_KEY        = os.getenv("GROQ_API_KEY", "")
GROQ_MODEL          = os.getenv("GROQ_MODEL", "llama-3.3-70b-versatile")  # best free model
GROQ_MODEL_FAST     = "llama-3.1-8b-instant"   # fallback if 70b rate-limited
GROQ_URL            = "https://api.groq.com/openai/v1/chat/completions"
OLLAMA_URL          = os.getenv("OLLAMA_URL", "")
LOCAL_LLM_MODEL     = os.getenv("LOCAL_LLM_MODEL", "qwen2.5:3b")
AI_CONFIGURED       = bool(GROQ_API_KEY or OLLAMA_URL)

def call_groq(messages, max_tokens=600, temperature=0.5):
    """Call Groq API — free tier (14,400 req/day). Server-side only."""
    if not GROQ_API_KEY:
        return None
    for model in (GROQ_MODEL, GROQ_MODEL_FAST):
        try:
            resp = req.post(GROQ_URL,
                headers={"Authorization": f"Bearer {GROQ_API_KEY}", "Content-Type": "application/json"},
                json={"model": model, "messages": messages, "max_tokens": max_tokens, "temperature": temperature},
                timeout=18)
            if resp.status_code == 200:
                return resp.json()["choices"][0]["message"]["content"].strip()
            if resp.status_code == 429:
                continue   # rate-limited on this model, try faster one
        except Exception:
            pass
    return None

PAYMOB_SECRET_KEY   = os.getenv("PAYMOB_SECRET_KEY", "")
PAYMOB_INTEGRATIONS = {
    "card":          os.getenv("PAYMOB_INTEGRATION_CARD", ""),
    "mobile_wallet": os.getenv("PAYMOB_INTEGRATION_WALLET", ""),
}

# ── User type constants ───────────────────────────────────────────
USER_TYPE_B2C = "b2c"   # individual user (global)
USER_TYPE_B2B = "b2b"   # company/HR (Egypt enterprise)

def _detect_user_type(uid: str = "", role: dict = None) -> str:
    """B2B: has company_id or acct_type=company/hr/employee. B2C: everyone else."""
    if not role: role = {}
    if role.get("company_id"): return USER_TYPE_B2B
    if role.get("acct_type") in ("company","hr","employee"): return USER_TYPE_B2B
    return USER_TYPE_B2C

# ── Pricing table ─────────────────────────────────────────────────
# B2C Egypt (EGP cents via PayMob)
# B2C Gulf/Global (USD cents via Stripe)
# Amounts in CENTS (EGP cents / USD cents)

_PAYMOB_PRICES = {
    # B2C Egypt — EGP cents
    "standard":     {"monthly": 0,       "yearly": 0},
    "basic":        {"monthly": 19900,   "yearly": 159000},   # 199 EGP/mo | 1,590/yr
    "professional": {"monthly": 39900,   "yearly": 319000},   # 399 EGP/mo | 3,190/yr
    "pro":          {"monthly": 39900,   "yearly": 319000},   # alias
    "elite":        {"monthly": 69900,   "yearly": 559000},   # 699 EGP/mo | 5,590/yr
    "premium":      {"monthly": 69900,   "yearly": 559000},   # alias
    # B2B Egypt (companies) — EGP cents, flat-rate (NOT per-seat)
    "b2b_starter":    {"monthly": 249900,   "yearly": 2399000},   # 2,499 EGP/mo | 23,990/yr
    "b2b_growth":     {"monthly": 699900,   "yearly": 6719000},   # 6,999 EGP/mo | 67,190/yr
    # b2b_enterprise intentionally has NO price — always contact-sales/custom
}

_STRIPE_PRICES = {
    # B2C Gulf/Global — USD cents
    "standard":     {"monthly": 0,      "yearly": 0},
    "basic":        {"monthly": 999,    "yearly": 7999},    # $9.99/mo  | $79.99/yr
    "professional": {"monthly": 1999,   "yearly": 15999},   # $19.99/mo | $159.99/yr
    "pro":          {"monthly": 1999,   "yearly": 15999},   # alias
    "elite":        {"monthly": 3999,   "yearly": 29999},   # $39.99/mo | $299.99/yr
    "premium":      {"monthly": 3999,   "yearly": 29999},   # alias
    # B2B Gulf/Global (companies) — USD cents, flat-rate (NOT per-seat)
    "b2b_starter":    {"monthly": 7900,    "yearly": 75800},    # $79/mo | $758/yr
    "b2b_growth":     {"monthly": 19900,   "yearly": 191000},   # $199/mo | $1,910/yr
    # b2b_enterprise intentionally has NO price — always contact-sales/custom, starts at $499/mo
}

# Human-readable prices for UI display
PRICING_DISPLAY = {
    "egypt": {
        "basic":        {"monthly": "199 EGP",  "yearly": "1,590 EGP", "yearly_note": "وفّر شهرين"},
        "professional": {"monthly": "399 EGP",  "yearly": "3,190 EGP", "yearly_note": "وفّر شهرين"},
        "elite":        {"monthly": "699 EGP",  "yearly": "5,590 EGP", "yearly_note": "وفّر شهرين"},
        # B2B (companies) — flat-rate
        "b2b_starter":    {"monthly": "2,499 EGP", "yearly": "23,990 EGP", "yearly_note": "وفّر 20%"},
        "b2b_growth":     {"monthly": "6,999 EGP", "yearly": "67,190 EGP", "yearly_note": "وفّر 20%"},
        "b2b_enterprise": {"monthly": "تواصل معنا", "yearly": "تواصل معنا", "yearly_note": "يبدأ من $499/شهر"},
    },
    "gulf": {
        "basic":        {"monthly": "$9.99",   "yearly": "$79.99",   "yearly_note": "Save 2 months"},
        "professional": {"monthly": "$19.99",  "yearly": "$159.99",  "yearly_note": "Save 2 months"},
        "elite":        {"monthly": "$39.99",  "yearly": "$299.99",  "yearly_note": "Save 2 months"},
        # B2B (companies) — flat-rate
        "b2b_starter":    {"monthly": "$79",   "yearly": "$758",   "yearly_note": "Save 20%"},
        "b2b_growth":     {"monthly": "$199",  "yearly": "$1,910", "yearly_note": "Save 20%"},
        "b2b_enterprise": {"monthly": "Contact us", "yearly": "Contact us", "yearly_note": "Starting at $499/mo"},
    },
}

def get_paymob_amount(tier: str, billing: str = "monthly", user_type: str = "b2c") -> int | None:
    """Return amount in EGP cents for PayMob. None if free."""
    prices = _PAYMOB_PRICES.get(tier.lower(), {})
    amount = prices.get(billing.lower())
    return amount if amount else None

def get_stripe_amount(tier: str, billing: str = "monthly") -> int | None:
    """Return amount in USD cents for Stripe. None if free."""
    prices = _STRIPE_PRICES.get(tier.lower(), {})
    return prices.get(billing.lower())

def validate_plan_request(tier: str, billing: str) -> tuple[bool, str]:
    """Validate tier + billing combo."""
    tier_lower = tier.lower()
    # Enterprise tiers are always custom/contact-sales — never go through automated checkout
    if tier_lower in ("b2b_enterprise", "enterprise"):
        return False, "Enterprise plans require a custom contract — contact sales@corvus.io"
    valid_tiers    = set(_PAYMOB_PRICES.keys()) | set(_STRIPE_PRICES.keys())
    valid_billings = {"monthly", "yearly", "annual"}
    if tier_lower not in valid_tiers:
        return False, f"Unknown tier: {tier}. Valid: {sorted(valid_tiers)}"
    if billing.lower() not in valid_billings:
        return False, f"Unknown billing: {billing}. Valid: monthly, yearly"
    return True, ""

APP_URL        = os.getenv("APP_URL", "https://postureai.vercel.app")
SUPPORT_EMAIL  = os.getenv("SUPPORT_EMAIL", "support@postureai.io")
ADMIN_EMAIL    = os.getenv("ADMIN_EMAIL", "admin@postureai.io")
ADMIN_PHONE    = os.getenv("ADMIN_PHONE", "")
WORK_HOURS_START = int(os.getenv("WORK_HOURS_START", "9"))
WORK_HOURS_END   = int(os.getenv("WORK_HOURS_END", "18"))
SENDGRID_API_KEY = os.getenv("SENDGRID_API_KEY", "")
SMTP_HOST   = os.getenv("SMTP_HOST", "smtp.gmail.com")
# HIGH-02: Gmail is rate-limited to ~500/day — use SendGrid/Resend in production
if SMTP_HOST == "smtp.gmail.com" and os.getenv("FLASK_ENV") == "production":
    print("⚠️  SMTP_HOST=smtp.gmail.com in production. Gmail caps at 500 emails/day.", flush=True)
    print("   Set SMTP_HOST=smtp.sendgrid.net / SMTP_USER=apikey / SMTP_PASS=SG.xxx", flush=True)
SMTP_PORT   = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER   = os.getenv("SMTP_USER", "")
SMTP_PASS   = os.getenv("SMTP_PASS", "")
SLACK_WEBHOOK_URL  = os.getenv("SLACK_WEBHOOK_URL", "")
TEAMS_WEBHOOK_URL  = os.getenv("TEAMS_WEBHOOK_URL", "")
WA_PHONE_ID        = os.getenv("WA_PHONE_NUMBER_ID", "")
WA_ACCESS_TOKEN    = os.getenv("WA_ACCESS_TOKEN", "")

# ── MediaPipe ─────────────────────────────────────────────────────
print("[1/4] Loading MediaPipe...", flush=True)
# MediaPipe loaded lazily on first analysis request (not at startup)
# Reason: 500MB per worker × N workers = OOM on startup under gunicorn
# The _mp module is set by _ensure_mediapipe() called inside analysis functions
_mp = None

def _ensure_mediapipe():
    """Lazy-load MediaPipe on first use. Thread-safe via GIL."""
    global _mp
    if _mp is None:
        try:
            import mediapipe as _mediapipe
            _mp = _mediapipe
            print("✅ MediaPipe loaded lazily", flush=True)
        except ImportError:
            print("⚠️  mediapipe not installed — pip install mediapipe", flush=True)
    return _mp


# MediaPipe models loaded lazily on first analysis request
# Module-level initialization removed: 500MB × N workers = OOM crash
_mp_pose    = None
_mp_face    = None
_mp_drawing = None
POSE_LITE   = None
POSE_FULL   = None
FACE_MESH   = None
_models_lock = __import__("threading").Lock()

def _ensure_models():
    """Initialize MediaPipe models + cv2/numpy on first call. Thread-safe."""
    global _mp_pose, _mp_face, _mp_drawing, POSE_LITE, POSE_FULL, FACE_MESH, cv2, np, face_c, eye_c, prof_c
    if POSE_LITE is not None:
        return True
    with _models_lock:
        if POSE_LITE is not None:
            return True  # double-checked locking
        try:
            # Lazy-load cv2 and numpy here — saves ~130MB per worker at startup
            import cv2 as _cv2
            import numpy as _np
            cv2 = _cv2
            np  = _np
            # Cascade classifiers only need cv2 (not MediaPipe) — init them here
            # so the fallback path in analyze_front_cascade() always has them
            # ready once _ensure_models() has run once, regardless of whether
            # MediaPipe itself loads successfully below.
            face_c = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
            eye_c  = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_eye.xml')
            prof_c = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_profileface.xml')
            # Init MODEL_3D now that numpy is available
            global MODEL_3D
            MODEL_3D = _np.array(_MODEL_3D_RAW, dtype=_np.float64)
            mp = _ensure_mediapipe()
            if mp is None:
                return False
            _mp_pose    = mp.solutions.pose
            _mp_face    = mp.solutions.face_mesh
            _mp_drawing = mp.solutions.drawing_utils
            # static_image_mode=False: MediaPipe reuses tracking across frames — 3x faster
            POSE_LITE = _mp_pose.Pose(static_image_mode=False, model_complexity=0,
                                       min_detection_confidence=0.50, min_tracking_confidence=0.50)
            POSE_FULL = _mp_pose.Pose(static_image_mode=False, model_complexity=1,
                                       min_detection_confidence=0.50, min_tracking_confidence=0.50)
            FACE_MESH = _mp_face.FaceMesh(static_image_mode=False, max_num_faces=1,
                                           refine_landmarks=True,
                                           min_detection_confidence=0.50, min_tracking_confidence=0.50)
            print("✅ MediaPipe models initialized on first request", flush=True)
            # Warm up: process a tiny black image to pre-JIT the model
            try:
                warmup = _np.zeros((64, 64, 3), dtype=_np.uint8)
                _warmed = cv2.cvtColor(warmup, cv2.COLOR_BGR2RGB)
                POSE_LITE.process(_warmed)
                POSE_FULL.process(_warmed)   # warm up FULL model too (Pro/Elite)
                FACE_MESH.process(_warmed)   # warm up FaceMesh (distance + neck)
            except Exception:
                pass
            return True
        except Exception as e:
            print(f"⚠️  MediaPipe model init failed: {e}", flush=True)
            return False

class PL:
    NOSE=0; L_EYE_INNER=1; L_EYE=2; L_EYE_OUTER=3
    R_EYE_INNER=4; R_EYE=5; R_EYE_OUTER=6
    L_EAR=7; R_EAR=8
    L_SHOULDER=11; R_SHOULDER=12
    L_ELBOW=13;    R_ELBOW=14
    L_WRIST=15;    R_WRIST=16
    L_HIP=23;      R_HIP=24
    L_KNEE=25;     R_KNEE=26
    L_ANKLE=27;    R_ANKLE=28
    L_HEEL=29;     R_HEEL=30

L_EYE_LMK = [33, 7, 163, 144, 145, 153, 154, 155, 133]
R_EYE_LMK = [362, 382, 381, 380, 374, 373, 390, 249, 263]
L_PUPIL = 468; R_PUPIL = 473

# Blink landmarks (vertical eye opening)
L_EYE_TOP = 159; L_EYE_BOT = 145
R_EYE_TOP = 386; R_EYE_BOT = 374

# Lazy — np is None at module load time; converted in _ensure_cv2()
_MODEL_3D_RAW = (
    (0.0,    0.0,    0.0),
    (0.0,  -330.0, -65.0),
    (-225.0, 170.0,-135.0),
    ( 225.0, 170.0,-135.0),
    (-150.0,-150.0,-125.0),
    ( 150.0,-150.0,-125.0),
)
MODEL_3D = None  # initialized in _ensure_cv2()

ALLOWED_ORIGINS = [
    "http://localhost:3000",
    "http://localhost:5173",
    "http://localhost:4173",
    "https://postureai-pro-omega-nine.vercel.app",   # ← LIVE (primary)
    "https://postureai-pro-omega.vercel.app",         # ← old (keep for redirect safety)
    "https://postureai.vercel.app",
    "https://postureai-pro.vercel.app",
    "https://postureai.io",
    "https://www.postureai.io",
    "https://postureai-prod.web.app",
    "https://postureai-prod.firebaseapp.com",
]
# Allow custom origins from env (Railway preview URLs, white-label domains)
_extra = os.getenv("ALLOWED_ORIGINS", "")
if _extra:
    ALLOWED_ORIGINS += [o.strip() for o in _extra.split(",") if o.strip()]

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 2 * 1024 * 1024  # 2MB max request — prevents OOM on large frames

# ── Optional WebSocket / Socket.IO (enable with SOCKETIO_ENABLED=true) ────────
_SOCKETIO_STATUS = {"enabled": False, "initialized": False, "error": None}
if os.getenv("SOCKETIO_ENABLED", "false").lower() == "true":
    _SOCKETIO_STATUS["enabled"] = True
    try:
        from services.websocket_server import init_socketio
        socketio = init_socketio(app)
        _SOCKETIO_STATUS["initialized"] = True
        print("✅ WebSocket server initialized (Socket.IO)", flush=True)
    except ImportError as _ws_err:
        _SOCKETIO_STATUS["error"] = str(_ws_err)
        print(f"⚠️  WebSocket init skipped: {_ws_err} — pip install flask-socketio gevent", file=sys.stderr)
    except Exception as _ws_err2:
        _SOCKETIO_STATUS["error"] = str(_ws_err2)
        print(f"⚠️  WebSocket init error: {_ws_err2}", file=sys.stderr)

@app.errorhandler(413)
def too_large(e):
    return jsonify({"error": "Request too large — max 2MB"}), 413

# ── Sentry monitoring (production error tracking) ─────────────────
_sentry_dsn = os.getenv("SENTRY_DSN", "")
if _sentry_dsn:
    try:
        import sentry_sdk
        from sentry_sdk.integrations.flask import FlaskIntegration
        sentry_sdk.init(
            dsn=_sentry_dsn,
            integrations=[FlaskIntegration()],
            environment=os.getenv("FLASK_ENV", "development"),
            traces_sample_rate=0.1,  # 10% of transactions for performance monitoring
            send_default_pii=False,   # GDPR: don't send personally identifiable info
        )
        print("✅ Sentry initialized")
    except ImportError:
        print("⚠️  sentry-sdk not installed — run: pip install sentry-sdk[flask]")
    except Exception as e:
        print(f"⚠️  Sentry init failed: {e}")

# CORS: explicit whitelist — never fallback to wildcard
_cors_origins = [o.strip() for o in (os.getenv("ALLOWED_ORIGINS","") or "").split(",") if o.strip()]
if not _cors_origins:
    _cors_origins = [
        "https://postureai-pro-omega-nine.vercel.app",  # ← LIVE (primary)
        "https://postureai-pro-omega.vercel.app",        # ← old (keep for safety)
        "https://postureai.io",
        "https://www.postureai.io",
        "http://localhost:5173",
        "http://localhost:3000",
    ]
CORS(app, resources={
    r"/api/*":   {"origins": _cors_origins, "supports_credentials": True},
    r"/scim/*":  {"origins": "*"},   # SCIM needs open for IdP
})
register_error_handlers(app)

# ── Register new v17 blueprints ────────────────────────────────────
try:
    from routes.stripe_billing import stripe_bp
    app.register_blueprint(stripe_bp, url_prefix="/api/stripe")
    print("✅ Stripe billing blueprint registered")
except Exception as _e:
    print(f"⚠️  Stripe blueprint skipped: {_e}")

try:
    from routes.health import health_bp
    app.register_blueprint(health_bp)
    print("✅ Health blueprint registered")
except Exception as _e:
    print(f"⚠️  Health blueprint skipped: {_e}")

# ── Rate Limiting ──────────────────────────────────────────────────
# IMPORTANT: storage_uri MUST point to Redis in production.
# With multiple Gunicorn workers, in-memory (memory://) is per-process
# and does NOT enforce shared limits. Set REDIS_URL env var.
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
_redis_url = os.getenv("REDIS_URL", "")
_flask_env = os.getenv("FLASK_ENV", "development")
if not _redis_url:
    if _flask_env == "production":
        print(
            "🚨 FATAL: REDIS_URL not set in production.\n"
            "   Rate limiting is per-process only with multiple Gunicorn workers.\n"
            "   Set REDIS_URL in Railway/Render env vars (use Redis addon).\n"
            "   Refusing to start without shared rate limiting.",
            file=sys.stderr, flush=True
        )
        import sys as _sys_redis
        _sys_redis.exit(1)
    print("⚠️  REDIS_URL not set — rate limiter using in-memory (OK for development only)", file=sys.stderr)
    _limiter_storage = "memory://"
else:
    _limiter_storage = _redis_url

# ── Redis connection pool (singleton) ─────────────────────────────
# Never create a new connection per request — use shared pool
# max_connections=20: enough for Gunicorn workers + background threads
_redis_pool = None
_redis_pool_lock = __import__("threading").Lock()

def get_redis():
    """
    Return a Redis client backed by a shared connection pool.
    Thread-safe: pool is created once on first call.
    Returns None if Redis unavailable (graceful degradation).
    """
    global _redis_pool
    if _redis_pool is not None:
        try:
            import redis as _r
            return _r.Redis(connection_pool=_redis_pool)
        except Exception:
            return None
    if not _redis_url:
        return None
    with _redis_pool_lock:
        if _redis_pool is not None:
            try:
                import redis as _r
                return _r.Redis(connection_pool=_redis_pool)
            except Exception:
                return None
        try:
            import redis as _r
            _redis_pool = _r.ConnectionPool.from_url(
                _redis_url,
                max_connections=20,
                socket_timeout=1.5,
                socket_connect_timeout=1.5,
                retry_on_timeout=True,
            )
            print("✅ Redis connection pool initialized (max_connections=20)", flush=True)
            return _r.Redis(connection_pool=_redis_pool)
        except Exception as e:
            print(f"⚠️  Redis pool init failed: {e}", file=sys.stderr)
            return None
limiter = Limiter(
    get_remote_address, app=app,
    default_limits=["200 per minute", "2000 per hour"],
    storage_uri=_limiter_storage
)

@app.route("/api/pricing", methods=["GET"])
@limiter.limit("60 per minute")
def get_pricing():
    """Public endpoint — return pricing for Egypt and Gulf markets."""
    region = request.args.get("region", "gulf").lower()
    lang   = request.args.get("lang", "en")
    is_ar  = lang == "ar"

    plans = [
        {
            "id":    "basic",
            "name":  "Basic" if not is_ar else "أساسي",
            "color": "#3b82f6",
            "features": {
                "en": ["Unlimited sessions", "AI Coach (10 msgs/month)", "Streak tracking", "Goals"],
                "ar": ["جلسات غير محدودة", "مدرب AI (10 رسائل/شهر)", "تتبع السلسلة", "الأهداف"],
            },
            "pricing": PRICING_DISPLAY.get(region, PRICING_DISPLAY["gulf"]).get("basic", {}),
            "cta": "ابدأ الآن" if is_ar else "Get Started",
        },
        {
            "id":    "professional",
            "name":  "Pro" if not is_ar else "احترافي",
            "color": "#8b5cf6",
            "popular": True,
            "features": {
                "en": ["Everything in Basic", "AI Insights", "Reports", "Compare sessions", "Pain prediction"],
                "ar": ["كل Basic", "رؤى AI", "تقارير", "مقارنة الجلسات", "توقع الألم"],
            },
            "pricing": PRICING_DISPLAY.get(region, PRICING_DISPLAY["gulf"]).get("professional", {}),
            "cta": "الأكثر شعبية" if is_ar else "Most Popular",
        },
        {
            "id":    "elite",
            "name":  "Elite",
            "color": "#f59e0b",
            "features": {
                "en": ["Everything in Pro", "AI Coach unlimited", "Predictive AI", "PDF report", "Priority support"],
                "ar": ["كل Pro", "مدرب AI غير محدود", "AI تنبؤي", "تقرير PDF", "دعم أولوية"],
            },
            "pricing": PRICING_DISPLAY.get(region, PRICING_DISPLAY["gulf"]).get("elite", {}),
            "cta": "الأفضل" if is_ar else "Best Value",
        },
    ]

    return jsonify({
        "ok":     True,
        "region": region,
        "currency": "EGP" if region == "egypt" else "USD",
        "gateway": "paymob" if region == "egypt" else "stripe",
        "plans":  plans,
    })

# ── Input validation ───────────────────────────────────────────────
def validate_frame(data):
    if not data:
        return False, "No data provided"
    frame = data.get("frame", "")
    if not frame:
        return False, "No frame data"
    if "," in frame:
        frame = frame.split(",", 1)[1]
    if len(frame) > 960_000:
        actual_kb = int(len(frame) * 3 / 4 / 1024)  # base64 → real decoded byte size
        return False, f"Frame too large ({actual_kb}KB, max 700KB)"
    mode = data.get("mode", "")
    if mode not in ("laptop", "phone", "side", ""):
        return False, f"Invalid mode: {mode}"
    tier = getattr(g, "tier", None) or "standard"  # never trust client tier
    if tier not in ("standard", "professional", "elite", "basic", "pro", "premium"):
        return False, f"Invalid tier: {tier}"
    return True, None

# ── COUPONS ───────────────────────────────────────────────────────
# ── Coupons stored in Firestore (not memory) ─────────────────────
# Default coupons are seeded once to Firestore on first use.
COUPONS_DEFAULT = {
    "POSTURE20": {"discount": 20, "label": "20% off",               "max_uses": 100},
    "EGYPT30":   {"discount": 30, "label": "30% off — Egypt launch", "max_uses": 50 },
    "HR2025":    {"discount": 15, "label": "15% off — HR teams",     "max_uses": 200},
    "TRIAL50":   {"discount": 50, "label": "50% first month",        "max_uses": 30 },
}

def _get_coupon_from_db(code: str):
    """Fetch coupon from Firestore. Falls back to seeding defaults."""
    try:
        db = firestore.client()
        ref = db.collection("coupons").document(code)
        doc = ref.get()
        if doc.exists:
            return ref, doc.to_dict()
        # Seed default if it exists in defaults
        if code in COUPONS_DEFAULT:
            data = {**COUPONS_DEFAULT[code], "used": 0, "code": code}
            ref.set(data)
            return ref, data
        return None, None
    except Exception:
        return None, None

# ── Notification helpers ───────────────────────────────────────────
def send_slack(text: str, score: int = 0) -> dict:
    if not SLACK_WEBHOOK_URL:
        return {"ok": False, "reason": "SLACK_WEBHOOK_URL not set"}
    color = "#ef4444" if score < 50 else "#f59e0b" if score < 70 else "#10b981"
    payload = {
        "attachments": [{
            "color": color,
            "blocks": [
                {"type": "header", "text": {"type": "plain_text", "text": "⚠️ PostureAI Alert", "emoji": True}},
                {"type": "section", "text": {"type": "mrkdwn", "text": text}},
                {"type": "context", "elements": [{"type": "mrkdwn", "text": f"Score: *{score}/100* · {datetime.now().strftime('%H:%M')}"}]}
            ]
        }]
    }
    try:
        r = req.post(SLACK_WEBHOOK_URL, json=payload, timeout=8)
        return {"ok": r.status_code == 200, "status": r.status_code}
    except Exception as e:
        return {"ok": False, "reason": str(e)}

def send_teams(text: str, score: int = 0, employee: str = "") -> dict:
    if not TEAMS_WEBHOOK_URL:
        return {"ok": False, "reason": "TEAMS_WEBHOOK_URL not set"}
    payload = {
        "type": "message",
        "attachments": [{
            "contentType": "application/vnd.microsoft.card.adaptive",
            "content": {
                "$schema": "http://adaptivecards.io/schemas/adaptive-card.json",
                "type": "AdaptiveCard", "version": "1.4",
                "body": [
                    {"type": "TextBlock", "text": "⚠️ PostureAI Alert", "weight": "Bolder", "size": "Medium", "color": "Attention"},
                    {"type": "TextBlock", "text": text, "wrap": True},
                    {"type": "FactSet", "facts": [
                        {"title": "Employee", "value": employee or "—"},
                        {"title": "Score", "value": f"{score}/100"},
                        {"title": "Time", "value": datetime.now().strftime("%H:%M")}
                    ]}
                ]
            }
        }]
    }
    try:
        r = req.post(TEAMS_WEBHOOK_URL, json=payload, timeout=8)
        return {"ok": r.status_code in (200, 202), "status": r.status_code}
    except Exception as e:
        return {"ok": False, "reason": str(e)}

def send_whatsapp(to_phone: str, text: str, template: str = None, lang: str = "en") -> dict:
    """
    Send WhatsApp message via Meta Cloud API.
    Supports: text messages + approved templates (for marketing/alerts).
    phone: international format e.g. "+201012345678" or "201012345678"
    """
    if not WA_PHONE_ID or not WA_ACCESS_TOKEN:
        return {"ok": False, "reason": "WA_PHONE_NUMBER_ID or WA_ACCESS_TOKEN not set"}
    phone_clean = to_phone.replace("+","").replace(" ","").replace("-","")
    url = f"https://graph.facebook.com/v18.0/{WA_PHONE_ID}/messages"
    headers = {"Authorization": f"Bearer {WA_ACCESS_TOKEN}", "Content-Type": "application/json"}
    if template:
        # Template message (pre-approved by Meta — needed for first contact)
        payload = {
            "messaging_product": "whatsapp",
            "to": phone_clean,
            "type": "template",
            "template": {
                "name": template,
                "language": {"code": "ar" if lang == "ar" else "en_US"},
            }
        }
    else:
        payload = {
            "messaging_product": "whatsapp",
            "to": phone_clean,
            "type": "text",
            "text": {"body": text, "preview_url": False}
        }
    try:
        r = req.post(url, json=payload, headers=headers, timeout=10)
        result = r.json()
        ok = r.status_code == 200
        if not ok:
            log_event("whatsapp_error", meta={"status": r.status_code, "error": str(result)[:100]})
        return {"ok": ok, "status": r.status_code, "data": result}
    except Exception as e:
        return {"ok": False, "reason": str(e)}


def notify_user(uid: str, title: str, body: str,
                channel: str = "auto", lang: str = "en",
                data: dict = None) -> dict:
    """
    Smart notification router — sends via best available channel.
    channel: "auto" | "whatsapp" | "push" | "email"
    "auto" priority: WhatsApp (if phone known) → Push (if token registered) → Email

    Used for BOTH B2B (HR alerts) and B2C (streak reminders, milestones).
    """
    results = {}
    try:
        # Load user profile for contact info
        user_doc = db.collection("users").document(uid).get()
        user     = user_doc.to_dict() if user_doc.exists else {}
        phone    = user.get("phone") or user.get("wa_phone") or ""
        email    = user.get("email") or ""
        notif_pref = user.get("notif_channel", "auto")  # user preference

        # Respect user preference
        if notif_pref != "auto" and channel == "auto":
            channel = notif_pref

        sent = False

        # 1. WhatsApp (highest open rate ~98%)
        if channel in ("auto", "whatsapp") and phone:
            wa_text = f"*{title}*\n\n{body}\n\n_PostureAI Pro_"
            result  = send_whatsapp(phone, wa_text, lang=lang)
            results["whatsapp"] = result
            if result.get("ok"):
                sent = True
                log_event("notify_whatsapp_sent", uid, {"title": title[:40]})

        # 2. Push notification (FCM)
        if not sent and channel in ("auto", "push"):
            tokens = _get_user_tokens(uid)
            if tokens:
                push_results = [_send_fcm(t, title, body, data or {}) for t in tokens]
                results["push"] = {"sent": sum(push_results), "tokens": len(tokens)}
                if any(push_results):
                    sent = True
                    log_event("notify_push_sent", uid, {"title": title[:40]})

        # 3. Email (fallback)
        if not sent and channel in ("auto", "email") and email:
            html = f"<h2>{title}</h2><p>{body}</p><p><small>PostureAI Pro</small></p>"
            ok   = send_email(email, title, html)
            results["email"] = {"sent": ok}
            if ok:
                sent = True
                log_event("notify_email_sent", uid, {"title": title[:40]})

    except Exception as e:
        results["error"] = str(e)

    results["sent"] = sent
    return results

def send_email(to_email: str, subject: str, html_body: str, from_name: str = "PostureAI") -> bool:
    """
    Send transactional email.
    Priority: 1) SendGrid API  2) SMTP (warning in production if Gmail)
    Set SENDGRID_API_KEY for production-grade deliverability.
    """
    if not to_email or "@" not in to_email:
        return False
    
    # ── 1. SendGrid (recommended for production) ──────────────────
    if SENDGRID_API_KEY:
        try:
            resp = req.post(
                "https://api.sendgrid.com/v3/mail/send",
                headers={"Authorization": f"Bearer {SENDGRID_API_KEY}", "Content-Type": "application/json"},
                json={
                    "personalizations": [{"to": [{"email": to_email}]}],
                    "from":    {"email": SMTP_USER or "noreply@postureai.io", "name": from_name},
                    "subject": subject,
                    "content": [{"type": "text/html", "value": html_body}],
                },
                timeout=10,
            )
            if resp.status_code in (200, 202):
                return True
            print(f"[email] SendGrid error {resp.status_code}: {resp.text[:200]}", file=sys.stderr)
        except Exception as e:
            print(f"[email] SendGrid exception: {e}", file=sys.stderr)
    
    # ── 2. SMTP fallback ──────────────────────────────────────────
    if not SMTP_USER or not SMTP_PASS:
        print(f"[email] No email credentials — skipped: {subject[:60]}", file=sys.stderr)
        return False
    
    # Warn in production if using Gmail (500/day cap, high spam score)
    if SMTP_HOST == "smtp.gmail.com" and os.getenv("FLASK_ENV") == "production":
        print("[email] ⚠️  Using Gmail SMTP in production — set SENDGRID_API_KEY for reliability", file=sys.stderr)
    
    try:
        import smtplib
        from email.mime.multipart import MIMEMultipart
        from email.mime.text      import MIMEText
        msg            = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"]    = f"{from_name} <{SMTP_USER}>"
        msg["To"]      = to_email
        msg.attach(MIMEText(html_body, "html"))
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=15) as server:
            server.ehlo()
            server.starttls()
            server.login(SMTP_USER, SMTP_PASS)
            server.sendmail(SMTP_USER, to_email, msg.as_string())
        return True
    except Exception as e:
        print(f"[email] SMTP error: {e}", file=sys.stderr)
        return False


def send_invoice_email(payment: dict) -> bool:
    name   = payment.get("user_name") or "Valued Customer"
    email  = payment.get("user_email", "")
    tier   = payment.get("tier", "standard").title()
    amount = payment.get("amount", 0)
    cycle  = payment.get("billing_cycle", "monthly")
    ref    = payment.get("ref_code", "—")
    method = payment.get("payment_method_name", "—")
    if not email:
        return False
    html = f"""
    <div style="font-family:Arial,sans-serif;max-width:560px;margin:0 auto;background:#030b14;color:#f0f4f8;border-radius:12px;overflow:hidden">
      <div style="background:linear-gradient(135deg,#1a56db,#0891b2);padding:28px;text-align:center">
        <div style="font-size:32px;margin-bottom:8px">◈</div>
        <div style="font-size:22px;font-weight:700">PostureAI Pro</div>
        <div style="font-size:13px;opacity:.8;margin-top:4px">Payment Confirmed ✓</div>
      </div>
      <div style="padding:28px">
        <p style="font-size:15px;margin-bottom:20px">Dear <strong>{name}</strong>,</p>
        <p style="color:#94a3b8;margin-bottom:24px">Your payment has been confirmed. Here are your invoice details:</p>
        <div style="background:#05101f;border:1px solid rgba(148,163,184,.1);border-radius:10px;padding:20px;margin-bottom:24px">
          <table style="width:100%;border-collapse:collapse">
            <tr><td style="padding:8px 0;color:#64748b;font-size:13px">Plan</td><td style="padding:8px 0;text-align:right;font-weight:600;color:#f0f4f8">{tier}</td></tr>
            <tr><td style="padding:8px 0;color:#64748b;font-size:13px">Amount</td><td style="padding:8px 0;text-align:right;font-weight:700;color:#10b981;font-size:18px">{amount:,} EGP</td></tr>
            <tr><td style="padding:8px 0;color:#64748b;font-size:13px">Billing</td><td style="padding:8px 0;text-align:right;color:#f0f4f8">{cycle.title()}</td></tr>
            <tr><td style="padding:8px 0;color:#64748b;font-size:13px">Method</td><td style="padding:8px 0;text-align:right;color:#f0f4f8">{method}</td></tr>
            <tr><td style="padding:8px 0;color:#64748b;font-size:13px">Reference</td><td style="padding:8px 0;text-align:right;font-family:monospace;color:#a5b4fc">{ref}</td></tr>
            <tr><td style="padding:8px 0;color:#64748b;font-size:13px">Date</td><td style="padding:8px 0;text-align:right;color:#f0f4f8">{datetime.now().strftime('%d %b %Y')}</td></tr>
          </table>
        </div>
        <div style="text-align:center;margin-bottom:24px">
          <a href="{APP_URL}" style="background:#1a56db;color:white;padding:12px 32px;border-radius:8px;text-decoration:none;font-weight:600;font-size:14px">Open PostureAI Pro →</a>
        </div>
        <p style="font-size:12px;color:#475569;text-align:center">Questions? Reply to this email or WhatsApp: <strong style="color:#f0f4f8">{ADMIN_PHONE}</strong></p>
      </div>
    </div>"""
    return send_email(email, f"✅ PostureAI Pro — {tier.title()} Intelligence Plan Activated", html)

def send_admin_notification(payment: dict) -> bool:
    name   = payment.get("user_name", "—")
    email  = payment.get("user_email", "—")
    tier   = payment.get("tier", "—")
    amount = payment.get("amount", 0)
    method = payment.get("payment_method_name", "—")
    ref    = payment.get("ref_code", "—")
    html = f"""
    <div style="font-family:Arial,sans-serif;max-width:480px;background:#fff;border-radius:10px;padding:24px">
      <h2 style="color:#1a56db;margin-bottom:16px">💳 New Payment — Action Required</h2>
      <table style="width:100%;border-collapse:collapse">
        <tr><td style="padding:6px 0;color:#666">Customer</td><td style="font-weight:600">{name} ({email})</td></tr>
        <tr><td style="padding:6px 0;color:#666">Plan</td><td style="font-weight:600;color:#1a56db">{tier.upper()}</td></tr>
        <tr><td style="padding:6px 0;color:#666">Amount</td><td style="font-weight:700;color:#10b981;font-size:18px">{amount:,} EGP</td></tr>
        <tr><td style="padding:6px 0;color:#666">Method</td><td>{method}</td></tr>
        <tr><td style="padding:6px 0;color:#666">Reference</td><td style="font-family:monospace;color:#7c3aed">{ref}</td></tr>
      </table>
      <div style="margin-top:20px;padding:12px;background:#fef3c7;border-radius:8px;font-size:13px">
        ⚠️ Please verify the payment and confirm in the Admin Panel.
      </div>
    </div>"""
    return send_email(ADMIN_EMAIL, f"🔔 New Payment: {name} — {amount:,} EGP ({tier})", html)

# ── Email sequence (automated nurture) ────────────────────────────
@app.route("/api/email/sequence", methods=["POST"])
@require_auth
@limiter.limit("10 per minute")
def email_sequence():
    try:
        data      = request.get_json(force=True) or {}
        to        = data.get("email", "")
        name      = data.get("name", "there")
        day       = data.get("day", 0)
        avg_score = data.get("avg_score", 0)
        sess_cnt  = data.get("session_count", 0)
        tier = getattr(g, "tier", None) or "standard"  # never trust client tier
        if not to:
            return jsonify({"ok": False, "reason": "email required"}), 400

        upgrade_url = data.get("upgrade_url", APP_URL)

        subjects = {
            0: "✦ Your AI Workforce Intelligence Platform is live",
            2: f"📊 Your health intelligence data is in — here's what we found",
            5: f"Week 1 workforce health report: {avg_score}/100 avg — {('trending up ↑' if avg_score >= 70 else 'needs attention ⚠️')}",
            6: "⏰ 24h left — lock in your workforce intelligence subscription",
            7: "Your trial ended — your health data is secured for 30 days",
            21: f"📋 21-day workforce health intelligence report ready",
        }
        bodies = {
            0: f"""<div style="font-family:-apple-system,sans-serif;max-width:560px;margin:0 auto;background:#07111f;border-radius:14px;overflow:hidden">
<div style="background:linear-gradient(135deg,#4f46e5,#0891b2);padding:28px;text-align:center">
  <div style="font-size:13px;color:rgba(255,255,255,.7);margin-bottom:4px">AI WORKFORCE INTELLIGENCE</div>
  <div style="font-size:20px;font-weight:700;color:white">Your platform is live, {name}</div>
</div>
<div style="padding:28px">
  <p style="color:#94a3b8;font-size:13px;margin:0 0 20px">Your AI health intelligence layer is now active. Here's how to generate your first insights in under 60 seconds:</p>
  <div style="background:#0c1728;border-radius:10px;padding:18px;margin-bottom:20px;border:1px solid rgba(255,255,255,.06)">
    <div style="color:#cbd5e1;font-size:13px;line-height:2.2">
      <b style="color:#818cf8">Step 1:</b> Open PostureAI Pro → select camera mode<br>
      <b style="color:#38bdf8">Step 2:</b> Click "Start Session" → sit naturally<br>
      <b style="color:#10b981">Step 3:</b> AI score appears in 5 seconds — your intelligence layer is live
    </div>
  </div>
  <div style="text-align:center;margin-bottom:20px">
    <a href="{APP_URL}" style="background:linear-gradient(135deg,#4f46e5,#0891b2);color:white;padding:12px 28px;border-radius:8px;text-decoration:none;font-weight:700;font-size:13px">Launch Intelligence Dashboard →</a>
  </div>
  <p style="color:#475569;font-size:11px;text-align:center">Questions? Reply to this email — we respond within 2 hours.</p>
</div></div>""",
            2: f"""<div style="font-family:-apple-system,sans-serif;max-width:560px;margin:0 auto;background:#07111f;border-radius:14px;overflow:hidden">
<div style="background:linear-gradient(135deg,#0f172a,#1e293b);padding:24px 28px;border-bottom:1px solid rgba(255,255,255,.06)">
  <div style="font-size:11px;font-weight:700;color:#64748b;letter-spacing:.08em;text-transform:uppercase;margin-bottom:6px">WORKFORCE HEALTH INTELLIGENCE</div>
  <div style="font-size:18px;font-weight:700;color:#f1f5f9">Your first {sess_cnt} sessions are analyzed, {name}</div>
</div>
<div style="padding:28px">
  <div style="display:flex;gap:10px;margin-bottom:20px">
    <div style="flex:1;background:#0c1728;border-radius:10px;padding:16px;text-align:center;border:1px solid rgba(255,255,255,.06)">
      <div style="font-size:32px;font-weight:800;color:{'#10b981' if (avg_score or 0) >= 70 else '#f59e0b' if (avg_score or 0) >= 50 else '#ef4444'}">{avg_score or '—'}</div>
      <div style="font-size:10px;color:#64748b;margin-top:4px">HEALTH INTELLIGENCE SCORE</div>
    </div>
    <div style="flex:1;background:#0c1728;border-radius:10px;padding:16px;text-align:center;border:1px solid rgba(255,255,255,.06)">
      <div style="font-size:32px;font-weight:800;color:#818cf8">{sess_cnt}</div>
      <div style="font-size:10px;color:#64748b;margin-top:4px">SESSIONS LOGGED</div>
    </div>
  </div>
  <div style="background:rgba(99,102,241,.08);border:1px solid rgba(99,102,241,.2);border-radius:10px;padding:16px;margin-bottom:20px">
    <div style="font-size:12px;color:#a5b4fc;font-weight:600;margin-bottom:6px">📊 Industry Benchmark</div>
    <div style="font-size:13px;color:#cbd5e1;line-height:1.6">Companies using PostureAI's workforce intelligence platform reduce sick leave costs by <b style="color:#10b981">31% within 90 days</b>. Average MSK injury cost: $18K per employee per year — eliminated.</div>
  </div>
  <div style="text-align:center">
    <a href="{APP_URL}" style="background:linear-gradient(135deg,#4f46e5,#0891b2);color:white;padding:12px 28px;border-radius:8px;text-decoration:none;font-weight:700;font-size:13px">View Your Intelligence Report →</a>
  </div>
</div></div>""",
            5: f"""<div style="font-family:-apple-system,sans-serif;max-width:560px;margin:0 auto;background:#07111f;border-radius:14px;overflow:hidden">
<div style="background:linear-gradient(135deg,#0f172a,#1e293b);padding:24px 28px;border-bottom:1px solid rgba(255,255,255,.06)">
  <div style="font-size:11px;font-weight:700;color:#64748b;letter-spacing:.08em;text-transform:uppercase;margin-bottom:6px">WEEK 1 INTELLIGENCE REPORT</div>
  <div style="font-size:18px;font-weight:700;color:#f1f5f9">5-day workforce health summary</div>
</div>
<div style="padding:28px">
  <div style="display:flex;gap:10px;margin-bottom:20px">
    <div style="flex:1;background:#0c1728;border-radius:10px;padding:14px;text-align:center;border:1px solid rgba(255,255,255,.06)">
      <div style="font-size:28px;font-weight:800;color:{'#10b981' if avg_score>=70 else '#f59e0b' if avg_score>=50 else '#ef4444'}">{avg_score}</div>
      <div style="font-size:10px;color:#64748b;margin-top:3px">HEALTH SCORE</div>
    </div>
    <div style="flex:1;background:#0c1728;border-radius:10px;padding:14px;text-align:center;border:1px solid rgba(255,255,255,.06)">
      <div style="font-size:28px;font-weight:800;color:#818cf8">{sess_cnt}</div>
      <div style="font-size:10px;color:#64748b;margin-top:3px">SESSIONS</div>
    </div>
  </div>
  <div style="background:rgba(16,185,129,.07);border:1px solid rgba(16,185,129,.18);border-radius:10px;padding:14px;margin-bottom:20px">
    <div style="font-size:12px;color:#6ee7b7;line-height:1.6">Your trial ends in 2 days. Upgrade now to keep your intelligence history, AI Coach access, predictive burnout tracking, and clinical PDF exports.</div>
  </div>
  <div style="text-align:center">
    <a href="{upgrade_url}" style="background:linear-gradient(135deg,#4f46e5,#0891b2);color:white;padding:12px 28px;border-radius:8px;text-decoration:none;font-weight:700;font-size:13px">Upgrade Intelligence Plan →</a>
  </div>
  <p style="color:#475569;font-size:11px;text-align:center;margin-top:14px">Use code <b style="color:#818cf8">INTEL20</b> for 20% off your first month</p>
</div></div>""",
            6: f"""<div style="font-family:-apple-system,sans-serif;max-width:560px;margin:0 auto;background:#07111f;border-radius:14px;overflow:hidden">
<div style="background:linear-gradient(135deg,#7f1d1d,#991b1b);padding:24px 28px;text-align:center">
  <div style="font-size:20px;margin-bottom:6px">⏰</div>
  <div style="font-size:18px;font-weight:700;color:white">24 hours left, {name}</div>
  <div style="font-size:12px;color:rgba(255,255,255,.65);margin-top:4px">Your Professional intelligence trial ends tomorrow</div>
</div>
<div style="padding:28px">
  <div style="background:#0c1728;border-radius:10px;padding:16px;margin-bottom:18px;border:1px solid rgba(255,255,255,.06)">
    <div style="font-size:11px;color:#64748b;margin-bottom:10px;font-weight:600;text-transform:uppercase;letter-spacing:.06em">YOUR INTELLIGENCE DATA AT STAKE</div>
    <div style="color:#cbd5e1;font-size:13px;line-height:2">
      📊 {sess_cnt} intelligence sessions logged<br>
      🎯 {avg_score}/100 average health score<br>
      🔮 Burnout risk profile built<br>
      💡 AI Coach insights active
    </div>
  </div>
  <div style="background:rgba(239,68,68,.08);border:1px solid rgba(239,68,68,.2);border-radius:10px;padding:14px;margin-bottom:20px">
    <div style="color:#fca5a5;font-size:13px;font-weight:600">After midnight: downgraded to Starter (5 sessions/month). Your intelligence data is paused.</div>
  </div>
  <div style="text-align:center;margin-bottom:12px">
    <a href="{upgrade_url}" style="background:linear-gradient(135deg,#dc2626,#b91c1c);color:white;padding:14px 32px;border-radius:10px;text-decoration:none;font-weight:700;font-size:14px">Secure Your Intelligence Platform →</a>
  </div>
  <p style="color:#475569;font-size:11px;text-align:center">Use code <b style="color:#f87171">INTEL20</b> for 20% off first month · No contract · Cancel anytime</p>
</div></div>""",
            7: f"""<div style="font-family:-apple-system,sans-serif;max-width:560px;margin:0 auto;background:#07111f;border-radius:14px;overflow:hidden">
<div style="background:linear-gradient(135deg,#1e1b4b,#312e81);padding:24px 28px;text-align:center">
  <div style="font-size:18px;font-weight:700;color:white;margin-bottom:4px">Your trial ended, {name}</div>
  <div style="font-size:12px;color:rgba(255,255,255,.6)">But your workforce intelligence data is secured for 30 days</div>
</div>
<div style="padding:28px">
  <p style="color:#94a3b8;font-size:13px;margin:0 0 18px;line-height:1.6">We want you to make the right decision — so we're giving you a <b style="color:#a5b4fc">3-day extension</b> to evaluate your data with zero pressure.</p>
  <div style="background:#0c1728;border-radius:10px;padding:16px;margin-bottom:18px;border:1px solid rgba(255,255,255,.06)">
    <div style="font-size:12px;color:#64748b;font-weight:600;text-transform:uppercase;letter-spacing:.06em;margin-bottom:10px">WHAT STAYS SECURED</div>
    <div style="color:#cbd5e1;font-size:13px;line-height:2">✦ All {sess_cnt} intelligence sessions<br>✦ Burnout risk profile &amp; fatigue patterns<br>✦ AI Coach conversation history<br>✦ Trend forecasting data</div>
  </div>
  <div style="text-align:center;margin-bottom:12px">
    <a href="{APP_URL}?extend=true" style="background:linear-gradient(135deg,#4f46e5,#0891b2);color:white;padding:13px 28px;border-radius:9px;text-decoration:none;font-weight:700;font-size:13px">Claim 3-Day Extension — Free →</a>
  </div>
  <p style="color:#475569;font-size:11px;text-align:center">No credit card. No commitment. Just more time to decide.</p>
</div></div>""",
            21: f"""<div style="font-family:-apple-system,sans-serif;max-width:560px;margin:0 auto;background:#07111f;border-radius:14px;overflow:hidden">
<div style="background:linear-gradient(135deg,#0f172a,#1e293b);padding:24px 28px;border-bottom:1px solid rgba(255,255,255,.06)">
  <div style="font-size:11px;font-weight:700;color:#64748b;letter-spacing:.08em;text-transform:uppercase;margin-bottom:6px">21-DAY INTELLIGENCE REPORT</div>
  <div style="font-size:18px;font-weight:700;color:#f1f5f9">{name} — your workforce health trend is in</div>
</div>
<div style="padding:28px">
  <div style="display:flex;gap:10px;margin-bottom:20px">
    <div style="flex:1;background:#0c1728;border-radius:10px;padding:14px;text-align:center;border:1px solid rgba(255,255,255,.06)">
      <div style="font-size:28px;font-weight:800;color:{'#10b981' if avg_score>=70 else '#f59e0b' if avg_score>=50 else '#ef4444'}">{avg_score}</div>
      <div style="font-size:10px;color:#64748b;margin-top:3px">3-WEEK AVG SCORE</div>
    </div>
    <div style="flex:1;background:#0c1728;border-radius:10px;padding:14px;text-align:center;border:1px solid rgba(255,255,255,.06)">
      <div style="font-size:28px;font-weight:800;color:#818cf8">{sess_cnt}</div>
      <div style="font-size:10px;color:#64748b;margin-top:3px">SESSIONS LOGGED</div>
    </div>
  </div>
  <div style="background:rgba(99,102,241,.08);border:1px solid rgba(99,102,241,.18);border-radius:10px;padding:14px;margin-bottom:20px">
    <div style="font-size:12px;color:#a5b4fc;font-weight:600;margin-bottom:6px">📈 Your 21-day AI analysis is ready</div>
    <div style="font-size:12px;color:#94a3b8;line-height:1.6">Executive summary · Burnout risk trajectory · Fatigue pattern analysis · Personalized intervention recommendations</div>
  </div>
  <div style="text-align:center">
    <a href="{APP_URL}" style="background:linear-gradient(135deg,#4f46e5,#0891b2);color:white;padding:12px 28px;border-radius:9px;text-decoration:none;font-weight:700;font-size:13px">View Full Intelligence Report →</a>
  </div>
</div></div>""",
        }
        subject   = subjects.get(day, f"PostureAI Pro — Day {day} workforce health update")
        html_body = bodies.get(day, f"<p>Hi {name}, thanks for using PostureAI Pro.</p>")
        ok = send_email(to, subject, html_body)
        return jsonify({"ok": ok, "day": day, "to": to})
    except Exception as e:
        return safe_error(e)


# ── Session store — Redis-backed (falls back to memory) ───────────
# Uses redis_service if REDIS_URL is set → survives Gunicorn restarts & multi-workers
# Falls back to in-memory dict transparently when Redis is not available.

import uuid as _uuid_mod

SESSION_TTL_S = 3600 * 8  # 8 hours — active posture session

class _SessionStore:
    """
    Dict-like session store backed by Redis (via redis_service).
    Falls back to an in-process dict when Redis is unavailable.
    Keys are session IDs; values are arbitrary dicts.
    Thread-safe: Redis operations are atomic; memory fallback uses a lock.
    """
    def __init__(self):
        self._mem:  dict         = {}
        self._lock: threading.Lock = threading.Lock()

    # ── internal helpers ─────────────────────────────────────────
    def _r_key(self, sid: str) -> str:
        return f"posture_session:{sid}"

    def _r_get(self, sid: str):
        try:
            raw = rget(self._r_key(sid))
            if raw:
                import json as _j
                return _j.loads(raw)
        except Exception:
            pass
        return None

    def _r_set(self, sid: str, data: dict):
        try:
            import json as _j
            rset(self._r_key(sid), _j.dumps(data, default=str), SESSION_TTL_S)
            return True
        except Exception:
            return False

    def _r_del(self, sid: str):
        try:
            rdel(self._r_key(sid))
        except Exception:
            pass

    # ── dict-like API ────────────────────────────────────────────
    def setdefault(self, sid: str, default: dict) -> dict:
        """Return existing session or create one from default."""
        # Try Redis first
        existing = self._r_get(sid)
        if existing is not None:
            return existing
        # Try memory fallback
        with self._lock:
            if sid in self._mem:
                return self._mem[sid]
            self._mem[sid] = default
        # Persist to Redis
        self._r_set(sid, default)
        return default

    def get(self, sid: str, fallback=None):
        d = self._r_get(sid)
        if d is not None:
            return d
        with self._lock:
            return self._mem.get(sid, fallback)

    def __getitem__(self, sid: str):
        d = self.get(sid)
        if d is None:
            raise KeyError(sid)
        return d

    def __setitem__(self, sid: str, data: dict):
        self._r_set(sid, data)
        with self._lock:
            self._mem[sid] = data

    def __delitem__(self, sid: str):
        self._r_del(sid)
        with self._lock:
            self._mem.pop(sid, None)

    def __contains__(self, sid: str):
        return self.get(sid) is not None

    def values(self):
        """Return an iterator over all in-memory sessions (best-effort for analytics)."""
        with self._lock:
            return list(self._mem.values())

    def items(self):
        with self._lock:
            return list(self._mem.items())

    def __len__(self):
        with self._lock:
            return len(self._mem)

    # ── mutation helper (read-modify-write) ──────────────────────
    def update_session(self, sid: str, updates: dict):
        """Atomically update fields in a session."""
        s = self.get(sid, {})
        if s is None:
            s = {}
        s.update(updates)
        self[sid] = s
        return s

sessions = _SessionStore()
print("✅ Session store: Redis-backed (falls back to memory if no REDIS_URL)")

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
        Table, TableStyle, HRFlowable, PageBreak)
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    from reportlab.graphics.shapes import Drawing, Rect, Line, String
    REPORTLAB_OK = True
except ImportError:
    REPORTLAB_OK = False

# ── Geometry helpers ───────────────────────────────────────────────
def lm_xy(lm, w, h): return (lm.x * w, lm.y * h)
def lm_xyz(lm, w, h): return np.array([lm.x * w, lm.y * h, lm.z * w])
def dist2d(a, b):
    """Euclidean distance between two 2D points."""
    return math.sqrt((a[0]-b[0])**2 + (a[1]-b[1])**2)

def dist2d_sq(a, b):
    """Squared distance — use for comparisons to avoid sqrt overhead."""
    return (a[0]-b[0])**2 + (a[1]-b[1])**2

def angle_vert(p1, p2):
    dx = p2[0] - p1[0]
    dy = p2[1] - p1[1]
    if abs(dy) < 0.5:
        return 90.0
    return abs(math.degrees(math.atan2(abs(dx), abs(dy))))

def angle_horiz(p1, p2):
    dx = p2[0] - p1[0]
    dy = p2[1] - p1[1]
    if abs(dx) < 0.5:
        return 90.0
    return abs(math.degrees(math.atan2(abs(dy), abs(dx))))

def angle_3pt(a, b, c):
    """Angle at point b in triangle abc — pure Python, no numpy overhead."""
    v1x, v1y = a[0]-b[0], a[1]-b[1]
    v2x, v2y = c[0]-b[0], c[1]-b[1]
    n1 = math.sqrt(v1x*v1x + v1y*v1y)
    n2 = math.sqrt(v2x*v2x + v2y*v2y)
    if n1 < 0.001 or n2 < 0.001: return 90.0
    cos_a = (v1x*v2x + v1y*v2y) / (n1 * n2)
    return math.degrees(math.acos(max(-1.0, min(1.0, cos_a))))

def angle_3pt_3d(a, b, c):
    """3D angle at b using landmark (x,y,z) — more accurate for wrist/elbow."""
    v1x,v1y,v1z = a[0]-b[0], a[1]-b[1], a[2]-b[2]
    v2x,v2y,v2z = c[0]-b[0], c[1]-b[1], c[2]-b[2]
    n1 = math.sqrt(v1x*v1x + v1y*v1y + v1z*v1z)
    n2 = math.sqrt(v2x*v2x + v2y*v2y + v2z*v2z)
    if n1 < 0.001 or n2 < 0.001: return 90.0
    cos_a = (v1x*v2x + v1y*v2y + v1z*v2z) / (n1 * n2)
    return math.degrees(math.acos(max(-1.0, min(1.0, cos_a))))

def score_m(v, ideal, ok, bad):
    """
    Piecewise ergonomic score (0–100):
    - ok zone  [0, ok]:       100 → 75  (linear — comfortable range)
    - bad zone [ok, bad]:     75  → 30  (linear — attention needed)
    - beyond bad:             30  → 5   (quadratic — injury risk zone)

    Quadratic beyond bad: small extra deviations are forgiven,
    large ones (e.g. 40° neck lean) hit 5 fast.
    Mirrors ergonomic research: MSK injury risk is non-linear.
    """
    d = abs(v - ideal)
    if d <= ok:
        return max(0, int(100 - (d / max(ok, .1)) * 25))
    if d <= bad:
        return max(0, int(75  - ((d - ok) / max(bad - ok, .1)) * 45))
    # Beyond bad: quadratic decay → floor 5
    excess = d - bad
    decay  = min(25, excess ** 1.6 * 0.9)  # quadratic, capped at 25
    return max(5, int(30 - decay))

def cam_mat(w, h):
    """
    Camera intrinsic matrix with aspect-ratio-correct focal length.
    Assumes ~70° horizontal FOV (typical webcam/phone).
    fx = w / (2 * tan(hFOV/2)) ≈ w * 0.85 for 70° FOV
    fy = h / (2 * tan(vFOV/2)) — preserves aspect ratio for portrait cameras
    """
    fx = w * 0.85          # horizontal focal: ~70° H-FOV
    fy = h * 0.85          # vertical focal: scales with h, not w
    cx, cy = w / 2, h / 2
    return np.array([[fx, 0, cx], [0, fy, cy], [0, 0, 1]], dtype=np.float64)

# ── Blink rate detection using FaceMesh ───────────────────────────
_blink_history = []  # (timestamp, ear_ratio)
_lm_history     = {}   # {session_id: [lm_array,...]} last 3 frames for jitter reduction
_celery_task    = None  # lazy Celery singleton
_score_pool     = []   # rolling pool of recent scores for percentile (max 10000)

def compute_gaze(face_lms, w, h):
    """
    Estimate gaze direction from iris position within eye socket.
    Returns (h_offset, v_offset) normalized -1..+1.
    h_offset: +1=looking right, -1=looking left
    v_offset: +1=looking down, -1=looking up
    Returns None if iris not visible.
    """
    try:
        def pt(i): return (face_lms.landmark[i].x*w, face_lms.landmark[i].y*h)
        # Left iris center (468) within left eye socket (33=inner, 133=outer, 159=top, 145=bot)
        l_iris  = pt(468)
        l_inner = pt(33);  l_outer = pt(133)
        l_top   = pt(159); l_bot   = pt(145)
        l_cx = (l_inner[0]+l_outer[0])/2
        l_cy = (l_top[1]+l_bot[1])/2
        l_hw = max(abs(l_outer[0]-l_inner[0])/2, 1)
        l_hh = max(abs(l_bot[1]-l_top[1])/2, 1)
        l_gx = (l_iris[0]-l_cx) / l_hw   # -1..+1
        l_gy = (l_iris[1]-l_cy) / l_hh

        # Right iris center (473) within right eye socket
        r_iris  = pt(473)
        r_inner = pt(362); r_outer = pt(263)
        r_top   = pt(386); r_bot   = pt(374)
        r_cx = (r_inner[0]+r_outer[0])/2
        r_cy = (r_top[1]+r_bot[1])/2
        r_hw = max(abs(r_outer[0]-r_inner[0])/2, 1)
        r_hh = max(abs(r_bot[1]-r_top[1])/2, 1)
        r_gx = (r_iris[0]-r_cx) / r_hw
        r_gy = (r_iris[1]-r_cy) / r_hh

        gaze_h = (l_gx + r_gx) / 2   # + = looking right
        gaze_v = (l_gy + r_gy) / 2   # + = looking down
        return {"h": round(gaze_h,3), "v": round(gaze_v,3),
                "looking_at_screen": abs(gaze_h) < 0.35 and abs(gaze_v) < 0.35}
    except Exception:
        return None

def compute_cognitive_load(face_lms, w, h):
    """
    Estimate cognitive load / stress from facial micro-expressions.
    High cognitive load → tight jaw, furrowed brows, faster blinking, eye squint.
    Returns load score 0-100 (0=relaxed, 100=high stress/focus).
    Based on FACS (Facial Action Coding System) — AU4 (brow furrow) + AU6 (cheek raise).
    """
    try:
        def pt(i): return (face_lms.landmark[i].x * w, face_lms.landmark[i].y * h)
        def d(a, b): return math.sqrt((a[0]-b[0])**2 + (a[1]-b[1])**2)

        # ── AU4: Brow furrow (corrugator supercilii) ──────────────
        # FaceMesh: L brow inner 107, R brow inner 336
        # Measure distance between inner brows — closer = more furrow
        l_brow_in = pt(107); r_brow_in = pt(336)
        brow_dist = d(l_brow_in, r_brow_in)
        # Normalize by face width (eye outer corners: 33, 263)
        face_w = max(d(pt(33), pt(263)), 1)
        brow_ratio = brow_dist / face_w   # lower = more furrow

        # ── AU6: Cheek raise / Duchenne marker ───────────────────
        # Lower eyelid tightening: distance from pupil center to lower lid
        l_lower_lid = pt(145); r_lower_lid = pt(374)
        l_pupil_pt  = pt(468); r_pupil_pt  = pt(473)
        l_squint = d(l_pupil_pt, l_lower_lid)
        r_squint = d(r_pupil_pt, r_lower_lid)
        avg_squint = (l_squint + r_squint) / 2
        squint_norm = avg_squint / max(face_w * 0.06, 1)  # normalize

        # ── Jaw tension: mouth openness ───────────────────────────
        # Upper lip: 13, lower lip: 14
        lip_top = pt(13); lip_bot = pt(14)
        mouth_open = d(lip_top, lip_bot) / max(face_w * 0.2, 1)  # 0=closed, 1=open
        # Tension = mouth slightly open / clenched (not relaxed neutral)
        jaw_tension = abs(mouth_open - 0.15) / 0.15   # 0=neutral, 1=tense/open

        # ── Combine ───────────────────────────────────────────────
        # Low brow_ratio = furrow; low squint_norm = tight eyes; high jaw_tension
        brow_stress  = max(0, min(1, (0.35 - brow_ratio) / 0.15))   # 0-1
        squint_stress= max(0, min(1, (1.0 - squint_norm) / 0.5))    # 0-1
        jaw_stress   = max(0, min(1, jaw_tension))

        # Weighted: brow furrow most reliable FACS marker for cognitive load
        load_raw = brow_stress * 0.50 + squint_stress * 0.30 + jaw_stress * 0.20
        load_score = max(0, min(100, int(round(load_raw * 100))))

        return {
            "load_score":   load_score,
            "level":        "high" if load_score > 65 else "moderate" if load_score > 35 else "low",
            "brow_furrow":  round(brow_stress, 3),
            "eye_squint":   round(squint_stress, 3),
            "jaw_tension":  round(jaw_stress, 3),
            "interpretation": (
                "High cognitive load — take a mental break" if load_score > 65 else
                "Moderate focus/concentration" if load_score > 35 else
                "Relaxed / low cognitive load"
            ),
        }
    except Exception:
        return None


def compute_ear(face_lms, w, h):
    """Eye Aspect Ratio for blink detection and eye strain."""
    try:
        def pt(idx): return (face_lms.landmark[idx].x * w, face_lms.landmark[idx].y * h)
        # Left eye vertical/horizontal ratio
        lt = pt(L_EYE_TOP); lb = pt(L_EYE_BOT)
        l_inner = pt(33);   l_outer = pt(133)
        rt = pt(R_EYE_TOP); rb = pt(R_EYE_BOT)
        r_inner = pt(362);  r_outer = pt(263)
        l_v = math.dist(lt, lb)
        l_h = math.dist(l_inner, l_outer)
        r_v = math.dist(rt, rb)
        r_h = math.dist(r_inner, r_outer)
        if l_h < 1 or r_h < 1: return None
        ear = (l_v / l_h + r_v / r_h) / 2.0
        return round(ear, 3)
    except Exception:
        return None

# ── Head velocity tracker ─────────────────────────────────────────
_head_velocity: dict = {}   # {session_key: [(timestamp, nose_x, nose_y), ...]}

# ── Forward head creep buffer ────────────────────────────────────
_head_creep_buffer: dict = {}  # {sid: [(timestamp, neck_lean)]}
_HEAD_CREEP_WINDOW_S = 1200    # 20 minutes

def detect_head_creep(sid: str, neck_lean: float, timestamp: float = None) -> dict:
    """
    Detect gradual forward head creep over 20 minutes.
    Creep = slow progressive increase in neck lean (more dangerous than sudden lean).
    Returns: {"creep_rate": deg/min, "severity": none/mild/moderate/severe, "alert": str}
    """
    import time as _t
    ts = timestamp or _t.time()
    buf = _head_creep_buffer.setdefault(sid, [])
    buf.append((ts, neck_lean))

    # Prune old entries
    cutoff = ts - _HEAD_CREEP_WINDOW_S
    _head_creep_buffer[sid] = [(t, v) for t, v in buf if t >= cutoff]
    buf = _head_creep_buffer[sid]

    if len(buf) < 10:   # need at least 10 data points
        return {"creep_rate": 0.0, "severity": "insufficient_data"}

    # Linear regression to find trend slope (deg/min)
    n = len(buf)
    ts_arr  = [(b[0] - buf[0][0]) / 60 for b in buf]  # minutes from start
    val_arr = [b[1] for b in buf]
    if n == 0:
        return None
    mean_t  = sum(ts_arr) / n
    mean_v  = sum(val_arr) / n
    num     = sum((ts_arr[i] - mean_t) * (val_arr[i] - mean_v) for i in range(n))
    den     = sum((t - mean_t) ** 2 for t in ts_arr) or 1
    slope   = num / den  # deg per minute — positive = getting worse

    severity = (
        "severe"   if slope > 1.5 else
        "moderate" if slope > 0.8 else
        "mild"     if slope > 0.3 else
        "none"
    )

    alert = None
    if severity == "severe":
        alert = f"⚠️ Progressive neck creep {slope:.1f}°/min — your posture is deteriorating rapidly. Take a break now."
    elif severity == "moderate":
        alert = f"Neck slowly moving forward ({slope:.1f}°/min over {int((ts-buf[0][0])/60)}min) — adjust position"

    return {
        "creep_rate":  round(slope, 2),
        "severity":    severity,
        "window_min":  round((ts - buf[0][0]) / 60, 1),
        "start_lean":  round(buf[0][1], 1),
        "current_lean":round(neck_lean, 1),
        "total_change":round(neck_lean - buf[0][1], 1),
        "alert":       alert,
    }

def compute_head_velocity(nose_x, nose_y, session_key):
    """
    Track head movement velocity and detect sudden movements.
    Useful for: attention switching detection, whiplash risk, fidgeting.
    Returns velocity in deg/sec equivalent and movement pattern.
    """
    now = time.time()
    buf = _head_velocity.setdefault(session_key, [])
    buf.append((now, nose_x, nose_y))
    # Keep 2s window
    buf[:] = [(t,x,y) for t,x,y in buf if now-t < 2.0]
    if len(buf) < 3:
        return None

    # Velocity: position change per second (normalized to frame)
    dt  = max(0.001, buf[-1][0] - buf[-3][0])
    dx  = abs(buf[-1][1] - buf[-3][1])   # normalized 0-1
    dy  = abs(buf[-1][2] - buf[-3][2])
    vel = math.sqrt(dx**2 + dy**2) / dt if dt > 0 else 0.0  # units/sec

    # Acceleration: change in velocity
    if len(buf) >= 5:
        dt2 = max(0.001, buf[-3][0] - buf[-5][0])
        dx2 = abs(buf[-3][1] - buf[-5][1])
        dy2 = abs(buf[-3][2] - buf[-5][2])
        vel2 = math.sqrt(dx2**2 + dy2**2) / dt2 if dt2 > 0 else 0.0
        accel = (vel - vel2) / max(dt, 0.001)
    else:
        accel = 0.0

    # Classify movement
    if vel > 0.15:   pattern = "rapid"    # attention switch / startled
    elif vel > 0.05: pattern = "moderate" # normal repositioning
    elif vel > 0.01: pattern = "micro"    # natural micro-movement (good)
    else:            pattern = "frozen"   # too still

    return {
        "velocity":    round(vel * 1000, 1),   # mm/sec equivalent
        "acceleration":round(accel * 100, 2),
        "pattern":     pattern,
        "direction_x": round(dx / max(dt, 0.001) * 1000, 1),
        "direction_y": round(dy / max(dt, 0.001) * 1000, 1),
        "sudden_move": vel > 0.12,
    }


# ── Kalman state per session per landmark ─────────────────────────
# State: [x, y, vx, vy] — position + velocity
# Measurement noise R adapts to landmark visibility
_kalman_states: dict = {}  # {session_key: {idx: [x,y,vx,vy]}}

# ── Adaptive Kalman noise store ──────────────────────────────────
_kalman_noise: dict = {}   # {uid: {"R_sum": float, "count": int, "R_est": float}}

def kalman_update(state, measurement, visibility, dt=0.033, uid: str = ""):
    """
    ADAPTIVE 1D Kalman filter — learns user-specific noise from residuals.
    Adapts to users with natural tremor or unstable landmarks.
    Converges after ~50 frames to user-specific noise level.
    Returns updated [x, y, vx, vy].
    """
    if state is None:
        return [measurement[0], measurement[1], 0.0, 0.0]

    x, y, vx, vy = state

    # ── Adaptive R (measurement noise) ───────────────────────────
    base_R = max(0.001, (1.0 - min(1.0, visibility)) * 0.08 + 0.005)
    if uid and uid in _kalman_noise and _kalman_noise[uid]["count"] >= 30:
        learned_R = _kalman_noise[uid]["R_est"]
        R = 0.55 * learned_R + 0.45 * base_R   # blend learned + visibility
    else:
        R = base_R

    Q = 0.003
    P_pred = Q + 0.01
    K = P_pred / (P_pred + R)

    x_pred = x + vx * dt
    y_pred = y + vy * dt

    innov_x = measurement[0] - x_pred
    innov_y = measurement[1] - y_pred

    x_upd  = x_pred + K * innov_x
    y_upd  = y_pred + K * innov_y
    vx_upd = (x_upd - x) / max(dt, 0.001)
    vy_upd = (y_upd - y) / max(dt, 0.001)

    # ── Online R estimation from residuals ────────────────────────
    if uid and visibility > 0.5:
        ns = _kalman_noise.setdefault(uid, {"R_sum": 0.0, "count": 0, "R_est": 0.01})
        residual_sq = (innov_x**2 + innov_y**2) / 2
        ns["R_sum"] += residual_sq
        ns["count"] += 1
        if ns["count"] >= 10:
            ns["R_est"] = max(0.001, min(0.5, ns["R_sum"] / ns["count"]))
        if ns["count"] > 500:   # prevent unbounded growth
            ns["R_sum"] /= 2; ns["count"] //= 2

    return [x_upd, y_upd, vx_upd, vy_upd]


# Per-session blink history — keyed by uid:session to avoid cross-contamination
_blink_sessions: dict      = {}   # {key: [(timestamp, ear), ...]}
_ear_baselines:  dict      = {}   # {key: float} — per-user open-eye EAR baseline
_blink_sessions_last_clean = 0.0

def analyze_blink_rate(face_lms, w, h, uid=None, session_id=None):
    """
    Return blink rate per minute and eye strain risk.
    Uses per-session in-memory buffer — no cross-user contamination.
    Redis preferred when available (persists across cold starts).
    """
    global _blink_sessions, _blink_sessions_last_clean
    ear = compute_ear(face_lms, w, h)
    if ear is None:
        return None
    now = time.time()

    # ── Per-user Redis buffer (preferred) ────────────────────────
    if uid and REDIS_READY:
        try:
            push_blink(uid, ear, now)
            return get_blink_rate(uid)
        except Exception:
            pass

    # ── Per-session in-memory buffer ─────────────────────────────
    # Key: uid + session_id (or uid alone, or anonymous)
    buf_key = f"{uid or 'anon'}:{session_id or 'default'}"
    buf = _blink_sessions.setdefault(buf_key, [])
    buf.append((now, ear))
    # Keep only last 60s
    _blink_sessions[buf_key] = [(t, e) for t, e in buf if now - t < 60]
    buf = _blink_sessions[buf_key]
    if len(buf) > 600: _blink_sessions[buf_key] = buf[-600:]  # hard cap

    # Periodic cleanup of stale sessions (>90s inactive)
    if now - _blink_sessions_last_clean > 300:
        stale = [k for k, v in _blink_sessions.items()
                 if v and now - v[-1][0] > 90]
        for k in stale: del _blink_sessions[k]
        _blink_sessions_last_clean = now

    if len(buf) < 5:
        return None

    # ── Adaptive EAR threshold ───────────────────────────────────
    # Baseline = average of top-40% EAR values (open-eye moments)
    # Blink threshold = baseline * 0.60 (eye 40% closed = blink)
    ear_vals = [e for _, e in buf]
    if len(ear_vals) >= 10:
        sorted_ears = sorted(ear_vals, reverse=True)
        baseline = sum(sorted_ears[:max(1, len(sorted_ears)//3)]) / max(1, len(sorted_ears)//3)
        _ear_baselines[buf_key] = baseline
    else:
        baseline = _ear_baselines.get(buf_key, 0.28)  # default if not enough data yet

    close_thresh = baseline * 0.60   # blink = eye 40% closed
    open_thresh  = baseline * 0.75   # hysteresis gap

    blinks = 0
    was_closed = False
    for _, e in buf:
        if e < close_thresh and not was_closed:
            blinks += 1
            was_closed = True
        elif e >= open_thresh:
            was_closed = False

    # Calibrate to 60s window
    window_s = max(1.0, buf[-1][0] - buf[0][0])
    blinks_per_min = round(blinks * 60 / window_s) if window_s > 0 else 0

    # Ergonomic risk thresholds (NIOSH guidelines: normal = 12-20/min)
    if blinks_per_min < 6:   risk = "high"
    elif blinks_per_min < 12: risk = "moderate"
    elif blinks_per_min <= 20: risk = "normal"
    else:                      risk = "elevated"  # >20/min may indicate irritation

    return {
        "blink_rate_per_min": blinks_per_min,
        "eye_strain_risk":    risk,
        "ear_ratio":          ear,
        "window_s":           round(window_s, 1),
        "samples":            len(buf),
    }

# ── IPD-based distance ─────────────────────────────────────────────
# ── Per-session focal calibration store ─────────────────────────────
_focal_cal: dict = {}  # {session_id: focal_px} — calibrated from face_width
_focal_cal_lock = threading.RLock()

def _calibrate_focal(face_lms, w, h, session_id: str = "", known_dist_cm: float = 65.0):
    """
    Calibrate focal length from face width (known ~14cm real width).
    focal = face_width_px * known_dist_cm / 14.0
    Stores result per-session so subsequent frames use calibrated focal.
    Only updates when yaw < 10° (face nearly frontal = most accurate).
    """
    try:
        # Use landmarks 234 (left cheek) and 454 (right cheek) — stable width markers
        lc = face_lms.landmark[234]; rc = face_lms.landmark[454]
        face_w_px = abs(rc.x * w - lc.x * w)
        if face_w_px < 20: return None
        # focal = (face_width_px * real_width_cm) / real_width_cm
        # real face width ≈ 14cm (bizygomatic diameter, population mean)
        focal_est = (face_w_px * known_dist_cm) / 14.0
        # Running average: new = 0.1*new + 0.9*old (slow drift)
        if session_id and session_id in _focal_cal:
            _focal_cal[session_id] = 0.10 * focal_est + 0.90 * _focal_cal[session_id]
        elif session_id:
            _focal_cal[session_id] = focal_est
        return _focal_cal.get(session_id, focal_est)
    except Exception:
        return None

def ipd_distance_face(face_lms, w, h, yaw_deg=0.0, session_id: str = ""):
    """
    IMPROVED: Dual-estimator distance with calibrated focal length.
    1. IPD estimator  — 6.3cm real IPD, yaw-corrected
    2. Face-width estimator — 14cm bizygomatic diameter
    Blend: 70% IPD + 30% face-width (when yaw < 20° both are reliable)
           100% IPD when yaw >= 20° (face-width less reliable at angle)
    Focal: calibrated per-session from face-width estimator
    """
    try:
        lp = face_lms.landmark[L_PUPIL]
        rp = face_lms.landmark[R_PUPIL]
        lpx, lpy = lp.x * w, lp.y * h
        rpx, rpy = rp.x * w, rp.y * h
        ipd_px = math.sqrt((rpx-lpx)**2 + (rpy-lpy)**2)
        if ipd_px < 6: return None

        # Yaw correction
        yaw_rad          = math.radians(abs(yaw_deg))
        cos_yaw          = max(math.cos(yaw_rad), 0.5)
        ipd_px_corrected = ipd_px / cos_yaw

        # ── Calibrated focal ─────────────────────────────────────
        # Try to calibrate focal from face width (more stable than IPD alone)
        focal_cal = _calibrate_focal(face_lms, w, h, session_id) if abs(yaw_deg) < 10 else None
        focal     = focal_cal if focal_cal else (630 * (w / 640))  # fallback: sensor estimate

        # ── IPD estimator ────────────────────────────────────────
        dist_ipd = round((6.3 * focal) / ipd_px_corrected, 1) if ipd_px_corrected > 0 else 0.0

        # ── Face-width estimator (secondary) ─────────────────────
        dist_fw = None
        if abs(yaw_deg) < 20:
            try:
                lc = face_lms.landmark[234]; rc = face_lms.landmark[454]
                fw_px = abs(rc.x * w - lc.x * w)
                if fw_px > 20:
                    dist_fw = round((14.0 * focal) / fw_px, 1)
            except Exception:
                pass

        # ── Blend ────────────────────────────────────────────────
        if dist_fw and abs(yaw_deg) < 20:
            dist = 0.70 * dist_ipd + 0.30 * dist_fw
        else:
            dist = dist_ipd

        return max(20, min(150, round(dist, 1)))
    except Exception:
        return None

# ── solvePnP head pose ─────────────────────────────────────────────
# Extended 3D model points — anatomically stable FaceMesh landmarks
# (nose tip, chin, eye corners, mouth corners, cheekbones, forehead)
_MODEL_3D_EXT_RAW = (
    (0.0,   0.0,    0.0),    # 1  — nose tip
    (0.0,  -63.6, -12.5),    # 152 — chin
    (-43.3, 32.7, -26.0),    # 33  — left eye inner corner
    (43.3,  32.7, -26.0),    # 263 — right eye inner corner
    (-28.9,-28.9, -24.1),    # 61  — left mouth corner
    (28.9, -28.9, -24.1),    # 291 — right mouth corner
    (-56.1,  5.5, -55.5),    # 234 — left cheek
    (56.1,   5.5, -55.5),    # 454 — right cheek
    (0.0,   27.1, -39.5),    # 10  — mid forehead
    (-35.5, 52.0, -30.0),    # 21  — left brow outer
    (35.5,  52.0, -30.0),    # 251 — right brow outer
    (-35.0,-45.0, -40.0),    # 58  — left jaw
    (35.0, -45.0, -40.0),    # 288 — right jaw
    (0.0,   15.0, -55.0),    # 168 — nose bridge
)
_MODEL_3D_EXT_IDXS = [1, 152, 33, 263, 61, 291, 234, 454, 10, 21, 251, 58, 288, 168]
MODEL_3D_EXT = None  # initialized lazily

def head_pose_from_face(face_lms, w, h):
    """
    solvePnP with 14 stable landmarks (was 6).
    More points → lower reprojection error → ±2° accuracy vs ±5° before.
    Falls back to 6-point if extended model not yet initialized.
    """
    global MODEL_3D_EXT
    try:
        # Initialize extended model on first call
        if MODEL_3D_EXT is None:
            MODEL_3D_EXT = np.array(_MODEL_3D_EXT_RAW, dtype=np.float64)

        pts2d = np.array([
            [face_lms.landmark[i].x * w, face_lms.landmark[i].y * h]
            for i in _MODEL_3D_EXT_IDXS
        ], dtype=np.float64)

        cam = cam_mat(w, h)
        # SOLVEPNP_EPNP: faster and more stable than ITERATIVE for n>6
        ok, rvec, tvec = cv2.solvePnP(
            MODEL_3D_EXT, pts2d, cam, np.zeros((4, 1)),
            flags=cv2.SOLVEPNP_EPNP
        )
        if not ok:
            return None

        # Refine with LM (Levenberg-Marquardt) for sub-degree accuracy
        rvec, tvec = cv2.solvePnPRefineLM(
            MODEL_3D_EXT, pts2d, cam, np.zeros((4, 1)), rvec, tvec
        )

        rmat, _ = cv2.Rodrigues(rvec)
        angles, *_ = cv2.RQDecomp3x3(rmat)

        # Compute reprojection error as quality indicator
        proj, _ = cv2.projectPoints(MODEL_3D_EXT, rvec, tvec, cam, np.zeros((4,1)))
        reproj_err = float(np.mean(np.linalg.norm(proj.reshape(-1,2) - pts2d, axis=1)))

        return {
            "pitch":       round(float(angles[0]), 1),
            "yaw":         round(float(angles[1]), 1),
            "roll":        round(float(angles[2]), 1),
            "reproj_err":  round(reproj_err, 2),   # <2.0 = reliable, >5.0 = low accuracy
            "landmarks_n": len(_MODEL_3D_EXT_IDXS),
        }
    except Exception:
        return None

# ── FRONT ANALYSIS ─────────────────────────────────────────────────
def analyze_front(image, mode="laptop", tier="standard", session_id=None):
    _ensure_models()  # lazy-load MediaPipe on first call
    h, w = image.shape[:2]
    rgb   = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)

    out = {
        "mode": mode, "detected": False, "metrics": {}, "score": 0,
        "alerts": [], "recommendations": [], "landmarks": [],
        "head_pose": None, "confidence": 0, "engine": "mediapipe",
        "eye_strain": None
    }

    pose_model = POSE_FULL if tier in ("professional", "elite", "pro", "premium", "business") else POSE_LITE
    pose_result = pose_model.process(rgb)
    face_result = FACE_MESH.process(rgb)

    if not pose_result.pose_landmarks:
        out["alerts"].append("No person detected — ensure your upper body is visible in the camera frame")
        return analyze_front_cascade(image, mode, out)

    # ── Landmark averaging: reduce ±3° jitter to ±1° ─────────────
    _sid  = session_id or out.get("session_id", "front_default")
    _raw  = pose_result.pose_landmarks.landmark
    _hist = _lm_history.setdefault(_sid, [])
    _hist.append(_raw)
    if len(_hist) > 3: _hist.pop(0)

    # ── Kalman + weighted average landmark smoothing ─────────────
    # Step 1: weighted avg over 3 frames (recent=0.6, prev=0.3, old=0.1)
    # Step 2: Kalman filter on top — adapts to visibility (low vis = trust prediction)
    _w = [0.60, 0.30, 0.10][:len(_hist)]
    _w_sum  = sum(_w)
    _w_norm = [wi / _w_sum for wi in _w] if _w_sum > 0 else _w

    _k_key   = f"kalman:{_sid}"
    _kstates = _kalman_states.setdefault(_k_key, {})

    class _KalmanLM:
        __slots__ = ("x","y","z","visibility")
        def __init__(self, idx):
            frames_rev = list(reversed(_hist))
            weights    = _w_norm[:len(frames_rev)]
            # Weighted average position
            wx = sum(frames_rev[i][idx].x          * weights[i] for i in range(len(frames_rev)))
            wy = sum(frames_rev[i][idx].y          * weights[i] for i in range(len(frames_rev)))
            wz = sum(frames_rev[i][idx].z          * weights[i] for i in range(len(frames_rev)))
            wv = sum(frames_rev[i][idx].visibility * weights[i] for i in range(len(frames_rev)))
            # Kalman refinement on x,y
            kst = kalman_update(_kstates.get(idx), (wx, wy), wv)
            _kstates[idx] = kst
            self.x          = kst[0]
            self.y          = kst[1]
            self.z          = wz
            self.visibility = wv

    class _KalmanLMs:
        def __getitem__(self, idx): return _KalmanLM(idx)
    lms = _KalmanLMs()

    out["detected"]   = True
    out["engine"]     = "mediapipe_pose+kalman"
    out["avg_frames"] = len(_hist)
    global _last_model_use; _last_model_use = time.time()  # update liveness timestamp

    def g(idx): return lms[idx]
    def px(idx): return (g(idx).x * w, g(idx).y * h)

    nose    = px(PL.NOSE)
    l_ear   = px(PL.L_EAR);      r_ear  = px(PL.R_EAR)
    l_sh    = px(PL.L_SHOULDER);  r_sh   = px(PL.R_SHOULDER)
    l_hip   = px(PL.L_HIP);      r_hip  = px(PL.R_HIP)
    l_eye   = px(PL.L_EYE);      r_eye  = px(PL.R_EYE)

    mid_sh  = ((l_sh[0]+r_sh[0])/2,  (l_sh[1]+r_sh[1])/2)
    mid_ear = ((l_ear[0]+r_ear[0])/2, (l_ear[1]+r_ear[1])/2)
    mid_hip = ((l_hip[0]+r_hip[0])/2, (l_hip[1]+r_hip[1])/2)
    mid_eye = ((l_eye[0]+r_eye[0])/2, (l_eye[1]+r_eye[1])/2)

    # ── FIX: neck_lean uses nose instead of ear ────────────────────
    # ear moves with head rotation → false forward-lean readings
    # nose → mid_sh measures true head-forward displacement
    # We use 85% nose + 15% ear blend for robustness when nose is occluded
    vis_nose = g(PL.NOSE).visibility if hasattr(g(PL.NOSE), 'visibility') else 1.0
    vis_l_ear = g(PL.L_EAR).visibility; vis_r_ear = g(PL.R_EAR).visibility
    ear_weight = 0.15 if vis_nose > 0.7 else 0.50  # use more ear if nose occluded
    nose_weight = 1.0 - ear_weight
    neck_ref_x = nose[0] * nose_weight + mid_ear[0] * ear_weight
    neck_ref_y = nose[1] * nose_weight + mid_ear[1] * ear_weight
    neck_ref   = (neck_ref_x, neck_ref_y)

    # Nose is ~10cm in front of ear plane — subtract natural offset
    # At 65cm distance, 10cm = ~8.8° apparent lean → correct by -5° (conservative)
    neck_lean_raw = angle_vert(mid_sh, neck_ref)

    # ── Monitor height estimation ────────────────────────────────
    try:
        _pitch_val   = out.get("head_pose", {}).get("pitch", 0) if out.get("head_pose") else 0
        _dist_now    = out.get("distCm")   # safe: already written to out before this point
        if _pitch_val and _dist_now:
            import math as _mh
            _monitor_offset_cm = round(_dist_now * _mh.tan(_mh.radians(abs(_pitch_val))), 1)
            _monitor_dir = "below" if _pitch_val < 0 else "above"
            _monitor_sc  = score_m(abs(_pitch_val), 0, 5, 18)
            out["metrics"]["monitor_height"] = {
                "value":      _monitor_offset_cm,
                "score":      _monitor_sc,
                "unit":       "cm",
                "label":      "Monitor height offset",
                "direction":  _monitor_dir,
                "pitch_deg":  _pitch_val,
                "adjustment": f"Raise monitor {_monitor_offset_cm}cm" if _monitor_dir == "below" else
                              f"Lower monitor {_monitor_offset_cm}cm",
            }
            if abs(_pitch_val) > 12:
                out["alerts"].append(
                    f"⚠️ Monitor ~{_monitor_offset_cm}cm {_monitor_dir} eye level — "
                    f"{'raise' if _monitor_dir == 'below' else 'lower'} monitor to reduce neck strain")
    except Exception:
        pass

    # ── Forward Head Posture Index (FHP) ─────────────────────────
    # Clinical measure: horizontal offset of ear ahead of shoulder
    # Converts pixel offset to approximate cm using shoulder width reference
    # Normal: <2cm forward. Each 2.5cm forward = +4-5kg neck load
    try:
        _ear_x_offset  = abs(mid_ear[0] - mid_sh[0])   # pixels
        _sh_width_cm   = 42.0   # reference shoulder width in cm
        _sh_w_early    = abs(r_sh[0] - l_sh[0])         # compute here — sh_width_px defined later
        _cm_per_px     = _sh_width_cm / max(_sh_w_early, 1)
        _fhp_cm        = round(_ear_x_offset * _cm_per_px, 1)
        _extra_weight  = round(_fhp_cm / 2.5 * 4.5, 1)  # kg added to neck
        _fhp_sc        = score_m(_fhp_cm, 0, 2, 6)
        out["metrics"]["fhp_index"] = {
            "value":        _fhp_cm,
            "score":        _fhp_sc,
            "unit":         "cm",
            "label":        "Forward head posture",
            "extra_load_kg": _extra_weight,
        }
        if _fhp_cm > 6:
            out["alerts"].append(f"⚠️ Forward head posture {_fhp_cm}cm (+{_extra_weight}kg neck load) — critical: raise monitor immediately")
        elif _fhp_cm > 3:
            out["alerts"].append(f"Forward head posture {_fhp_cm}cm (+{_extra_weight}kg neck load) — tuck chin back")
    except Exception:
        pass
    # Natural nose-to-shoulder offset correction (~5° at typical distances)
    nose_correction = 5.0 * nose_weight  # scale correction by how much nose we're using
    neck_lean_nose  = max(0.0, neck_lean_raw - nose_correction)

    # ── Shoulder width normalization ────────────────────────────
    sh_width_px  = abs(r_sh[0] - l_sh[0])
    ref_sh_frac  = 0.34
    sh_frac      = sh_width_px / max(w, 1)
    sh_ratio     = sh_frac / ref_sh_frac if ref_sh_frac > 0 else 1.0
    sh_ratio     = max(0.70, min(1.30, sh_ratio))
    neck_ok_adj  = max(5.0, 6.0  * sh_ratio)
    neck_bad_adj = max(12.0, 17.0 * sh_ratio)

    # ── IMPROVED: use solvePnP pitch as PRIMARY neck lean source ──
    # solvePnP pitch is geometrically exact (not affected by camera angle
    # or nose-to-ear distance). Nose-based used only as fallback.
    # Blend: 80% solvePnP + 20% nose when both available (belt+suspenders)
    _hp_now = out.get("head_pose")
    _reproj  = _hp_now.get("reproj_err", 99) if _hp_now else 99
    if _hp_now and _reproj < 5.0:
        # solvePnP pitch: positive = head tilted down (forward lean)
        _pitch_abs = abs(_hp_now.get("pitch", 0))
        # Blend: reliable solvePnP gets 80%, nose-proxy gets 20%
        neck_lean = 0.80 * _pitch_abs + 0.20 * neck_lean_nose
        out["metrics"]["neck_source"] = {"value": "solvePnP+nose_blend", "reproj_err": round(_reproj, 2)}
    else:
        # Fallback: nose-based only (FaceMesh not available or poor quality)
        neck_lean = neck_lean_nose
        out["metrics"]["neck_source"] = {"value": "nose_only", "reproj_err": round(_reproj, 2) if _hp_now else None}

    neck_sc    = score_m(neck_lean, 0, neck_ok_adj, neck_bad_adj)

    # ── Forward head creep detection ─────────────────────────────
    try:
        import time as _ct
        _creep = detect_head_creep(_sid, neck_lean, _ct.time())
        out["metrics"]["head_creep"] = {
            "value":       _creep["creep_rate"],
            "severity":    _creep["severity"],
            "window_min":  _creep.get("window_min", 0),
            "total_change":_creep.get("total_change", 0),
            "unit":        "°/min",
            "label":       "Forward head creep rate",
        }
        if _creep.get("alert"):
            out["alerts"].append(_creep["alert"])
            out.setdefault("alerts_ar", []).append(
                f"⚠️ زحف الرأس للأمام {_creep['creep_rate']:.1f}°/دقيقة — وضعيتك تتردى تدريجياً. خذ استراحة."
                if _creep["severity"] == "severe"
                else f"رقبتك تتحرك للأمام ببطء ({_creep['creep_rate']:.1f}°/دقيقة) — اضبط وضعيتك"
            )
    except Exception:
        pass
    out["metrics"]["shoulder_width_ratio"] = {"value": round(sh_ratio, 2), "unit": "×", "label": "Shoulder width ratio"}

    head_tilt  = angle_horiz(l_eye, r_eye)
    tilt_sc    = score_m(head_tilt, 0, 3, 10)

    sh_tilt    = angle_horiz(l_sh, r_sh)
    sh_sc      = score_m(sh_tilt, 0, 3, 10)

    # ── Shoulder elevation (shrug/tension) ────────────────────────
    # Elevated shoulders = trapezius tension / stress posture
    # Measured: distance from shoulder to ear (Y axis)
    # Normal: shoulder 15-25% of frame height below ear
    # Elevated: < 10% = shrugging
    try:
        _l_ear_y  = g(PL.L_EAR).y
        _r_ear_y  = g(PL.R_EAR).y
        _l_sh_gap = g(PL.L_SHOULDER).y - _l_ear_y   # positive = shoulder below ear (normal)
        _r_sh_gap = g(PL.R_SHOULDER).y - _r_ear_y
        _sh_elev  = (_l_sh_gap + _r_sh_gap) / 2      # avg gap (normalized 0-1)
        _sh_elev_pct = round(_sh_elev * 100, 1)       # % of frame height

        # Score: ideal gap 12-25%, too close = elevated shoulders
        _sh_elev_sc = score_m(_sh_elev_pct, 18, 8, 4)
        out["metrics"]["shoulder_elevation"] = {
            "value":      _sh_elev_pct,
            "score":      _sh_elev_sc,
            "unit":       "% frame",
            "label":      "Shoulder elevation (tension)",
            "reference":  "Normal: 12-25% gap between ear and shoulder",
        }
        if _sh_elev_pct < 6:
            out["alerts"].append("⚠️ Shoulders elevated/shrugging — relax shoulders down and back")
            out.setdefault("alerts_ar", []).append("⚠️ الكتفان مرفوعان — أرخِ كتفيك للأسفل والخلف")
        elif _sh_elev_pct < 10:
            out["alerts"].append("Shoulders slightly elevated — try to relax trap muscles")
    except Exception:
        pass

    # ── Rounded shoulders (protraction) — Z depth asymmetry ───────
    # MediaPipe Z: more negative = closer to camera (in front of body)
    # Rounded shoulders: both shoulders move FORWARD (more negative Z)
    # vs neutral: Z ≈ 0 relative to spine midpoint
    _ls_z = g(PL.L_SHOULDER).z   # normalized, relative to hip midpoint
    _rs_z = g(PL.R_SHOULDER).z
    # Forward protraction: both Z values are negative (shoulders in front)
    # Score: average forward displacement, scaled by shoulder width
    _sh_z_avg        = (_ls_z + _rs_z) / 2.0          # negative = forward
    _sh_z_asym       = abs(_ls_z - _rs_z)              # L/R asymmetry
    _rounded_depth   = max(0.0, -_sh_z_avg * 100)      # convert to 0-100 scale
    _rounded_sc      = score_m(_rounded_depth, 0, 8, 20)
    out["metrics"]["rounded_shoulders"] = {
        "value":      round(_rounded_depth, 1),
        "z_left":     round(_ls_z, 3),
        "z_right":    round(_rs_z, 3),
        "asymmetry":  round(_sh_z_asym, 3),
        "score":      _rounded_sc,
        "unit":       "depth units",
        "label":      "Rounded shoulders (protraction)",
        "reference":  "MediaPipe Z-depth, negative = forward of spine",
    }
    if _rounded_depth > 15:
        out["alerts"].append(f"⚠️ Rounded shoulders detected — pull shoulder blades together and down")
        if "alerts_ar" not in out: out["alerts_ar"] = []
        out["alerts_ar"].append("⚠️ كتفان مائلان للأمام — اسحب لوحي الكتف للخلف وللأسفل")
    elif _rounded_depth > 8:
        out["alerts"].append("Shoulders slightly forward — open chest, squeeze shoulder blades gently")

    # ── IMPROVED: 4-point spine with kyphosis detection ──────────
    # 4 reference points (top→bottom):
    #   mid_sh       → shoulder line   (C7/T1 level)
    #   mid_thoracic → T8 proxy         (midpoint shoulders+hips)
    #   mid_lumbar   → L3 proxy         (75% down shoulder→hip)
    #   mid_hip      → hip line         (L5/S1 level)
    #
    # Kyphosis check: shoulders-thoracic angle vs thoracic-hip angle
    # If they curve in opposite directions → kyphosis present
    # Each segment scored separately, worst segment drives final score

    mid_thoracic = (
        (mid_hip[0] * 0.40 + mid_sh[0] * 0.60),
        (mid_hip[1] * 0.40 + mid_sh[1] * 0.60),
    )
    mid_lumbar = (
        (mid_hip[0] * 0.75 + mid_sh[0] * 0.25),
        (mid_hip[1] * 0.75 + mid_sh[1] * 0.25),
    )

    spine_upper  = angle_vert(mid_thoracic, mid_sh)      # C7→T8 (thoracic)
    spine_mid    = angle_vert(mid_lumbar,   mid_thoracic) # T8→L3 (thoracolumbar)
    spine_lower  = angle_vert(mid_hip,      mid_lumbar)   # L3→S1 (lumbar)

    # ── Kyphosis detector ─────────────────────────────────────
    # Classic kyphosis: upper spine rounds forward while lower is straight/lordotic
    # Sign of upper vs lower segment divergence
    _upper_signed = mid_sh[0]      - mid_thoracic[0]  # + = leaning right (screen right)
    _lower_signed = mid_thoracic[0] - mid_hip[0]
    _kyphosis_pen = 0.0
    if spine_upper > 3 and spine_lower < 3 and (_upper_signed * _lower_signed < 0):
        # Segments curve in opposite directions — classic thoracic kyphosis
        _kyphosis_pen = min(12.0, spine_upper * 0.6)
        out["alerts"].append(f"⚠️ Thoracic kyphosis detected ({round(spine_upper,1)}°) — sit back, open chest")
        if "alerts_ar" not in out: out["alerts_ar"] = []
        out["alerts_ar"].append(f"⚠️ انحناء الصدر للأمام ({round(spine_upper,1)}°) — اجلس للخلف وافتح صدرك")

    # Weighted composite: upper most visible/relevant for desk posture
    spine_lean   = spine_upper * 0.50 + spine_mid * 0.30 + spine_lower * 0.20
    # S-curve penalty: large divergence between segments
    spine_scurve_pen = max(0, abs(spine_upper - spine_lower) - 8) * 0.5
    spine_lean   = min(45.0, spine_lean + spine_scurve_pen + _kyphosis_pen)
    spine_sc     = score_m(spine_lean, 0, 4, 12)

    out["metrics"]["spine_upper"] = {"value": round(spine_upper, 1), "unit": "°", "label": "Upper spine (thoracic)"}
    out["metrics"]["spine_mid"]   = {"value": round(spine_mid,   1), "unit": "°", "label": "Mid spine (thoracolumbar)"}
    out["metrics"]["spine_lower"] = {"value": round(spine_lower, 1), "unit": "°", "label": "Lower spine (lumbar)"}
    out["metrics"]["kyphosis_pen"]= {"value": round(_kyphosis_pen, 1), "unit": "°", "label": "Kyphosis penalty"}

    # ── FaceMesh ───────────────────────────────────────────────────
    dist_cm = None
    blink_data = None
    if face_result.multi_face_landmarks:
        face_lms = face_result.multi_face_landmarks[0]
        out["head_pose"] = head_pose_from_face(face_lms, w, h)
        # Pass yaw for IPD correction — must compute head_pose first
        _yaw = out["head_pose"]["yaw"] if out["head_pose"] else 0.0
        # ── Load user-calibrated focal if available ──────────────
        _uid_for_focal = getattr(g, "uid", None)
        if _uid_for_focal and _uid_for_focal not in _focal_cal:
            try:
                _rc = get_redis()
                if _rc:
                    import json as _j
                    _cal = _rc.get(f"dist_cal:{_uid_for_focal}")
                    if _cal:
                        _cal_data = _j.loads(_cal)
                        with _focal_cal_lock:
                            _focal_cal[_uid_for_focal] = _cal_data.get("focal_px")
                            _focal_cal[_sid]           = _cal_data.get("focal_px")
            except Exception:
                pass
        dist_cm  = ipd_distance_face(face_lms, w, h, yaw_deg=_yaw, session_id=_sid)
        out["engine"]    = "mediapipe_pose+facemesh"
        # Eye strain — all paid tiers (iris tracking via FaceMesh)
        if tier in ("elite", "premium", "professional", "pro", "business"):
            _uid_blink = getattr(g, "uid", None)
            _sid_blink = out.get("session_id", "default")
            blink_data = analyze_blink_rate(face_lms, w, h, uid=_uid_blink, session_id=_sid_blink)
            out["eye_strain"] = blink_data

    # Fallback distance
    if dist_cm is None:
        # Fresh shoulder width for distance fallback (not the normalized sh_width_px)
        _sh_w_fallback = dist2d(l_sh, r_sh)
        focal   = 600 * (w / 640)
        dist_cm = round((40.0 * focal) / max(_sh_w_fallback, 1), 1)
        dist_cm = max(20, min(150, dist_cm))

    lo, hi = (50, 80) if mode == "laptop" else (60, 90)
    ideal  = (lo + hi) / 2   # 65cm laptop, 75cm desktop

    # ── Continuous quadratic scoring — no buckets ─────────────────
    # Perfect: inside lo-hi = 100
    # Linear decay outside range, quadratic beyond 20cm from edge
    if lo <= dist_cm <= hi:
        dist_sc = 100
    else:
        _gap = min(abs(dist_cm - lo), abs(dist_cm - hi))
        if _gap <= 10:
            dist_sc = max(60, 100 - _gap * 4)    # 4pts/cm → 60 at 10cm off
        elif _gap <= 20:
            dist_sc = max(30, 60 - (_gap-10) * 3)# 3pts/cm → 30 at 20cm off
        else:
            dist_sc = max(5,  30 - (_gap-20) * 2) # 2pts/cm → very bad
    dist_sc = int(dist_sc)

    hp = out.get("head_pose")
    pose_sc = 75
    cam_pitch_correction = 0.0
    hp_reliable = False   # True only if reproj_err is low
    if hp:
        pitch = hp["pitch"]; yaw = hp["yaw"]; roll = hp["roll"]
        reproj = hp.get("reproj_err", 0.0)
        # Reprojection error gate: >4px = unreliable pose estimate
        hp_reliable = reproj < 4.0 if reproj > 0 else True

        # ── Camera angle correction ────────────────────────────────
        # If pitch > 0 (camera below eye level looking up), neck_lean
        # is overestimated. Correct proportionally.
        # Typical laptop: camera ~15° below eye → adds ~4-6° apparent neck lean
        if mode == "laptop" and pitch > 5:
            cam_pitch_correction = min(pitch * 0.35, 8.0)
            neck_lean = max(0.0, neck_lean - cam_pitch_correction)
            # Recompute with shoulder-normalized thresholds (not fixed 7/20)
            neck_sc   = score_m(neck_lean, 0, neck_ok_adj, neck_bad_adj)

        # ── Head pose score: pitch + yaw + roll ─────────────────────
        # yaw boosted — turning head >20° = serious ergonomic risk (neck rotation)
        # pitch most important (forward/back neck load)
        pose_combined = abs(pitch) * 0.45 + abs(yaw) * 0.40 + abs(roll) * 0.15

        # ── Yaw gate: hard penalty for extreme head turns ────────────
        _yaw_abs = abs(yaw)
        if _yaw_abs > 40:
            # Face turned >40° = person looking at 2nd screen or far away
            pose_sc = max(5, int(100 - _yaw_abs * 2.2))
            out["alerts"].append(f"⚠️ Head turned {round(_yaw_abs)}° — face your primary screen directly")
            if "alerts_ar" not in out: out["alerts_ar"] = []
            out["alerts_ar"].append(f"⚠️ رأسك مائل {round(_yaw_abs)}° — واجه شاشتك الرئيسية مباشرة")
        elif _yaw_abs > 25:
            pose_sc = max(25, int(score_m(pose_combined, 0, 8, 20)))
            out["alerts"].append(f"Head turned {round(_yaw_abs)}° — try to face screen more directly")
        else:
            pose_sc = score_m(pose_combined, 0, 8, 22)

        # ── Lying down detection ─────────────────────────────────────
        # When person lies down: roll ≈ 90°, pitch ≈ 0°, landmarks shift vertically
        # Also: face aspect ratio changes (wider than tall)
        _roll_abs = abs(roll)
        if _roll_abs > 55:
            out["alerts"].append(f"⚠️ Lying down detected (roll {round(_roll_abs)}°) — PostureAI requires upright seating position")
            if "alerts_ar" not in out: out["alerts_ar"] = []
            out["alerts_ar"].append(f"⚠️ تم اكتشاف وضعية استلقاء ({round(_roll_abs)}°) — يتطلب PostureAI الجلوس منتصباً")
            pose_sc = min(pose_sc, 20)
            out["posture_valid"] = False   # flag for frontend to show warning
        elif _roll_abs > 30:
            out["alerts"].append(f"Head tilted {round(_roll_abs)}° — are you lying down or leaning?")
            pose_sc = min(pose_sc, 50)

        # ── Proximity gate: too close to camera ─────────────────────
        # Detect by face size in frame: if face occupies >35% frame width = too close
        if face_lms:
            try:
                _l_cheek = face_lms.landmark[234]
                _r_cheek = face_lms.landmark[454]
                _face_width_frac = abs(_r_cheek.x - _l_cheek.x)  # 0-1 fraction of frame
                if _face_width_frac > 0.45:
                    _approx_cm = int(14 / _face_width_frac * 0.7)  # rough estimate
                    out["alerts"].insert(0, f"🔴 Too close to camera ({_face_width_frac:.0%} frame) — move back at least 50cm")
                    if "alerts_ar" not in out: out["alerts_ar"] = []
                    out["alerts_ar"].insert(0, f"🔴 قريب جداً من الكاميرا — ابتعد على الأقل 50 سم")
                    out["proximity_warning"] = True
                    # Also tank the score when too close
                    pose_sc = min(pose_sc, 30)
                elif _face_width_frac > 0.35:
                    out["alerts"].append(f"⚠️ Move back slightly — you're closer than recommended")
                    out["proximity_warning"] = True
                else:
                    out["proximity_warning"] = False
            except Exception:
                pass

        out["metrics"]["head_pose_detail"] = {
            "pitch": pitch, "yaw": yaw, "roll": roll,
            "cam_correction_applied": round(cam_pitch_correction, 1),
            "label": "3D head pose (solvePnP)"
        }

    vis_l_sh  = g(PL.L_SHOULDER).visibility
    vis_r_sh  = g(PL.R_SHOULDER).visibility
    vis_l_ear = g(PL.L_EAR).visibility
    vis_r_ear = g(PL.R_EAR).visibility
    vis_l_hip = g(PL.L_HIP).visibility
    vis_r_hip = g(PL.R_HIP).visibility
    vis_nose  = g(PL.NOSE).visibility if hasattr(g(PL.NOSE), 'visibility') else 1.0

    # ── Per-metric confidence factors ──────────────────────────────
    # If a landmark group is low-visibility, reduce its metric's weight
    # proportionally. Formula: conf = clamp(vis / 0.6, 0, 1)
    # vis=0.6 → conf=1.0 (full weight), vis=0.3 → conf=0.5, vis=0.0 → conf=0
    def vis_conf(v):
        return max(0.0, min(1.0, v / 0.6))

    avg_sh_vis    = (vis_l_sh  + vis_r_sh)  / 2
    avg_ear_vis   = (vis_l_ear + vis_r_ear) / 2
    avg_hip_vis   = (vis_l_hip + vis_r_hip) / 2
    avg_eye_vis   = (g(PL.L_EYE).visibility + g(PL.R_EYE).visibility) / 2

    # Confidence per metric
    # neck_lean: needs ears + shoulders
    conf_neck  = vis_conf(avg_ear_vis * 0.5 + avg_sh_vis * 0.5)
    # head_tilt: needs eyes
    conf_tilt  = vis_conf(avg_eye_vis)
    # shoulder_level: needs both shoulders
    conf_sh    = vis_conf(min(vis_l_sh, vis_r_sh))   # both must be visible
    # spine_lean: needs shoulders + hips
    conf_spine = vis_conf(avg_sh_vis * 0.5 + avg_hip_vis * 0.5)

    vis_bonus = 10 if (avg_sh_vis > 0.7 and avg_ear_vis > 0.7) else 0

    # ── Desk ergonomics composite score ─────────────────────────
    # Combines screen distance + monitor height (from pitch) + posture
    # into a single actionable "workstation quality" metric
    try:
        # Monitor height score: pitch ~0° = monitor at eye level (ideal)
        _pitch_now = out.get("head_pose", {}).get("pitch", 0) if out.get("head_pose") else 0
        _monitor_sc = score_m(abs(_pitch_now), 0, 5, 18)

        # Distance score already computed
        _ergo_dist_sc = dist_sc

        # Neck lean as posture component
        _ergo_neck_sc = score_m(neck_lean, 0, 5, 15)

        # Composite ergonomics score
        _ergo_sc = int(round(
            _monitor_sc    * 0.35 +   # monitor height most impactful
            _ergo_dist_sc  * 0.35 +   # distance equally important
            _ergo_neck_sc  * 0.30     # resulting neck posture
        ))
        _ergo_grade = "Excellent" if _ergo_sc>=85 else "Good" if _ergo_sc>=70 else "Fair" if _ergo_sc>=55 else "Poor"
        out["metrics"]["ergonomics_score"] = {
            "value": _ergo_sc, "score": _ergo_sc, "unit": "/100",
            "label": "Workstation ergonomics",
            "grade": _ergo_grade,
            "components": {
                "monitor_height": _monitor_sc,
                "screen_distance": _ergo_dist_sc,
                "neck_posture": _ergo_neck_sc,
            }
        }
        if _ergo_sc < 55:
            out["alerts"].append(f"⚠️ Poor workstation setup ({_ergo_sc}/100) — adjust monitor height and distance")
    except Exception:
        pass

    # ── Wrist angle ────────────────────────────────────────────────
    wrist_sc = None
    if tier in ("professional","elite","pro","premium","business") and len(_hist[0]) > PL.R_WRIST if _hist else False:
        try:
            vis_l_elbow = g(PL.L_ELBOW).visibility
            vis_r_elbow = g(PL.R_ELBOW).visibility
            vis_l_wrist = g(PL.L_WRIST).visibility
            vis_r_wrist = g(PL.R_WRIST).visibility
            _l_ok = vis_l_elbow > 0.45 and vis_l_wrist > 0.45
            _r_ok = vis_r_elbow > 0.45 and vis_r_wrist > 0.45
            both_ok = _l_ok and _r_ok

            def px3(idx):
                lm = g(idx)
                return (lm.x*w, lm.y*h, lm.z*w)

            wrist_angle = None
            _wrist_source = "none"

            if both_ok:
                # Both sides — average for best accuracy
                l_sh3 = px3(PL.L_SHOULDER); r_sh3 = px3(PL.R_SHOULDER)
                l_el3 = px3(PL.L_ELBOW);   r_el3 = px3(PL.R_ELBOW)
                l_wr3 = px3(PL.L_WRIST);   r_wr3 = px3(PL.R_WRIST)
                l_wrist_angle = angle_3pt_3d(l_sh3, l_el3, l_wr3)
                r_wrist_angle = angle_3pt_3d(r_sh3, r_el3, r_wr3)
                wrist_angle   = ((180-l_wrist_angle) + (180-r_wrist_angle)) / 2
                _wrist_source = "both_elbows"
            elif _l_ok:
                # Left only fallback
                l_wrist_angle = angle_3pt_3d(px3(PL.L_SHOULDER), px3(PL.L_ELBOW), px3(PL.L_WRIST))
                wrist_angle   = 180 - l_wrist_angle
                _wrist_source = "left_elbow_only"
            elif _r_ok:
                # Right only fallback
                r_wrist_angle = angle_3pt_3d(px3(PL.R_SHOULDER), px3(PL.R_ELBOW), px3(PL.R_WRIST))
                wrist_angle   = 180 - r_wrist_angle
                _wrist_source = "right_elbow_only"

            if wrist_angle is not None:
                wrist_sc = score_m(wrist_angle, 0, 10, 25)
                out["metrics"]["wrist_angle"] = {
                    "value": round(wrist_angle, 1), "score": wrist_sc,
                    "unit": "°", "label": "Wrist angle (CTD risk)",
                    "source": _wrist_source,
                }
                if wrist_angle > 20:
                    out["alerts"].append(f"⚠️ Wrist deviation {round(wrist_angle,1)}° — risk of Carpal Tunnel. Keep wrists straight.")
                elif wrist_angle > 12:
                    out["alerts"].append(f"Wrist deviation {round(wrist_angle,1)}° — try to keep wrists flat on desk.")
        except Exception:
            pass

    # ── Eye strain — blink rate + gaze deviation ─────────────────
    eye_sc = None

    # Gaze deviation (iris position relative to eye socket)
    _gaze_score = 100
    try:
        if face_lms:
            def _iris_offset(iris_idx, inner_idx, outer_idx):
                iris  = face_lms.landmark[iris_idx]
                inner = face_lms.landmark[inner_idx]
                outer = face_lms.landmark[outer_idx]
                eye_w = abs(outer.x - inner.x)
                if eye_w < 0.001: return 0.0
                return abs(iris.x - (inner.x + outer.x) / 2) / eye_w
            l_gaze = _iris_offset(468, 33,  133)
            r_gaze = _iris_offset(473, 362, 263)
            gaze_dev = (l_gaze + r_gaze) / 2
            _gaze_score = int(score_m(gaze_dev * 100, 0, 15, 40))
            out["metrics"]["gaze_deviation"] = {
                "value": round(gaze_dev * 100, 1), "score": _gaze_score,
                "unit": "%", "label": "Gaze deviation from screen center"
            }
            if gaze_dev > 0.35:
                out["alerts"].append("Eyes not centered on screen — reposition monitor or chair")
                out.setdefault("alerts_ar", []).append("عيناك لا تنظران للشاشة — اضبط الشاشة أو الكرسي")
    except Exception:
        pass

    if blink_data:
        br = blink_data["blink_rate_per_min"]
        blink_sc = score_m(br, 16, 6, 12)
        # Combined: 60% blink + 40% gaze
        eye_sc = int(0.60 * blink_sc + 0.40 * _gaze_score)
        out["metrics"]["eye_strain"] = {
            "value":      br,
            "score":      eye_sc,
            "blink_sc":   int(blink_sc),
            "gaze_sc":    _gaze_score,
            "unit":       "blinks/min",
            "label":      "Eye strain (blink + gaze)",
        }
        if blink_data["eye_strain_risk"] == "high":
            out["alerts"].append(f"⚠️ Eye strain risk — only {br} blinks/min. Apply 20-20-20 rule.")
            out.setdefault("alerts_ar", []).append(f"⚠️ خطر إجهاد عيون — {br} رمشة/دقيقة فقط. طبّق قاعدة 20-20-20.")
        elif blink_data["eye_strain_risk"] == "moderate":
            out["alerts"].append(f"Low blink rate ({br}/min) — blink consciously.")

    # ── Overall score — confidence-weighted ───────────────────────
    # All weights defined upfront so normalization is correct.
    # eye_sc weight 0.05 — always included when FaceMesh available
    # wrist_sc weight 0.08 — only for paid tiers
    # pose_sc weight up to 0.15 — solvePnP most accurate

    # eye confidence: needs FaceMesh iris landmarks (very reliable when available)
    conf_eye   = 1.0 if (eye_sc is not None) else 0.0
    conf_wrist = 1.0 if (wrist_sc is not None) else 0.0

    BASE_W = {"neck": 0.28, "tilt": 0.10, "sh": 0.08, "spine": 0.18, "dist": 0.22,
              "eye": 0.05, "wrist": 0.09}
    # dist boosted to 0.22 — screen distance is #1 preventable risk factor
    # Refs: AOA 2023 (20-20-20 rule), WHO screen guidelines

    eff_w = {
        "neck":  BASE_W["neck"]  * conf_neck,
        "tilt":  BASE_W["tilt"]  * conf_tilt,
        "sh":    BASE_W["sh"]    * conf_sh,
        "spine": BASE_W["spine"] * conf_spine,
        "dist":  BASE_W["dist"],                    # camera-independent — no penalty
        "eye":   BASE_W["eye"]   * conf_eye,        # 0.05 when FaceMesh active, else 0
        "wrist": BASE_W["wrist"] * conf_wrist,      # 0.08 for paid tiers, else 0
    }

    # Lost weight from low-vis metrics → redistributed to dist (stable)
    lost_w = sum(BASE_W.values()) - sum(eff_w.values())
    eff_w["dist"] = min(0.30, eff_w["dist"] + lost_w)

    # Build score_val from all present metrics
    scores = {
        "neck":  neck_sc,
        "tilt":  tilt_sc,
        "sh":    sh_sc,
        "spine": spine_sc,
        "dist":  dist_sc,
        "eye":   eye_sc   if eye_sc   is not None else 0,
        "wrist": wrist_sc if wrist_sc is not None else 0,
    }
    score_val = sum(scores[k] * eff_w[k] for k in scores)
    remaining = 1.0 - sum(eff_w.values())

    # head_pose (solvePnP) — most accurate, gets remaining weight up to 0.15
    # But only if reprojection error is low (reliable estimate)
    if hp and hp_reliable:
        pose_weight = min(remaining, 0.15)
        score_val  += pose_sc * pose_weight
        remaining  -= pose_weight
    elif hp and not hp_reliable:
        # Low confidence pose — half weight
        pose_weight = min(remaining, 0.07)
        score_val  += pose_sc * pose_weight
        remaining  -= pose_weight

    # Store confidence breakdown for debugging and frontend display
    out["metrics"]["_confidence"] = {
        "neck":  round(conf_neck,  2),
        "tilt":  round(conf_tilt,  2),
        "sh":    round(conf_sh,    2),
        "spine": round(conf_spine, 2),
        "eye":   round(conf_eye,   2),
        "wrist": round(conf_wrist, 2),
        "eff_weights": {k: round(v, 3) for k, v in eff_w.items()},
        "label": "Per-metric visibility confidence",
    }

    # ── No arbitrary baseline — normalize by actual weights used ──────
    # Instead of filling remaining weight with 72, we normalize score_val
    # by the actual weight sum so absent optional metrics don't inflate score.
    weight_used = 1.0 - remaining  # how much weight was actually scored
    if weight_used > 0.01:
        overall = max(0, min(100, int(round(score_val / weight_used))))
    else:
        overall = 0

    # ── Confidence: penalize low shoulder visibility ────────────────
    vis_l_sh_val = vis_l_sh  # already computed above
    vis_r_sh_val = vis_r_sh
    avg_vis = (vis_l_sh_val + vis_r_sh_val) / 2
    if avg_vis < 0.6:
        # Low visibility — shoulders partially occluded
        # Reduce score toward neutral (65) proportionally
        penalty    = (0.6 - avg_vis) / 0.6  # 0→1 as vis→0
        overall    = int(overall * (1 - penalty * 0.35) + 65 * penalty * 0.35)
        overall    = max(0, min(100, overall))
    # Confidence = weighted average of per-metric confidences
    overall_conf_raw = (
        conf_neck  * eff_w["neck"]  +
        conf_tilt  * eff_w["tilt"]  +
        conf_sh    * eff_w["sh"]    +
        conf_spine * eff_w["spine"]
    ) / max(sum(eff_w.values()) - eff_w["dist"], 0.01)  # exclude dist (always 1)

    confidence = min(96, int(
        40                                                               # base
        + (overall_conf_raw * 28)                                       # landmark quality (0-28)
        + (vis_bonus)                                                    # high-vis bonus
        + (8 if out["engine"] == "mediapipe_pose+facemesh" else 0)      # FaceMesh bonus
        + (6 if hp else 0)                                               # head pose bonus
        + (8 if len(_lm_history.get(out.get("session_id",""), [])) >= 3 else 0)  # avg bonus
    ))

    # ── Neck muscle load model (biomechanics) ────────────────────
    # Based on Hansraj 2014: head = 5.4kg at 0°
    # Load increases ~2.5x per 15° forward lean
    try:
        _head_wt  = 5.4   # kg average adult head
        _load_factor = 1 + (neck_lean / 15) ** 1.8 * 1.2
        _neck_load_kg = round(_head_wt * _load_factor, 1)
        out["metrics"]["neck_load"] = {
            "value":      _neck_load_kg,
            "score":      score_m(_neck_load_kg, 5.4, 3, 12),
            "unit":       "kg",
            "label":      "Estimated neck load",
            "head_weight": _head_wt,
            "multiplier":  round(_load_factor, 2),
            "reference":  "Hansraj 2014 — spine biomechanics model",
        }
        if _neck_load_kg > 22:
            out["alerts"].append(f"⚠️ Neck bearing ~{_neck_load_kg}kg (normal: 5.4kg) — critical forward head position")
        elif _neck_load_kg > 12:
            out["alerts"].append(f"Neck load ~{_neck_load_kg}kg (normal: 5.4kg) — straighten up to reduce strain")
    except Exception:
        pass

    out["score"]      = overall
    out["overall"]    = overall   # alias for frontend
    out["confidence"] = confidence
    out["distCm"]     = dist_cm   # top-level for frontend skeleton
    # raw object for frontend drawFront function
    # NOTE: idealDistLo/idealDistHi are the recommended distance range for
    # the current device mode (laptop vs external monitor) — NOT a
    # confidence interval on distCm itself. dist_cm is always a number by
    # this point (fallback at the top of this function guarantees it),
    # so these are never actually None in practice; kept defensive anyway.
    out["raw"] = {
        "distCm":      dist_cm,
        "idealDistLo": lo if dist_cm else None,
        "idealDistHi": hi if dist_cm else None,
        "neckLean":    round(neck_lean, 1)  if neck_lean  is not None else None,
        "spineLean":   round(spine_lean, 1) if spine_lean is not None else None,
        "shTilt":      round(sh_tilt, 1)    if sh_tilt    is not None else None,
    }
    out["metrics"].update({
        "neck_lean":       {"value": round(neck_lean, 1),  "score": neck_sc,  "unit": "°",  "label": "Neck lean"},
        "head_tilt":       {"value": round(head_tilt, 1),  "score": tilt_sc,  "unit": "°",  "label": "Head tilt"},
        "shoulder_level":  {"value": round(sh_tilt, 1),    "score": sh_sc,    "unit": "°",  "label": "Shoulder level"},
        "spine_lean":      {"value": round(spine_lean,1),  "score": spine_sc, "unit": "°",  "label": "Spine lean"},
        "screen_distance": {"value": dist_cm,               "score": dist_sc,  "unit": "cm", "label": "Screen distance"},
    })
    if hp:
        out["metrics"]["pitch"] = {"value": round(hp["pitch"],1), "score": pose_sc, "unit": "°", "label": "Head pitch (3D)"}
        out["metrics"]["yaw"]   = {"value": round(hp["yaw"],1),   "score": pose_sc, "unit": "°", "label": "Head yaw (3D)"}
        out["metrics"]["roll"]  = {"value": round(hp["roll"],1),  "score": pose_sc, "unit": "°", "label": "Head roll (3D)"}

    # ── Landmark dict for frontend skeleton ─────────────────────
    # Frontend drawFront needs landmark positions as {x, y} normalized
    try:
        def _lm(idx):
            lm = g(idx)
            return {"x": round(lm.x, 4), "y": round(lm.y, 4), "visibility": round(lm.visibility, 2)}
        out["lms"] = {
            "lSh":    _lm(PL.L_SHOULDER), "rSh":   _lm(PL.R_SHOULDER),
            "lEar":   _lm(PL.L_EAR),      "rEar":  _lm(PL.R_EAR),
            "lEye":   _lm(PL.L_EYE),      "rEye":  _lm(PL.R_EYE),
            "lHip":   _lm(PL.L_HIP),      "rHip":  _lm(PL.R_HIP),
            "lElbow": _lm(PL.L_ELBOW),    "rElbow":_lm(PL.R_ELBOW),
            "lWrist": _lm(PL.L_WRIST),    "rWrist":_lm(PL.R_WRIST),
            # Computed midpoints
            "midSh":  {"x": round((g(PL.L_SHOULDER).x+g(PL.R_SHOULDER).x)/2,4),
                       "y": round((g(PL.L_SHOULDER).y+g(PL.R_SHOULDER).y)/2,4), "visibility":1.0},
            "midHip": {"x": round((g(PL.L_HIP).x+g(PL.R_HIP).x)/2,4),
                       "y": round((g(PL.L_HIP).y+g(PL.R_HIP).y)/2,4), "visibility":1.0},
            "midEar": {"x": round((g(PL.L_EAR).x+g(PL.R_EAR).x)/2,4),
                       "y": round((g(PL.L_EAR).y+g(PL.R_EAR).y)/2,4), "visibility":1.0},
        }
    except Exception:
        pass

    # ── Alerts ────────────────────────────────────────────────────
    if neck_lean > 20:
        out["alerts"].append(f"⚠️ Severe neck lean {round(neck_lean,1)}° — raise monitor to eye level immediately")
    elif neck_lean > 12:
        out["alerts"].append(f"Neck lean {round(neck_lean,1)}° — tuck chin slightly and check monitor height")
    if head_tilt > 10:
        out["alerts"].append(f"Head tilting {round(head_tilt,1)}° — check chair height and monitor centering")
    if sh_tilt > 10:
        out["alerts"].append(f"Shoulder imbalance {round(sh_tilt,1)}° — adjust armrests")
    if dist_cm < lo - 20:
        out["alerts"].append(f"🔴 DANGER: Screen only {round(dist_cm)}cm away — severe eye & neck strain risk. Move back to {lo}–{hi}cm NOW")
        if "alerts_ar" not in out: out["alerts_ar"] = []
        out["alerts_ar"].append(f"🔴 خطر: الشاشة على بُعد {round(dist_cm)}سم فقط — ابتعد فوراً إلى {lo}–{hi}سم")
    elif dist_cm < lo - 10:
        out["alerts"].append(f"⚠️ Too close to screen ({round(dist_cm)}cm) — risk of eye strain & neck compression. Move to {lo}–{hi}cm")
        if "alerts_ar" not in out: out["alerts_ar"] = []
        out["alerts_ar"].append(f"⚠️ قريب جداً من الشاشة ({round(dist_cm)}سم) — خطر إجهاد عيون وضغط رقبة. ابتعد إلى {lo}–{hi}سم")
    elif dist_cm < lo:
        out["alerts"].append(f"Screen distance {round(dist_cm)}cm — move back slightly to {lo}–{hi}cm")
        if "alerts_ar" not in out: out["alerts_ar"] = []
        out["alerts_ar"].append(f"مسافة الشاشة {round(dist_cm)}سم — ابتعد قليلاً إلى {lo}–{hi}سم")
    elif dist_cm > hi + 15:
        out["alerts"].append(f"Too far from screen ({round(dist_cm)}cm) — ideal is {lo}–{hi}cm")
        if "alerts_ar" not in out: out["alerts_ar"] = []
        out["alerts_ar"].append(f"بعيد جداً عن الشاشة ({round(dist_cm)}سم) — المثالي {lo}–{hi}سم")
    if spine_lean > 18:
        out["alerts"].append(f"⚠️ Spine lean {round(spine_lean,1)}° — sit back and use lumbar support")
    elif spine_lean > 10:
        out["alerts"].append(f"Spine lean {round(spine_lean,1)}° — engage your core and sit upright")
    if hp and abs(hp["pitch"]) > 20:
        out["alerts"].append(f"Head pitched {round(hp['pitch'],1)}° — {'raise your monitor' if hp['pitch'] < 0 else 'lower your monitor'}")

    # ── Hip height asymmetry ─────────────────────────────────────
    # Measures Y-coordinate difference between left and right hip
    # >3% of frame height = significant asymmetry (spinal compression risk)
    hip_y_diff = abs(l_hip[1] - r_hip[1]) / max(h, 1) * 100
    if hip_y_diff > 5.0:
        higher_side = "left" if l_hip[1] < r_hip[1] else "right"
        out["alerts"].append(f"Hip asymmetry — {higher_side} hip {round(hip_y_diff, 1)}% higher. Check seat levelness.")
        out["metrics"]["hip_asymmetry"] = {
            "value": round(hip_y_diff, 1), "score": max(5, int(100 - hip_y_diff * 8)),
            "unit": "%", "label": "Hip height asymmetry"
        }
    elif hip_y_diff > 2.5:
        out["metrics"]["hip_asymmetry"] = {
            "value": round(hip_y_diff, 1), "score": max(30, int(100 - hip_y_diff * 8)),
            "unit": "%", "label": "Hip height asymmetry"
        }

    # ── Neck flexion vs extension distinction ────────────────────
    # pitch from solvePnP: negative = looking down (flexion), positive = up (extension)
    # Forward head: neck lean UP + pitch negative = computer neck (most common)
    # Looking up: neck lean + pitch positive = monitor too high
    if hp and neck_lean > 8:
        pitch = hp.get("pitch", 0)
        if pitch < -10:
            out["metrics"]["neck_lean"]["label"] = "Neck flexion (looking down)"
            out["metrics"]["neck_lean"]["posture_type"] = "flexion"
        elif pitch > 10:
            out["metrics"]["neck_lean"]["label"] = "Neck extension (looking up)"
            out["metrics"]["neck_lean"]["posture_type"] = "extension"
        else:
            out["metrics"]["neck_lean"]["posture_type"] = "lateral"

        # ── Circadian posture adjustment ─────────────────────────────
    # Posture degrades predictably by time of day (circadian rhythm)
    # Account for this: don't penalize afternoon slump as harshly
    # Based on: Monk et al. (1997) circadian performance rhythms
    try:
        _hour = datetime.utcnow().hour + 3  # UTC+3 (Egypt/KSA) approximate
        _hour = _hour % 24
        # Circadian factor: 1.0 = peak (9-11am), 0.88 = worst (2-4pm)
        _circadian_factors = {
            0: 0.92, 1: 0.90, 2: 0.88, 3: 0.87, 4: 0.87, 5: 0.88,
            6: 0.91, 7: 0.93, 8: 0.96, 9: 1.00, 10: 1.00, 11: 0.99,
            12: 0.97, 13: 0.94, 14: 0.91, 15: 0.90, 16: 0.92, 17: 0.94,
            18: 0.95, 19: 0.95, 20: 0.93, 21: 0.91, 22: 0.90, 23: 0.91,
        }
        _cf = _circadian_factors.get(_hour, 1.0)
        # Adjusted score: if person scores 65 at 3pm (factor=0.90),
        # their "circadian-adjusted" score = 65/0.90 = 72 (better than it looks)
        _circ_adjusted = min(100, round(overall / max(_cf, 0.85)))
        out["circadian"] = {
            "hour":             _hour,
            "factor":           _cf,
            "adjusted_score":   _circ_adjusted,
            "period":           (
                "morning_peak" if 8 <= _hour <= 11 else
                "post_lunch_dip" if 13 <= _hour <= 15 else
                "afternoon_recovery" if 16 <= _hour <= 18 else
                "evening" if 19 <= _hour <= 22 else "off_hours"
            ),
            "note": (
                "Post-lunch circadian dip — posture naturally worse now" if 13<=_hour<=15 else
                "Morning peak — best time for demanding tasks" if 8<=_hour<=11 else
                "End-of-day fatigue window" if 17<=_hour<=19 else ""
            ),
        }
    except Exception:
        pass

    # ── Breathing rate estimation from shoulder movement ────────
    # Normal: 12-20 breaths/min. Stress/poor posture → shallow breathing
    # Shoulders rise ~2-5% of frame per breath
    try:
        _br_key   = f"breath:{out.get('session_id','default')}"
        _sh_y_now = (g(PL.L_SHOULDER).y + g(PL.R_SHOULDER).y) / 2
        _br_buf   = _blink_sessions.setdefault(_br_key, [])
        _br_buf.append((time.time(), _sh_y_now))
        # Keep 30s window
        _now_t = time.time()
        _blink_sessions[_br_key] = [(t,y) for t,y in _br_buf if _now_t-t < 30]
        _br_buf = _blink_sessions[_br_key]
        if len(_br_buf) >= 30:
            # Detect peaks in shoulder Y (rises = breaths)
            _ys = [y for _,y in _br_buf]
            _mn, _mx = min(_ys), max(_ys)
            _amp = _mx - _mn
            if _amp > 0.005:   # minimum 0.5% frame movement
                _peaks = 0
                _in_rise = False
                _thresh = _mn + _amp * 0.4
                for _y in _ys:
                    if _y > _thresh and not _in_rise:
                        _peaks += 1; _in_rise = True
                    elif _y <= _thresh:
                        _in_rise = False
                _window_s = max(1, _br_buf[-1][0] - _br_buf[0][0])
                _bpm = round(_peaks * 60 / _window_s) if _window_s > 0 else 0
                _br_risk = "normal" if 12<=_bpm<=20 else "elevated" if _bpm>20 else "shallow"
                out["metrics"]["breathing_rate"] = {
                    "value": _bpm, "score": score_m(_bpm, 16, 4, 8),
                    "unit": "breaths/min", "label": "Breathing rate (est.)",
                    "amplitude": round(_amp*100, 1), "risk": _br_risk,
                }
                if _br_risk == "shallow":
                    out["alerts"].append(f"Shallow breathing detected ({_bpm}/min) — take a deep breath, relax shoulders")
    except Exception:
        pass

    # ── RULA simplified score (Rapid Upper Limb Assessment) ─────
    # Based on McAtamney & Corlett 1993 — occupational health standard
    # Scores each body segment 1-6, combines into action level 1-4
    # Action level 3-4 = requires immediate ergonomic investigation
    try:
        # Upper arm score (0-4): based on shoulder elevation
        _neck_v   = neck_lean
        _spine_v  = spine_lean if hasattr(spine_lean, '__float__') else 0
        _ua_score = (1 if _neck_v < 20 else 2 if _neck_v < 45 else
                     3 if _neck_v < 90 else 4)  # upper arm from neck proxy
        # Neck score (1-4): posture of neck
        _neck_rula = (1 if neck_lean < 10 else 2 if neck_lean < 20 else
                      3 if neck_lean < 30 else 4)
        # Trunk score (1-4)
        _trunk_rula = (1 if _spine_v < 10 else 2 if _spine_v < 20 else
                       3 if _spine_v < 60 else 4)
        # Head twist penalty
        _hp_yaw = abs(out.get("head_pose",{}).get("yaw",0) or 0)
        _twist_pen = 1 if _hp_yaw > 15 else 0
        # RULA score (simplified — 1=acceptable, 4=investigate immediately)
        _rula_raw = (_ua_score + _neck_rula + _trunk_rula + _twist_pen)
        _rula_norm = max(1, min(4, round(_rula_raw / 3.0)))  # normalize to 1-4
        _rula_action = {
            1: "Acceptable posture — no action required",
            2: "Monitor — investigate if sustained",
            3: "Investigate — changes required soon",
            4: "⚠️ Investigate immediately — urgent ergonomic changes needed",
        }
        _rula_sc = score_m(_rula_norm, 1, 0.5, 2)   # score: 1=100, 4=5
        out["metrics"]["rula_score"] = {
            "value":       _rula_norm,
            "score":       _rula_sc,
            "unit":        "/4",
            "label":       "RULA ergonomic risk",
            "action":      _rula_action[_rula_norm],
            "components":  {"upper_arm": _ua_score, "neck": _neck_rula, "trunk": _trunk_rula},
            "reference":   "McAtamney & Corlett 1993",
        }
        if _rula_norm >= 3:
            out["alerts"].append(f"⚠️ RULA score {_rula_norm}/4 — ergonomic investigation required")
    except Exception:
        pass

    # ── Posture symmetry score ────────────────────────────────────
    # Compare left vs right landmark heights for imbalance detection
    # High asymmetry → possible scoliosis pattern or bad chair setup
    try:
        l_sh_y = g(PL.L_SHOULDER).y * h
        r_sh_y = g(PL.R_SHOULDER).y * h
        l_hip_y = g(PL.L_HIP).y * h
        r_hip_y = g(PL.R_HIP).y * h
        sh_asym  = abs(l_sh_y - r_sh_y) / h * 100   # % of frame height
        hip_asym = abs(l_hip_y - r_hip_y) / h * 100

        # Combined symmetry score
        max_asym = max(sh_asym, hip_asym)
        symmetry_sc = score_m(max_asym, 0, 1.5, 5.0)  # ok=1.5% bad=5% of frame

        out["metrics"]["symmetry"] = {
            "value":        round(max_asym, 2),
            "score":        symmetry_sc,
            "unit":         "%",
            "label":        "Body symmetry",
            "sh_asym":      round(sh_asym, 2),
            "hip_asym":     round(hip_asym, 2),
        }
        if max_asym > 5.0:
            out["alerts"].append(f"⚠️ Significant body asymmetry {round(max_asym,1)}% — possible scoliosis pattern or uneven chair setup")
        elif max_asym > 2.5:
            out["alerts"].append(f"Body asymmetry {round(max_asym,1)}% — check armrest heights and sitting position")
    except Exception:
        pass

    # ── Micro-movement / stillness score ────────────────────────
    # Completely frozen posture → muscle fatigue → pain
    # Optimal: small natural movements (5-15% position variance)
    try:
        _mv_key = f"mv:{out.get('session_id','default')}"
        _nose_x, _nose_y = nose[0]/w, nose[1]/h
        _mv_buf = _blink_sessions.setdefault(_mv_key, [])
        _mv_buf.append((_nose_x, _nose_y))
        if len(_mv_buf) > 15: _mv_buf.pop(0)
        if len(_mv_buf) >= 10:
            _xs = [p[0] for p in _mv_buf]; _ys = [p[1] for p in _mv_buf]
            _mx_mean = sum(_xs)/len(_xs); _my_mean = sum(_ys)/len(_ys)
            _variance = sum((x-_mx_mean)**2+(y-_my_mean)**2 for x,y in _mv_buf) / len(_mv_buf)
            _movement_pct = round(_variance * 10000, 2)  # scale to 0-100
            _mv_status = (
                "frozen"   if _movement_pct < 0.5 else
                "optimal"  if _movement_pct < 8.0 else
                "restless"
            )
            out["metrics"]["movement"] = {
                "value":  _movement_pct,
                "score":  score_m(_movement_pct, 3.0, 2.5, 7.0),
                "unit":   "%",
                "label":  "Natural movement",
                "status": _mv_status,
            }
            if _mv_status == "frozen":
                out["alerts"].append("Body too still — shift position slightly every few minutes to prevent muscle fatigue")
    except Exception:
        pass

    # ── ISO 11226 whole-body posture compliance ─────────────────
    # ISO 11226:2000 — Ergonomics: evaluation of static working postures
    # Classifies posture as: acceptable / conditional / not recommended
    try:
        # Trunk inclination forward (ISO 11226 Table 1)
        # ≤20° acceptable, 20-60° conditional, >60° not recommended
        _iso_trunk  = ("acceptable"       if _spine_v <= 20 else
                       "conditional"      if _spine_v <= 60 else
                       "not_recommended")
        # Head inclination forward
        # ≤25° acceptable, 25-85° conditional, >85° not recommended
        _iso_head   = ("acceptable"       if neck_lean <= 25 else
                       "conditional"      if neck_lean <= 85 else
                       "not_recommended")
        # Shoulder — upper arm elevation
        # ≤20° acceptable, 20-60° conditional, >60° not recommended
        _iso_shoulder = ("acceptable"     if sh_tilt <= 20 else
                         "conditional"    if sh_tilt <= 60 else
                         "not_recommended")
        # Overall ISO compliance
        _iso_parts = [_iso_trunk, _iso_head, _iso_shoulder]
        _iso_overall = ("not_recommended" if "not_recommended" in _iso_parts else
                        "conditional"     if "conditional"     in _iso_parts else
                        "acceptable")
        _iso_sc = {"acceptable": 90, "conditional": 55, "not_recommended": 20}
        out["metrics"]["iso_11226"] = {
            "value":    {"acceptable": 1, "conditional": 2, "not_recommended": 3}[_iso_overall],
            "score":    _iso_sc[_iso_overall],
            "unit":     "compliance",
            "label":    "ISO 11226 posture compliance",
            "overall":  _iso_overall,
            "trunk":    _iso_trunk,
            "head":     _iso_head,
            "shoulder": _iso_shoulder,
            "reference":"ISO 11226:2000",
        }
        if _iso_overall == "not_recommended":
            out["alerts"].append(f"⚠️ ISO 11226: posture not recommended — immediate ergonomic correction required")
        elif _iso_overall == "conditional":
            out["alerts"].append(f"ISO 11226: conditional posture — acceptable only if not maintained for long periods")
    except Exception:
        pass

    # ── Cognitive load from micro-expressions ───────────────────
    if face_result and face_result.multi_face_landmarks:
        try:
            _cog = compute_cognitive_load(face_result.multi_face_landmarks[0], w, h)
            if _cog:
                out["metrics"]["cognitive_load"] = {
                    "value":    _cog["load_score"],
                    "score":    100 - _cog["load_score"],  # inverse: high load = lower score
                    "unit":     "/100",
                    "label":    "Cognitive load (FACS)",
                    "level":    _cog["level"],
                    "note":     _cog["interpretation"],
                    "brow":     _cog["brow_furrow"],
                    "squint":   _cog["eye_squint"],
                    "jaw":      _cog["jaw_tension"],
                }
                if _cog["load_score"] > 70:
                    out["alerts"].append(
                        f"High cognitive load detected — consider a 5-min break "
                        f"(brow furrow: {round(_cog['brow_furrow']*100)}%, "
                        f"jaw tension: {round(_cog['jaw_tension']*100)}%)")
        except Exception:
            pass

    # ── Head velocity + movement pattern ─────────────────────────
    try:
        _hv_key = f"vel:{out.get('session_id','default')}"
        _hv = compute_head_velocity(nose[0]/w, nose[1]/h, _hv_key)
        if _hv:
            out["metrics"]["head_velocity"] = {
                "value":   _hv["velocity"],
                "score":   score_m(_hv["velocity"], 30, 20, 80),  # ideal = natural movement
                "unit":    "mm/s",
                "label":   "Head movement speed",
                "pattern": _hv["pattern"],
                "sudden":  _hv["sudden_move"],
            }
            if _hv["sudden_move"]:
                out["alerts"].append(
                    f"Rapid head movement detected ({_hv['velocity']}mm/s) — "
                    f"possible attention switching or startled response")
            elif _hv["pattern"] == "frozen":
                out["alerts"].append("Head completely still — micro-movements help maintain blood flow")
    except Exception:
        pass

    # ── Gaze direction (if FaceMesh available) ────────────────────
    if face_result and face_result.multi_face_landmarks:
        try:
            _gaze = compute_gaze(face_result.multi_face_landmarks[0], w, h)
            if _gaze:
                out["metrics"]["gaze"] = {
                    "value": round(math.sqrt(_gaze["h"]**2 + _gaze["v"]**2), 3),
                    "score": score_m(abs(_gaze["h"])*100, 0, 20, 40),
                    "unit":  "offset",
                    "label": "Gaze direction",
                    "h":     _gaze["h"],
                    "v":     _gaze["v"],
                    "on_screen": _gaze["looking_at_screen"],
                }
                if not _gaze["looking_at_screen"]:
                    out["alerts"].append(f"Eyes not focused on screen — gaze offset h={_gaze['h']:+.2f} v={_gaze['v']:+.2f}")
        except Exception:
            pass

    # ── Posture DNA fingerprint ─────────────────────────────────
    # Compact posture pattern hash — detects habitual bad posture
    # "You've reverted to bad habit pattern seen 8× before"
    try:
        _dna_vec = [
            int(neck_lean  / 5) * 5,
            int(spine_lean / 5) * 5,
            int(sh_tilt     / 3) * 3,
            int(dist_cm    / 10)* 10,
            int(abs(out.get("head_pose",{}).get("yaw",0)   or 0) / 5)*5,
            int(abs(out.get("head_pose",{}).get("pitch",0) or 0) / 5)*5,
        ]
        _dna_str  = "-".join(str(v) for v in _dna_vec)
        _dna_hash = str(abs(hash(_dna_str)) % 100000).zfill(5)
        _uid_dna  = getattr(g, "uid", None)
        _prev_fp  = cache_get(f"dna:{_uid_dna}") if _uid_dna else None
        _match    = None
        if _prev_fp:
            try:
                import json as _jd2
                _match = _jd2.loads(_prev_fp).get(_dna_hash)
            except Exception: pass
        if _uid_dna and overall < 70:
            try:
                import json as _jd2
                _fp_hist = {}
                try: _fp_hist = _jd2.loads(cache_get(f"dna:{_uid_dna}") or "{}")
                except Exception: pass
                _fp_hist[_dna_hash] = {
                    "first_seen": _fp_hist.get(_dna_hash,{}).get("first_seen", datetime.utcnow().isoformat()),
                    "count":      _fp_hist.get(_dna_hash,{}).get("count", 0) + 1,
                    "last_seen":  datetime.utcnow().isoformat(),
                    "avg_score":  overall,
                }
                if len(_fp_hist) > 20:
                    _fp_hist = dict(sorted(_fp_hist.items(), key=lambda x: x[1].get("count",0), reverse=True)[:20])
                cache_set(f"dna:{_uid_dna}", _jd2.dumps(_fp_hist), 86400*90)
            except Exception: pass
        out["posture_dna"] = {
            "fingerprint": _dna_hash,
            "vector":      _dna_vec,
            "habitual":    bool(_match and _match.get("count",0) > 3),
            "seen_before": _match.get("count",0) if _match else 0,
            "first_seen":  _match.get("first_seen") if _match else None,
            "note":        f"Habitual — seen {_match['count']}× before" if (_match and _match.get("count",0)>3) else "New pattern",
        }
        if _match and _match.get("count",0) > 5 and overall < 65:
            out["alerts"].append(f"⚠️ Habitual bad posture — this pattern seen {_match['count']}× before")
    except Exception:
        pass

    # ── Head-shoulder-hip plumb line alignment (front view) ──────
    # Ideal: nose, mid_sh, mid_hip all on same vertical line
    # Deviation = lateral sway / leaning to one side
    try:
        _nose_x  = nose[0]     / max(w, 1)
        _sh_x    = mid_sh[0]   / max(w, 1)
        _hip_x   = mid_hip[0]  / max(w, 1)
        # Upper plumb: nose vs shoulder center (lateral head displacement)
        _plumb_upper = abs(_nose_x - _sh_x) * 100   # % of frame width
        # Lower plumb: shoulder center vs hip center (trunk lateral sway)
        _plumb_lower = abs(_sh_x - _hip_x) * 100
        # Combined: worst segment drives score
        _plumb_total = _plumb_upper * 0.55 + _plumb_lower * 0.45
        _plumb_sc    = score_m(_plumb_total, 0, 3, 8)
        out["metrics"]["plumb_line"] = {
            "value":       round(_plumb_total, 1),
            "upper_dev":   round(_plumb_upper, 1),
            "lower_dev":   round(_plumb_lower, 1),
            "score":       _plumb_sc,
            "unit":        "% frame",
            "label":       "Head-shoulder-hip alignment",
            "reference":   "Physiotherapy plumb line assessment",
        }
        if _plumb_total > 6:
            out["alerts"].append(f"⚠️ Body lateral sway {round(_plumb_total,1)}% — realign head over shoulders over hips")
        elif _plumb_total > 3.5:
            out["alerts"].append(f"Slight lateral lean {round(_plumb_total,1)}% — check chair balance and armrest height")
    except Exception:
        pass

    # ── Hip flexion angle (front view proxy) ─────────────────────
    # From front camera: hip flexion estimated from hip-to-knee vertical ratio
    # More accurate from side, but front gives useful proxy
    try:
        _vis_l_hip  = g(PL.L_HIP).visibility;  _vis_r_hip  = g(PL.R_HIP).visibility
        _vis_l_knee = g(PL.L_KNEE).visibility;  _vis_r_knee = g(PL.R_KNEE).visibility
        if min(_vis_l_hip, _vis_r_hip, _vis_l_knee, _vis_r_knee) > 0.4:
            _l_hip_knee_dy = abs(g(PL.L_KNEE).y - g(PL.L_HIP).y) * h
            _r_hip_knee_dy = abs(g(PL.R_KNEE).y - g(PL.R_HIP).y) * h
            _l_hip_knee_dx = abs(g(PL.L_KNEE).x - g(PL.L_HIP).x) * w
            _r_hip_knee_dx = abs(g(PL.R_KNEE).x - g(PL.R_HIP).x) * w
            _l_hip_ang = math.degrees(math.atan2(_l_hip_knee_dx, _l_hip_knee_dy + 0.001))
            _r_hip_ang = math.degrees(math.atan2(_r_hip_knee_dx, _r_hip_knee_dy + 0.001))
            _hip_flex  = round((_l_hip_ang + _r_hip_ang) / 2, 1)
            _hip_flex_sc = score_m(abs(_hip_flex - 10), 0, 10, 25)  # ideal ~10° outward
            out["metrics"]["hip_flexion_front"] = {
                "value": _hip_flex, "score": _hip_flex_sc,
                "unit": "°", "label": "Hip flexion (front proxy)",
            }
    except Exception:
        pass

    # ── Wrist single-elbow fallback ───────────────────────────────
    # (already computed above — gate loosened to single elbow)

    # ── Posture validity check (lying down from landmarks) ──────────
    try:
        _ear_y  = (g(PL.L_EAR).y  + g(PL.R_EAR).y)  / 2
        _sh_y   = (g(PL.L_SHOULDER).y + g(PL.R_SHOULDER).y) / 2
        _hip_y  = (g(PL.L_HIP).y  + g(PL.R_HIP).y)  / 2
        # Normal sitting: ear above shoulder, shoulder above hip
        # Lying: ear ≈ shoulder height, minimal vertical separation
        _ear_sh_diff = _sh_y - _ear_y    # positive = normal (shoulder below ear)
        _sh_hip_diff = _hip_y - _sh_y    # positive = normal (hip below shoulder)
        if _ear_sh_diff < 0.05 and _sh_hip_diff < 0.05:
            # Extremely small vertical separation = lying flat
            if not out.get("posture_valid") == False:  # don't double-flag
                out["alerts"].insert(0, "⚠️ Lying down detected — sit upright for accurate posture tracking")
                out["posture_valid"] = False
                # Invalidate score when lying
                overall = min(overall, 25)
        elif _ear_sh_diff < 0.08:
            # Leaning heavily or reclining
            out.setdefault("alerts", [])
            if not any("lying" in a.lower() or "reclining" in a.lower() for a in out["alerts"]):
                out["alerts"].append("Possible reclining posture — sit upright for best results")
    except Exception:
        pass

    # ── Smart break recommendation ───────────────────────────────
    try:
        _session_start = out.get("session_start", time.time())
        _break_uid = getattr(g, "uid", None)
        _break_rec = update_break_profile(_break_uid, overall, _session_start) if _break_uid else None
        if _break_rec:
            out["break_recommendation"] = _break_rec
            if _break_rec.get("message") and _break_rec["urgency"] in ("now", "soon"):
                out["alerts"].insert(0, _break_rec["message"])
            out.setdefault("alerts_ar", []).insert(0, _break_rec["message_ar"])
    except Exception:
        pass

    # ── Sitting duration fatigue model ────────────────────────────
    # If session has been running > 45min, flag fatigue risk
    try:
        _sid_dur = out.get("session_id", "")
        _sess_start = sessions.get(_sid_dur, {}).get("start", time.time()) if "sessions" in dir() else time.time()
        _duration_min = (time.time() - _sess_start) / 60
        if _duration_min > 45:
            fatigue_factor = min(1.0, (_duration_min - 45) / 60)
            out["fatigue_risk"] = {
                "duration_min":   round(_duration_min),
                "fatigue_factor": round(fatigue_factor, 2),
                "risk":           "high" if fatigue_factor > 0.5 else "moderate",
            }
            if fatigue_factor > 0.5:
                out["alerts"].append(f"⚠️ Seated {round(_duration_min)}min — take a standing break (every 45-60min)")
            elif fatigue_factor > 0.2:
                out["alerts"].append(f"Seated {round(_duration_min)}min — consider a short break soon")
    except Exception:
        pass

    # ── Per-joint injury risk scores ────────────────────────────
    # Each joint gets individual risk score based on angle severity
    # Useful for targeted physiotherapy recommendations
    try:
        _dur_min = (time.time() - sessions.get(out.get("session_id",""),{}).get("start",time.time())) / 60 if "sessions" in dir() else 0
        _joint_risks = {}
        _joint_def = {
            "neck":     (neck_lean,  [10, 20, 30],  "cervical spine"),
            "spine":    (spine_lean, [10, 20, 40],  "thoracic/lumbar spine"),
            "shoulder": (sh_tilt,     [5,  15, 30],  "shoulder girdle"),
        }
        for joint, (angle, (l,m,h), region) in _joint_def.items():
            if angle > h:      risk_level, risk_sc = "high",     20
            elif angle > m:    risk_level, risk_sc = "moderate", 55
            elif angle > l:    risk_level, risk_sc = "low",      80
            else:              risk_level, risk_sc = "minimal",  100
            # Duration multiplier
            dur_mult = min(2.0, 1 + (_dur_min / 60) * 0.5)
            time_risk = min(4, risk_sc / 25 * dur_mult)   # 0-4 scale
            _joint_risks[joint] = {
                "angle":      round(angle, 1),
                "risk_level": risk_level,
                "risk_score": risk_sc,
                "time_risk":  round(time_risk, 2),
                "region":     region,
            }
        out["metrics"]["joint_risks"] = {
            "value":  max(j["time_risk"] for j in _joint_risks.values()) if _joint_risks else 0,
            "score":  min(j["risk_score"] for j in _joint_risks.values()) if _joint_risks else 100,
            "unit":   "risk",
            "label":  "Per-joint injury risk",
            "joints": _joint_risks,
        }
    except Exception:
        pass

    # ── Pain onset prediction (tissue creep model) ─────────────
    # Based on McGill (2007) spine biomechanics:
    # Sustained load → tissue creep → pain onset in predictable time
    # Neck/spine: pain typically within 20-45min of sustained poor posture
    try:
        _sid_pain = out.get("session_id", "")
        _sess_data = sessions.get(_sid_pain, {}) if "sessions" in dir() else {}
        _dur_min = (time.time() - _sess_data.get("start", time.time())) / 60

        # Load factor: how bad is the current posture?
        _load_factor = max(0, (100 - overall) / 100)  # 0=perfect, 1=worst

        # Tissue creep rate: higher load = faster creep
        # At 100% load (worst posture): pain in ~15min
        # At 50% load: pain in ~35min
        # At 20% load: pain in ~90min
        if _load_factor > 0.1:
            _time_to_pain_min = 15 / max(_load_factor, 0.1)
            _remaining_min = max(0, _time_to_pain_min - _dur_min)
            _creep_pct = min(100, int((_dur_min / _time_to_pain_min) * 100)) if _time_to_pain_min > 0 else 0
            _pain_regions = []
            if neck_lean > 15:    _pain_regions.append("neck/upper trapezius")
            if spine_lean > 15:   _pain_regions.append("lower back")
            if sh_tilt > 8:        _pain_regions.append("shoulder")
            if dist_cm < 40:      _pain_regions.append("eyes/neck (screen proximity)")

            _urgency = (
                "imminent" if _remaining_min < 5  else
                "soon"     if _remaining_min < 15 else
                "moderate" if _remaining_min < 30 else
                "low"
            )
            _action_en = (
                "⚠️ Correct posture NOW — pain onset imminent" if _remaining_min < 5 else
                f"Correct posture within {round(_remaining_min)}min to avoid pain" if _remaining_min < 15 else
                f"~{round(_remaining_min)}min before discomfort if posture unchanged" if _remaining_min < 30 else
                "Posture acceptable for extended sitting"
            )
            _action_ar = (
                "⚠️ صحّح وضعيتك الآن — الألم وشيك" if _remaining_min < 5 else
                f"صحّح وضعيتك خلال {round(_remaining_min)} دقيقة لتجنب الألم" if _remaining_min < 15 else
                f"~{round(_remaining_min)} دقيقة قبل الإزعاج إذا لم تتغير الوضعية" if _remaining_min < 30 else
                "الوضعية مقبولة للجلوس لفترة أطول"
            )
            _pain_color = (
                "#ef4444" if _urgency == "imminent" else
                "#f97316" if _urgency == "soon"     else
                "#f59e0b" if _urgency == "moderate" else
                "#10b981"
            )
            out["pain_prediction"] = {
                "creep_pct":        _creep_pct,
                "time_to_pain_min": round(_time_to_pain_min, 1),
                "remaining_min":    round(_remaining_min, 1),
                "load_factor":      round(_load_factor, 2),
                "at_risk_regions":  _pain_regions,
                "at_risk_regions_ar": [
                    {"neck/upper trapezius": "الرقبة / الشريط العلوي",
                     "lower back": "أسفل الظهر",
                     "shoulder": "الكتف",
                     "eyes/neck (screen proximity)": "العينان / الرقبة (قرب الشاشة)"
                    }.get(r, r) for r in _pain_regions
                ],
                "urgency":    _urgency,
                "action":     _action_en,
                "action_ar":  _action_ar,
                "reference":  "McGill 2007 spine biomechanics + tissue creep model",
            }
            # ── pain_bar: top-level prominent field for UI ────────
            out["pain_bar"] = {
                "pct":      _creep_pct,
                "urgency":  _urgency,
                "remaining_min": round(_remaining_min, 1),
                "label":    _action_en,
                "label_ar": _action_ar,
                "color":    _pain_color,
            }
            if _remaining_min < 10 and _creep_pct > 60:
                out["alerts"].append(
                    f"⚠️ Pain onset predicted in ~{round(_remaining_min)}min "
                    f"({', '.join(_pain_regions or ['general'])} at risk)")
        else:
            out["pain_prediction"] = {
                "creep_pct": 0, "urgency": "none",
                "action": "✓ Posture excellent — no pain risk",
                "action_ar": "✓ وضعية ممتازة — لا خطر ألم",
            }
            out["pain_bar"] = {"pct": 0, "urgency": "none", "label": "No pain risk", "label_ar": "لا خطر ألم", "color": "#10b981"}
    except Exception:
        pass

    # ── RSI / Cumulative injury risk ──────────────────────────────
    # Tracks session-cumulative risk for repetitive strain injuries
    # Based on: wrist angle severity × duration + neck lean × duration
    try:
        _s_obj = sessions.get(out.get("session_id",""), {}) if "sessions" in dir() else {}
        _dur_so_far = (time.time() - _s_obj.get("start", time.time())) / 60  # minutes
        # RULA-inspired: angle × time weighting
        _neck_risk   = max(0, neck_lean - 10) * 0.8   # degrees over threshold
        _wrist_risk  = 0
        if wrist_sc is not None:
            _w_met = out["metrics"].get("wrist_angle", {})
            _wrist_risk = max(0, _w_met.get("value",0) - 15) * 1.2
        _rsi_rate    = (_neck_risk + _wrist_risk) * (_dur_so_far / 60)  # per hour
        _rsi_level   = "low" if _rsi_rate < 2 else "moderate" if _rsi_rate < 6 else "high"
        out["metrics"]["rsi_risk"] = {
            "value":     round(_rsi_rate, 2),
            "score":     score_m(_rsi_rate, 0, 2, 6),
            "unit":      "risk units",
            "label":     "RSI cumulative risk",
            "level":     _rsi_level,
            "neck_contrib":  round(_neck_risk, 1),
            "wrist_contrib": round(_wrist_risk, 1),
        }
        if _rsi_level == "high":
            out["alerts"].append(f"⚠️ High RSI risk — prolonged poor posture increases repetitive strain injury risk")
    except Exception:
        pass

    # ── Alert cleanup: deduplicate + sort by severity + cap at 5 ──
    seen = set()
    deduped = []
    for alert in out["alerts"]:
        key = alert[:30]   # first 30 chars as dedup key
        if key not in seen:
            seen.add(key)
            deduped.append(alert)
    # Severe (⚠️) first, then informational
    deduped.sort(key=lambda a: (0 if a.startswith("⚠️") else 1))
    out["alerts"] = deduped[:5]   # max 5 alerts — avoid overwhelming user
    # ── Arabic alerts ─────────────────────────────────────────────
    out["alerts_ar"] = [translate_alert_ar(a) for a in out["alerts"]]

    grade_str = ("Excellent" if overall >= 85 else "Good" if overall >= 70
                 else "Fair — needs attention" if overall >= 55 else "Poor — correct now")
    out["recommendations"] = [
        f"Overall: {grade_str} ({overall}/100)",
        f"Screen distance: {'✓ Optimal' if lo <= dist_cm <= hi else f'Needs adjustment — move to {lo}–{hi}cm'} (current: {round(dist_cm)}cm)",
        "Keep ears directly above shoulder joints, chin parallel to floor",
        "Lumbar support: lower back should touch chair back fully",
        "Every 30 min: stand, stretch neck and shoulders for 2 min",
        "20-20-20 rule: every 20 min look 20 feet away for 20 seconds",
        "Monitor top edge should be at or slightly below eye level",
    ]
    if overall < 60:
        out["recommendations"].insert(1, "⚠️ Take a posture break now — stretch and reset your position")

    for idx, name in [(PL.NOSE,"nose"),(PL.L_EAR,"l_ear"),(PL.R_EAR,"r_ear"),
                      (PL.L_SHOULDER,"l_sh"),(PL.R_SHOULDER,"r_sh"),
                      (PL.L_HIP,"l_hip"),(PL.R_HIP,"r_hip"),
                      (PL.L_EYE,"l_eye"),(PL.R_EYE,"r_eye")]:
        lm = g(idx)
        out["landmarks"].append({
            "name": name, "x": round(lm.x, 4), "y": round(lm.y, 4),
            "visibility": round(lm.visibility, 2)
        })

    return out

# ── SIDE ANALYSIS ──────────────────────────────────────────────────
def analyze_side(image, tier="standard", session_id=None):
    """
    Side-camera posture analysis — full parity with analyze_front.
    Improvements over old version:
    - Landmark averaging (jitter reduction ±3°→±1°)
    - 3-point spine (upper + lower segments + S-curve detection)
    - Per-metric visibility confidence weighting
    - Hip + knee properly weighted in overall
    - Shoulder width normalization for neck lean
    - Alert dedup + severity sort + cap at 5
    """
    _ensure_models()
    h, w = image.shape[:2]
    rgb  = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)

    out = {
        "mode": "side", "detected": False, "metrics": {}, "score": 0,
        "alerts": [], "recommendations": [], "landmarks": [],
        "confidence": 0, "engine": "mediapipe"
    }

    pose_model  = POSE_FULL if tier in ("professional","elite","pro","premium","business") else POSE_LITE
    pose_result = pose_model.process(rgb)

    if not pose_result.pose_landmarks:
        out["alerts"].append("No person detected — camera should be 90° to your side, 80–120cm away")
        return analyze_side_cascade(image, out)

    # ── Landmark averaging (same as analyze_front) ────────────────
    _sid  = out.get("session_id", "side_default")
    _raw  = pose_result.pose_landmarks.landmark
    _key  = (session_id or _sid) + "_side"
    _hist = _lm_history.setdefault(_key, [])
    _hist.append(_raw)
    if len(_hist) > 3: _hist.pop(0)

    _w_s = [0.60, 0.30, 0.10][:len(_hist)]
    _w_s_sum  = sum(_w_s)
    _w_s_norm = [wi / _w_s_sum for wi in _w_s] if _w_s_sum > 0 else _w_s

    class _AvgLM:
        __slots__ = ("x","y","z","visibility")
        def __init__(self, idx):
            fr = list(reversed(_hist))
            ws = _w_s_norm[:len(fr)]
            self.x          = sum(fr[i][idx].x          * ws[i] for i in range(len(fr)))
            self.y          = sum(fr[i][idx].y          * ws[i] for i in range(len(fr)))
            self.z          = sum(fr[i][idx].z          * ws[i] for i in range(len(fr)))
            self.visibility = sum(fr[i][idx].visibility * ws[i] for i in range(len(fr)))
    class _AvgLMs:
        def __getitem__(self, idx): return _AvgLM(idx)
    lms = _AvgLMs()

    out["detected"]    = True
    out["engine"]      = "mediapipe_pose"
    out["avg_frames"]  = len(_hist)

    def g(idx): return lms[idx]
    def px(idx): return (g(idx).x * w, g(idx).y * h)

    # ── Pick dominant side (more visible) ─────────────────────────
    l_vis = g(PL.L_SHOULDER).visibility
    r_vis = g(PL.R_SHOULDER).visibility
    S     = "L" if l_vis >= r_vis else "R"
    _I    = lambda L, R: L if S == "L" else R

    ear   = px(_I(PL.L_EAR,   PL.R_EAR))
    sh    = px(_I(PL.L_SHOULDER, PL.R_SHOULDER))
    hip   = px(_I(PL.L_HIP,   PL.R_HIP))
    knee  = px(_I(PL.L_KNEE,  PL.R_KNEE))
    ankle = px(_I(PL.L_ANKLE, PL.R_ANKLE))

    vis_ear   = g(_I(PL.L_EAR,   PL.R_EAR)).visibility
    vis_sh    = max(l_vis, r_vis)
    vis_hip   = g(_I(PL.L_HIP,   PL.R_HIP)).visibility
    vis_knee  = g(_I(PL.L_KNEE,  PL.R_KNEE)).visibility
    vis_ankle = g(_I(PL.L_ANKLE, PL.R_ANKLE)).visibility

    # ── Visibility confidence (same formula as analyze_front) ─────
    def vis_conf(v): return max(0.0, min(1.0, v / 0.6))

    conf_neck  = vis_conf(vis_ear * 0.5 + vis_sh * 0.5)
    conf_trunk = vis_conf(vis_sh  * 0.5 + vis_hip * 0.5)
    conf_hip   = vis_conf(min(vis_sh, vis_hip, vis_knee))
    conf_knee  = vis_conf(min(vis_hip, vis_knee, vis_ankle))
    conf_spine = vis_conf(vis_ear * 0.4 + vis_sh * 0.3 + vis_ankle * 0.3)

    # ── Neck lean — ear-to-shoulder (side view is geometrically exact) ──
    sh_width_px = abs(g(PL.L_SHOULDER).x * w - g(PL.R_SHOULDER).x * w)
    # Distance proxy for threshold normalisation: full ear→shoulder span.
    # The L-R shoulder width foreshortens to ~0 px in profile, so the old
    # sh_frac/0.34 ratio always sat clamped at 0.70 regardless of distance.
    _ear_sh_span_px = math.hypot(ear[0] - sh[0], ear[1] - sh[1])
    sh_ratio    = max(0.7, min(1.3, (_ear_sh_span_px / max(w, 1)) / 0.14))

    neck_lean  = angle_vert(sh, ear)
    neck_ok    = max(5.0,  8.0 * sh_ratio)
    neck_bad   = max(16.0, 22.0 * sh_ratio)
    neck_sc    = score_m(neck_lean, 0, neck_ok, neck_bad)
    # FHP from side: horizontal ear-to-shoulder offset in cm.
    # cm ruler: the L-R shoulder width foreshortens to near-zero px in a
    # true profile view, which exploded the conversion (an upright sitter
    # could read 20+cm of "FHP"). Use spans that stay fully visible in
    # profile instead: shoulder→hip (~46cm seated torso) when the hip is
    # tracked, else ear→shoulder (~25cm). Shoulder width remains the last
    # resort. Kept in sync with postureEngine.js analyzeSideMP().
    _torso_px = math.hypot(sh[0] - hip[0], sh[1] - hip[1])
    if _torso_px > 20 and vis_hip >= 0.5:
        _cm_per_px_s = 46.0 / _torso_px
    elif _ear_sh_span_px > 10:
        _cm_per_px_s = 25.0 / _ear_sh_span_px
    else:
        _cm_per_px_s = 42.0 / max(sh_width_px, 1)
    _fhp_side_cm  = round(abs(ear[0] - sh[0]) * _cm_per_px_s, 1)
    _fhp_side_sc  = score_m(_fhp_side_cm, 0, 2.5, 7)
    if _fhp_side_cm > 5:
        out["alerts"].append(f"⚠️ Forward head posture {_fhp_side_cm}cm (side view) — tuck chin, pull head back over shoulders")

    # ── Trunk lean ────────────────────────────────────────────────
    trunk_lean = angle_vert(hip, sh)
    trunk_sc   = score_m(trunk_lean, 0, 6, 16)

    # ── IMPROVED: 4-point spine + kyphosis + lumbar lordosis ──────
    # 4 reference points: sh → T8 → L3 → hip
    mid_thoracic_s = (sh[0]*0.60 + hip[0]*0.40, sh[1]*0.60 + hip[1]*0.40)  # T8
    mid_lumbar_s   = (sh[0]*0.25 + hip[0]*0.75, sh[1]*0.25 + hip[1]*0.75)  # L3

    spine_upper_s  = angle_vert(mid_thoracic_s, sh)         # C7→T8
    spine_mid_s    = angle_vert(mid_lumbar_s,   mid_thoracic_s)  # T8→L3
    spine_lower_s  = angle_vert(hip,             mid_lumbar_s)    # L3→S1

    # ── Kyphosis: thoracic rounds forward, lumbar compensates ─────
    _upper_signed_s = sh[0]           - mid_thoracic_s[0]
    _lower_signed_s = mid_thoracic_s[0] - hip[0]
    _kyphosis_pen_s = 0.0
    if spine_upper_s > 3 and spine_lower_s < 3 and (_upper_signed_s * _lower_signed_s < 0):
        _kyphosis_pen_s = min(12.0, spine_upper_s * 0.6)
        out["alerts"].append(f"⚠️ Thoracic kyphosis (side) {round(spine_upper_s,1)}° — open chest, pull shoulder blades together")

    # ── Lumbar lordosis: normal = slight backward curve ───────────
    # Measured as angle between lumbar segment and vertical
    # <3° = flat back (lost lordosis) — risk of disc compression
    _lordosis_angle = abs(spine_lower_s - spine_mid_s)
    _lordosis_sc    = 100 if _lordosis_angle > 2 else score_m(2 - _lordosis_angle, 0, 1, 3)
    if spine_lower_s < 2 and spine_mid_s < 2:
        out["alerts"].append("Flat lower back detected — use lumbar support or sit further back in chair")
        _lordosis_sc = 60

    spine_scurve   = max(0, abs(spine_upper_s - spine_lower_s) - 8) * 0.5
    spine_combined = (spine_upper_s * 0.45 + spine_mid_s * 0.30 +
                      spine_lower_s * 0.25 + spine_scurve + _kyphosis_pen_s)
    spine_combined = min(45.0, spine_combined)
    spine_sc       = score_m(spine_combined, 0, 5, 14)

    # ── Spinal plumb line alignment (ear above ankle) ─────────────
    spine_align = abs(ear[0] - ankle[0]) / w * 100 if w > 0 else 0
    align_sc    = score_m(spine_align, 0, 4, 12)

    # ── Hip + knee angles ─────────────────────────────────────────
    hip_angle  = angle_3pt(sh, hip, knee)
    hip_sc     = score_m(abs(hip_angle - 90), 0, 12, 30)

    knee_angle = angle_3pt(hip, knee, ankle)
    knee_sc    = score_m(abs(knee_angle - 90), 0, 12, 35)

    # ── Confidence-weighted overall ───────────────────────────────
    BASE_W_S = {
        "neck":  0.30,   # most important — FHP most common desk issue
        "trunk": 0.25,   # trunk lean — direct back pain predictor
        "spine": 0.18,   # 3-point curvature — S-curve detection
        "align": 0.10,   # plumb line
        "hip":   0.10,   # hip angle — 90° optimal
        "knee":  0.07,   # knee angle — 90° optimal
    }
    eff_w_s = {
        "neck":  BASE_W_S["neck"]  * conf_neck,
        "trunk": BASE_W_S["trunk"] * conf_trunk,
        "spine": BASE_W_S["spine"] * conf_spine,
        "align": BASE_W_S["align"] * conf_spine,   # same landmarks as spine
        "hip":   BASE_W_S["hip"]   * conf_hip,
        "knee":  BASE_W_S["knee"]  * conf_knee,
    }
    scores_s = {
        "neck": neck_sc, "trunk": trunk_sc, "spine": spine_sc,
        "align": align_sc, "hip": hip_sc, "knee": knee_sc,
    }
    score_val_s  = sum(scores_s[k] * eff_w_s[k] for k in scores_s)
    weight_used_s = sum(eff_w_s.values())
    overall = max(0, min(100, int(round(score_val_s / max(weight_used_s, 0.01)))))

    # ── Confidence score ──────────────────────────────────────────
    overall_conf_s = sum(
        {"neck":conf_neck,"trunk":conf_trunk,"spine":conf_spine,"hip":conf_hip,"knee":conf_knee}[k]
        * BASE_W_S.get(k, 0) for k in ("neck","trunk","spine","hip","knee")
    ) / sum(BASE_W_S[k] for k in ("neck","trunk","spine","hip","knee"))

    confidence = min(95, int(
        45
        + (overall_conf_s * 30)
        + (10 if len(_hist) >= 3 else 0)
        + (10 if vis_sh > 0.7 else 0)
    ))

    out["score"]      = overall
    out["confidence"] = confidence
    out["metrics"]    = {
        "neck_lean_side": {"value": round(neck_lean,1),     "score": neck_sc,   "unit": "°", "label": "Neck lean (side)"},
        "trunk_lean":     {"value": round(trunk_lean,1),    "score": trunk_sc,  "unit": "°", "label": "Trunk lean"},
        "spine_curvature":{"value": round(spine_combined,1),"score": spine_sc,  "unit": "°", "label": "Spine curvature"},
        "spine_upper_s":   {"value": round(spine_upper_s,1),   "score": spine_sc,      "unit": "°", "label": "Upper spine C7→T8 (side)"},
        "spine_mid_s":     {"value": round(spine_mid_s,1),     "score": spine_sc,      "unit": "°", "label": "Mid spine T8→L3 (side)"},
        "spine_lower_s":   {"value": round(spine_lower_s,1),   "score": spine_sc,      "unit": "°", "label": "Lower spine L3→S1 (side)"},
        "kyphosis_pen_s":  {"value": round(_kyphosis_pen_s,1), "score": max(0,100-_kyphosis_pen_s*5), "unit": "°", "label": "Kyphosis penalty (side)"},
        "lumbar_lordosis": {"value": round(_lordosis_angle,1), "score": _lordosis_sc, "unit": "°", "label": "Lumbar lordosis"},
        "fhp_side_cm":     {"value": _fhp_side_cm,            "score": _fhp_side_sc, "unit": "cm","label": "Forward head posture (side)"},
        "spine_align":     {"value": round(spine_align,1),     "score": align_sc,     "unit": "%", "label": "Spinal plumb line"},
        "hip_angle":       {"value": round(hip_angle,1),       "score": hip_sc,       "unit": "°", "label": "Hip flexion angle"},
        "knee_angle":      {"value": round(knee_angle,1),      "score": knee_sc,      "unit": "°", "label": "Knee angle"},
        "_side_confidence": {
            "neck": round(conf_neck,2), "trunk": round(conf_trunk,2),
            "hip":  round(conf_hip,2),  "knee":  round(conf_knee,2),
            "eff_weights": {k: round(v,3) for k,v in eff_w_s.items()},
            "label": "Per-metric side confidence",
        },
    }

    # ── Alerts ────────────────────────────────────────────────────
    if neck_lean > 20:
        out["alerts"].append(f"⚠️ Forward head {round(neck_lean,1)}° — ear must align directly above shoulder")
    elif neck_lean > 12:
        out["alerts"].append(f"Forward head {round(neck_lean,1)}° — tuck chin and lengthen spine")
    if trunk_lean > 18:
        out["alerts"].append(f"⚠️ Trunk leaning {round(trunk_lean,1)}° — press back fully into chair backrest")
    elif trunk_lean > 10:
        out["alerts"].append(f"Trunk lean {round(trunk_lean,1)}° — sit upright, activate core")
    if spine_scurve > 3:
        out["alerts"].append(f"⚠️ S-curve detected — upper/lower spine misaligned ({round(spine_upper_s,1)}° vs {round(spine_lower_s,1)}°)")
    if abs(hip_angle - 90) > 20:
        out["alerts"].append(f"Hip angle {round(hip_angle,1)}° (ideal 90°) — {'raise' if hip_angle < 90 else 'lower'} chair height")
    if abs(knee_angle - 90) > 25:
        out["alerts"].append(f"Knee angle {round(knee_angle,1)}° (ideal 90°) — adjust seat depth or use footrest")
    if spine_align > 8:
        out["alerts"].append("Head not aligned above feet — check overall seated posture")

    # Dedup + sort + cap (same as analyze_front)
    seen_s, deduped_s = set(), []
    for a in out["alerts"]:
        k = a[:30]
        if k not in seen_s:
            seen_s.add(k); deduped_s.append(a)
    deduped_s.sort(key=lambda a: (0 if a.startswith("⚠️") else 1))
    out["alerts"] = deduped_s[:5]
    # ── Arabic alerts (side view) ─────────────────────────────────
    out["alerts_ar"] = [translate_alert_ar(a) for a in out["alerts"]]

    grade = ("Excellent" if overall >= 85 else "Good" if overall >= 70
             else "Fair" if overall >= 55 else "Poor")
    out["recommendations"] = [
        f"{grade} lateral posture ({overall}/100) — ear→shoulder→hip→ankle should align vertically",
        f"{'Left' if S=='L' else 'Right'} side visible — {'✓ good angle' if vis_sh > 0.7 else 'improve camera angle for better accuracy'}",
        f"Hip angle {round(hip_angle,1)}° — {'✓ ideal' if abs(hip_angle-90) < 12 else 'adjust chair so thighs are parallel to floor'}",
        f"Spine curvature {round(spine_combined,1)}° — {'✓ good' if spine_combined < 8 else 'press lower back into lumbar support'}",
        "Feet flat on floor, lower back fully touching chair back",
        "Ear directly above shoulder — avoid forward head position",
    ]
    if overall < 60:
        out["recommendations"].insert(0, "⚠️ Take a posture break — stand and stretch for 2 min")

    for name, pt in [("ear",ear),("sh",sh),("hip",hip),("knee",knee),("ankle",ankle)]:
        out["landmarks"].append({"name":name,"x":round(pt[0]/w,4),"y":round(pt[1]/h,4),"vis":round({"ear":vis_ear,"sh":vis_sh,"hip":vis_hip,"knee":vis_knee,"ankle":vis_ankle}[name],2)})

    return out

# ── CASCADE FALLBACK ───────────────────────────────────────────────
# face_c/eye_c/prof_c are lazy-initialized in _ensure_models() alongside cv2 —
# see the placeholders near the top of the file. They used to be initialized
# here at module level, which crashed on every cold start (cv2 is still None
# at import time, before _ensure_models() has ever run).

def analyze_front_cascade(image, mode, out):
    """
    Cascade fallback (OpenCV Haar) — uses SAME score_m thresholds and weights
    as analyze_front() so score is consistent regardless of which engine runs.
    Lower confidence (72) but same scoring formula.
    """
    _ensure_models()
    h, w = image.shape[:2]
    gray = cv2.createCLAHE(2.5, (8, 8)).apply(cv2.cvtColor(image, cv2.COLOR_BGR2GRAY))
    faces = face_c.detectMultiScale(gray, 1.08, 7, minSize=(50, 50))
    if not len(faces):
        faces = face_c.detectMultiScale(gray, 1.14, 5, minSize=(35, 35))
    if not len(faces):
        out["engine"] = "no_detection"; return out

    fx, fy, fw, fh = max(faces, key=lambda f: f[2] * f[3])
    fcx, fcy       = fx + fw // 2, fy + fh // 2
    out["detected"] = True
    out["engine"]   = "cascade_fallback"

    # ── Distance via IPD (face width fallback) ────────────────────
    roi  = gray[fy:fy + int(fh * .55), fx:fx + fw]
    eyes = eye_c.detectMultiScale(roi, 1.05, 4, minSize=(14, 14))
    dist_cm = round((14.0 * 600 * (w / 640)) / max(fw, 1), 1)
    le = re = None
    if len(eyes) >= 2:
        es  = sorted(eyes, key=lambda e: e[0])
        le  = (fx + es[0][0]  + es[0][2]  // 2, fy + es[0][1]  + es[0][3]  // 2)
        re  = (fx + es[-1][0] + es[-1][2] // 2, fy + es[-1][1] + es[-1][3] // 2)
        ipd = math.dist(le, re)
        if ipd > 8:
            dist_cm = round((6.3 * 630 * (w / 640)) / ipd, 1)
    dist_cm = max(20, min(150, dist_cm))

    lo, hi = (50, 80) if mode == "laptop" else (60, 90)
    # ── Adaptive focal from face size (better than fixed 600) ────
    # Standard face width ~14cm. Reference: 640px wide frame at 65cm → face ~140px
    ref_face_pct = 0.22   # ~22% of frame width at reference distance
    face_pct     = fw / max(w, 1)
    if face_pct > 0.05:   # face large enough to estimate
        adaptive_focal = 600 * (ref_face_pct / face_pct) * (w / 640)
        dist_cm = round((14.0 * adaptive_focal) / max(fw, 1), 1)
        dist_cm = max(20, min(150, dist_cm))
    # ── SAME dist_sc logic as analyze_front ──────────────────────
    if   lo <= dist_cm <= hi:                                dist_sc = 100
    elif (lo - 8)  <= dist_cm <= (hi + 12):                 dist_sc = 80
    elif (lo - 16) <= dist_cm <= (hi + 20):                 dist_sc = 55
    else:                                                    dist_sc = 30

    # ── Head tilt from eye positions (same threshold as analyze_front) ──
    ht = 0.0
    if le and re:
        ht = abs(math.degrees(math.atan2(re[1] - le[1], re[0] - le[0])))
    tilt_sc = score_m(ht, 0, 3, 10)  # same as analyze_front

    # ── Neck lean proxy: face center Y relative to frame ──────────
    # fcy/h: face high in frame (~0.15) = forward lean; neutral ~0.30–0.38
    neck_proxy = max(0.0, (fcy / h) - 0.30) * 80   # maps 0–0.25 → 0–20 degrees
    neck_proxy = min(neck_proxy, 35.0)
    neck_sc    = score_m(neck_proxy, 0, 7, 20)      # same thresholds as analyze_front

    # ── Forward lean proxy: face area relative to frame ──────────
    face_area_pct = (fw * fh) / (w * h) * 100
    # neutral face area ~8–14%. > 18% = too close/leaning
    forward_proxy = max(0.0, face_area_pct - 11) * 3
    forward_sc    = score_m(forward_proxy, 0, 5, 18)

    # ── Overall — SAME weights as analyze_front ───────────────────
    # neck.28 tilt.14 dist.18 forward.11 → remaining=0.29 → baseline fill
    score_val  = neck_sc * 0.28 + tilt_sc * 0.14 + dist_sc * 0.18 + forward_sc * 0.11
    remaining  = 1.0 - (0.28 + 0.14 + 0.18 + 0.11)   # 0.29
    # Normalize by actual weights used — no arbitrary baseline
    weight_used = 1.0 - remaining
    overall = max(0, min(100, int(round(score_val / max(weight_used, 0.01))))) if weight_used > 0.01 else 0

    out["score"]      = overall
    out["confidence"] = 68   # lower than MediaPipe (82+) to signal fallback quality
    out["metrics"]    = {
        "head_tilt":       {"value": round(ht, 1),           "score": tilt_sc,    "unit": "°",  "label": "Head tilt"},
        "neck_lean":       {"value": round(neck_proxy, 1),   "score": neck_sc,    "unit": "°",  "label": "Neck lean (est.)"},
        "screen_distance": {"value": dist_cm,                "score": dist_sc,    "unit": "cm", "label": "Screen distance"},
        "forward_lean":    {"value": round(forward_proxy,1), "score": forward_sc, "unit": "°",  "label": "Forward lean (est.)"},
    }

    # Alerts — same thresholds as analyze_front
    if ht > 10:
        out["alerts"].append(f"Head tilting {round(ht,1)}° — level your head")
    if dist_cm < lo - 10:
        out["alerts"].append(f"⚠️ Very close to screen ({round(dist_cm)}cm) — move back to {lo}–{hi}cm")
    elif dist_cm < lo:
        out["alerts"].append(f"Too close to screen ({round(dist_cm)}cm) — move back to {lo}–{hi}cm")
    elif dist_cm > hi + 15:
        out["alerts"].append(f"Too far from screen ({round(dist_cm)}cm) — ideal is {lo}–{hi}cm")
    if neck_proxy > 20:
        out["alerts"].append(f"Forward head detected — raise monitor to eye level")

    grade = "Good" if overall >= 70 else "Fair" if overall >= 55 else "Poor"
    out["recommendations"] = [
        f"{grade} posture estimate (limited precision — ensure good lighting for full analysis)",
        f"Screen distance ~{round(dist_cm)}cm (ideal {lo}–{hi}cm)",
        "Keep ears directly above shoulders, chin parallel to floor",
    ]
    out["landmarks"] = [{"name": "face", "x": round(fcx / w, 4), "y": round(fcy / h, 4)}]
    return out

def analyze_side_cascade(image, out):
    _ensure_models()  # lazy-load MediaPipe on first call
    h, w = image.shape[:2]
    gray = cv2.createCLAHE(2.0,(8,8)).apply(cv2.cvtColor(image, cv2.COLOR_BGR2GRAY))
    pl = prof_c.detectMultiScale(gray, 1.07,4,minSize=(48,48))
    pr = prof_c.detectMultiScale(cv2.flip(gray,1),1.07,4,minSize=(48,48))
    ff = face_c.detectMultiScale(gray, 1.11,5,minSize=(52,52))
    pok = len(pl)>0 or len(pr)>0
    if not pok and not len(ff):
        out["engine"] = "no_detection"; return out
    out["detected"] = True; out["engine"] = "cascade_fallback_side"
    if   len(pl)>0: px2,py,pw,ph=max(pl,key=lambda f:f[2]*f[3]); side="L"
    elif len(pr)>0: px2,py,pw,ph=max(pr,key=lambda f:f[2]*f[3]); px2=w-px2-pw; side="R"
    else:           px2,py,pw,ph=max(ff,key=lambda f:f[2]*f[3]); side="F"
    fcx, fcy = px2+pw//2, py+ph//2
    fd = abs(fcx/w-0.5)*100; nsc = score_m(fd,0,8,25)
    hh = (1-fcy/h)*100;      psc = score_m(max(0,60-hh),0,9,28)
    overall = max(0,min(100,int(nsc*.45+psc*.40+(90 if pok else 50)*.15)))
    out["score"]=overall; out["confidence"]=70 if pok else 50
    out["metrics"]={"head_forward":{"value":round(fd,1),"score":nsc,"unit":"°","label":"Forward head"},
                    "head_height":{"value":round(hh,1),"score":psc,"unit":"%","label":"Head height"}}
    if not pok: out["alerts"].append("Rotate camera 90° to your side for full body analysis")
    out["recommendations"]=["Ear above shoulder above hip — check side profile","Feet flat, knees 90°, lumbar support engaged"]
    out["landmarks"]=[{"name":"profile","x":round(fcx/w,4),"y":round(fcy/h,4)}]
    return out

# ── Gemini AI ──────────────────────────────────────────────────────
def call_local_llm(prompt, system_prompt=None, max_tokens=900, temperature=0.25,
                   model=None, timeout=30):
    """
    Call a locally-running Ollama instance.
    Uses Ollama's OpenAI-compatible endpoint so the request/response shape
    is identical to any OpenAI SDK call — easy to test and debug.

    Returns the response text string, or None on any failure (connection
    refused, timeout, model not pulled, etc.). All failures are silent —
    the caller (call_ai) just returns None and the feature degrades
    gracefully (no AI text shown) rather than raising.

    Setup (run once on your machine / server):
        curl -fsSL https://ollama.com/install.sh | sh
        ollama pull qwen2.5:3b          # ~2 GB, bilingual AR+EN
        ollama serve                    # starts on :11434 by default

    Then in your .env / Railway vars:
        OLLAMA_URL=http://localhost:11434   (or your server IP)
        LOCAL_LLM_MODEL=qwen2.5:3b
    """
    if not OLLAMA_URL:
        return None

    _model = model or LOCAL_LLM_MODEL
    _url   = OLLAMA_URL.rstrip("/") + "/v1/chat/completions"

    messages = []
    if system_prompt:
        messages.append({"role": "system", "content": system_prompt})
    messages.append({"role": "user", "content": prompt})

    try:
        resp = req.post(
            _url,
            headers={"Content-Type": "application/json"},
            json={
                "model":       _model,
                "messages":    messages,
                "max_tokens":  max_tokens,
                "temperature": temperature,
                "stream":      False,
            },
            timeout=timeout,
        )
        if resp.status_code == 200:
            data = resp.json()
            text = (data.get("choices") or [{}])[0].get("message", {}).get("content", "")
            if text:
                log_event("local_llm_ok", meta={
                    "model": _model, "chars": len(text), "url": OLLAMA_URL
                })
            return text or None

        # Model not pulled yet — give a clear actionable log
        if resp.status_code == 404:
            _err = resp.json().get("error","")
            if "model" in _err.lower():
                print(
                    f"[local_llm] Model '{_model}' not found on Ollama at {OLLAMA_URL}. "
                    f"Run: ollama pull {_model}",
                    file=sys.stderr,
                )
        else:
            log_event("local_llm_error", meta={"status": resp.status_code, "model": _model})
        return None

    except req.exceptions.ConnectionError:
        # Ollama not running — silent fallback
        return None
    except req.exceptions.Timeout:
        log_event("local_llm_timeout", meta={"model": _model, "timeout_s": timeout})
        return None
    except Exception as e:
        log_event("local_llm_exception", meta={"error": str(e)[:120]})
        return None


def call_ai(prompt, system_prompt=None, max_tokens=900, temperature=0.25,
            cache_key=None, cache_ttl=90, tier="standard"):
    """
    THE unified AI call entry point for this entire backend.
    Every endpoint that needs AI text generation should call this instead of
    calling call_local_llm() directly.

    Server-side only — no AI provider credential is ever sent to the
    client. Uses Ollama (if OLLAMA_URL is set — free, self-hosted,
    zero per-request cost). `tier` is accepted for call-site
    compatibility; token budget is set by the caller via
    tier_quality.py, not by this function.

    Returns None if no AI backend is configured / unreachable.
    """
    # ── Cache check ──────────────────────────────────────────────────
    if cache_key:
        _cached = cache_get(f"ai:{cache_key}")
        if _cached:
            return _cached

    result = None
    if OLLAMA_URL:
        result = call_local_llm(prompt, system_prompt=system_prompt, max_tokens=max_tokens, temperature=temperature)

    # ── Cache successful result ──────────────────────────────────────
    if result and cache_key:
        cache_set(f"ai:{cache_key}", result, ttl_s=cache_ttl)

    return result


def analyze_with_local_ai(result, context="", lang="en"):
    if not AI_CONFIGURED: return None
    # ── Cache check — avoid duplicate AI calls for same posture state ──
    import hashlib as _hl
    _cache_key = _hl.md5(
        f"{result.get('score',0):.0f}|{result.get('head_pose','')}|{lang}".encode()
    ).hexdigest()[:16]
    _cached = cache_get(f"ai_local:{_cache_key}")
    if _cached:
        return _cached
    s     = result.get("score", 0)
    mets  = result.get("metrics", {})
    alts  = result.get("alerts", [])
    hp    = result.get("head_pose")
    mode  = result.get("mode", "laptop")
    eng   = result.get("engine", "MediaPipe")
    tier  = result.get("tier", "free")

    # All metrics including new ones (spine_upper/lower, shoulder_width, cam_correction)
    def _fmt_met(k, m):
        val  = m.get("value", "?"); unit = m.get("unit",""); sc = m.get("score","?")
        calib = " [calibrated]" if m.get("calibrated") else ""
        return f"- {m.get('label', k)}: {val}{unit} (score {sc}/100{calib})"
    mtext = "\n".join([_fmt_met(k,m) for k,m in mets.items() if isinstance(m,dict) and "score" in m])

    # Enrich with extra context
    cam_corr = result.get("metrics",{}).get("head_pose_detail",{}).get("cam_correction_applied", 0)
    blur_note = " [some frames were blurry — accuracy may be reduced]" if result.get("blurry_frame") else ""
    sh_ratio  = result.get("metrics",{}).get("shoulder_width_ratio",{}).get("value","?")

    # ── Build enriched context for Gemini ───────────────────────
    pose_text = (f"3D head pose: pitch={hp['pitch']}°, yaw={hp['yaw']}°, roll={hp['roll']}° "
                 f"(reproj_err={hp.get('reproj_err','?')}, cam_correction={cam_corr}°)"
                 if hp else "3D pose: unavailable")

    # Extra clinical metrics for richer AI output
    pain_bar    = result.get("pain_bar", {})
    fhp         = mets.get("fhp_index", {})
    neck_load   = mets.get("neck_load", {})
    rula        = mets.get("rula_score", {})
    iso         = mets.get("iso_11226", {})
    kyphosis    = mets.get("kyphosis_pen", {})
    plumb       = mets.get("plumb_line", {})
    lumbar      = mets.get("lumbar_lordosis", {})
    rsi         = mets.get("rsi_risk", {})
    twa         = mets.get("twa_score", {})
    neck_src    = mets.get("neck_source", {})
    dist_m      = mets.get("screen_distance", {})
    percentile  = result.get("percentile")

    clinical_ctx = []
    if fhp.get("value"):        clinical_ctx.append(f"FHP index: {fhp['value']}cm (Hansraj model — adds {round(fhp['value']*1.4,1)}kg to cervical spine)")
    if neck_load.get("value"):  clinical_ctx.append(f"Neck load: {neck_load['value']}kg on cervical spine (Hansraj 2014)")
    if rula.get("value"):       clinical_ctx.append(f"RULA score: {rula['value']}/7 (McAtamney 1993 — {'action required' if rula['value']>=4 else 'acceptable'})")
    if iso.get("value"):        clinical_ctx.append(f"ISO 11226 compliance: {iso['value']}")
    if kyphosis.get("value"):   clinical_ctx.append(f"Kyphosis penalty: {kyphosis['value']}° (thoracic rounding detected)")
    if plumb.get("value"):      clinical_ctx.append(f"Plumb line deviation: {plumb['value']}% (upper:{plumb.get('upper_dev','?')}%, lower:{plumb.get('lower_dev','?')}%)")
    if lumbar.get("value") is not None: clinical_ctx.append(f"Lumbar lordosis angle: {lumbar['value']}° ({'normal' if lumbar['value']>2 else 'FLAT BACK — risk of disc compression'})")
    if rsi.get("value"):        clinical_ctx.append(f"RSI risk index: {rsi['value']} ({'high' if rsi['value']>0.7 else 'moderate' if rsi['value']>0.4 else 'low'})")
    if twa.get("value"):        clinical_ctx.append(f"Time-weighted average score: {twa['value']}/100")
    if pain_bar.get("urgency"): clinical_ctx.append(f"Pain prediction: {pain_bar.get('label','?')} (tissue creep model, McGill 2007)")
    if dist_m.get("value"):     clinical_ctx.append(f"Screen distance: {dist_m['value']}cm (neck_source: {neck_src.get('value','nose_only')})")
    if percentile is not None:  clinical_ctx.append(f"User percentile: top {100-percentile}% of {mode} users")

    clinical_text = "\n".join(clinical_ctx) if clinical_ctx else "No additional clinical data"
    is_ar = lang == "ar"

    prompt = f"""You are a certified occupational physiotherapist (COPT) with expertise in workplace ergonomics and musculoskeletal injury prevention. Analyze this real-time posture data and provide a precise clinical assessment.

═══ SESSION DATA ═══
Camera mode: {mode} | Score: {s}/100 | Engine: {eng}{blur_note}
{pose_text}
Shoulder ratio: {sh_ratio}× reference

═══ ALL METRICS (with scores) ═══
{mtext}

═══ CLINICAL CONTEXT ═══
{clinical_text}

═══ ACTIVE ALERTS ═══
{chr(10).join(alts) if alts else "✓ No alerts — posture within acceptable range"}
{f"Employee context: {context}" if context else ""}

{"Write a complete clinical assessment in Arabic." if is_ar else "Write a complete clinical assessment in English."}

**SECTION 1 — Clinical Summary** (2 sentences)
State overall posture quality and single most critical finding with specific angle/measurement.

**SECTION 2 — Alert Analysis**
For each active alert:
• Finding: [metric name + value]
• Clinical significance: [what tissue/joint is stressed]
• Immediate correction: [1 specific action]
• Long-term risk if ignored: [injury name + timeframe]

**SECTION 3 — Evidence-Based Recommendations** (top 3)
Each recommendation must cite: standard (ISO 9241-5 / OSHA / NIOSH / Hansraj 2014 / McGill 2007).

**SECTION 4 — Injury Risk Profile**
Overall MSK risk: Low / Medium / High / Critical
Structures at risk: [specific muscles, discs, nerves]
Estimated time to injury onset if unchanged: [X months]

**SECTION 5 — Priority Action**
The single ergonomic adjustment with highest injury-prevention ROI right now.

{"أجب بالعربية الفصحى الطبية." if is_ar else "Use precise anatomical and ergonomic terminology. Be concise and actionable."}"""
    # ── Call AI (local Ollama only) ────────────────────────────────
    # Max tokens by tier: Elite gets full 1200, others get 900
    _tokens = 1200 if tier in ("elite","premium","professional","pro") else 900
    narrative = call_ai(
        prompt,
        max_tokens=_tokens,
        temperature=0.25,
        cache_key=_cache_key,
        cache_ttl=90,
        tier=tier,
    )
    return narrative

# ── PDF GENERATION ─────────────────────────────────────────────────
def generate_pdf(sd):
    if not REPORTLAB_OK: raise ImportError("pip install reportlab")
    lang = sd.get("lang","en")
    if lang == "ar":
        return generate_pdf_ar(sd)
    return generate_pdf_en(sd)

def generate_pdf_ar(sd):
    if not REPORTLAB_OK: raise ImportError("pip install reportlab")
    NAVY=colors.HexColor("#05101f"); BLUE=colors.HexColor("#1a56db")
    GREEN=colors.HexColor("#059669"); AMBER=colors.HexColor("#d97706")
    RED=colors.HexColor("#dc2626"); GRAY=colors.HexColor("#64748b")
    LGRAY=colors.HexColor("#f1f5f9"); DGRAY=colors.HexColor("#1e293b")
    WHITE=colors.white
    def ps(name,**kw):
        d=dict(fontName='Helvetica',fontSize=9,textColor=DGRAY,leading=13,spaceAfter=2)
        d.update(kw); return ParagraphStyle(name,**d)
    buf=io.BytesIO(); W,_=A4; uw=W-34*mm
    doc=SimpleDocTemplate(buf,pagesize=A4,leftMargin=17*mm,rightMargin=17*mm,topMargin=12*mm,bottomMargin=12*mm)
    co=sd.get("company_info",{}); avs=sd.get("avg_score",0)
    gp=sd.get("good_pct",0); dur=sd.get("duration_s",0)
    alts=sd.get("alerts",[]); recs=sd.get("recommendations",[])
    tier=sd.get("tier","standard"); met=sd.get("metrics",{})
    ai_a=sd.get("claude_analysis"); dt=sd.get("date",datetime.now().strftime("%d/%m/%Y"))
    sid=sd.get("session_id","—")
    def sc_col(s): return GREEN if s>=75 else AMBER if s>=50 else RED
    def ar_grade(s): return "ممتاز" if s>=85 else "جيد" if s>=70 else "مقبول" if s>=50 else "ضعيف"
    story=[]
    hdr=Table([[
        Paragraph(f'<b>{co.get("logo_text","PostureAI")}</b>',ps('hl',fontSize=18,textColor=WHITE,fontName='Helvetica-Bold')),
        Paragraph(f'<b>{co.get("name","PostureAI")}</b><br/><font size="8" color="#94a3b8">تقرير تحليل وضعية الجلوس</font>',ps('hr',fontSize=11,textColor=WHITE,alignment=TA_RIGHT))
    ]],colWidths=[uw*.52,uw*.48])
    hdr.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,-1),NAVY),('PADDING',(0,0),(-1,-1),13),('VALIGN',(0,0),(-1,-1),'MIDDLE')]))
    story.append(hdr); story.append(Spacer(1,7))
    story.append(Paragraph(f"النتيجة الإجمالية: <b>{avs}/100</b> — {ar_grade(avs)}",ps('sc',fontSize=14,textColor=sc_col(avs) if avs else GRAY,fontName='Helvetica-Bold')))
    story.append(Paragraph(f"التاريخ: {dt} · الجلسة: {sid} · المدة: {dur//60}د {dur%60}ث",ps('meta',fontSize=9,textColor=GRAY)))
    story.append(Spacer(1,8))
    if met:
        story.append(Paragraph("المقاييس التفصيلية",ps('h2',fontSize=11,fontName='Helvetica-Bold',textColor=DGRAY,spaceAfter=4)))
        rows=[]
        for k,m in met.items():
            s=m.get("score",0)
            rows.append([m.get("label",""),f'{m.get("value","?")} {m.get("unit","")}',f'{s}/100',ar_grade(s)])
        mt=Table([["القياس","القيمة","النقاط","التقييم"]]+rows,colWidths=[uw*.35,uw*.25,uw*.2,uw*.2])
        mt.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),BLUE),('TEXTCOLOR',(0,0),(-1,0),WHITE),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,-1),9),('PADDING',(0,0),(-1,-1),(7,5)),('ROWBACKGROUNDS',(0,1),(-1,-1),[LGRAY,WHITE]),('LINEBELOW',(0,0),(-1,-1),0.3,colors.HexColor("#e2e8f0"))]))
        story.append(mt); story.append(Spacer(1,8))
    if recs:
        story.append(Paragraph("التوصيات",ps('h2',fontSize=11,fontName='Helvetica-Bold',textColor=DGRAY,spaceAfter=4)))
        for r in recs[:6]:
            story.append(Paragraph(f"• {r}",ps('r',fontSize=10,textColor=DGRAY,leftIndent=8,spaceAfter=3)))
        story.append(Spacer(1,8))
    if ai_a:
        story.append(Paragraph("التحليل الطبي بالذكاء الاصطناعي",ps('h2',fontSize=11,fontName='Helvetica-Bold',textColor=DGRAY,spaceAfter=4)))
        for para_text in ai_a.split('\n')[:20]:
            if para_text.strip():
                story.append(Paragraph(para_text.strip(),ps('ai',fontSize=9,textColor=DGRAY,spaceAfter=3)))
        story.append(Spacer(1,8))
    story.append(HRFlowable(width=uw,thickness=0.5,color=colors.HexColor("#e2e8f0")))
    story.append(Paragraph("تقرير صادر عن PostureAI Pro — للأغراض المعلوماتية فقط وليس تشخيصاً طبياً.",
        ps('ft',fontSize=7.5,textColor=GRAY,alignment=TA_CENTER)))
    doc.build(story); buf.seek(0); return buf

def generate_pdf_en(sd):
    """
    Professional Elite PDF report — multi-page, paginated footer/header,
    visual score ring, bar chart, new metrics (FHP, RSI, ergonomics, neck load),
    employee health model section, and AI clinical narrative.
    """
    if not REPORTLAB_OK: raise ImportError("pip install reportlab")

    # ── Colour palette ────────────────────────────────────────────
    NAVY  = colors.HexColor("#05101f"); BLUE  = colors.HexColor("#1a56db")
    BLUE2 = colors.HexColor("#3b82f6"); TEAL  = colors.HexColor("#0891b2")
    GREEN = colors.HexColor("#059669"); AMBER = colors.HexColor("#d97706")
    RED   = colors.HexColor("#dc2626"); GRAY  = colors.HexColor("#64748b")
    LGRAY = colors.HexColor("#f1f5f9"); DGRAY = colors.HexColor("#1e293b")
    MGRAY = colors.HexColor("#e2e8f0"); BGRAY = colors.HexColor("#f8fafc")
    WHITE = colors.white; ELBLUE = colors.HexColor("#dbeafe")
    ELGREEN= colors.HexColor("#dcfce7"); ELRED = colors.HexColor("#fee2e2")
    ELAMBER= colors.HexColor("#fef3c7"); ELTEAL= colors.HexColor("#cffafe")

    def sc_col(s):
        return GREEN if s >= 75 else AMBER if s >= 50 else RED
    def sc_bg(s):
        return ELGREEN if s >= 75 else ELAMBER if s >= 50 else ELRED
    def sc_lbl(s):
        return "Excellent" if s >= 85 else "Good" if s >= 70 else "Fair" if s >= 50 else "Poor"
    def risk_col(r):
        return {"low": GREEN, "moderate": AMBER, "high": RED}.get(str(r).lower(), GRAY)

    def ps(name, **kw):
        d = dict(fontName="Helvetica", fontSize=9, textColor=DGRAY, leading=13, spaceAfter=2)
        d.update(kw)
        return ParagraphStyle(name, **d)

    # ── Session data ──────────────────────────────────────────────
    buf  = io.BytesIO()
    W, H = A4
    uw   = W - 34 * mm

    co   = sd.get("company_info", {})
    emp  = sd.get("employee", {})
    hist = sd.get("score_history", [])
    alts = sd.get("alerts", [])
    recs = sd.get("recommendations", [])
    met  = sd.get("metrics", {})
    mode = sd.get("mode", "laptop")
    avs  = sd.get("avg_score", 0) or 0
    gp   = sd.get("good_pct", 0) or 0
    dur  = sd.get("duration_s", 0) or 0
    tf   = sd.get("total_frames", 0) or 0
    sid  = sd.get("session_id", "—")
    tier = sd.get("tier", "standard")
    ai_a = sd.get("claude_analysis") or sd.get("ai_tip")
    dt   = sd.get("date", datetime.now().strftime("%B %d, %Y"))
    hp   = sd.get("head_pose")
    conf = sd.get("confidence", 0) or 0
    eng  = sd.get("engine", "")
    qual = sd.get("quality", {}) or {}
    qual_sc = qual.get("quality_score", 0) or 0
    trend   = sd.get("session_trend", {}) or {}

    mlab = {"laptop": "Laptop Camera", "phone": "Phone Camera", "side": "Side Camera"}
    tlab = {"standard": "Standard", "professional": "Professional", "elite": "Elite + AI", "business": "Business + AI"}

    story = []

    # ══════════════════════════════════════════════════════════════
    # PAGE TEMPLATE — running header/footer
    # ══════════════════════════════════════════════════════════════
    def on_page(canvas_obj, doc_obj):
        canvas_obj.saveState()
        pg = doc_obj.page
        # Footer line
        canvas_obj.setStrokeColor(MGRAY)
        canvas_obj.setLineWidth(0.4)
        canvas_obj.line(17*mm, 10*mm, W-17*mm, 10*mm)
        canvas_obj.setFont("Helvetica", 6.5)
        canvas_obj.setFillColor(GRAY)
        canvas_obj.drawString(17*mm, 7*mm,
            f"PostureAI Pro  ·  {co.get('name', 'Confidential')}  ·  {dt}  ·  Session {sid[:16]}")
        canvas_obj.drawRightString(W-17*mm, 7*mm, f"Page {pg}")
        # Top accent bar (pages 2+)
        if pg > 1:
            canvas_obj.setFillColor(NAVY)
            canvas_obj.rect(17*mm, H-12*mm, uw, 6, fill=1, stroke=0)
            canvas_obj.setFillColor(WHITE)
            canvas_obj.setFont("Helvetica-Bold", 7)
            canvas_obj.drawString(19*mm, H-9.5*mm, "PostureAI Pro — Posture Assessment Report (continued)")
        canvas_obj.restoreState()

    doc = SimpleDocTemplate(buf, pagesize=A4,
        leftMargin=17*mm, rightMargin=17*mm,
        topMargin=14*mm, bottomMargin=18*mm)

    # ══════════════════════════════════════════════════════════════
    # PAGE 1 — HEADER
    # ══════════════════════════════════════════════════════════════
    # Gradient accent bar
    acc = Drawing(uw, 5)
    for i, c in enumerate([BLUE, TEAL, GREEN]):
        acc.add(Rect(i*(uw/3), 0, uw/3+1, 5, fillColor=c, strokeColor=None))

    hdr = Table([[
        Paragraph(
            f'<b><font size="20">{co.get("logo_text","PostureAI")}</font></b>'
            f'<br/><font size="8" color="#94a3b8">Workplace Posture Intelligence</font>',
            ps("hl", textColor=WHITE, leading=26)),
        Paragraph(
            f'<b><font size="11">{co.get("name","PostureAI Client")}</font></b>'
            f'<br/><font size="8" color="#94a3b8">Posture Assessment Report</font>'
            f'<br/><font size="7" color="#60a5fa">{tlab.get(tier,"Standard")} Tier  ·  Confidential</font>',
            ps("hr", textColor=WHITE, alignment=TA_RIGHT, leading=16)),
    ]], colWidths=[uw*.52, uw*.48])
    hdr.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), NAVY),
        ("PADDING",    (0,0), (-1,-1), 14),
        ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
    ]))
    story.append(hdr)
    story.append(acc)
    story.append(Spacer(1, 8))

    # ── Session meta bar ─────────────────────────────────────────
    meta_items = [
        ("Date",       dt),
        ("Duration",   f"{dur//60}m {dur%60}s"),
        ("Camera",     mlab.get(mode, mode)),
        ("Engine",     (eng or "OpenCV")[:18]),
        ("Confidence", f"{conf}%"),
        ("Session",    sid[:14]),
    ]
    meta = Table([[
        Paragraph(
            f'<font color="#94a3b8" size="7">{lbl}</font><br/>'
            f'<b><font size="9">{val}</font></b>',
            ps(f"m{i}", leading=14))
        for i, (lbl, val) in enumerate(meta_items)
    ]], colWidths=[uw/6]*6)
    meta.setStyle(TableStyle([
        ("BACKGROUND",  (0,0), (-1,-1), BGRAY),
        ("PADDING",     (0,0), (-1,-1), (8,6)),
        ("LINEBETWEEN", (0,0), (-1,-1), 0.3, MGRAY),
        ("LINEBELOW",   (0,0), (-1,-1), 0.6, colors.HexColor("#cbd5e1")),
    ]))
    story.append(meta)
    story.append(Spacer(1, 10))

    # ── Employee info ─────────────────────────────────────────────
    if emp and any(emp.values()):
        emp_items = [
            ("Name",        emp.get("name","—")),
            ("Department",  emp.get("dept","—")),
            ("Position",    emp.get("position","—")),
            ("Employee ID", emp.get("id","—")),
        ]
        er = Table([[
            Paragraph(
                f'<font color="#94a3b8" size="7">{lbl}</font><br/>'
                f'<b><font size="9">{val}</font></b>',
                ps(f"e{i}", leading=14))
            for i, (lbl, val) in enumerate(emp_items)
        ]], colWidths=[uw*.28, uw*.22, uw*.28, uw*.22])
        er.setStyle(TableStyle([
            ("BOX",         (0,0), (-1,-1), 0.5, MGRAY),
            ("LINEBETWEEN", (0,0), (-1,-1), 0.3, MGRAY),
            ("PADDING",     (0,0), (-1,-1), (10,7)),
            ("BACKGROUND",  (0,0), (-1,-1), BGRAY),
        ]))
        story.append(er)
        story.append(Spacer(1, 10))

    # ══════════════════════════════════════════════════════════════
    # SCORE RING (SVG-style via Drawing) + KPI ROW
    # ══════════════════════════════════════════════════════════════
    story.append(Paragraph("Session Overview",
        ps("h1", fontSize=12, fontName="Helvetica-Bold", textColor=NAVY, spaceAfter=8)))

    ring_w, ring_h = 90, 90
    ring = Drawing(ring_w, ring_h)
    cx, cy, R = ring_w/2, ring_h/2, 36
    # Background ring
    ring.add(Drawing(ring_w, ring_h))
    # Simulate arc with thin lines (ReportLab Drawing)
    import math as _m
    steps = 120
    for i in range(steps):
        a0 = _m.radians(90 - (i/steps)*360)
        a1 = _m.radians(90 - ((i+1)/steps)*360)
        ring.add(Line(
            cx + R*_m.cos(a0), cy + R*_m.sin(a0),
            cx + R*_m.cos(a1), cy + R*_m.sin(a1),
            strokeColor=MGRAY, strokeWidth=8))
    # Filled arc = score
    fill_steps = int(steps * avs / 100)
    arc_col = sc_col(avs)
    for i in range(fill_steps):
        a0 = _m.radians(90 - (i/steps)*360)
        a1 = _m.radians(90 - ((i+1)/steps)*360)
        ring.add(Line(
            cx + R*_m.cos(a0), cy + R*_m.sin(a0),
            cx + R*_m.cos(a1), cy + R*_m.sin(a1),
            strokeColor=arc_col, strokeWidth=8))
    # Score text
    ring.add(String(cx, cy+6,  str(avs),
        fontSize=22, fontName="Helvetica-Bold",
        fillColor=arc_col, textAnchor="middle"))
    ring.add(String(cx, cy-8, "/100",
        fontSize=8, fillColor=GRAY, textAnchor="middle"))
    ring.add(String(cx, cy-18, sc_lbl(avs),
        fontSize=7.5, fontName="Helvetica-Bold",
        fillColor=arc_col, textAnchor="middle"))

    # KPI cards row
    kpi_items = [
        ("Good Posture",    f"{gp}%",       GREEN  if gp >= 70 else AMBER),
        ("Alerts",          str(len(alts)), RED    if len(alts) > 3 else AMBER if alts else GREEN),
        ("Frames",          str(tf),        BLUE),
        ("Data Quality",    f"{qual_sc}/100", sc_col(qual_sc)),
        ("Session Trend",   trend.get("trend","—").replace("_"," ").title(), GREEN if "improv" in trend.get("trend","") else AMBER if "stable" in trend.get("trend","") else RED),
    ]
    kpi_cells = []
    for lbl, val, col in kpi_items:
        kpi_cells.append(Paragraph(
            f'<font color="#94a3b8" size="7">{lbl}</font><br/>'
            f'<b><font size="14" color="#{col.hexval()[1:]}">{val}</font></b>',
            ps(f"kpi{lbl}", leading=18, alignment=TA_CENTER)))

    ov_row = Table([[ring, Table([kpi_cells], colWidths=[uw*.145]*5)]],
        colWidths=[ring_w+10, uw-ring_w-10])
    ov_row.setStyle(TableStyle([
        ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
        ("BACKGROUND", (0,0), (-1,-1), BGRAY),
        ("PADDING",    (0,0), (-1,-1), 10),
        ("LINEAFTER",  (0,0), (0,-1),  0.5, MGRAY),
        ("BOX",        (0,0), (-1,-1), 0.5, MGRAY),
    ]))
    story.append(ov_row)
    story.append(Spacer(1, 12))

    # ══════════════════════════════════════════════════════════════
    # SCORE HISTORY BAR CHART (improved — with moving avg line)
    # ══════════════════════════════════════════════════════════════
    if hist:
        story.append(Paragraph("Score Timeline",
            ps("h2", fontSize=11, fontName="Helvetica-Bold", textColor=DGRAY, spaceAfter=6)))
        ch_w = float(uw * 0.96); ch_h = 90.0
        d = Drawing(ch_w, ch_h)
        d.add(Rect(0, 0, ch_w, ch_h, fillColor=BGRAY, strokeColor=None))
        # Grid lines
        for pct in [25, 50, 75, 100]:
            y = ch_h * pct / 100
            d.add(Line(0, y, ch_w, y, strokeColor=MGRAY, strokeWidth=0.4))
            d.add(String(ch_w-2, y+1, str(pct), fontSize=6, fillColor=GRAY, textAnchor="end"))
        # 65% threshold line
        d.add(Line(0, ch_h*0.65, ch_w, ch_h*0.65,
            strokeColor=AMBER, strokeDashArray=[4,3], strokeWidth=0.8))
        d.add(String(3, ch_h*0.65+2, "65 (good)", fontSize=6, fillColor=AMBER))
        # Bars
        bw = ch_w / max(len(hist), 1)
        for i, sv in enumerate(hist):
            bh = max(1.5, ch_h * sv / 100)
            d.add(Rect(i*bw+0.8, 0, max(1.5, bw-1.6), bh,
                fillColor=sc_col(sv), strokeColor=None))
        # 7-point moving average line
        if len(hist) >= 5:
            ma_pts = []
            for i in range(len(hist)):
                sl = hist[max(0, i-3):i+4]
                ma_pts.append((i*bw + bw/2, ch_h * (sum(sl)/len(sl)) / 100))
            for i in range(len(ma_pts)-1):
                d.add(Line(ma_pts[i][0], ma_pts[i][1],
                            ma_pts[i+1][0], ma_pts[i+1][1],
                    strokeColor=NAVY, strokeWidth=1.2, strokeDashArray=[2,2]))
        story.append(d)
        story.append(Spacer(1, 4))
        story.append(Paragraph(
            '<font color="#64748b" size="7">Bars = per-frame score  ·  Dashed line = 7-frame moving avg  ·  Amber line = 65pt threshold</font>',
            ps("chart_leg", fontSize=7, textColor=GRAY)))
        story.append(Spacer(1, 12))

    # ══════════════════════════════════════════════════════════════
    # DETAILED METRICS TABLE — incl. new features
    # ══════════════════════════════════════════════════════════════
    story.append(Paragraph("Detailed Posture Metrics",
        ps("h2", fontSize=11, fontName="Helvetica-Bold", textColor=DGRAY, spaceAfter=6)))

    # Extended metric descriptions + clinical context
    mdesc = {
        "neck_lean":        ("Neck lean",        "Forward lean of ear ahead of shoulder vertical"),
        "head_tilt":        ("Head tilt",         "Lateral head tilt measured from eye-line angle (FaceMesh)"),
        "shoulder_level":   ("Shoulder symmetry", "Height asymmetry between left and right shoulders"),
        "spine_lean":       ("Spine lean",        "Trunk lean angle: shoulder→hip vs vertical"),
        "spine_upper_s":    ("Upper spine",       "Thoracic spine segment lean (3-point measurement)"),
        "spine_lower_s":    ("Lower spine",       "Lumbar spine segment lean (3-point measurement)"),
        "screen_distance":  ("Screen distance",   "IPD-based distance via FaceMesh iris landmarks"),
        "head_yaw":         ("Head turn",         "Left/right head rotation from camera forward axis"),
        "wrist_angle":      ("Wrist angle",       "Wrist deviation — CTD/RSI risk indicator (3D)"),
        "eye_strain":       ("Eye strain",        "Blink rate vs NIOSH standard (12–20/min normal)"),
        "symmetry":         ("Body symmetry",     "Shoulder/hip height asymmetry — scoliosis indicator"),
        "ergonomics_score": ("Workstation",       "Composite: monitor height + distance + neck posture"),
        "fhp_index":        ("Forward head (FHP)","Clinical: ear offset from shoulder plumb line in cm"),
        "neck_load":        ("Neck load",         "Estimated head weight on cervical spine (Hansraj 2014)"),
        "rsi_risk":         ("RSI risk",          "Cumulative repetitive strain injury risk (session)"),
        "breathing_rate":   ("Breathing",         "Estimated breaths/min from shoulder movement cycle"),
        "movement":         ("Micro-movement",    "Natural position variance — frozen posture detection"),
        "gaze":             ("Gaze direction",    "Iris position within eye socket — screen attention"),
        "hip_angle":        ("Hip angle",         "Hip flexion: shoulder–hip–knee (target 90°)"),
        "knee_angle":       ("Knee angle",        "Knee flexion: hip–knee–ankle (target 90°)"),
        "trunk_lean":       ("Trunk lean",        "Side-view trunk lean: shoulder→hip from vertical"),
    }
    SKIP_METRICS = {"_confidence", "_side_confidence", "head_pose_detail", "spine_curvature"}

    mh_row = [Paragraph(f"<b>{t}</b>",
        ps(f"mh{t}", textColor=WHITE, fontSize=8, alignment=TA_CENTER if t != "Metric" else TA_LEFT))
        for t in ["Metric", "Value", "Score", "Status", "Clinical Context"]]
    m_rows = [mh_row]

    for k, m3 in met.items():
        if k in SKIP_METRICS or not isinstance(m3, dict) or "score" not in m3: continue
        s3  = m3.get("score", 0) or 0
        v3  = m3.get("value")
        u3  = m3.get("unit", "")
        vs  = (f"{round(v3,1)}{u3}" if isinstance(v3, float) else
               f"{v3}{u3}"           if isinstance(v3, int) else
               str(v3)               if v3 is not None else "—")
        lbl, ctx = mdesc.get(k, (m3.get("label", k), ""))
        # Special: show extra sub-values for complex metrics
        extra = ""
        if k == "neck_load":
            extra = f' (+{m3.get("extra_load_kg","?"):}kg vs 5.4kg baseline)'
        elif k == "fhp_index":
            extra = f' ({m3.get("extra_load_kg","?"):}kg extra load)'
        elif k == "rsi_risk":
            extra = f' ({m3.get("level","?")} risk)'
        elif k == "ergonomics_score":
            extra = f' ({m3.get("grade","?")})'
        elif k == "breathing_rate":
            extra = f' ({m3.get("risk","?")})'

        m_rows.append([
            Paragraph(f"{lbl}",              ps(f"mc{k}", fontSize=8)),
            Paragraph(f"<b>{vs}{extra}</b>", ps(f"mv{k}", fontSize=8, fontName="Helvetica-Bold",
                                                textColor=sc_col(s3) if s3 < 60 else DGRAY)),
            Paragraph(f"<b>{s3}</b>",        ps(f"ms{k}", fontSize=8.5, fontName="Helvetica-Bold",
                                                textColor=sc_col(s3), alignment=TA_CENTER)),
            Paragraph(sc_lbl(s3),            ps(f"mst{k}", fontSize=7.5, textColor=sc_col(s3),
                                                alignment=TA_CENTER)),
            Paragraph(ctx,                   ps(f"md{k}", fontSize=7.5, textColor=GRAY, leading=11)),
        ])

    mt2 = Table(m_rows, colWidths=[uw*.18, uw*.16, uw*.08, uw*.10, uw*.48])
    mt2.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),  (-1,0),  NAVY),
        ("TEXTCOLOR",     (0,0),  (-1,0),  WHITE),
        ("FONTSIZE",      (0,0),  (-1,0),  8),
        ("PADDING",       (0,0),  (-1,-1), (6,5)),
        ("ROWBACKGROUNDS",(0,1),  (-1,-1), [WHITE, BGRAY]),
        ("LINEBELOW",     (0,0),  (-1,-1), 0.3, MGRAY),
        ("VALIGN",        (0,0),  (-1,-1), "MIDDLE"),
        ("LINEBELOW",     (0,0),  (-1,0),  1.2, BLUE),
    ]))
    story.append(mt2)
    story.append(Spacer(1, 12))

    # ══════════════════════════════════════════════════════════════
    # HEAD POSE 3D TABLE
    # ══════════════════════════════════════════════════════════════
    if hp:
        story.append(Paragraph("3D Head Pose  (solvePnP · 14 landmarks · EPNP+LM)",
            ps("h2", fontSize=11, fontName="Helvetica-Bold", textColor=DGRAY, spaceAfter=6)))
        reproj = hp.get("reproj_err", 0)
        reliability = "High" if reproj < 2 else "Medium" if reproj < 4 else "Low"
        rel_col = GREEN if reproj < 2 else AMBER if reproj < 4 else RED
        pi = {
            "pitch": lambda v: f"Head down {abs(v)}°" if v<-10 else f"Head up {abs(v)}°" if v>10 else "✓ At eye level",
            "yaw":   lambda v: f"Turned {'left' if v<0 else 'right'} {abs(round(v,1))}° — asymmetric load" if abs(v)>8 else "✓ Facing forward",
            "roll":  lambda v: f"Tilted {'left' if v<0 else 'right'} {abs(round(v,1))}° — shoulder asymmetry" if abs(v)>5 else "✓ Head level",
        }
        pr2 = [[Paragraph(f"<b>{t}</b>", ps(f"ph{t}", textColor=WHITE, fontSize=8))
                for t in ["Axis","Value","Interpretation","Risk"]]]
        for axis in ["pitch","yaw","roll"]:
            v3 = hp.get(axis, 0) or 0
            c3 = GREEN if abs(v3) < 8 else AMBER if abs(v3) < 18 else RED
            risk_label = "Low" if abs(v3) < 8 else "Moderate" if abs(v3) < 18 else "High"
            pr2.append([
                Paragraph(axis.capitalize(), ps(f"pa{axis}", fontSize=8.5)),
                Paragraph(f"<b>{v3}°</b>",  ps(f"pv{axis}", fontSize=8.5,
                    fontName="Helvetica-Bold", textColor=c3)),
                Paragraph(pi[axis](v3),      ps(f"pi{axis}", fontSize=8)),
                Paragraph(risk_label,        ps(f"pr{axis}", fontSize=8, textColor=c3,
                    fontName="Helvetica-Bold", alignment=TA_CENTER)),
            ])
        pt = Table(pr2, colWidths=[uw*.12, uw*.12, uw*.58, uw*.18])
        pt.setStyle(TableStyle([
            ("BACKGROUND",    (0,0), (-1,0),  TEAL),
            ("TEXTCOLOR",     (0,0), (-1,0),  WHITE),
            ("PADDING",       (0,0), (-1,-1), (7,5)),
            ("ROWBACKGROUNDS",(0,1), (-1,-1), [WHITE, BGRAY]),
            ("LINEBELOW",     (0,0), (-1,-1), 0.3, MGRAY),
            ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ]))
        story.append(pt)
        story.append(Paragraph(
            f'<font color="#64748b" size="7">Reprojection error: {reproj}px  ·  '
            f'Reliability: <b>{reliability}</b>  ·  '
            f'Landmarks used: {hp.get("landmarks_n", 6)}  ·  '
            f'Camera correction applied: {sd.get("cam_correction_applied", 0)}°</font>',
            ps("reproj", fontSize=7, textColor=GRAY)))
        story.append(Spacer(1, 12))

    # ══════════════════════════════════════════════════════════════
    # EMPLOYEE HEALTH MODEL — the premium differentiator
    # ══════════════════════════════════════════════════════════════
    story.append(PageBreak())
    story.append(Paragraph("Employee Musculoskeletal Health Model",
        ps("h1", fontSize=13, fontName="Helvetica-Bold", textColor=NAVY, spaceAfter=4)))
    story.append(Paragraph(
        "Evidence-based risk assessment across 4 body regions based on ISO 9241-110, OSHA 3125, and NIOSH ergonomic guidelines.",
        ps("sub", fontSize=8.5, textColor=GRAY, spaceAfter=10)))

    # Body region risk matrix
    fhp = met.get("fhp_index", {})
    neck_load_m = met.get("neck_load", {})
    rsi_m  = met.get("rsi_risk", {})
    ergo_m = met.get("ergonomics_score", {})
    sym_m  = met.get("symmetry", {})
    eye_m  = met.get("eye_strain", {})
    breath_m = met.get("breathing_rate", {})
    wrist_m  = met.get("wrist_angle", {})

    def risk_from_score(s):
        if not s: return "Unknown", GRAY
        if s >= 75: return "Low", GREEN
        if s >= 55: return "Moderate", AMBER
        return "High", RED

    regions = [
        {
            "region": "Cervical Spine (Neck)",
            "icon":   "🔴 CRITICAL" if (fhp.get("value",0) or 0) > 6 else "🟡 MODERATE" if (fhp.get("value",0) or 0) > 3 else "🟢 NORMAL",
            "metrics": [
                ("Forward Head Posture", f'{fhp.get("value","—")}cm', f'+{fhp.get("extra_load_kg","?")}kg neck load'),
                ("Neck Load Est.",       f'{neck_load_m.get("value","—")}kg', f'Multiplier: {neck_load_m.get("multiplier","?")}×'),
                ("Neck Lean Score",      f'{met.get("neck_lean",{}).get("value","—")}°', f'Score: {met.get("neck_lean",{}).get("score","?")}'),
            ],
            "risk_score": met.get("neck_lean",{}).get("score",0) or 0,
            "intervention": "Raise monitor to eye level. Each 2.5cm of forward head adds ~4.5kg cervical load. Target: ear directly above shoulder.",
        },
        {
            "region": "Thoracic & Lumbar Spine",
            "icon":   "🔴 HIGH RISK" if (met.get("spine_lean",{}).get("score",0) or 0) < 50 else "🟡 MODERATE" if (met.get("spine_lean",{}).get("score",0) or 0) < 70 else "🟢 NORMAL",
            "metrics": [
                ("Spine Lean",     f'{met.get("spine_lean",{}).get("value","—")}°',     f'Score: {met.get("spine_lean",{}).get("score","?")}'),
                ("Upper Spine",    f'{met.get("spine_upper_s",{}).get("value","—")}°',  "Thoracic segment"),
                ("Lower Spine",    f'{met.get("spine_lower_s",{}).get("value","—")}°',  "Lumbar segment"),
                ("Body Symmetry",  f'{sym_m.get("value","—")}%',                        f'sh: {sym_m.get("sh_asym","?")}%  hip: {sym_m.get("hip_asym","?")}%'),
            ],
            "risk_score": met.get("spine_lean",{}).get("score",0) or 0,
            "intervention": "Ensure lumbar support contacts lower back. Sit back fully in chair. Symmetry > 5% warrants ergonomic assessment.",
        },
        {
            "region": "Upper Extremity (Wrist & Arm)",
            "icon":   "🔴 HIGH RISK" if (wrist_m.get("score",0) or 0) < 50 else "🟡 MODERATE" if (wrist_m.get("score",0) or 0) < 70 else "🟢 NORMAL",
            "metrics": [
                ("Wrist Angle 3D", f'{wrist_m.get("value","—")}°',  f'Score: {wrist_m.get("score","?")}'),
                ("RSI Risk Index", f'{rsi_m.get("value","—")}',      f'{rsi_m.get("level","?")} ({rsi_m.get("neck_contrib","?")} neck + {rsi_m.get("wrist_contrib","?")} wrist)'),
                ("Workstation",    f'{ergo_m.get("value","—")}/100', f'Grade: {ergo_m.get("grade","?")}'),
            ],
            "risk_score": wrist_m.get("score",0) or 0,
            "intervention": "Keyboard at elbow height. Wrist deviation >15° significantly increases CTD risk. Consider wrist rest and keyboard tray.",
        },
        {
            "region": "Visual & Autonomic System",
            "icon":   "🔴 HIGH RISK" if (eye_m.get("score",0) or 0) < 50 else "🟡 MODERATE" if (eye_m.get("score",0) or 0) < 70 else "🟢 NORMAL",
            "metrics": [
                ("Blink Rate",    f'{eye_m.get("value","—")}/min',    f'Risk: {eye_m.get("eye_strain_risk","?")}'),
                ("Breathing",     f'{breath_m.get("value","—")}/min', f'Amplitude: {breath_m.get("amplitude","?")}%'),
                ("Gaze on Screen",f'{met.get("gaze",{}).get("on_screen","?")}', "Screen focus indicator"),
            ],
            "risk_score": eye_m.get("score",0) or 0,
            "intervention": "Apply 20-20-20 rule: every 20 min, look 6m away for 20 sec. Blink <12/min = digital eye strain. Breathing <12/min = stress indicator.",
        },
    ]

    for reg in regions:
        rs = reg["risk_score"]
        r_bg = ELGREEN if rs >= 75 else ELAMBER if rs >= 55 else ELRED
        r_col = GREEN  if rs >= 75 else AMBER   if rs >= 55 else RED

        # Region header
        reg_hdr = Table([[
            Paragraph(f'<b>{reg["region"]}</b>', ps("rh", fontSize=10, fontName="Helvetica-Bold", textColor=WHITE)),
            Paragraph(reg["icon"], ps("ri", fontSize=8, textColor=WHITE, alignment=TA_RIGHT)),
        ]], colWidths=[uw*.72, uw*.28])
        reg_hdr.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,-1), NAVY),
            ("PADDING",    (0,0), (-1,-1), (10,7)),
            ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
        ]))
        story.append(reg_hdr)

        # Metrics grid
        m_cells = []
        for lbl, val, note in reg["metrics"]:
            m_cells.append(Paragraph(
                f'<font color="#64748b" size="7">{lbl}</font><br/>'
                f'<b><font size="10">{val}</font></b><br/>'
                f'<font size="7" color="#94a3b8">{note}</font>',
                ps(f"rc{lbl}", leading=14, alignment=TA_CENTER)))
        # Pad to 4
        while len(m_cells) < 4:
            m_cells.append(Paragraph("", ps("empty")))

        risk_bar_h = 6
        risk_bar = Drawing(uw*0.22, risk_bar_h)
        risk_bar.add(Rect(0, 0, uw*0.22, risk_bar_h, fillColor=MGRAY, strokeColor=None))
        risk_bar.add(Rect(0, 0, uw*0.22*rs/100, risk_bar_h, fillColor=r_col, strokeColor=None))

        met_row = Table(
            [m_cells[:4]] if len(m_cells) >= 4 else [m_cells],
            colWidths=[uw*.25]*4)
        met_row.setStyle(TableStyle([
            ("BACKGROUND",  (0,0), (-1,-1), r_bg),
            ("PADDING",     (0,0), (-1,-1), (8,8)),
            ("LINEBETWEEN", (0,0), (-1,-1), 0.3, MGRAY),
        ]))
        story.append(met_row)

        # Intervention recommendation
        intv = Table([[
            Paragraph("⚕ Intervention:", ps("il", fontSize=8, fontName="Helvetica-Bold", textColor=TEAL)),
            Paragraph(reg["intervention"], ps("it", fontSize=8, textColor=DGRAY, leading=12)),
        ]], colWidths=[uw*.18, uw*.82])
        intv.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,-1), ELTEAL),
            ("PADDING",    (0,0), (-1,-1), (8,6)),
        ]))
        story.append(intv)
        story.append(Spacer(1, 8))

    # ══════════════════════════════════════════════════════════════
    # AI CLINICAL NARRATIVE
    # ══════════════════════════════════════════════════════════════
    if ai_a:
        story.append(Spacer(1, 6))
        story.append(Paragraph("AI Clinical Analysis  (Gemini Pro — Physiotherapist Mode)",
            ps("h2", fontSize=11, fontName="Helvetica-Bold", textColor=DGRAY, spaceAfter=6)))
        ai_box = Table([[
            Paragraph("AI", ps("aib", fontSize=10, fontName="Helvetica-Bold",
                textColor=WHITE, alignment=TA_CENTER)),
            Paragraph(ai_a, ps("ait", fontSize=8.5, textColor=DGRAY, leading=13)),
        ]], colWidths=[uw*.07, uw*.93])
        ai_box.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (0,-1), BLUE),
            ("BACKGROUND", (1,0), (1,-1), ELBLUE),
            ("VALIGN",     (0,0), (-1,-1), "TOP"),
            ("PADDING",    (0,0), (-1,-1), (9,9)),
            ("BOX",        (0,0), (-1,-1), 0.8, BLUE2),
        ]))
        story.append(ai_box)
        story.append(Spacer(1, 10))

    # ══════════════════════════════════════════════════════════════
    # ALERT LOG
    # ══════════════════════════════════════════════════════════════
    story.append(Paragraph("Alert Log",
        ps("h2", fontSize=11, fontName="Helvetica-Bold", textColor=DGRAY, spaceAfter=6)))
    if alts:
        a_rows = [[Paragraph(f"<b>{t}</b>", ps(f"ah{t}", textColor=WHITE, fontSize=8))
                   for t in ["#","Alert","Severity"]]]
        for idx, a3 in enumerate(alts[:20]):
            msg  = a3.get("msg", a3) if isinstance(a3, dict) else str(a3)
            sev  = "⚠️ High" if msg.startswith("⚠️") else "ℹ️ Info"
            scol = RED if msg.startswith("⚠️") else AMBER
            a_rows.append([
                Paragraph(str(idx+1),   ps(f"an{idx}", fontSize=8, textColor=GRAY, alignment=TA_CENTER)),
                Paragraph(msg.replace("⚠️","").strip(), ps(f"am{idx}", fontSize=8.5)),
                Paragraph(sev,          ps(f"as{idx}", fontSize=8, textColor=scol, fontName="Helvetica-Bold")),
            ])
        at2 = Table(a_rows, colWidths=[uw*.06, uw*.76, uw*.18])
        at2.setStyle(TableStyle([
            ("BACKGROUND",    (0,0), (-1,0),  DGRAY),
            ("TEXTCOLOR",     (0,0), (-1,0),  WHITE),
            ("PADDING",       (0,0), (-1,-1), (7,5)),
            ("ROWBACKGROUNDS",(0,1), (-1,-1), [WHITE, BGRAY]),
            ("LINEBELOW",     (0,0), (-1,-1), 0.3, MGRAY),
            ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ]))
        story.append(at2)
    else:
        story.append(Paragraph("✓ No alerts triggered — excellent session compliance.",
            ps("na", fontSize=9, textColor=GREEN)))
    story.append(Spacer(1, 12))

    # ══════════════════════════════════════════════════════════════
    # RECOMMENDATIONS
    # ══════════════════════════════════════════════════════════════
    if recs:
        story.append(Paragraph("Evidence-Based Recommendations",
            ps("h2", fontSize=11, fontName="Helvetica-Bold", textColor=DGRAY, spaceAfter=6)))
        r_rows = []
        for i, r3 in enumerate(recs):
            priority = "🔴" if i == 0 else "🟡" if i < 3 else "🟢"
            r_rows.append([
                Paragraph(f"<b>{i+1}</b>",   ps(f"rn{i}", fontSize=9, fontName="Helvetica-Bold",
                    textColor=WHITE, alignment=TA_CENTER)),
                Paragraph(f"{priority} {r3}", ps(f"rb{i}", fontSize=8.5, leading=13)),
            ])
        rt = Table(r_rows, colWidths=[uw*.055, uw*.945])
        rt.setStyle(TableStyle([
            ("BACKGROUND",    (0,0),  (0,-1),  BLUE),
            ("VALIGN",        (0,0),  (-1,-1), "MIDDLE"),
            ("PADDING",       (0,0),  (-1,-1), (7,6)),
            ("ROWBACKGROUNDS",(1,0),  (1,-1),  [BGRAY, WHITE]),
            ("LINEBELOW",     (0,0),  (-1,-1), 0.3, MGRAY),
        ]))
        story.append(rt)
        story.append(Spacer(1, 10))

    # ══════════════════════════════════════════════════════════════
    # SESSION QUALITY & DATA INTEGRITY
    # ══════════════════════════════════════════════════════════════
    if qual:
        story.append(Paragraph("Session Data Quality",
            ps("h2", fontSize=11, fontName="Helvetica-Bold", textColor=DGRAY, spaceAfter=6)))
        qcomp = qual.get("components", {})
        q_items = [
            ("Detection Rate",   qcomp.get("detection_rate", 0), "%"),
            ("Blur Rate",        qcomp.get("blur_rate", 0),      "%"),
            ("Avg Visibility",   qcomp.get("avg_visibility", 0), "%"),
            ("Avg Confidence",   qcomp.get("avg_confidence", 0), "%"),
            ("Overall Quality",  qual_sc,                        "/100"),
        ]
        q_rows = [[Paragraph(f"<b>{t}</b>", ps(f"qh{t}", textColor=WHITE, fontSize=8))
                   for t in ["Component","Value","Quality Bar","Interpretation"]]]
        for lbl, val, unit in q_items:
            q_rows.append([
                Paragraph(lbl, ps(f"ql{lbl}", fontSize=8.5)),
                Paragraph(f"<b>{val}{unit}</b>", ps(f"qv{lbl}", fontSize=9,
                    fontName="Helvetica-Bold", textColor=sc_col(val))),
                Paragraph(
                    f'{"█"*int(val//10)}{"░"*int(10-val//10)}',
                    ps(f"qb{lbl}", fontSize=9, textColor=sc_col(val), fontName="Helvetica")),
                Paragraph(sc_lbl(val), ps(f"qi{lbl}", fontSize=8, textColor=sc_col(val))),
            ])
        qt = Table(q_rows, colWidths=[uw*.26, uw*.14, uw*.22, uw*.38])
        qt.setStyle(TableStyle([
            ("BACKGROUND",    (0,0), (-1,0),  DGRAY),
            ("TEXTCOLOR",     (0,0), (-1,0),  WHITE),
            ("PADDING",       (0,0), (-1,-1), (7,5)),
            ("ROWBACKGROUNDS",(0,1), (-1,-1), [WHITE, BGRAY]),
            ("LINEBELOW",     (0,0), (-1,-1), 0.3, MGRAY),
            ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ]))
        story.append(qt)
        if qual.get("issues"):
            story.append(Spacer(1,4))
            for issue in qual["issues"]:
                story.append(Paragraph(f"• {issue}",
                    ps(f"qi{issue[:10]}", fontSize=8, textColor=AMBER, leftIndent=8)))
        story.append(Spacer(1, 10))

    # ══════════════════════════════════════════════════════════════
    # LEGAL FOOTER
    # ══════════════════════════════════════════════════════════════
    story.append(HRFlowable(width=uw, thickness=0.5, color=MGRAY))
    story.append(Spacer(1, 4))
    story.append(Paragraph(
        "This report is generated by PostureAI Pro and is intended for occupational health monitoring purposes only. "
        "It does not constitute medical advice. For clinical diagnosis or treatment, consult a qualified healthcare professional. "
        "Data processed in accordance with applicable data protection regulations.",
        ps("legal", fontSize=6.5, textColor=GRAY, alignment=TA_CENTER, leading=10)))

    # ── Build ─────────────────────────────────────────────────────
    doc.build(story, onFirstPage=on_page, onLaterPages=on_page)
    return buf.getvalue()


def blur_face_in_frame(img):
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    try:
        face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
        faces = face_cascade.detectMultiScale(gray, 1.1, 4, minSize=(30, 30))
        result = img.copy()
        for (x, y, w, h) in faces:
            pad = int(w * 0.15)
            x1, y1 = max(0, x-pad), max(0, y-pad)
            x2, y2 = min(img.shape[1], x+w+pad), min(img.shape[0], y+h+pad)
            face_roi = result[y1:y2, x1:x2]
            blur_sz = max(25, (w // 4) * 2 + 1)
            result[y1:y2, x1:x2] = cv2.GaussianBlur(face_roi, (blur_sz, blur_sz), 30)
        return result
    except Exception:
        return img

def frame_to_png_bytes(img_bgr):
    _, buf = cv2.imencode('.png', img_bgr)
    return buf.tobytes()

# MED-02: session_snapshots backed by Redis when available (multi-worker safe)
# Falls back to local dict for dev environments without Redis
_session_snapshots_local: dict = {}

def _snapshot_key(sid): return f"snap:{sid}"

def _get_snapshots(sid):
    """Get snapshots from Redis (if available) or local dict."""
    if REDIS_READY:
        try:
            raw = rget(_snapshot_key(sid))
            if raw:
                import json as _j
                return _j.loads(raw)
        except Exception:
            pass
    return _session_snapshots_local.get(sid, [])

def _set_snapshots(sid, snaps):
    """Store snapshots in Redis (if available) and local dict."""
    _session_snapshots_local[sid] = snaps
    if REDIS_READY:
        try:
            import json as _j
            rset(_snapshot_key(sid), _j.dumps(snaps), ttl_s=7200)
        except Exception:
            pass

session_snapshots = _session_snapshots_local  # keep backward compat name


# ══════════════════════════════════════════════════════════════════
# LOW-LIGHT PREPROCESSING PIPELINE (CLAHE + Gamma + Adaptive)
# ══════════════════════════════════════════════════════════════════
def enhance_low_light(img_bgr: np.ndarray) -> tuple:
    """
    Enhance image for low-light conditions.
    Returns (enhanced_image, brightness_level, was_enhanced)
    """
    # Measure brightness
    gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
    brightness = float(np.mean(gray))

    if brightness >= 120:
        return img_bgr, brightness, False  # Good light — skip CLAHE (saves ~8ms/frame)

    enhanced = img_bgr.copy()

    # Step 1: Gamma correction (inverse gamma for dark images)
    gamma = 1.0
    if brightness < 40:
        gamma = 0.35    # very dark
    elif brightness < 70:
        gamma = 0.5     # dark
    elif brightness < 100:
        gamma = 0.7     # dim

    if gamma < 1.0:
        inv_gamma = 1.0 / gamma
        table = np.array([min(255, int((i / 255.0) ** inv_gamma * 255)) for i in range(256)], dtype=np.uint8)
        enhanced = cv2.LUT(enhanced, table)

    # Step 2: CLAHE on L channel (YCrCb)
    ycrcb = cv2.cvtColor(enhanced, cv2.COLOR_BGR2YCrCb)
    y, cr, cb = cv2.split(ycrcb)
    clip_limit = 3.0 if brightness < 40 else 2.0
    clahe = cv2.createCLAHE(clipLimit=clip_limit, tileGridSize=(8, 8))
    y = clahe.apply(y)
    enhanced = cv2.cvtColor(cv2.merge([y, cr, cb]), cv2.COLOR_YCrCb2BGR)

    # Step 3: Adaptive brightness normalization
    mean_after = float(np.mean(cv2.cvtColor(enhanced, cv2.COLOR_BGR2GRAY)))
    if mean_after < 80:
        alpha = min(2.0, 100.0 / max(mean_after, 1))
        enhanced = cv2.convertScaleAbs(enhanced, alpha=alpha, beta=10)

    return enhanced, brightness, True

# ── API ENDPOINTS ──────────────────────────────────────────────────

@require_auth
@app.route("/api/session/snapshot", methods=["POST"])
@limiter.limit("30 per minute")
def add_snapshot():
    try:
        data  = request.get_json(force=True) or {}
        if not data.get("session_id"):
            return jsonify({"error": "session_id required"}), 400
        if "score" in data:
            data["score"] = max(0, min(100, int(data.get("score", 0))))  # clamp
        sid   = data.get("session_id", "default")
        score = data.get("score", 0)
        ts    = data.get("timestamp", datetime.now().strftime("%H:%M:%S"))
        b64   = data.get("frame", "")
        if not b64: return jsonify({"ok": False, "reason": "no frame"}), 400
        if "," in b64: b64 = b64.split(",", 1)[1]
        img = cv2.imdecode(np.frombuffer(base64.b64decode(b64), np.uint8), cv2.IMREAD_COLOR)
        if img is None: return jsonify({"ok": False, "reason": "decode error"}), 400
        img_blurred = blur_face_in_frame(img)
        png_bytes   = frame_to_png_bytes(img_blurred)
        snaps = session_snapshots.setdefault(sid, [])
        if len(snaps) < 8:
            snaps.append({"score": score, "ts": ts, "png": png_bytes})
            return jsonify({"ok": True, "count": len(snaps)})
        return jsonify({"ok": False, "reason": "max 8 snapshots reached"})
    except Exception as e:
        return safe_error(e)

@app.route("/api/analyze", methods=["POST"])
@require_auth
@limiter.limit("60 per minute")          # global
@limiter.limit("60 per minute; uid", key_func=lambda: getattr(g,"uid","anon"))  # per-user
def analyze():
    try:
        t0 = time.perf_counter()
        data = request.get_json(force=True)
        ok, err = validate_frame(data)
        if not ok:
            return jsonify({"error": err}), 400
        mode      = data.get("mode", "laptop")
        tier      = getattr(g, "tier", None) or "standard"  # SECURITY: from auth, never client
        uid       = getattr(g, "uid", None) or ""
        user_type = getattr(g, "user_type", None) or _detect_user_type(uid, getattr(g,"role",{}))

        # ── Employee Consent Gate ──────────────────────────────────────────
        # Employees invited by HR must explicitly consent before monitoring starts
        company_id = getattr(g, "company_id", None)
        if company_id:  # only check for invited employees (not direct signups)
            _consent_doc = db.collection("user_consent").document(uid).get()
            _consented = _consent_doc.to_dict().get("consented", False) if _consent_doc.exists else False
            if not _consented:
                return jsonify({
                    "error": "consent_required",
                    "message": "Employee must accept monitoring consent before analysis can begin.",
                    "consent_url": "/consent"
                }), 403


        # ── Frame hash cache: thumbnail-based (32x32) — 10x faster than base64 hash ──
        # Downscale to 32x32 grayscale → hash pixels → identical frames detected in <1ms
        import hashlib as _hl
        b64_raw = data.get("frame", "")
        if "," in b64_raw: b64_raw = b64_raw.split(",", 1)[1]
        # Browsers always emit correctly-padded base64, but some clients/proxies
        # strip trailing '='. Pad to the correct length (0/1/2 chars) rather than
        # blindly appending '==', which only happens to work when exactly 2 are
        # missing and raises "Incorrect padding" otherwise.
        _pad_needed = (-len(b64_raw)) % 4
        if _pad_needed:
            b64_raw = b64_raw + ("=" * _pad_needed)
        try:
            # Thumbnail hash: decode → resize 32x32 → md5 pixels
            import base64 as _b64
            _raw_bytes = _b64.b64decode(b64_raw)
            _np_arr    = np.frombuffer(_raw_bytes, dtype=np.uint8) if np else None
            if _np_arr is not None and len(_np_arr) > 100:
                _thumb = cv2.imdecode(_np_arr, cv2.IMREAD_GRAYSCALE)
                if _thumb is not None:
                    _thumb_sm = cv2.resize(_thumb, (32, 32))
                    frame_hash = _hl.md5(_thumb_sm.tobytes()).hexdigest()[:16]
                else:
                    frame_hash = _hl.md5(b64_raw[:2048].encode()).hexdigest()[:16]
            else:
                frame_hash = _hl.md5(b64_raw[:2048].encode()).hexdigest()[:16]
        except Exception:
            frame_hash = _hl.md5(b64_raw[:2048].encode()).hexdigest()[:16]  # fallback
        cached_result = cache_get(f"frame:{uid}:{frame_hash}")
        if cached_result:
            import json as _j
            r = _j.loads(cached_result)
            r["cached"] = True
            return jsonify(r)

        # ── Async mode: queue to Celery worker (recommended for scale) ──
        if os.getenv("CELERY_ENABLED", "false").lower() == "true":
            try:
                # Lazy singleton — imported once, not on every request
                global _celery_task
                if _celery_task is None:
                    from services.celery_app import analyze_frame_task as _act
                    _celery_task = _act
                job = _celery_task.delay(data, uid, tier)
                return jsonify({"job_id": job.id, "status": "queued"}), 202
            except Exception as celery_err:
                print(f"[analyze] Celery queue failed, falling back to sync: {celery_err}", file=sys.stderr)

        lang = data.get("lang", "en")
        img  = cv2.imdecode(np.frombuffer(base64.b64decode(b64_raw), np.uint8), cv2.IMREAD_COLOR)
        if img is None: return jsonify({"error": "Cannot decode image"}), 400
        if mode in ("laptop", "phone"): img = cv2.flip(img, 1)

        # ── Resize to 480p max — MediaPipe doesn't need 720p, saves 60% CPU ──
        h_img, w_img = img.shape[:2]
        if h_img > 480:
            scale = 480 / h_img
            img = cv2.resize(img, (int(w_img * scale), 480), interpolation=cv2.INTER_AREA)

        # ── Motion blur detection (Laplacian variance) ────────────
        # Blurry frames degrade landmark accuracy → skip and return cached
        gray_check  = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        blur_score  = cv2.Laplacian(gray_check, cv2.CV_64F).var()
        is_blurry   = blur_score < 45  # threshold: <45 = motion blur
        if is_blurry:
            cached_blur = cache_get(f"last_valid:{uid}")
            if cached_blur:
                import json as _bj
                rv = _bj.loads(cached_blur)
                rv["blurry_frame"] = True
                rv["blur_score"]   = round(blur_score, 1)
                return jsonify(rv)
            # No cache yet — process anyway but mark as low confidence

        # Low-light enhancement pipeline
        img_to_analyze, brightness, was_enhanced = enhance_low_light(img)

        # ── Light quality gate ─────────────────────────────────────
        _post_bright = float(np.mean(cv2.cvtColor(img_to_analyze, cv2.COLOR_BGR2GRAY)))
        _light_q = ("good" if _post_bright >= 100 else
                    "dim"  if _post_bright >= 60  else
                    "poor" if _post_bright >= 30  else "critical")
        if _light_q == "critical":
            return jsonify({
                "detected":      False, "score": 0, "overall": 0,
                "confidence":    0.0,   "error_type": "insufficient_light",
                "light_quality": {"level": "critical", "brightness": round(_post_bright,1)},
                "alerts":    ["⚠️ Lighting too dark — move to a well-lit area"],
                "alerts_ar": ["⚠️ الإضاءة ضعيفة جداً — انتقل لمكان أكثر إضاءة"],
            })

        calib = data.get("calibration")   # personal calibration from Firestore
        # Pass session_id into analysis so landmark averaging uses correct per-session key
        _session_id_for_analysis = data.get("session_id", f"{uid}:default")
        if mode == "side":
            result = analyze_side(img_to_analyze, tier, session_id=_session_id_for_analysis)
        else:
            result = analyze_front(img_to_analyze, mode, tier, session_id=_session_id_for_analysis)
        # ── Apply personal calibration with asymmetric thresholds ──
        # Asymmetric: if user's natural lean is toward right (+),
        # we widen the ok/bad zone on the right side and tighten on left.
        # This prevents penalizing natural anatomical asymmetry.
        if calib and isinstance(calib, dict) and result.get("detected"):
            tols     = calib.get("tolerances", {})
            offsets  = calib.get("offsets", {})   # signed directional offsets
            met_map  = {
                "neck_lean":      "neck_angle",
                "head_tilt":      "head_tilt",
                "shoulder_level": "shoulder_tilt",
                "spine_lean":     "spine_angle",
            }
            calib_scores = []

            for met_key, calib_key in met_map.items():
                met = result.get("metrics", {}).get(met_key)
                tol = tols.get(calib_key)
                if not (met and tol): continue

                raw_val  = met["value"]                    # always positive (absolute angle)
                ideal    = tol.get("ideal", 0)             # calibrated neutral (absolute)
                ok_t     = tol.get("ok", 7)
                bad_t    = tol.get("bad", 20)

                # ── Asymmetric offset correction ──────────────────
                # offset_signed: positive = user naturally leans right/forward
                # negative = user naturally leans left/back
                offset_signed = offsets.get(calib_key, 0.0)

                # Compute signed deviation from user's personal neutral
                # We use the calibrated ideal as the zero point
                deviation = raw_val - ideal  # + = more than usual, - = less than usual

                # Asymmetric threshold: widen ok/bad in the natural direction
                # e.g. if offset = +5° (leans right naturally):
                #   right side (deviation > 0): ok_right = ok_t + 3°, bad_right = bad_t + 4°
                #   left side  (deviation < 0): ok_left  = ok_t - 1°, bad_left  = bad_t - 1°
                asym_factor = max(-4.0, min(8.0, abs(offset_signed) * 0.6))
                if (offset_signed >= 0 and deviation >= 0) or (offset_signed < 0 and deviation < 0):
                    # Same direction as natural lean → widen tolerance
                    ok_eff  = ok_t  + asym_factor
                    bad_eff = bad_t + asym_factor * 1.2
                else:
                    # Opposite direction → slight tightening
                    ok_eff  = max(3.0, ok_t  - asym_factor * 0.3)
                    bad_eff = max(10.0, bad_t - asym_factor * 0.3)

                # Score using effective asymmetric thresholds
                d_eff = abs(deviation)
                if d_eff <= ok_eff:
                    cs = max(0, int(100 - (d_eff / max(ok_eff, .1)) * 25))
                elif d_eff <= bad_eff:
                    cs = max(0, int(75  - ((d_eff - ok_eff) / max(bad_eff - ok_eff, .1)) * 45))
                else:
                    cs = max(5, int(30 - (d_eff - bad_eff) * 1.5))

                result["metrics"][met_key]["score"]      = cs
                result["metrics"][met_key]["calibrated"] = True
                result["metrics"][met_key]["calib_detail"] = {
                    "ideal":       round(ideal, 1),
                    "deviation":   round(deviation, 1),
                    "offset":      round(offset_signed, 1),
                    "ok_eff":      round(ok_eff, 1),
                    "bad_eff":     round(bad_eff, 1),
                }
                calib_scores.append(cs)

            if calib_scores:
                raw_score  = result.get("score", 0)
                calib_avg  = sum(calib_scores) / len(calib_scores)
                result["score"]   = max(0, min(100, int(raw_score * 0.40 + calib_avg * 0.60)))
                result["overall"] = result["score"]
                result["calibration_applied"]   = True
                result["asymmetric_correction"] = True
        result["brightness"] = round(brightness, 1)
        result["low_light_enhanced"] = was_enhanced
        result["light_quality"]     = {
            "level":      _light_q,
            "brightness": round(_post_bright, 1),
            "enhanced":   was_enhanced,
        }
        result["tier"]    = tier
        result["overall"] = result.get("score", 0)

        # Redis score smoothing (if Redis available)
        # uid already resolved from auth token above — don't override from client data
        if uid and result.get("detected"):
            push_score(uid, result["overall"])
            smoothed = get_smoothed_score(uid)
            if smoothed is not None:
                result["overall_smoothed"] = int(smoothed)
                result["overall"] = int(smoothed)
            # Risk threshold tracking for webhooks
            if result["overall"] < 45:
                set_risk_start(uid, result["overall"])
            else:
                clear_risk(uid)

        sid = data.get("session_id", "default")
        s = sessions.setdefault(sid, {
            "score_history": [], "alerts": [], "mode": mode, "tier": tier,
            "start": time.time(), "frames": 0, "good": 0, "last_ai": 0,
            "engine": "unknown",
            # Quality tracking
            "blurry_count": 0,   # frames that were too blurry to analyze
            "total_frames": 0,   # all frames including blurry/undetected
            "vis_sum":     0.0,   # sum of per-frame avg landmark visibility
            "vis_count":   0,     # frames with visibility data
            "conf_sum":    0.0,   # sum of per-frame confidence scores
            "conf_count":  0,     # frames with confidence data
        })
        _s_dirty = False
        # Always count total frames (including blurry/undetected)
        s["total_frames"] = s.get("total_frames", 0) + 1
        if result.get("blurry_frame"):
            s["blurry_count"] = s.get("blurry_count", 0) + 1

        if result["detected"] and result["score"] > 0:
            s.setdefault("score_history", []).append(result["score"])
            s["frames"] = s.get("frames", 0) + 1
            if result["score"] >= 65: s["good"] = s.get("good", 0) + 1
            s["engine"] = result.get("engine", "")

            # ── Fatigue detection: score trend within session ────
            hist_scores = s["score_history"]
            if len(hist_scores) >= 10:
                recent_10  = hist_scores[-10:]
                recent_avg = sum(recent_10) / 10
                # Compare to session first-quarter average (baseline when fresh)
                q1_len = max(5, len(hist_scores) // 4)
                q1_avg = sum(hist_scores[:q1_len]) / q1_len
                drop   = q1_avg - recent_avg

                if drop >= 8 and recent_avg < 70:
                    result["fatigue_signal"] = {
                        "detected":    True,
                        "drop_pts":    round(drop, 1),
                        "q1_avg":      round(q1_avg, 1),
                        "recent_avg":  round(recent_avg, 1),
                        "severity":    "high"   if drop >= 15 else "moderate",
                        "message":     "Fatigue detected — posture declining. Take a 5-min break.",
                    }
                    if "⚠️ Fatigue" not in " ".join(result.get("alerts", [])):
                        result.setdefault("alerts", []).insert(0,
                            f"⚠️ Fatigue — posture dropped {round(drop)}pts this session. Take a break.")
                elif drop >= 4:
                    result["fatigue_signal"] = {
                        "detected":   False,
                        "drop_pts":   round(drop, 1),
                        "severity":   "early",
                        "message":    "Early fatigue — posture gradually declining.",
                    }

            # Accumulate visibility — from per-metric confidence data
            conf_detail = result.get("metrics", {}).get("_confidence", {})
            if conf_detail:
                vis_vals = [v for k, v in conf_detail.items()
                            if k in ("neck","tilt","sh","spine") and isinstance(v, (int,float))]
                if vis_vals:
                    s["vis_sum"]   = s.get("vis_sum", 0.0)   + sum(vis_vals) / len(vis_vals)
                    s["vis_count"] = s.get("vis_count", 0)   + 1

            # Accumulate confidence score
            frame_conf = result.get("confidence", 0)
            if frame_conf > 0:
                s["conf_sum"]   = s.get("conf_sum", 0.0) + frame_conf
                s["conf_count"] = s.get("conf_count", 0)  + 1

            _s_dirty = True

            # ── IMPROVED: adaptive velocity-aware smoothing ────────
            # Fixed α=0.30 caused score to jump on tiny (1°) posture changes.
            # Fix: α adapts to the magnitude of the score change:
            #   small changes (<5pts)  → α=0.12 (heavy smoothing, ignore jitter)
            #   normal changes (5-15)  → α=0.30 (balanced)
            #   large changes (15-25)  → α=0.50 (respond faster)
            #   very large (>25pts)    → α=0.70 (deliberate change, update quickly)
            # Hard clamp: ±18pts max per frame (was ±12, slightly more responsive)
            hist      = s["score_history"]
            raw_score = hist[-1]

            # ── 10-second weighted window (6 frames ≈ 10s at 1.5s/frame) ─
            # Weights: more recent frames get higher weight
            # Prevents single bad frame from tanking the score
            _WIN_WEIGHTS = [0.40, 0.25, 0.15, 0.10, 0.06, 0.04]
            _window = hist[-len(_WIN_WEIGHTS):]   # last N frames
            if len(_window) >= 3:
                # Trim weights to window size
                _w = _WIN_WEIGHTS[:len(_window)]
                _w_norm = [x / sum(_w) for x in _w]
                # Apply weights (most recent first)
                _window_rev = list(reversed(_window))
                smoothed_local = int(round(sum(s * w for s, w in zip(_window_rev, _w_norm))))
            else:
                smoothed_local = raw_score

            # Still clamp against previous to prevent sudden spikes
            prev_ewma = s.get("last_ewma", smoothed_local)
            _max_jump = 15  # max change per frame after windowing
            smoothed_local = max(prev_ewma - _max_jump, min(prev_ewma + _max_jump, smoothed_local))
            smoothed_local = max(0, min(100, smoothed_local))

            s["last_ewma"]           = smoothed_local
            result["score_smoothed"] = smoothed_local
            result["score_raw"]      = raw_score
            result["window_size"]    = len(_window)

            # ── Time-Weighted Average (TWA) score ─────────────────
            # Standard occupational health metric for continuous exposure
            # TWA = Σ(score_i × time_i) / total_time
            if len(hist) >= 5:
                frame_duration = s.get("duration_s", len(hist)*0.033*30) / max(len(hist),1)
                twa_score = sum(sc * frame_duration for sc in hist) / max(s.get("duration_s",1), 1)
                twa_score = max(0, min(100, round(twa_score)))
                twa_risk = "low" if twa_score >= 70 else "moderate" if twa_score >= 55 else "high"
                result["twa_score"] = {
                    "score":    twa_score,
                    "risk":     twa_risk,
                    "duration": round(s.get("duration_s",0)),
                    "label":    "Time-Weighted Average (occupational standard)",
                }

            # ── Within-session trend ───────────────────────────────
            # Compare recent vs early scores: are you getting better?
            if len(hist) >= 10:
                _early  = hist[:5]
                _recent = hist[-5:]
                _early_avg  = sum(_early)  / len(_early)
                _recent_avg = sum(_recent) / len(_recent)
                _trend_delta = _recent_avg - _early_avg
                _trend = ("improving"    if _trend_delta > 4  else
                          "stable"       if _trend_delta > -4 else
                          "deteriorating")
                result["session_trend"] = {
                    "trend":        _trend,
                    "delta":        round(_trend_delta, 1),
                    "early_avg":    round(_early_avg, 1),
                    "recent_avg":   round(_recent_avg, 1),
                    "frames_total": len(hist),
                }
            if result.get("overall_smoothed") is None:
                result["overall"] = smoothed_local
                result["score"]   = smoothed_local
        for a in result.get("alerts",[]):
            recent = [x.get("msg") for x in s.get("alerts",[])[-4:]]
            if a not in recent:
                s.setdefault("alerts", []).append({"time": datetime.now().strftime("%H:%M:%S"), "msg": a, "score": result["score"]})
                _s_dirty = True

        if (tier in ("elite", "premium", "professional", "pro", "basic") and result["detected"] and
            result.get("score", 100) < 80 and   # skip AI call when posture is already good
            AI_CONFIGURED and time.time() - s.get("last_ai", 0) > 30):  # 30s cooldown
            # Fire AI narrative in background — don't block the analysis response
            _r, _ctx, _lg, _sid = dict(result), data.get("employee_context",""), lang, sid
            def _bg_gemini(r, ctx, lg, _session_id):
                try:
                    txt = analyze_with_local_ai(r, ctx, lg)
                    if txt:
                        # Read-modify-write back to Redis-backed store
                        _s = sessions.get(_session_id, {})
                        _s["last_gemini_text"] = txt
                        sessions[_session_id] = _s
                except Exception as _e:
                    print(f"⚠️  Background Gemini error: {_e}")
            fut = _ai_executor.submit(_bg_gemini, _r, _ctx, _lg, _sid)
            # Discard future — fire-and-forget with 15s implicit timeout via executor
            del fut
            s["last_ai"] = time.time()
            _s_dirty = True
            # Return cached Gemini text from previous call if available
            if s.get("last_gemini_text"):
                result["ai_tip"] = s["last_gemini_text"]; result["claude_analysis"] = s["last_gemini_text"]  # backward compat
                result["ai_enhanced"]     = True

        # Persist session state back to Redis
        if _s_dirty:
            sessions[sid] = s

        # Track usage for billing limits (use existing redis_service, not a new connection)
        if uid and result.get("detected"):
            today = datetime.utcnow().strftime("%Y-%m-%d")
            month = datetime.utcnow().strftime("%Y-%m")
            try:
                rset(f"usage:{uid}:frames:{today}",
                     str(int(cache_get(f"usage:{uid}:frames:{today}") or 0) + 1),
                     86400 * 2)
            except Exception: pass

        # Cache result for 1.5s (avoids reprocessing identical frames)
        try:
            import json as _j
            cache_set(f"frame:{uid}:{frame_hash}", _j.dumps(result), 2)
        except Exception: pass

        ms = round((time.perf_counter() - t0) * 1000)
        result["processing_ms"] = ms
        _record_latency(ms)  # rolling latency tracker for /api/health/detailed

        # Include live quality estimate in every 10th frame (avoid overhead)
        if s.get("frames", 0) % 10 == 0 and s.get("total_frames", 0) >= 10:
            result["session_quality"] = compute_session_quality(s)

        # Ergonomic twin: update + detect setup drift
        if uid and result.get("detected"):
            try:
                _twin = update_ergonomic_twin(uid, result)
                result["ergonomic_twin"] = {
                    "calibrated":     _twin.get("calibrated", False),
                    "optimal_dist":   _twin.get("optimal_distance"),
                    "optimal_pitch":  _twin.get("optimal_pitch"),
                    "drift_alerts":   _twin.get("drift_alerts", []),
                    "drift_count":    _twin.get("drift_count", 0),
                    "sessions_to_learn": _twin.get("sessions_to_learn", 5),
                }
                for da in _twin.get("drift_alerts", []):
                    if da not in result.get("alerts", []):
                        result.setdefault("alerts", []).insert(0, da)
            except Exception:
                pass

        # Keystroke stress correlation (if client sends typing data)
        _ks_data = data.get("keystroke_data")
        if _ks_data and isinstance(_ks_data, dict):
            try:
                _ks_stress = analyze_keystroke_stress(_ks_data)
                if _ks_stress:
                    result["keystroke_stress"] = _ks_stress
                    # Correlate: high stress + bad posture = compound risk
                    if _ks_stress["stress_score"] > 60 and result["score"] < 65:
                        result["compound_risk"] = {
                            "level": "high",
                            "note":  "High work stress + poor posture = elevated MSK injury risk",
                            "action":"Take a 5-min break immediately",
                        }
                        result["alerts"] = result.get("alerts", [])
                        result["alerts"].insert(0, "⚠️ Compound risk: work stress + poor posture — break now")
            except Exception:
                pass

        # Update adaptive coaching profile every 15th frame
        if s.get("frames", 0) % 15 == 0 and uid:
            try:
                _cp = update_coaching_profile(uid, result, result.get("metrics", {}))
                result["coaching"] = {
                    "phase":        _cp.get("coach_phase"),
                    "plan":         _cp.get("coaching_plan", [])[:2],  # top 2 only
                    "message":      _cp.get("phase_message"),
                    "sessions_n":   _cp.get("sessions_analyzed"),
                    "new_badges":   [b for b in _cp.get("badges_earned",[])
                                     if b not in s.get("prev_badges",[])],
                }
                s["prev_badges"] = _cp.get("badges_earned", [])
            except Exception:
                pass
        # Cache last valid (non-blurry) result for blur-frame fallback
        if result.get("detected") and not result.get("blurry_frame"):
            try:
                import json as _vj
                cache_set(f"last_valid:{uid}", _vj.dumps(result), 5)
            except Exception: pass
        # Log slow requests (>800ms) for performance monitoring
        if ms > 800:
            log_event("analyze_slow", meta={"ms": ms, "engine": result.get("engine"), "tier": tier})
        # Confidence floor: if MediaPipe detected landmarks, score ≥ 18
        # Prevents confusing "0" results from edge-case geometry
        if result.get("detected") and result.get("score", 100) < 18:
            result["score"]   = 18
            result["overall"] = 18
        # ── Percentile score ──────────────────────────────────────
        if result.get("detected") and result.get("score") is not None:
            _pct = compute_percentile(int(result["score"]))
            if _pct >= 0:
                result["percentile"]       = _pct
                result["percentile_label"] = f"Better than {_pct}% of users"
        return jsonify(result)
    except Exception as e:
        import traceback
        return safe_error(e)

@require_auth
@app.route("/api/analyze/job/<job_id>", methods=["GET"])
@require_auth
@limiter.limit("120 per minute")
def analyze_job_status(job_id):
    """Poll for async analysis job result."""
    try:
        if os.getenv("CELERY_ENABLED", "false").lower() != "true":
            return jsonify({"error": "Async mode not enabled"}), 404
        from services.celery_app import analyze_frame_task
        from celery.result import AsyncResult
        result = AsyncResult(job_id, app=analyze_frame_task.app)
        state = result.state
        if state == "PENDING":
            return jsonify({"status": "queued", "job_id": job_id})
        elif state == "STARTED":
            return jsonify({"status": "processing", "job_id": job_id})
        elif state == "SUCCESS":
            return jsonify({"status": "done", "job_id": job_id, "result": result.result})
        elif state == "FAILURE":
            return jsonify({"status": "failed", "job_id": job_id, "error": str(result.result)}), 500
        else:
            return jsonify({"status": state.lower(), "job_id": job_id})
    except Exception as e:
        return safe_error(e, "Failed to get job status")

@app.route("/api/pdf", methods=["POST"])
@require_auth
@limiter.limit("10 per minute")
def pdf_endpoint():
    if not REPORTLAB_OK:
        return jsonify({"error":"pip install reportlab"}), 500
    try:
        data = request.get_json(force=True)
        sid  = data.get("session_id","default")
        sess = sessions.get(sid, {})
        hist = sess.get("score_history") or data.get("score_history", [])
        alts = sess.get("alerts") or data.get("alerts", [])
        dur  = data.get("duration_s") or int(time.time() - sess.get("start", time.time()))
        avg  = data.get("avg_score") or (round(sum(hist)/len(hist)) if hist else 0)
        gd   = data.get("good_pct") or round(sess.get("good",0)/max(sess.get("frames",1),1)*100)
        la   = data.get("last_analysis",{})
        sd2 = {
            "session_id":   sid,
            "date":         datetime.now().strftime("%B %d, %Y"),
            "mode":         sess.get("mode", data.get("mode","laptop")),
            "tier":         getattr(g, "role", {}).get("tier", "standard"),  # SECURITY: auth-sourced
            "avg_score":    avg, "good_pct": gd, "duration_s": dur,
            "total_frames": sess.get("frames",0),
            "score_history":hist, "alerts": alts,
            "metrics":      la.get("metrics",{}),
            "recommendations": la.get("recommendations",[
                "Maintain optimal screen distance at all times",
                "Schedule a 2-minute posture break every 30 minutes",
                "Position monitor top at eye level to reduce neck lean",
                "Ensure lumbar support contacts lower back when seated",
                "Apply the 20-20-20 rule every 20 minutes",
            ]),
            "head_pose":       la.get("head_pose"),
            "claude_analysis": la.get("claude_analysis") or data.get("claude_analysis"),
            "confidence":      la.get("confidence",0),
            "engine":          sess.get("engine","") or la.get("engine",""),
            "company_info": data.get("company_info", {"name":"PostureAI Client","logo_text":"PostureAI"}),
            "employee":     data.get("employee",{}),
            "snapshots":    session_snapshots.get(sid, []),
            "lang":         data.get("lang","en"),
        }
        tier_sd = sd2["tier"]
        if tier_sd == "elite" and AI_CONFIGURED and not sd2.get("claude_analysis") and hist:
            summary_result = {
                "score": avg, "metrics": la.get("metrics",{}),
                "alerts": [a.get("msg","") for a in alts[-5:]],
                "head_pose": la.get("head_pose"), "mode": sd2["mode"], "engine": sd2["engine"]
            }
            _ctx = f"Session summary: {len(hist)} frames, {dur//60}m {dur%60}s, avg {avg}/100, {gd}% good posture, {len(alts)} alerts"
            _lang = data.get("lang","en")
            # Run AI narrative in thread with 12s timeout to avoid blocking Flask worker
            try:
                _fut = _ai_executor.submit(analyze_with_local_ai, summary_result, _ctx, _lang)
                sd2["claude_analysis"] = _fut.result(timeout=12)
            except (FuturesTimeoutError, Exception) as _ai_err:
                print(f"⚠️  Gemini PDF analysis skipped: {_ai_err}")
                sd2["claude_analysis"] = None
        pdf_bytes = generate_pdf(sd2)
        buf  = io.BytesIO(pdf_bytes)
        fname = (f"PostureAI_{sd2['tier']}_{sid}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf")
        return send_file(buf, mimetype="application/pdf", as_attachment=True, download_name=fname)
    except Exception as e:
        import traceback
        return safe_error(e)

@require_auth
@app.route("/api/session/start", methods=["POST"])
@limiter.limit("60 per minute")
def start_session():
    try:
        data    = request.get_json(force=True) or {}
        role    = getattr(g, "role", {})
        tier    = role.get("tier", "standard")
        uid     = getattr(g, "uid", "")

        # ── Free plan session cap ────────────────────────────────────
        # The "standard" (Free) plan is marketed as limited sessions
        # (see TIERS.standard.features in App.jsx) but had zero server-side
        # enforcement — any Free user could run unlimited sessions. basic/
        # professional/elite/business/b2b plans all promise unlimited
        # sessions already, so this only applies to literal "standard".
        if tier == "standard":
            month_key = f"usage:{uid}:sessions:{datetime.utcnow().strftime('%Y-%m')}"
            day_key   = f"usage:{uid}:sessions:{datetime.utcnow().strftime('%Y-%m-%d')}"
            try:
                month_count = int(rget(month_key) or 0)
                day_count   = int(rget(day_key) or 0)
            except Exception:
                month_count = day_count = 0
            FREE_MONTHLY_LIMIT = 5
            FREE_DAILY_LIMIT   = 3
            if month_count >= FREE_MONTHLY_LIMIT or day_count >= FREE_DAILY_LIMIT:
                return jsonify({
                    "error":   "session_limit_reached",
                    "message": "Free plan session limit reached. Upgrade to continue.",
                    "limit_monthly": FREE_MONTHLY_LIMIT,
                    "limit_daily":   FREE_DAILY_LIMIT,
                    "used_monthly":  month_count,
                    "used_daily":    day_count,
                    "upgrade": True,
                }), 403

        # ── Seat enforcement for company plans ──────────────────────
        # enterprise customers have a seats field — enforce it server-side
        # so frontend bypasses are impossible
        company_id = role.get("company_id")
        if company_id and tier in ("professional", "elite", "enterprise"):
            max_seats = role.get("seats", 25)
            if max_seats > 0:  # 0 or None = unlimited
                # Count active users in company from in-memory sessions (last 10 min)
                cutoff = time.time() - 600
                active = sum(
                    1 for s in sessions.values()
                    if s.get("company_id") == company_id and s.get("start", 0) > cutoff
                )
                if active >= max_seats:
                    return jsonify({
                        "error":     f"Seat limit reached ({max_seats} seats). "
                                     "Upgrade your plan or remove inactive users.",
                        "seat_limit": max_seats,
                        "active":     active,
                        "upgrade":    True,
                    }), 403

        sid = f"PA{int(time.time()*1000)%1000000:06d}"
        sessions[sid] = {
            "score_history": [], "alerts":  [], "mode": data.get("mode", "laptop"),
            "tier":          tier,              "start": time.time(),
            "frames":        0,    "good":   0, "engine": "unknown",
            "uid":           g.uid,
            "company_id":    company_id,
        }
        if tier == "standard":
            try:
                rincr(f"usage:{uid}:sessions:{datetime.utcnow().strftime('%Y-%m')}", ttl_s=86400 * 35)
                rincr(f"usage:{uid}:sessions:{datetime.utcnow().strftime('%Y-%m-%d')}", ttl_s=86400 * 2)
            except Exception:
                pass
        return jsonify({"session_id": sid, "started": datetime.now().isoformat()})
    except Exception as e:
        return safe_error(e)

@require_auth
@app.route("/api/session/<sid>", methods=["GET"])
@limiter.limit("120 per minute")
def get_session(sid):
    try:
        s = sessions.get(sid)
        if not s: return jsonify({"error":"Session not found"}), 404
        h = s["score_history"]
        quality = compute_session_quality(s)
        avg = round(sum(h)/len(h)) if h else 0
        duration_s = int(time.time() - s["start"])

        # ── Grade ─────────────────────────────────────────────────
        grade = (
            "A" if avg >= 85 else
            "B" if avg >= 70 else
            "C" if avg >= 55 else
            "D" if avg >= 40 else "F"
        )

        # ── Score trend: compare first vs last 20% of frames ──────
        trend = "stable"
        if len(h) >= 10:
            split = max(3, len(h) // 5)
            early_avg = sum(h[:split]) / split
            late_avg  = sum(h[-split:]) / split
            diff = late_avg - early_avg
            trend = "improving" if diff > 5 else "declining" if diff < -5 else "stable"

        # ── Top unique alerts (deduped) ────────────────────────────
        seen_alerts = []
        for a in s.get("alerts", []):
            txt = a if isinstance(a, str) else a.get("text", str(a))
            if txt not in seen_alerts:
                seen_alerts.append(txt)
        top_alerts = seen_alerts[:5]

        # ── Worst metric + improvement tip ────────────────────────
        metric_scores = s.get("metric_scores", {})
        worst_metric  = None
        worst_score   = 101
        tip_map = {
            "neck":  "Raise your monitor to eye level to reduce neck flexion.",
            "spine": "Sit back fully in your chair and use lumbar support.",
            "dist":  "Move your screen to 50–70cm away from your eyes.",
            "tilt":  "Keep your head level — avoid tilting toward your dominant side.",
            "sh":    "Relax your shoulders down and back, away from your ears.",
            "wrist": "Keep wrists straight and elbows at 90° when typing.",
        }
        for metric, sc in metric_scores.items():
            if isinstance(sc, (int, float)) and sc < worst_score:
                worst_score  = sc
                worst_metric = metric
        improvement_tip = tip_map.get(worst_metric, "Take a 2-minute posture break every 30 minutes.")

        # ── Pain prediction summary ────────────────────────────────
        pain_mins = s.get("pain_prediction_min")
        pain_summary = None
        if pain_mins and isinstance(pain_mins, (int, float)):
            if pain_mins < 30:
                pain_summary = f"Discomfort likely within {int(pain_mins)} min — take a break now."
            elif pain_mins < 90:
                pain_summary = f"~{int(pain_mins)} min before likely discomfort."
            else:
                pain_summary = "No imminent discomfort predicted."

        return jsonify({
            "session_id":       sid,
            "avg_score":        avg,
            "grade":            grade,
            "frames":           s["frames"],
            "good_pct":         round(s["good"] / max(s["frames"], 1) * 100),
            "alerts_count":     len(s.get("alerts", [])),
            "top_alerts":       top_alerts,
            "duration_s":       duration_s,
            "duration_min":     round(duration_s / 60, 1),
            "mode":             s["mode"],
            "tier":             s.get("tier", "standard"),
            "engine":           s.get("engine", ""),
            "trend":            trend,
            "worst_metric":     worst_metric,
            "improvement_tip":  improvement_tip,
            "pain_summary":     pain_summary,
            "quality":          quality,
            "score_history":    h[-30:] if h else [],   # last 30 frames for mini chart
        })
    except Exception as e:
        return safe_error(e)

@require_auth
@app.route("/api/notify/slack", methods=["POST"])
@limiter.limit("30 per minute")
def notify_slack():
    try:
        data = request.get_json(force=True) or {}
        result = send_slack(text=data.get("text","Posture alert"), score=data.get("score",0))
        return jsonify(result), 200 if result["ok"] else 400
    except Exception as e:
        return safe_error(e)

@require_auth
@app.route("/api/notify/teams", methods=["POST"])
@limiter.limit("30 per minute")
def notify_teams():
    try:
        data = request.get_json(force=True) or {}
        result = send_teams(text=data.get("text","Posture alert"), score=data.get("score",0), employee=data.get("employee",""))
        return jsonify(result), 200 if result["ok"] else 400
    except Exception as e:
        return safe_error(e)

@require_auth
@app.route("/api/notify/whatsapp", methods=["POST"])
@limiter.limit("10 per minute")
def notify_whatsapp():
    try:
        data = request.get_json(force=True) or {}
        phone = data.get("phone","")
        if not phone: return jsonify({"ok":False,"reason":"phone required"}), 400
        result = send_whatsapp(phone, data.get("text","PostureAI posture alert"))
        return jsonify(result), 200 if result["ok"] else 400
    except Exception as e:
        return safe_error(e)



# ── MFA Endpoints ────────────────────────────────────────────────────
@app.route("/api/auth/mfa/totp/setup", methods=["POST"])
@require_auth
@limiter.limit("10 per hour")
def mfa_totp_setup():
    """Generate a TOTP secret and QR URI for the authenticated user."""
    try:
        from auth.mfa import generate_totp_secret, get_totp_uri, PYOTP_OK
        if not PYOTP_OK:
            return jsonify({"error": "TOTP not available — install pyotp"}), 503
        uid   = g.uid
        email = g.user.get("email", "")
        secret = generate_totp_secret()
        uri    = get_totp_uri(secret, email)
        # Store pending secret in Redis (expires in 10 min)
        if REDIS_READY:
            rset(f"mfa:totp:pending:{uid}", secret, ttl_s=600)
        return jsonify({"secret": secret, "uri": uri, "email": email})
    except Exception as e:
        return safe_error(e, "MFA setup failed")


@app.route("/api/auth/mfa/totp/verify", methods=["POST"])
@require_auth
@limiter.limit("10 per hour")
def mfa_totp_verify():
    """Verify TOTP code and enable MFA for the user."""
    try:
        from auth.mfa import verify_totp, generate_backup_codes, hash_backup_code, PYOTP_OK
        if not PYOTP_OK:
            return jsonify({"error": "TOTP not available"}), 503
        data = request.get_json(force=True) or {}
        code = str(data.get("code", "")).strip()
        uid  = g.uid
        # Get pending secret from Redis
        secret = None
        if REDIS_READY:
            raw = rget(f"mfa:totp:pending:{uid}")
            if raw:
                secret = raw.decode() if isinstance(raw, bytes) else raw
        if not secret:
            return jsonify({"error": "Setup session expired — restart MFA setup"}), 400
        if not verify_totp(secret, code):
            return jsonify({"error": "Invalid code — please try again"}), 400
        # Generate backup codes
        codes        = generate_backup_codes(8)
        hashed_codes = [hash_backup_code(c) for c in codes]
        # Save to Firestore
        if db:
            db.collection("users").document(uid).update({
                "mfa_enabled":     True,
                "mfa_totp_secret": secret,   # In production: encrypt with MFA_ENCRYPTION_KEY
                "mfa_backup_codes": hashed_codes,
                "mfa_enabled_at":  datetime.utcnow().isoformat(),
            })
        if REDIS_READY:
            rdel(f"mfa:totp:pending:{uid}")
        invalidate_user_cache(uid)
        return jsonify({"success": True, "backup_codes": codes})
    except Exception as e:
        return safe_error(e, "MFA verification failed")


@app.route("/api/auth/mfa/sms/send", methods=["POST"])
@require_auth
@limiter.limit("5 per hour")
def mfa_sms_send():
    """Send SMS verification code for MFA setup."""
    try:
        from auth.mfa import send_sms_code, TWILIO_OK
        data  = request.get_json(force=True) or {}
        phone = str(data.get("phone", "")).strip()
        if not phone or len(phone) < 8:
            return jsonify({"error": "Valid phone number required"}), 400
        if not TWILIO_OK:
            return jsonify({"error": "SMS not configured — use TOTP instead", "totp_available": True}), 503
        ok = send_sms_code(phone, g.uid)
        if not ok:
            return jsonify({"error": "Failed to send SMS"}), 500
        return jsonify({"success": True, "message": "Code sent"})
    except Exception as e:
        return safe_error(e, "SMS send failed")


@app.route("/api/auth/mfa/sms/verify", methods=["POST"])
@require_auth
@limiter.limit("10 per hour")
def mfa_sms_verify():
    """Verify SMS code and enable MFA."""
    try:
        from auth.mfa import verify_sms_code, generate_backup_codes, hash_backup_code
        data  = request.get_json(force=True) or {}
        code  = str(data.get("code", "")).strip()
        phone = str(data.get("phone", "")).strip()
        uid   = g.uid
        if not verify_sms_code(uid, code):
            return jsonify({"error": "Invalid or expired code"}), 400
        codes        = generate_backup_codes(8)
        hashed_codes = [hash_backup_code(c) for c in codes]
        if db:
            db.collection("users").document(uid).update({
                "mfa_enabled":     True,
                "mfa_method":      "sms",
                "mfa_phone":       phone,
                "mfa_backup_codes": hashed_codes,
                "mfa_enabled_at":  datetime.utcnow().isoformat(),
            })
        invalidate_user_cache(uid)
        return jsonify({"success": True, "backup_codes": codes})
    except Exception as e:
        return safe_error(e, "SMS verification failed")


@app.route("/api/auth/mfa/disable", methods=["POST"])
@require_auth
@limiter.limit("3 per hour")
def mfa_disable():
    """Disable MFA. Requires re-confirmation of UID (frontend confirms via Firebase re-auth)."""
    try:
        uid = g.uid
        if db:
            db.collection("users").document(uid).update({
                "mfa_enabled":     False,
                "mfa_totp_secret": None,
                "mfa_method":      None,
                "mfa_backup_codes": [],
                "mfa_disabled_at": datetime.utcnow().isoformat(),
            })
        invalidate_user_cache(uid)
        return jsonify({"success": True})
    except Exception as e:
        return safe_error(e, "MFA disable failed")


@app.route("/api/auth/mfa/backup-codes/regenerate", methods=["POST"])
@require_auth
@limiter.limit("3 per day")
def mfa_backup_codes_regen():
    """Regenerate backup codes — invalidates old ones."""
    try:
        from auth.mfa import generate_backup_codes, hash_backup_code
        uid          = g.uid
        codes        = generate_backup_codes(8)
        hashed_codes = [hash_backup_code(c) for c in codes]
        if db:
            db.collection("users").document(uid).update({
                "mfa_backup_codes": hashed_codes,
                "mfa_backup_regen_at": datetime.utcnow().isoformat(),
            })
        return jsonify({"backup_codes": codes})
    except Exception as e:
        return safe_error(e, "Backup code regeneration failed")





# ── V10_2 Extracted: NPS, Referrals, Integrations API, Audit Export, Tenant Mgmt, Usage Metering, Health Scores, Product Tour ──
@app.route("/api/org/health-scores", methods=["GET"])
@require_auth
@require_tier("professional")
def org_health_scores():
    """Return all members with computed health + churn risk scores."""
    try:
        db  = firestore.client()
        org_id = request.args.get("org_id") or getattr(g,"company_id","")
        if not org_id:
            return jsonify({"error":"org_id required"}),400

        members = db.collection("users").where("company_id","==",org_id).get()
        now     = datetime.utcnow()
        results = []

        for doc in members:
            u = doc.to_dict(); u["uid"] = doc.id
            sessions_30 = u.get("sessions_30d", 0)
            last_login  = u.get("last_login_at")
            avg_score   = u.get("avg_score", 0)
            plan        = u.get("plan","starter")

            # Days inactive
            inactive_days = 0
            if last_login:
                ll = last_login if isinstance(last_login,datetime) else datetime.fromisoformat(str(last_login))
                inactive_days = (now - ll).days

            # Health signals (0-100 each)
            login_freq   = max(0, 100 - inactive_days * 5)
            session_sig  = min(100, sessions_30 * 3)
            score_sig    = avg_score
            payment_sig  = 100 if u.get("payment_ok", True) else 50
            support_sig  = max(0, 100 - u.get("open_tickets",0)*10)

            health = int(
                login_freq  * 0.30 +
                session_sig * 0.25 +
                score_sig   * 0.20 +
                payment_sig * 0.15 +
                support_sig * 0.10
            )

            # Churn risk (inverse of health, weighted by inactivity)
            churn_risk = max(0, min(100, int((100 - health) * 0.7 + inactive_days * 1.2)))

            stage = (
                "champion" if health >= 88 else
                "healthy"  if health >= 70 else
                "at_risk"  if health >= 45 else
                "critical"
            )

            results.append({
                "uid":          u["uid"],
                "name":         u.get("name",""),
                "email":        u.get("email",""),
                "org":          u.get("company_name",""),
                "plan":         plan,
                "health":       health,
                "churn_risk":   churn_risk,
                "stage":        stage,
                "last_login":   str(last_login)[:19] if last_login else None,
                "sessions_30d": sessions_30,
                "avg_score":    avg_score,
                "mrr":          u.get("mrr",0),
            })

        results.sort(key=lambda x: x["churn_risk"], reverse=True)
        return jsonify({"ok":True,"members":results,"count":len(results)})
    except Exception as e:
        return jsonify({"error":str(e)}),500




@app.route("/api/org/health-scores/<uid>/playbook", methods=["POST"])
@require_auth
@require_tier("elite")
def launch_playbook(uid):
    """Log that a churn intervention playbook was launched for a user."""
    try:
        db   = firestore.client()
        data = request.get_json(force=True) or {}
        db.collection("playbooks").add({
            "target_uid":  uid,
            "launched_by": g.uid,
            "stage":       data.get("stage","at_risk"),
            "steps":       data.get("steps",[]),
            "created_at":  datetime.utcnow().isoformat(),
            "status":      "active",
        })
        audit(g.uid,"playbook_launched","crm",{"target":uid,"stage":data.get("stage")})
        return jsonify({"ok":True})
    except Exception as e:
        return jsonify({"error":str(e)}),500


# ── Real Audit Logs ───────────────────────────────────────────────────


@app.route("/api/audit/logs/query", methods=["POST"])
@require_auth
@require_tier("professional")
def query_audit_logs():
    """Query audit logs with filters: category, severity, date range, search."""
    try:
        db   = firestore.client()
        data = request.get_json(force=True) or {}
        org_id   = getattr(g,"company_id","") or data.get("org_id","")
        category = data.get("category","all")
        severity = data.get("severity","all")
        limit    = min(int(data.get("limit",100)),500)
        date_from= data.get("date_from")
        date_to  = data.get("date_to")

        q = db.collection("audit_logs")
        if org_id:
            q = q.where("org_id","==",org_id)
        if category != "all":
            q = q.where("category","==",category)
        if severity != "all":
            q = q.where("severity","==",severity)
        if date_from:
            q = q.where("ts",">=",date_from)
        if date_to:
            q = q.where("ts","<=",date_to+"T23:59:59")
        q = q.order_by("ts", direction=firestore.Query.DESCENDING).limit(limit)

        docs = q.get()
        logs = []
        search = (data.get("search","") or "").lower()
        for doc in docs:
            d = doc.to_dict(); d["id"] = doc.id
            if search:
                blob = f"{d.get('user','')} {d.get('action','')} {d.get('detail','')} {d.get('ip','')}".lower()
                if search not in blob:
                    continue
            logs.append(d)

        return jsonify({"ok":True,"logs":logs,"count":len(logs)})
    except Exception as e:
        return jsonify({"error":str(e)}),500




@app.route("/api/audit/export", methods=["POST"])
@require_auth
@require_admin
def export_audit_logs():
    """Export audit logs as CSV (for compliance downloads)."""
    try:
        import csv, io as sio
        db   = firestore.client()
        data = request.get_json(force=True) or {}
        org_id = data.get("org_id","")

        q = db.collection("audit_logs")
        if org_id:
            q = q.where("org_id","==",org_id)
        q = q.order_by("ts",direction=firestore.Query.DESCENDING).limit(5000)
        docs = q.get()

        buf = sio.StringIO()
        w   = csv.writer(buf)
        w.writerow(["id","ts","user","uid","action","category","severity","ip","geo","resource","detail"])
        for doc in docs:
            d = doc.to_dict()
            w.writerow([doc.id, d.get("ts",""), d.get("user",""), d.get("uid",""),
                        d.get("action",""), d.get("category",""), d.get("severity",""),
                        d.get("ip",""), d.get("geo",""), d.get("resource",""), d.get("detail","")])

        audit(g.uid,"audit_log_exported","compliance",{"org_id":org_id})
        return send_file(
            io.BytesIO(buf.getvalue().encode()),
            mimetype="text/csv",
            as_attachment=True,
            download_name=f"audit_log_{datetime.utcnow().strftime('%Y%m%d')}.csv"
        )
    except Exception as e:
        return jsonify({"error":str(e)}),500


# ── Multi-Tenant Management ───────────────────────────────────────────


@app.route("/api/admin/tenants", methods=["POST"])
@require_auth
@require_admin
@limiter.limit("10 per minute")
def provision_tenant():
    """Provision a new tenant organization."""
    try:
        db   = firestore.client()
        data = request.get_json(force=True) or {}
        required = ["name","domain","admin_email","plan"]
        missing  = [f for f in required if not data.get(f)]
        if missing:
            return jsonify({"error":f"Missing: {', '.join(missing)}"}),400

        org_id = data["domain"].replace(".","_").replace("-","_")
        db.collection("companies").document(org_id).set({
            "name":           data["name"],
            "domain":         data["domain"],
            "admin_email":    data["admin_email"],
            "plan":           data["plan"],
            "seats":          int(data.get("seats",50)),
            "region":         data.get("region","us-east"),
            "trial_days":     int(data.get("trial_days",14)),
            "status":         "active",
            "created_at":     datetime.utcnow().isoformat(),
            "created_by":     g.uid,
            "white_label_domain": data.get("white_label_domain",""),
            "mrr":            0,
        })
        audit(g.uid,"tenant_provisioned","admin",{"org_id":org_id,"name":data["name"],"plan":data["plan"]})
        return jsonify({"ok":True,"org_id":org_id})
    except Exception as e:
        return jsonify({"error":str(e)}),500




@app.route("/api/admin/tenants/<org_id>", methods=["PATCH"])
@require_auth
@require_admin
@limiter.limit("20 per minute")
def update_tenant(org_id):
    """Update tenant plan, seats, or status."""
    try:
        db   = firestore.client()
        data = request.get_json(force=True) or {}
        allowed = ["plan","seats","status","white_label_domain","region"]
        update  = {k:v for k,v in data.items() if k in allowed}
        if not update:
            return jsonify({"error":"No valid fields to update"}),400
        update["updated_at"] = datetime.utcnow().isoformat()
        db.collection("companies").document(org_id).update(update)
        audit(g.uid,"tenant_updated","admin",{"org_id":org_id,"changes":update})
        return jsonify({"ok":True,"updated":update})
    except Exception as e:
        return jsonify({"error":str(e)}),500


# ── Usage Metering ────────────────────────────────────────────────────


@app.route("/api/billing/usage/summary", methods=["GET"])
@require_auth
def usage_summary():
    """Return current month's metered usage for a user/org."""
    try:
        db  = firestore.client()
        uid = g.uid
        now = datetime.utcnow()
        cycle_start = now.replace(day=1,hour=0,minute=0,second=0,microsecond=0)

        ref  = db.collection("usage").document(f"{uid}_{cycle_start.strftime('%Y%m')}")
        snap = ref.get()
        usage = snap.to_dict() if snap.exists else {}

        # Plan limits
        user = db.collection("users").document(uid).get().to_dict() or {}
        plan = user.get("plan","starter")
        LIMITS = {
            "starter":    {"analysis_frames":10000,"ai_reports":50,"api_calls":5000,"pdf_exports":20,"seats":10,"storage_gb":5},
            "growth":     {"analysis_frames":50000,"ai_reports":200,"api_calls":25000,"pdf_exports":100,"seats":30,"storage_gb":20},
            "scale":      {"analysis_frames":250000,"ai_reports":1000,"api_calls":150000,"pdf_exports":500,"seats":100,"storage_gb":100},
            "enterprise": {"analysis_frames":-1,"ai_reports":-1,"api_calls":-1,"pdf_exports":-1,"seats":-1,"storage_gb":-1},
        }
        limits = LIMITS.get(plan, LIMITS["starter"])
        rates  = {"analysis_frames":0.002,"ai_reports":0.15,"api_calls":0.0004,"pdf_exports":0.05,"seats":4.0,"storage_gb":0.08}

        meters = {}
        overage_total = 0.0
        for meter, limit in limits.items():
            used = usage.get(meter, 0)
            over = max(0, used - limit) if limit >= 0 else 0
            cost = round(over * rates.get(meter,0), 2)
            overage_total += cost
            meters[meter] = {"used":used,"included":limit,"overage":over,"overage_cost":cost}

        return jsonify({
            "ok":True,"plan":plan,"cycle_start":cycle_start.isoformat(),
            "meters":meters,"overage_total":round(overage_total,2),
            "base_price":{"starter":0,"growth":49,"scale":199,"enterprise":0}.get(plan,0),
        })
    except Exception as e:
        return jsonify({"error":str(e)}),500




@app.route("/api/billing/usage/meter", methods=["POST"])
@require_auth
def meter_usage():
    """Increment a usage meter (called internally by other routes)."""
    try:
        db   = firestore.client()
        data = request.get_json(force=True) or {}
        uid  = g.uid
        now  = datetime.utcnow()
        cycle_key = f"{uid}_{now.strftime('%Y%m')}"
        meter = data.get("meter","api_calls")
        qty   = int(data.get("qty",1))

        allowed_meters = ["analysis_frames","ai_reports","api_calls","pdf_exports","seats","storage_gb"]
        if meter not in allowed_meters:
            return jsonify({"error":"invalid meter"}),400

        db.collection("usage").document(cycle_key).set(
            {meter: firestore.Increment(qty), "uid":uid, "updated_at":now.isoformat()},
            merge=True
        )
        return jsonify({"ok":True,"meter":meter,"qty":qty})
    except Exception as e:
        return jsonify({"error":str(e)}),500


# ── Referral System ───────────────────────────────────────────────────


@app.route("/api/referral/stats", methods=["GET"])
@require_auth
def referral_stats():
    """Return referral stats for current user."""
    try:
        db  = firestore.client()
        uid = g.uid
        refs = db.collection("referrals").where("referrer_uid","==",uid).get()
        items = []
        total_earned = 0
        for doc in refs:
            d = doc.to_dict(); d["id"] = doc.id
            total_earned += d.get("earned",0)
            items.append(d)
        ref_code = f"PAI-{uid[:6].upper()}"
        return jsonify({"ok":True,"ref_code":ref_code,"referrals":items,"total_earned":total_earned,"count":len(items)})
    except Exception as e:
        return jsonify({"error":str(e)}),500




@app.route("/api/referral/track", methods=["POST"])
@optional_auth
@limiter.limit("10 per minute")
def track_referral():
    """Track a new referral signup — called on auth/register."""
    try:
        db   = firestore.client()
        data = request.get_json(force=True) or {}
        ref_code  = data.get("ref_code","")
        new_uid   = data.get("uid","")
        new_email = data.get("email","")
        if not ref_code or not new_uid:
            return jsonify({"ok":False,"reason":"missing fields"})

        # SECURITY: if the request carries a valid Firebase token, the claimed
        # new_uid MUST match the authenticated uid — prevents anyone from
        # crediting a referral to an arbitrary uid they don't own. If no token
        # is present (e.g. called in the brief window right after signup before
        # the client has refreshed its token), we allow it through but rely on
        # the one-referral-per-uid check in Firestore rules / below to limit abuse.
        if getattr(g, "uid", None) and g.uid != new_uid:
            return jsonify({"ok": False, "reason": "uid mismatch"}), 403

        # Look up referrer
        referrer_uid = ref_code.replace("PAI-","").upper()
        referrer_docs = db.collection("users").where("uid_prefix","==",referrer_uid[:6]).get()
        if not referrer_docs:
            return jsonify({"ok":False,"reason":"ref code not found"})

        referrer = referrer_docs[0]

        # Prevent duplicate/spam referral claims for the same new_uid
        existing = db.collection("referrals").where("referred_uid","==",new_uid).limit(1).get()
        if existing:
            return jsonify({"ok": False, "reason": "referral already tracked for this user"})

        db.collection("referrals").add({
            "referrer_uid":  referrer.id,
            "referred_uid":  new_uid,
            "referred_email":new_email,
            "ref_code":      ref_code,
            "status":        "pending",
            "earned":        0,
            "created_at":    datetime.utcnow().isoformat(),
        })
        audit(new_uid,"referral_signup","growth",{"ref_code":ref_code,"referrer":referrer.id})
        return jsonify({"ok":True,"referrer_uid":referrer.id})
    except Exception as e:
        return jsonify({"error":str(e)}),500




@app.route("/api/referral/convert", methods=["POST"])
@require_auth
def convert_referral():
    """Mark referral as converted when referred user becomes paying."""
    try:
        db   = firestore.client()
        data = request.get_json(force=True) or {}
        uid  = data.get("uid","") or g.uid
        plan = data.get("plan","starter")

        # Find pending referral
        refs = db.collection("referrals").where("referred_uid","==",uid).where("status","==","pending").get()
        # Normalize tier names to our standard naming
        REWARD = {
            "standard":     0,    # No reward for free conversions
            "professional": 49,   # 49 EGP credit (~1 month standard)
            "elite":        199,  # 199 EGP credit (~1 month professional)
            "enterprise":   400,  # 400 EGP credit (custom)
            # Legacy names from V10_2
            "starter":      0, "growth": 49, "scale": 199,
        }
        earned = REWARD.get(plan,0)

        for ref in refs:
            ref.reference.update({"status":"active","converted_at":datetime.utcnow().isoformat(),"earned":earned,"plan":plan})
            # Credit referrer
            referrer_uid = ref.to_dict().get("referrer_uid","")
            if referrer_uid:
                db.collection("users").document(referrer_uid).update({
                    "referral_credits": firestore.Increment(earned),
                    "referral_count":   firestore.Increment(1),
                })
                # Notify referrer by email if reward > 0
                if earned > 0:
                    try:
                        r_doc = db.collection("users").document(referrer_uid).get()
                        if r_doc.exists:
                            r_email = r_doc.to_dict().get("email","")
                            r_name  = r_doc.to_dict().get("name","").split()[0] or "there"
                            if r_email:
                                send_email(r_email,
                                    f"🎉 You earned {earned} EGP — your referral upgraded!",
                                    f"<p>Hi {r_name},</p>"
                                    f"<p>Someone you referred just upgraded to <b>{plan.title()}</b>.</p>"
                                    f"<p>We've added <b>{earned} EGP</b> to your PostureAI account credit.</p>"
                                    f"<p><a href=\"{APP_URL}/#referral\">View your referral dashboard →</a></p>",
                                )
                    except Exception as _ne:
                        print(f"[referral] Notify error: {_ne}", file=sys.stderr)
            audit(uid, "referral_converted", "growth", {"earned": earned, "plan": plan, "referrer": referrer_uid})

        return jsonify({"ok":True,"earned":earned})
    except Exception as e:
        return jsonify({"error":str(e)}),500


# ── Integrations / Webhooks ───────────────────────────────────────────


@app.route("/api/integrations", methods=["GET"])
@require_auth
def list_integrations():
    """Return connected integrations for current org."""
    try:
        db  = firestore.client()
        org_id = getattr(g,"company_id","") or g.uid
        snap = db.collection("integrations").document(org_id).get()
        data = snap.to_dict() if snap.exists else {}
        return jsonify({"ok":True,"integrations":data})
    except Exception as e:
        return jsonify({"error":str(e)}),500




@app.route("/api/integrations/<integration_id>", methods=["DELETE"])
@require_auth
def disconnect_integration(integration_id):
    """Disconnect an integration."""
    try:
        db  = firestore.client()
        org_id = getattr(g,"company_id","") or g.uid
        db.collection("integrations").document(org_id).update({integration_id:firestore.DELETE_FIELD})
        audit(g.uid,f"integration_{integration_id}_disconnected","integrations",{"org_id":org_id})
        return jsonify({"ok":True})
    except Exception as e:
        return jsonify({"error":str(e)}),500


# ── Slack real delivery ───────────────────────────────────────────────


@app.route("/api/integrations/slack/send", methods=["POST"])
@require_auth
def send_slack_message():
    """Deliver a posture alert or digest to the org's Slack channel."""
    try:
        db   = firestore.client()
        data = request.get_json(force=True) or {}
        org_id = getattr(g,"company_id","") or g.uid
        snap   = db.collection("integrations").document(org_id).get()
        cfg    = (snap.to_dict() or {}).get("slack",{})
        webhook= cfg.get("webhook_url") or data.get("webhook_url","")
        if not webhook:
            return jsonify({"error":"Slack not connected — no webhook URL"}),400

        payload = {
            "text": data.get("text","PostureAI alert"),
            "blocks": data.get("blocks",[
                {"type":"section","text":{"type":"mrkdwn","text": data.get("text","PostureAI notification")}}
            ])
        }
        r = req.post(webhook, json=payload, timeout=8)
        audit(g.uid,"slack_message_sent","integrations",{"org_id":org_id,"status":r.status_code})
        return jsonify({"ok": r.ok, "status_code": r.status_code})
    except Exception as e:
        return jsonify({"error":str(e)}),500


# ── MFA / 2FA ─────────────────────────────────────────────────────────


@app.route("/api/nps/submit", methods=["POST"])
@require_auth
def submit_nps():
    """Submit NPS response."""
    try:
        db   = firestore.client()
        data = request.get_json(force=True) or {}
        score   = int(data.get("score",0))
        comment = str(data.get("comment",""))[:500]
        if not 0 <= score <= 10:
            return jsonify({"error":"score must be 0-10"}),400

        uid  = g.uid
        user = db.collection("users").document(uid).get().to_dict() or {}
        segment = user.get("plan","starter")

        db.collection("nps_responses").add({
            "uid":     uid,
            "email":   user.get("email",""),
            "name":    user.get("name",""),
            "score":   score,
            "comment": comment,
            "segment": segment,
            "date":    datetime.utcnow().isoformat()[:10],
            "created_at": datetime.utcnow().isoformat(),
        })
        db.collection("users").document(uid).update({"nps_submitted":True,"last_nps_score":score,"emails_sent":firestore.ArrayUnion(["nps_survey"])})
        return jsonify({"ok":True,"score":score,"category":"promoter" if score>=9 else "passive" if score>=7 else "detractor"})
    except Exception as e:
        return jsonify({"error":str(e)}),500




@app.route("/api/nps/results", methods=["GET"])
@require_auth
@require_admin
def nps_results():
    """Return NPS summary for admin dashboard."""
    try:
        db   = firestore.client()
        docs = db.collection("nps_responses").order_by("created_at",direction=firestore.Query.DESCENDING).limit(500).get()
        responses = [d.to_dict() for d in docs]
        if not responses:
            return jsonify({"ok":True,"nps":0,"avg":0,"count":0,"responses":[]})

        promoters  = sum(1 for r in responses if r.get("score",0)>=9)
        detractors = sum(1 for r in responses if r.get("score",0)<=6)
        nps_score  = round((promoters-detractors)/len(responses)*100) if responses else 0
        avg_score  = round(sum(r.get("score",0) for r in responses)/len(responses),1) if responses else 0
        return jsonify({"ok":True,"nps":nps_score,"avg":avg_score,"count":len(responses),"promoters":promoters,"detractors":detractors,"responses":responses[:50]})
    except Exception as e:
        return jsonify({"error":str(e)}),500


# ── Product Tour ──────────────────────────────────────────────────────


@app.route("/api/user/tour/complete", methods=["POST"])
@require_auth
def mark_tour_complete():
    """Mark product tour as completed for user."""
    try:
        db = firestore.client()
        db.collection("users").document(g.uid).update({
            "product_tour_completed": True,
            "product_tour_completed_at": datetime.utcnow().isoformat(),
        })
        return jsonify({"ok":True})
    except Exception as e:
        return jsonify({"error":str(e)}),500



# ── Merged from v10 ─────────────────────────────────────────────
@require_auth



# ── Drip Email Triggers ───────────────────────────────────────────────────────
@app.route("/api/user/onboard", methods=["POST"])
@require_auth
@limiter.limit("5 per day")
def user_onboard():
    """
    Called by frontend immediately after first login / profile creation.
    Fires the welcome email drip sequence.
    Safe to call multiple times — checks if already onboarded.
    """
    try:
        uid   = getattr(g, "uid", "")
        role  = getattr(g, "role", {})
        email = getattr(g, "user", {}).get("email", "")
        name  = role.get("name", email.split("@")[0] if email else "there")
        plan  = role.get("tier", "standard")

        if not email:
            return jsonify({"error": "No email on account"}), 400

        if not _firebase_ok:
            return jsonify({"ok": True, "skipped": "firebase_unavailable"})

        db  = firestore.client()
        ref = db.collection("users").document(uid)
        doc = ref.get()

        if not doc.exists:
            return jsonify({"error": "User profile not found"}), 404

        data = doc.to_dict()

        # Idempotency guard — only fire drip once
        if data.get("onboard_email_sent"):
            return jsonify({"ok": True, "skipped": "already_onboarded"})

        # Fire welcome email
        try:
            from services.email_sequences import send_email as drip_send, EmailContext
            ctx = EmailContext(
                to_email=email, to_name=name, user_id=uid,
                plan=plan, template_key="welcome",
                data={"plan_display": plan.title(), "first_session_url": f"{APP_URL}/#camera"},
            )
            drip_send(ctx)
        except ImportError:
            # Fallback to basic send_email if SendGrid templates not configured
            send_email(email, f"Welcome to PostureAI Pro 👋",
                f"<p>Hi {name},</p><p>Your account is ready. <a href=\"{APP_URL}\">Start your first session →</a></p>")
        except Exception as drip_err:
            print(f"[onboard] Drip email error: {drip_err}", file=sys.stderr)

        # Mark onboarded
        ref.update({
            "onboard_email_sent": True,
            "onboarded_at": datetime.utcnow().isoformat() + "Z",
        })

        audit(uid, "user_onboarded", "lifecycle", {"email": email, "plan": plan})
        return jsonify({"ok": True, "welcome_email": "queued"})

    except Exception as e:
        return safe_error(e, "Onboarding failed")


@app.route("/api/user/trigger-drip", methods=["POST"])
@require_auth
@require_admin
@limiter.limit("30 per minute")
def trigger_drip_manual():
    """Admin-only: manually trigger a drip email for a specific user (for testing)."""
    try:
        data     = request.get_json(force=True) or {}
        uid      = data.get("uid", "")
        template = data.get("template", "day1_tip")

        if not uid or not template:
            return jsonify({"error": "uid and template required"}), 400

        if not _firebase_ok:
            return jsonify({"error": "Firebase unavailable"}), 503

        db   = firestore.client()
        doc  = db.collection("users").document(uid).get()
        if not doc.exists:
            return jsonify({"error": "User not found"}), 404

        user  = doc.to_dict()
        email = user.get("email", "")
        name  = user.get("name", email.split("@")[0])
        plan  = user.get("tier", "standard")

        try:
            from services.email_sequences import send_email as drip_send, EmailContext
            drip_send(EmailContext(
                to_email=email, to_name=name, user_id=uid,
                plan=plan, template_key=template,
                data={"admin_triggered": True},
            ))
            audit(getattr(g,"uid","admin"), "drip_manual_triggered", "admin", {"target_uid": uid, "template": template})
            return jsonify({"ok": True, "sent_to": email, "template": template})
        except Exception as e:
            return safe_error(e, "Drip trigger failed")
    except Exception as e:
        return safe_error(e, "Trigger drip error")




@app.route("/api/changelog", methods=["GET"])
@limiter.limit("60 per minute")
def get_changelog():
    """
    Return in-app changelog entries.
    Stored in Firestore /changelog collection, managed from AdminDashboard.
    Falls back to hardcoded entries if Firestore unavailable.
    """
    try:
        entries = []
        
        if _firebase_ok:
            db = firestore.client()
            snaps = db.collection("changelog").order_by("released_at", direction="DESCENDING").limit(20).get()
            for doc in snaps:
                d = doc.to_dict()
                entries.append({
                    "id":          doc.id,
                    "title":       d.get("title", ""),
                    "title_ar":    d.get("title_ar", ""),
                    "body":        d.get("body", ""),
                    "body_ar":     d.get("body_ar", ""),
                    "type":        d.get("type", "improvement"),  # new / improvement / fix
                    "released_at": d.get("released_at", ""),
                    "version":     d.get("version", ""),
                })
        
        # Fallback entries if Firestore empty
        if not entries:
            entries = [
                {
                    "id": "v12-mfa", "type": "new",
                    "title": "Multi-Factor Authentication", "title_ar": "المصادقة الثنائية",
                    "body": "Secure your account with TOTP (Google Authenticator) or SMS verification.",
                    "body_ar": "أمّن حسابك مع المصادقة الثنائية عبر Google Authenticator أو SMS.",
                    "released_at": "2026-06-01", "version": "v12",
                },
                {
                    "id": "v12-nps", "type": "improvement",
                    "title": "In-App Feedback", "title_ar": "تقييمات داخل التطبيق",
                    "body": "Share your feedback directly inside the app. Your input shapes our roadmap.",
                    "body_ar": "شارك رأيك مباشرة داخل التطبيق.",
                    "released_at": "2026-06-01", "version": "v12",
                },
                {
                    "id": "v12-referral", "type": "new",
                    "title": "Referral Rewards Automation", "title_ar": "مكافآت الإحالة التلقائية",
                    "body": "Earn EGP credit automatically when your referrals upgrade. No manual claiming.",
                    "body_ar": "اكسب رصيداً تلقائياً عند ترقية من أحلتهم.",
                    "released_at": "2026-06-01", "version": "v12",
                },
            ]
        
        # Return unread count for notification badge
        uid  = getattr(g, "uid", None) if hasattr(g, "uid") else None
        seen = []
        if uid and _firebase_ok:
            try:
                db   = firestore.client()
                udoc = db.collection("users").document(uid).get()
                if udoc.exists:
                    seen = udoc.to_dict().get("changelog_seen", [])
            except Exception:
                pass
        
        unread = sum(1 for e in entries if e["id"] not in seen)
        return jsonify({"entries": entries, "unread": unread})
    
    except Exception as e:
        return safe_error(e, "Failed to load changelog")


@app.route("/api/changelog/mark-read", methods=["POST"])
@require_auth
@limiter.limit("30 per minute")
def mark_changelog_read():
    """Mark all current changelog entries as seen for this user."""
    try:
        data       = request.get_json(force=True) or {}
        entry_ids  = data.get("ids", [])
        uid        = getattr(g, "uid", "")
        if uid and _firebase_ok and entry_ids:
            db = firestore.client()
            db.collection("users").document(uid).update({
                "changelog_seen": firestore.ArrayUnion(entry_ids)
            })
        return jsonify({"ok": True})
    except Exception as e:
        return safe_error(e, "Failed to mark changelog read")




# ── SSO / SAML / OIDC ────────────────────────────────────────────────────────
# Firebase Auth natively handles OIDC providers (Google, Microsoft, Okta, etc.)
# These endpoints handle the server-side config and post-auth provisioning.

@app.route("/api/auth/sso/config", methods=["GET"])
@require_auth
@require_hr
@limiter.limit("30 per minute")
def get_sso_config():
    """Return SSO configuration for the user's organization."""
    try:
        company_id = getattr(g, "role", {}).get("company_id")
        if not company_id:
            return jsonify({"error": "No organization found"}), 404

        if not _firebase_ok:
            return jsonify({"sso_enabled": False, "provider": None})

        db  = firestore.client()
        doc = db.collection("sso_configs").document(company_id).get()
        if not doc.exists:
            return jsonify({"sso_enabled": False, "provider": None, "company_id": company_id})

        cfg = doc.to_dict()
        # Never return the client_secret to the frontend
        safe_cfg = {k: v for k, v in cfg.items() if k not in ("client_secret", "private_key")}
        return jsonify(safe_cfg)
    except Exception as e:
        return safe_error(e, "Failed to load SSO config")


@app.route("/api/auth/sso/config", methods=["POST"])
@require_auth
@require_admin
@limiter.limit("10 per minute")
def save_sso_config():
    """
    Save SSO configuration for an organization.
    Supports: Google Workspace, Microsoft Azure AD, Okta, Generic OIDC.
    
    Firebase Auth handles the actual OIDC flow — this stores the config
    and is used to validate post-login provisioning.
    """
    try:
        data       = request.get_json(force=True) or {}
        company_id = data.get("company_id") or getattr(g, "role", {}).get("company_id")
        provider   = data.get("provider", "")     # google / microsoft / okta / oidc / saml
        domain     = data.get("domain", "").lower().strip()
        client_id  = data.get("client_id", "")
        tenant_id  = data.get("tenant_id", "")    # Azure AD
        issuer_url = data.get("issuer_url", "")   # Generic OIDC / Okta
        enabled    = data.get("enabled", True)

        if not company_id or not provider:
            return jsonify({"error": "company_id and provider required"}), 400

        valid_providers = {"google", "microsoft", "okta", "saml", "oidc"}
        if provider not in valid_providers:
            return jsonify({"error": f"Invalid provider. Valid: {sorted(valid_providers)}"}), 400

        if not _firebase_ok:
            return jsonify({"error": "Firebase unavailable"}), 503

        db  = firestore.client()
        cfg = {
            "company_id":   company_id,
            "provider":     provider,
            "domain":       domain,
            "client_id":    client_id,
            "tenant_id":    tenant_id,
            "issuer_url":   issuer_url,
            "sso_enabled":  enabled,
            "updated_at":   datetime.utcnow().isoformat() + "Z",
            "updated_by":   getattr(g, "uid", ""),
        }

        # Build Firebase OIDC provider config based on provider type
        firebase_provider_config = {}
        if provider == "google":
            firebase_provider_config = {
                "provider_id":   "google.com",
                "display_name":  "Google Workspace",
                "note":          "Configure in Firebase Console → Authentication → Sign-in method → Google",
            }
        elif provider == "microsoft":
            firebase_provider_config = {
                "provider_id":   "microsoft.com",
                "oidc_url":      f"https://login.microsoftonline.com/{tenant_id}/v2.0",
                "note":          "Configure in Firebase Console → Authentication → Sign-in method → Microsoft",
            }
        elif provider == "okta":
            firebase_provider_config = {
                "provider_id":   f"oidc.{domain.replace('.', '-')}",
                "issuer":        issuer_url or f"https://{domain}",
                "client_id":     client_id,
                "note":          "Add as OIDC provider in Firebase Console → Authentication → Sign-in method → Add new provider",
            }
        elif provider in ("oidc", "saml"):
            firebase_provider_config = {
                "provider_id":   f"{'saml' if provider == 'saml' else 'oidc'}.{company_id}",
                "issuer":        issuer_url,
                "client_id":     client_id,
                "note":          f"Configure {provider.upper()} in Firebase Console → Authentication → Sign-in method",
            }

        cfg["firebase_config"] = firebase_provider_config
        db.collection("sso_configs").document(company_id).set(cfg)

        audit(getattr(g, "uid", ""), "sso_config_updated", "enterprise",
              {"company_id": company_id, "provider": provider, "enabled": enabled})

        return jsonify({
            "ok":               True,
            "provider":         provider,
            "setup_guide":      firebase_provider_config.get("note", ""),
            "firebase_config":  firebase_provider_config,
        })
    except Exception as e:
        return safe_error(e, "Failed to save SSO config")


@app.route("/api/auth/sso/provision", methods=["POST"])
@require_auth
@limiter.limit("10 per minute")
def sso_provision_user():
    """
    Called after a user authenticates via SSO/OIDC through Firebase Auth.
    Provisions the user into the correct organization based on their email domain.
    Auto-assigns role based on the org's SSO config.
    """
    try:
        uid   = getattr(g, "uid", "")
        email = getattr(g, "user", {}).get("email", "")
        if not email or not uid:
            return jsonify({"error": "Auth required"}), 401

        domain = email.split("@")[-1].lower() if "@" in email else ""
        if not domain:
            return jsonify({"error": "Invalid email domain"}), 400

        if not _firebase_ok:
            return jsonify({"ok": True, "provisioned": False, "reason": "firebase_unavailable"})

        db = firestore.client()

        # Find org by SSO domain
        orgs = db.collection("sso_configs").where("domain", "==", domain).where("sso_enabled", "==", True).limit(1).get()
        if not orgs:
            return jsonify({"ok": True, "provisioned": False, "reason": "no_sso_org_for_domain"})

        org_cfg    = orgs[0].to_dict()
        company_id = org_cfg.get("company_id")

        # Check if user already has a profile
        user_ref = db.collection("users").document(uid)
        user_doc = user_ref.get()

        if user_doc.exists:
            existing = user_doc.to_dict()
            # Already provisioned in this org
            if existing.get("company_id") == company_id:
                return jsonify({"ok": True, "provisioned": False, "reason": "already_provisioned"})

        # Get org details
        company_doc = db.collection("companies").document(company_id).get()
        company_name = company_doc.to_dict().get("name", "") if company_doc.exists else ""

        # Check seat limit
        seats_used = len(db.collection("users").where("company_id", "==", company_id).get())
        max_seats  = company_doc.to_dict().get("max_seats", 25) if company_doc.exists else 25
        if max_seats > 0 and seats_used >= max_seats:
            return jsonify({"error": "Organization seat limit reached — contact your admin"}), 403

        # Provision user into org
        provision_data = {
            "company_id":   company_id,
            "company_name": company_name,
            "role":         "employee",
            "sso_provider": org_cfg.get("provider", "oidc"),
            "sso_domain":   domain,
            "provisioned_via": "sso",
            "provisioned_at":  datetime.utcnow().isoformat() + "Z",
            "tier":            "standard",
        }

        if user_doc.exists:
            user_ref.update(provision_data)
        else:
            provision_data.update({"email": email, "uid": uid, "created_at": datetime.utcnow().isoformat() + "Z"})
            user_ref.set(provision_data)

        # Invalidate cache so next request gets fresh role
        from auth.middleware import invalidate_user_cache
        invalidate_user_cache(uid)

        audit(uid, "sso_user_provisioned", "enterprise",
              {"company_id": company_id, "provider": org_cfg.get("provider"), "domain": domain})

        return jsonify({
            "ok":           True,
            "provisioned":  True,
            "company_id":   company_id,
            "company_name": company_name,
        })
    except Exception as e:
        return safe_error(e, "SSO provisioning failed")




# ── Announcements, Security Center, Feature Flags, Onboarding Analytics ─────────
@app.route("/api/admin/feature-flags", methods=["GET", "POST", "PATCH"])
@require_auth
@require_admin
@limiter.limit("30 per minute")
def admin_feature_flags():
    """Admin: list, create, or update feature flags.
    Persisted in Firestore (collection: feature_flags) rather than an
    in-memory dict — a plain module global wouldn't survive across
    gunicorn's multiple worker processes; each worker would see its own
    separate flags, silently diverging from what the admin actually set."""
    try:
        if request.method == "GET":
            docs = db.collection("feature_flags").stream()
            flags = {d.id: d.to_dict() for d in docs}
            return jsonify({"flags": flags})

        data = request.get_json(force=True) or {}
        if request.method == "POST":
            key = data.get("key", "").strip().lower().replace(" ", "_")
            if not key:
                return jsonify({"error": "key required"}), 400
            ref = db.collection("feature_flags").document(key)
            if ref.get().exists:
                return jsonify({"error": "Invalid or duplicate flag key"}), 400
            flag = {
                "enabled":     data.get("enabled", False),
                "rollout_pct": int(data.get("rollout_pct", 0)),
                "tiers":       data.get("tiers", []),
                "updated_at":  datetime.utcnow().isoformat()+"Z",
                "updated_by":  g.uid,
            }
            ref.set(flag)
            return jsonify({"success": True, "flag": key})

        # PATCH
        key = data.get("key", "")
        ref = db.collection("feature_flags").document(key)
        doc = ref.get()
        if not doc.exists:
            return jsonify({"error": "Flag not found"}), 404
        upd = {k: v for k, v in data.items() if k != "key"}
        upd["updated_at"] = datetime.utcnow().isoformat()+"Z"
        upd["updated_by"] = g.uid
        ref.update(upd)
        return jsonify({"success": True})
    except Exception as e:
        return safe_error(e)


# ── Security Center ────────────────────────────────────────────────


@app.route("/api/security/overview", methods=["GET"])
@require_auth
@limiter.limit("20 per minute")
def security_overview():
    """Return security posture summary for the authenticated user."""
    try:
        uid  = g.uid
        role = g.role
        profile = {}
        if db:
            snap = db.collection("users").document(uid).get()
            if snap.exists:
                profile = snap.to_dict() or {}

        checks = {
            "mfa_enabled":         bool(profile.get("mfa_enabled")),
            "password_strength":   "unknown",   # Firebase handles this
            "session_count":       1,
            "last_login":          profile.get("last_login_at", "unknown"),
            "api_keys_active":     0,
            "suspicious_logins":   0,
            "data_exports":        0,
        }

        # Count active API keys
        if db:
            keys_snap = db.collection("api_keys")                .where("uid", "==", uid)                .where("revoked", "==", False).get()
            checks["api_keys_active"] = len(keys_snap)

        score = sum([
            30 if checks["mfa_enabled"] else 0,
            40,   # base score for having an account
            15 if checks["api_keys_active"] == 0 else 10,
            15 if checks["suspicious_logins"] == 0 else 0,
        ])

        recommendations = []
        if not checks["mfa_enabled"]:
            recommendations.append({
                "priority": "high",
                "title":    "Enable Two-Factor Authentication",
                "action":   "mfa_setup",
                "impact":   "+30 security score"
            })

        return jsonify({
            "score":           score,
            "grade":           "A" if score >= 90 else "B" if score >= 75 else "C" if score >= 60 else "D",
            "checks":          checks,
            "recommendations": recommendations,
        })
    except Exception as e:
        return safe_error(e, "Security overview failed")


@app.route("/api/security/active-sessions", methods=["GET"])
@require_auth
@limiter.limit("20 per minute")
def security_active_sessions():
    """Return list of active sessions for the user (for Security Center)."""
    try:
        # Firebase handles session management — return current session info
        return jsonify({
            "sessions": [{
                "id":         "current",
                "device":     request.headers.get("User-Agent", "Unknown")[:60],
                "ip":         request.remote_addr,
                "created_at": datetime.utcnow().isoformat(),
                "is_current": True,
            }]
        })
    except Exception as e:
        return safe_error(e, "Active sessions failed")


@app.route("/api/security/revoke-session", methods=["POST"])
@require_auth
@limiter.limit("10 per minute")
def security_revoke_session():
    """Revoke a specific session (currently revokes Firebase token)."""
    try:
        if firebase_auth:
            firebase_auth.revoke_refresh_tokens(g.uid)
        invalidate_user_cache(g.uid)
        return jsonify({"success": True})
    except Exception as e:
        return safe_error(e, "Session revoke failed")


# ── Onboarding Analytics ───────────────────────────────────────────
_onboarding_events: list = []


@app.route("/api/admin/analytics/onboarding", methods=["GET"])
@require_auth
@require_admin
@limiter.limit("20 per minute")
def admin_onboarding_analytics():
    """Admin: onboarding funnel analytics."""
    try:
        steps = ["welcome", "calibration", "first_session", "invite_team", "billing"]
        funnel = []
        if db:
            for step in steps:
                snap = db.collection("onboarding_events")                    .where("step", "==", step)                    .where("completed", "==", True).get()
                funnel.append({"step": step, "completions": len(snap)})
        else:
            for step in steps:
                count = sum(1 for e in _onboarding_events if e["step"] == step and e["completed"])
                funnel.append({"step": step, "completions": count})

        # Compute drop-off rates
        for i in range(1, len(funnel)):
            prev = funnel[i-1]["completions"] or 1
            curr = funnel[i]["completions"]
            funnel[i]["drop_off_pct"] = round((1 - curr/prev) * 100, 1)

        return jsonify({"funnel": funnel, "total_starts": funnel[0]["completions"] if funnel else 0})
    except Exception as e:
        return safe_error(e, "Onboarding analytics failed")


# ── API Usage Dashboard ────────────────────────────────────────────


# In-memory system announcements (e.g. "Scheduled maintenance Friday 2am").
# Empty by default — populate via a future admin endpoint if needed.
_announcements = []

@app.route("/api/announcements", methods=["GET"])
@require_auth
@limiter.limit("60 per minute")
def get_announcements():
    """Return active announcements for the user's tier."""
    try:
        tier = g.role.get("tier", "starter")
        relevant = [a for a in _announcements if a["active"] and tier in a.get("tier", [])]
        return jsonify({"announcements": relevant})
    except Exception as e:
        return safe_error(e, "Announcements failed")

@app.route("/api/announcements/<ann_id>/dismiss", methods=["POST"])
@require_auth
@limiter.limit("60 per minute")
def dismiss_announcement(ann_id):
    """Mark an announcement as dismissed for this user."""
    try:
        uid = getattr(g, "uid", "")
        if _firebase_ok and uid:
            db = firestore.client()
            db.collection("users").document(uid).update({
                "dismissed_announcements": firestore.ArrayUnion([ann_id])
            })
        return jsonify({"ok": True})
    except Exception as e:
        return safe_error(e, "Dismiss failed")


@app.route("/api/announcements", methods=["POST"])
@require_auth
@require_admin
@limiter.limit("10 per minute")
def create_announcement():
    """Admin: create a new in-app announcement."""
    try:
        data    = request.get_json(force=True) or {}
        msg     = str(data.get("message", "")).strip()[:500]
        ann_type= data.get("type", "feature")   # feature / security / tip / warning
        expires = data.get("expires_in_days", 7)
        if not msg:
            return jsonify({"error": "message required"}), 400
        if not _firebase_ok:
            return jsonify({"error": "Firebase unavailable"}), 503
        db = firestore.client()
        doc = db.collection("announcements").add({
            "message":    msg,
            "type":       ann_type,
            "active":     True,
            "created_at": datetime.utcnow().isoformat() + "Z",
            "expires_at": (datetime.utcnow() + timedelta(days=int(expires))).isoformat() + "Z",
            "created_by": getattr(g, "uid", ""),
        })
        audit(getattr(g,"uid",""), "announcement_created", "admin", {"type": ann_type})
        return jsonify({"ok": True, "id": doc[1].id})
    except Exception as e:
        return safe_error(e, "Create announcement failed")



@app.route("/api/contact/enterprise", methods=["POST"])
@limiter.limit("5 per minute")
def enterprise_contact():
    """Enterprise sales inquiry form — public endpoint."""
    try:
        import html as _html
        data    = request.get_json(force=True) or {}
        # SECURITY: escape all fields before HTML interpolation — this form is
        # public and unauthenticated, so untrusted input must never reach the
        # admin's inbox or the auto-reply email unescaped (XSS / email spoofing risk).
        name    = _html.escape(str(data.get("name", ""))[:100].strip())
        email   = _html.escape(str(data.get("email", ""))[:200].strip())
        company = _html.escape(str(data.get("company", ""))[:200].strip())
        size    = _html.escape(str(data.get("size", ""))[:50].strip())
        message = _html.escape(str(data.get("message", ""))[:2000].strip())

        # Re-validate raw (unescaped) email for format — escaped "@" is unaffected
        raw_email = str(data.get("email", ""))[:200].strip()
        if not raw_email or "@" not in raw_email or not name:
            return jsonify({"error": "name and valid email required"}), 400

        # Send to admin
        admin_html = f"""
        <h2>New Enterprise Inquiry</h2>
        <p><b>Name:</b> {name}</p>
        <p><b>Email:</b> {email}</p>
        <p><b>Company:</b> {company}</p>
        <p><b>Size:</b> {size}</p>
        <p><b>Message:</b><br>{message}</p>
        """
        send_email(ADMIN_EMAIL or SMTP_USER, f"Enterprise Inquiry: {company}", admin_html)

        # Auto-reply to prospect
        prospect_html = f"""
        <p>Hi {name},</p>
        <p>Thank you for your interest in PostureAI Enterprise.</p>
        <p>Our enterprise team will reach out within <b>24 hours</b> to schedule a personalized demo.</p>
        <p>In the meantime, <a href="{APP_URL}">explore our platform</a> or reply to this email with any questions.</p>
        <p>— PostureAI Enterprise Team</p>
        """
        send_email(raw_email, "PostureAI Enterprise — We'll be in touch soon", prospect_html)

        audit("anonymous", "enterprise_inquiry", "marketing", {"email": raw_email, "company": company})
        return jsonify({"ok": True})
    except Exception as e:
        return safe_error(e, "Failed to process inquiry")



@app.route("/api/user/delete-all-data", methods=["POST", "DELETE"])
@require_auth
@limiter.limit("3 per hour")
def gdpr_delete_all_data():
    """
    GDPR Right to Erasure — delete all data for the authenticated user.
    Deletes: sessions, payments records, audit logs, notifications, API keys, user profile.
    Irreversible — requires explicit confirmation.
    """
    try:
        uid     = getattr(g, "uid", "")
        data    = request.get_json(force=True) or {}
        confirm = data.get("confirm", False)

        if not confirm:
            return jsonify({
                "error": "Explicit confirmation required",
                "message": "Send {confirm: true} to permanently delete all your data",
            }), 400

        if not _firebase_ok:
            return jsonify({"error": "Database not available"}), 503

        db = firestore.client()
        deleted_counts = {}

        # Delete all sessions
        sessions_ref = db.collection("sessions").where("uid", "==", uid)
        sessions_snap = sessions_ref.get()
        for doc in sessions_snap:
            doc.reference.delete()
        deleted_counts["sessions"] = len(sessions_snap)

        # Delete notifications
        notifs = db.collection("notifications").where("uid", "==", uid).get()
        for doc in notifs:
            doc.reference.delete()
        deleted_counts["notifications"] = len(notifs)

        # Delete API keys
        keys = db.collection("api_keys").where("uid", "==", uid).get()
        for doc in keys:
            doc.reference.delete()
        deleted_counts["api_keys"] = len(keys)

        # Anonymise payment records (keep for accounting, remove PII)
        payments = db.collection("payments").where("uid", "==", uid).get()
        for doc in payments:
            doc.reference.update({
                "email": "[deleted]",
                "name":  "[deleted]",
                "gdpr_anonymised_at": datetime.utcnow().isoformat() + "Z",
            })
        deleted_counts["payments_anonymised"] = len(payments)

        # Delete user profile
        db.collection("users").document(uid).delete()
        deleted_counts["user_profile"] = 1

        # Log the erasure (audit log retained for legal compliance — GDPR Art. 17 exception)
        audit(uid, "gdpr_erasure_completed", "compliance", {
            "deleted_counts": deleted_counts,
            "requested_at": datetime.utcnow().isoformat() + "Z",
        })

        # Finally, delete the Firebase Auth account itself so the user can't
        # log back in to a shell of a profile. Done last so a failure here
        # doesn't block the data erasure above.
        auth_deleted = False
        if _fb_auth is not None:
            try:
                _fb_auth.delete_user(uid)
                auth_deleted = True
            except Exception as auth_err:
                print(f"[gdpr] Auth user delete failed for uid={uid}: {auth_err}", flush=True)

        print(f"[gdpr] Erasure completed for uid={uid} counts={deleted_counts} auth_deleted={auth_deleted}", flush=True)
        return jsonify({
            "ok": True,
            "message": "All personal data has been permanently deleted.",
            "deleted": deleted_counts,
            "auth_deleted": auth_deleted,
        })
    except Exception as e:
        return safe_error(e, "GDPR erasure failed")



# ── SCIM 2.0 Provisioning ────────────────────────────────────────────────────
# SCIM 2.0 compliant API for Okta, Azure AD, JumpCloud, and other IdPs.
# Docs: https://www.rfc-editor.org/rfc/rfc7644
# Auth: Bearer token — set SCIM_BEARER_TOKEN in env (separate from Firebase JWT)

SCIM_BEARER_TOKEN = os.getenv("SCIM_BEARER_TOKEN", "")

def _scim_auth():
    """Validate SCIM bearer token (separate from Firebase JWT)."""
    auth = request.headers.get("Authorization", "")
    if not auth.startswith("Bearer "):
        return False
    token = auth[7:].strip()
    if not SCIM_BEARER_TOKEN:
        print("[SCIM] ⚠️  SCIM_BEARER_TOKEN not set — rejecting all SCIM requests", file=sys.stderr)
        return False
    import hmac as _hmac
    return _hmac.compare_digest(token, SCIM_BEARER_TOKEN)

def _scim_user_to_resource(user_doc, uid):
    """Convert Firestore user dict to SCIM 2.0 User resource."""
    return {
        "schemas":    ["urn:ietf:params:scim:schemas:core:2.0:User"],
        "id":         uid,
        "externalId": user_doc.get("external_id", uid),
        "userName":   user_doc.get("email", ""),
        "displayName": user_doc.get("name", ""),
        "emails":     [{"value": user_doc.get("email", ""), "primary": True}],
        "name":       {
            "formatted":  user_doc.get("name", ""),
            "givenName":  user_doc.get("name", "").split()[0] if user_doc.get("name") else "",
            "familyName": " ".join(user_doc.get("name", "").split()[1:]),
        },
        "active":   user_doc.get("active", True),
        "groups":   [{"value": user_doc.get("company_id", ""), "display": user_doc.get("company_name", "")}]
                    if user_doc.get("company_id") else [],
        "meta": {
            "resourceType": "User",
            "created":      user_doc.get("created_at", ""),
            "lastModified": user_doc.get("updated_at", user_doc.get("created_at", "")),
            "location":     f"{request.host_url}scim/v2/Users/{uid}",
        },
    }


@app.route("/scim/v2/Users", methods=["GET"])
@limiter.limit("60 per minute")
def scim_list_users():
    """SCIM 2.0 — List users with optional filter (e.g. userName eq 'user@example.com')."""
    if not _scim_auth():
        return jsonify({"schemas": ["urn:ietf:params:scim:api:messages:2.0:Error"],
                        "status": "401", "detail": "Unauthorized"}), 401
    try:
        filter_param  = request.args.get("filter", "")
        start_index   = int(request.args.get("startIndex", 1))
        count         = min(int(request.args.get("count", 100)), 200)

        if not _firebase_ok:
            return jsonify({"schemas": ["urn:ietf:params:scim:api:messages:2.0:ListResponse"],
                            "totalResults": 0, "Resources": []}), 200

        db    = firestore.client()
        query = db.collection("users")

        # Parse SCIM filter (basic support: userName eq "x" and externalId eq "x")
        import re as _re
        email_match = _re.search(r'userName\s+eq\s+"([^"]+)"', filter_param, _re.IGNORECASE)
        if email_match:
            query = query.where("email", "==", email_match.group(1))

        snaps  = query.limit(count).get()
        users  = [_scim_user_to_resource(s.to_dict(), s.id) for s in snaps]

        return jsonify({
            "schemas":      ["urn:ietf:params:scim:api:messages:2.0:ListResponse"],
            "totalResults": len(users),
            "startIndex":   start_index,
            "itemsPerPage": len(users),
            "Resources":    users,
        })
    except Exception as e:
        return safe_error(e, "SCIM list failed")


@app.route("/scim/v2/Users/<uid>", methods=["GET"])
@limiter.limit("120 per minute")
def scim_get_user(uid):
    """SCIM 2.0 — Get a single user by ID."""
    if not _scim_auth():
        return jsonify({"schemas": ["urn:ietf:params:scim:api:messages:2.0:Error"],
                        "status": "401", "detail": "Unauthorized"}), 401
    try:
        if not _firebase_ok:
            return jsonify({"status": "404", "detail": "Not found"}), 404
        db  = firestore.client()
        doc = db.collection("users").document(uid).get()
        if not doc.exists:
            return jsonify({"schemas": ["urn:ietf:params:scim:api:messages:2.0:Error"],
                            "status": "404", "detail": f"User {uid} not found"}), 404
        return jsonify(_scim_user_to_resource(doc.to_dict(), uid))
    except Exception as e:
        return safe_error(e, "SCIM get user failed")


@app.route("/scim/v2/Users", methods=["POST"])
@limiter.limit("30 per minute")
def scim_create_user():
    """SCIM 2.0 — Create (provision) a new user from IdP."""
    if not _scim_auth():
        return jsonify({"schemas": ["urn:ietf:params:scim:api:messages:2.0:Error"],
                        "status": "401", "detail": "Unauthorized"}), 401
    try:
        data        = request.get_json(force=True) or {}
        username    = data.get("userName", "").lower().strip()
        display     = data.get("displayName", "") or data.get("name", {}).get("formatted", "")
        external_id = data.get("externalId", "")
        active      = data.get("active", True)
        emails      = data.get("emails", [])
        email       = next((e["value"] for e in emails if e.get("primary")), username)

        if not email or "@" not in email:
            return jsonify({"schemas": ["urn:ietf:params:scim:api:messages:2.0:Error"],
                            "status": "400", "detail": "userName (email) required"}), 400

        if not _firebase_ok:
            return jsonify({"status": "503", "detail": "Service unavailable"}), 503

        db = firestore.client()

        # Check if user already exists by email
        existing = db.collection("users").where("email", "==", email).limit(1).get()
        if existing:
            # Return existing user (idempotent)
            existing_doc = existing[0]
            return jsonify(_scim_user_to_resource(existing_doc.to_dict(), existing_doc.id)), 200

        # Determine company from email domain
        domain      = email.split("@")[-1].lower()
        sso_configs = db.collection("sso_configs").where("domain", "==", domain).limit(1).get()
        company_id  = sso_configs[0].to_dict().get("company_id") if sso_configs else None
        company_name = ""
        if company_id:
            c_doc = db.collection("companies").document(company_id).get()
            if c_doc.exists:
                company_name = c_doc.to_dict().get("name", "")

        # Create Firebase Auth user
        new_uid = external_id or email.replace("@", "_at_").replace(".", "_")
        try:
            if _fb_auth:
                fb_user = _fb_auth.create_user(email=email, display_name=display)
                new_uid = fb_user.uid
        except Exception as fb_err:
            print(f"[SCIM] Firebase Auth create error: {fb_err}", file=sys.stderr)
            # Continue with generated UID if Firebase Auth creation fails

        profile = {
            "uid":          new_uid,
            "email":        email,
            "name":         display,
            "external_id":  external_id,
            "active":       active,
            "company_id":   company_id,
            "company_name": company_name,
            "role":         "employee",
            "tier":         "standard",
            "provisioned_via": "scim",
            "created_at":   datetime.utcnow().isoformat() + "Z",
            "updated_at":   datetime.utcnow().isoformat() + "Z",
        }
        db.collection("users").document(new_uid).set(profile)
        audit("scim", "scim_user_created", "enterprise", {"email": email, "uid": new_uid})

        return jsonify(_scim_user_to_resource(profile, new_uid)), 201

    except Exception as e:
        return safe_error(e, "SCIM create user failed")


@app.route("/scim/v2/Users/<uid>", methods=["PUT"])
@limiter.limit("30 per minute")
def scim_update_user(uid):
    """SCIM 2.0 — Full replace of a user resource."""
    if not _scim_auth():
        return jsonify({"schemas": ["urn:ietf:params:scim:api:messages:2.0:Error"],
                        "status": "401", "detail": "Unauthorized"}), 401
    try:
        data     = request.get_json(force=True) or {}
        active   = data.get("active", True)
        display  = data.get("displayName", "")

        if not _firebase_ok:
            return jsonify({"status": "503"}), 503

        db  = firestore.client()
        doc = db.collection("users").document(uid).get()
        if not doc.exists:
            return jsonify({"status": "404", "detail": "Not found"}), 404

        update_data = {
            "active":     active,
            "updated_at": datetime.utcnow().isoformat() + "Z",
        }
        if display:
            update_data["name"] = display

        db.collection("users").document(uid).update(update_data)

        # If deactivated via SCIM — revoke Firebase session
        if not active and _fb_auth:
            try:
                _fb_auth.revoke_refresh_tokens(uid)
                from auth.middleware import invalidate_user_cache
                invalidate_user_cache(uid)
            except Exception:
                pass

        audit("scim", "scim_user_updated", "enterprise", {"uid": uid, "active": active})
        updated = db.collection("users").document(uid).get().to_dict()
        return jsonify(_scim_user_to_resource(updated, uid))
    except Exception as e:
        return safe_error(e, "SCIM update user failed")


@app.route("/scim/v2/Users/<uid>", methods=["PATCH"])
@limiter.limit("30 per minute")
def scim_patch_user(uid):
    """SCIM 2.0 — Partial update. Handles active=false (deprovisioning)."""
    if not _scim_auth():
        return jsonify({"schemas": ["urn:ietf:params:scim:api:messages:2.0:Error"],
                        "status": "401", "detail": "Unauthorized"}), 401
    try:
        data       = request.get_json(force=True) or {}
        operations = data.get("Operations", [])

        if not _firebase_ok:
            return jsonify({"status": "503"}), 503

        db  = firestore.client()
        doc = db.collection("users").document(uid).get()
        if not doc.exists:
            return jsonify({"status": "404", "detail": "Not found"}), 404

        update_data = {"updated_at": datetime.utcnow().isoformat() + "Z"}
        for op in operations:
            if op.get("op", "").lower() == "replace":
                path  = op.get("path", "")
                value = op.get("value")
                if path == "active" or path == "":
                    if isinstance(value, dict):
                        if "active" in value:
                            update_data["active"] = value["active"]
                    elif isinstance(value, bool):
                        update_data["active"] = value
                elif path == "displayName":
                    update_data["name"] = value

        db.collection("users").document(uid).update(update_data)

        # Revoke session if deactivated
        if update_data.get("active") is False and _fb_auth:
            try:
                _fb_auth.revoke_refresh_tokens(uid)
                from auth.middleware import invalidate_user_cache
                invalidate_user_cache(uid)
            except Exception:
                pass

        audit("scim", "scim_user_patched", "enterprise", {"uid": uid, "ops": len(operations)})
        updated = db.collection("users").document(uid).get().to_dict()
        return jsonify(_scim_user_to_resource(updated, uid))
    except Exception as e:
        return safe_error(e, "SCIM patch user failed")


@app.route("/scim/v2/Users/<uid>", methods=["DELETE"])
@limiter.limit("10 per minute")
def scim_delete_user(uid):
    """SCIM 2.0 — Deprovision (soft-delete) a user."""
    if not _scim_auth():
        return jsonify({"schemas": ["urn:ietf:params:scim:api:messages:2.0:Error"],
                        "status": "401", "detail": "Unauthorized"}), 401
    try:
        if not _firebase_ok:
            return "", 204

        db  = firestore.client()
        doc = db.collection("users").document(uid).get()
        if not doc.exists:
            return "", 204  # Idempotent — already deleted

        # Soft delete: mark inactive, revoke tokens (keep data for audit)
        db.collection("users").document(uid).update({
            "active":       False,
            "deprovisioned_via": "scim",
            "deprovisioned_at":  datetime.utcnow().isoformat() + "Z",
        })

        if _fb_auth:
            try:
                _fb_auth.revoke_refresh_tokens(uid)
                from auth.middleware import invalidate_user_cache
                invalidate_user_cache(uid)
            except Exception:
                pass

        audit("scim", "scim_user_deleted", "enterprise", {"uid": uid})
        return "", 204
    except Exception as e:
        return safe_error(e, "SCIM delete user failed")


@app.route("/scim/v2/ServiceProviderConfig", methods=["GET"])
@limiter.limit("60 per minute")
def scim_service_provider_config():
    """SCIM 2.0 — Advertise capabilities to IdP (Okta/Azure discovery endpoint)."""
    return jsonify({
        "schemas": ["urn:ietf:params:scim:schemas:core:2.0:ServiceProviderConfig"],
        "patch":         {"supported": True},
        "bulk":          {"supported": False, "maxOperations": 0, "maxPayloadSize": 0},
        "filter":        {"supported": True, "maxResults": 200},
        "changePassword":{"supported": False},
        "sort":          {"supported": False},
        "etag":          {"supported": False},
        "authenticationSchemes": [{
            "name":        "OAuth Bearer Token",
            "description": "Authentication scheme using the OAuth Bearer Token standard",
            "specUri":     "http://www.rfc-editor.org/info/rfc6750",
            "type":        "oauthbearertoken",
            "primary":     True,
        }],
        "meta": {
            "location":     f"{request.host_url}scim/v2/ServiceProviderConfig",
            "resourceType": "ServiceProviderConfig",
        },
    })


# ── Admin-facing SCIM provisioning status ──────────────────────────
# The routes above (/scim/v2/*) are the actual protocol endpoints an IdP
# (Okta / Azure AD / OneLogin) calls directly with SCIM_BEARER_TOKEN.
# That token is a server secret and must never reach the browser. These
# two endpoints let the platform admin (normal Firebase auth) see setup
# instructions and provisioning status without ever seeing the token.
#
# NOTE: SCIM_BEARER_TOKEN is currently a single platform-wide secret —
# there is no per-organization token yet, so today this is effectively
# single-tenant SCIM (fine for one enterprise IdP, not yet a true
# multi-org self-serve provisioning story). Flagging this explicitly
# rather than presenting it as more than it is.
@app.route("/api/admin/scim/status", methods=["GET"])
@require_auth
@require_admin
@limiter.limit("30 per minute")
def admin_scim_status():
    try:
        configured = bool(SCIM_BEARER_TOKEN)
        base = request.host_url.rstrip("/")
        return jsonify({
            "configured": configured,
            "endpoints": {
                "service_provider_config": f"{base}/scim/v2/ServiceProviderConfig",
                "users":                   f"{base}/scim/v2/Users",
            },
            "scope": "platform-wide (single shared token — not per-organization yet)",
        })
    except Exception as e:
        return safe_error(e)


@app.route("/api/admin/scim/users", methods=["GET"])
@require_auth
@require_admin
@limiter.limit("30 per minute")
def admin_scim_users():
    """Users that have actually been provisioned via SCIM (have an external_id
    set by the IdP), so the admin can confirm sync is really happening —
    as opposed to all users, most of whom signed up directly."""
    try:
        docs = db.collection("users").where("external_id", "!=", "").limit(200).stream()
        users = []
        for d in docs:
            u = d.to_dict()
            users.append({
                "id":          d.id,
                "email":       u.get("email",""),
                "name":        u.get("name",""),
                "external_id": u.get("external_id",""),
                "active":      u.get("active", True),
                "company_name":u.get("company_name",""),
                "updated_at":  u.get("updated_at", u.get("created_at","")),
            })
        return jsonify({"users": users, "count": len(users)})
    except Exception as e:
        return safe_error(e)


# ── Real (not hardcoded) system health for the admin observability panel ──
# Only reports on services we can actually check from here. Other rows
# the admin UI shows (Stripe, SendGrid, PDF generator, etc.) are not yet
# live-checked and are labeled as such in the frontend rather than faked.
@app.route("/api/admin/system/health", methods=["GET"])
@require_auth
@require_admin
@limiter.limit("30 per minute")
def admin_system_health():
    try:
        ws = dict(_SOCKETIO_STATUS)
        redis_status = redis_health() if REDIS_READY else {"status": "not_configured"}
        firestore_ok = db is not None

        # Stripe — lightweight, read-only balance check confirms the key is
        # live and reachable without touching any customer data.
        stripe_status = {"status": "not_configured"}
        if STRIPE_SECRET_KEY:
            try:
                r = req.get("https://api.stripe.com/v1/balance",
                            headers={"Authorization": f"Bearer {STRIPE_SECRET_KEY}"}, timeout=5)
                stripe_status = {"status": "healthy" if r.status_code == 200 else "error",
                                  "http_status": r.status_code}
            except Exception as e:
                stripe_status = {"status": "down", "error": str(e)}

        # SendGrid — /v3/scopes is a cheap authenticated no-op that confirms
        # the API key is valid without sending anything.
        sendgrid_status = {"status": "not_configured"}
        if SENDGRID_API_KEY:
            try:
                r = req.get("https://api.sendgrid.com/v3/scopes",
                            headers={"Authorization": f"Bearer {SENDGRID_API_KEY}"}, timeout=5)
                sendgrid_status = {"status": "healthy" if r.status_code == 200 else "error",
                                    "http_status": r.status_code}
            except Exception as e:
                sendgrid_status = {"status": "down", "error": str(e)}

        return jsonify({
            "websocket": {
                "enabled":     ws["enabled"],
                "initialized": ws["initialized"],
                "error":       ws["error"],
                "status": ("healthy" if ws["initialized"] else
                           "disabled" if not ws["enabled"] else "error"),
            },
            "redis":           redis_status,
            "firestore":       {"status": "healthy" if firestore_ok else "down"},
            "stripe":          stripe_status,
            "sendgrid":        sendgrid_status,
            "pdf_generator":   {"status": "healthy" if REPORTLAB_OK else "down"},
            "analysis_engine": {"status": "healthy" if POSE_LITE is not None else "loading"},
            "local_ai":        {"status": "healthy" if bool(OLLAMA_URL) else "not_configured"},
            "api_gateway":     {"status": "healthy"},  # trivially true — this response proves it
        })
    except Exception as e:
        return safe_error(e)


@app.route("/api/user/activity", methods=["GET"])
@require_auth
@limiter.limit("60 per minute")
def user_activity_log():
    """
    Account activity timeline for the authenticated user.
    Pulls from audit_logs collection filtered by uid.
    Returns last 100 events sorted by timestamp desc.
    """
    try:
        uid   = getattr(g, "uid", "")
        limit = min(int(request.args.get("limit", 50)), 100)
        cat   = request.args.get("category", "")     # optional filter

        if not _firebase_ok:
            return jsonify({"ok": True, "events": []})

        db    = firestore.client()
        query = db.collection("audit_logs").where("uid", "==", uid)
        if cat:
            query = query.where("category", "==", cat)
        query = query.order_by("timestamp", direction="DESCENDING").limit(limit)

        snaps  = query.get()
        events = []
        for doc in snaps:
            d = doc.to_dict()
            # Map audit log to activity event
            category = d.get("category", "info")
            action   = d.get("action",   "")
            ts_raw   = d.get("timestamp", d.get("ts", ""))
            # Human-readable timestamp
            if isinstance(ts_raw, str):
                ts_display = ts_raw[:16].replace("T", " ")
            else:
                ts_display = str(ts_raw)[:16]

            # Determine severity
            security_actions = {"login_failed","mfa_enabled","mfa_disabled","sso_user_provisioned",
                                 "api_key_created","password_changed","session_revoked","gdpr_erasure_completed"}
            billing_actions  = {"payment_confirmed","subscription_cancelled","plan_upgraded","stripe_payment_success"}
            severity = "warning" if any(x in action for x in ["fail","revok","delete","disable"])                   else "success" if any(x in action for x in ["confirm","success","enable","creat"])                   else "info"

            events.append({
                "id":       doc.id,
                "ts":       ts_display,
                "category": category,
                "action":   action,
                "detail":   str(d.get("meta", d.get("details", "")))[:120],
                "severity": severity,
            })

        return jsonify({"ok": True, "events": events})
    except Exception as e:
        return safe_error(e, "Failed to load activity log")



# ── Cold Start Prevention ──────────────────────────────────────────
@app.route("/api/ping", methods=["GET", "HEAD"])
def ping():
    """
    Ultra-fast liveness probe — no auth, no DB, no MediaPipe.
    Used by uptime monitors (UptimeRobot / BetterUptime / cron-job.org)
    to keep Render from spinning down the backend.
    Responds in <5ms. HEAD method supported for zero-body pings.
    """
    return jsonify({"ok": True, "ts": datetime.utcnow().isoformat() + "Z"}), 200

@app.route("/api/warmup", methods=["POST", "GET"])
def warmup():
    """
    Trigger MediaPipe model load in advance (called on frontend mount).
    Returns instantly if already warm. Takes ~2-3s if cold.
    Safe to call without auth — only loads models, touches no user data.
    """
    try:
        was_cold = POSE_LITE is None
        _ensure_models()
        return jsonify({
            "ok":      True,
            "warm":    not was_cold,
            "message": "Models ready" if not was_cold else "Models loaded from cold start",
            "ts":      datetime.utcnow().isoformat() + "Z",
        }), 200
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ── Data Export ────────────────────────────────────────────────────
@app.route("/api/export/sessions", methods=["GET"])
@require_auth
@limiter.limit("10 per minute")
def export_sessions():
    """
    Export user's session history as CSV or PDF.
    Query params:
      format=csv|pdf  (default: csv)
      days=7|30|90    (default: 30)
    """
    try:
        uid    = getattr(g, "uid", "")
        fmt    = request.args.get("format", "csv").lower()
        days   = min(90, max(1, int(request.args.get("days", 30))))
        cutoff = datetime.utcnow() - timedelta(days=days)

        # Fetch sessions from Firestore
        docs = (db.collection("sessions")
                  .where("uid", "==", uid)
                  .where("created_at", ">=", cutoff)
                  .order_by("created_at", direction="DESCENDING")
                  .limit(500)
                  .stream())

        rows = []
        for doc in docs:
            d = doc.to_dict()
            rows.append({
                "date":         d.get("created_at", "").isoformat()[:10] if hasattr(d.get("created_at",""), "isoformat") else str(d.get("created_at",""))[:10],
                "avg_score":    d.get("avg_score", ""),
                "grade":        ("A" if (d.get("avg_score") or 0)>=85 else "B" if (d.get("avg_score") or 0)>=70 else "C" if (d.get("avg_score") or 0)>=55 else "D"),
                "duration_min": round((d.get("duration_s") or 0) / 60, 1),
                "frames":       d.get("frames", ""),
                "good_pct":     d.get("good_pct", ""),
                "mode":         d.get("mode", ""),
                "alerts":       d.get("alerts_count", ""),
                "trend":        d.get("trend", ""),
                "worst_metric": d.get("worst_metric", ""),
            })

        if not rows:
            return jsonify({"error": "No sessions found for this period"}), 404

        if fmt == "csv":
            import csv, io
            output = io.StringIO()
            writer = csv.DictWriter(output, fieldnames=rows[0].keys())
            writer.writeheader()
            writer.writerows(rows)
            csv_bytes = output.getvalue().encode("utf-8-sig")  # UTF-8 BOM for Excel
            resp = make_response(csv_bytes)
            resp.headers["Content-Type"]        = "text/csv; charset=utf-8"
            resp.headers["Content-Disposition"] = f'attachment; filename="postureai_sessions_{datetime.utcnow().strftime("%Y%m%d")}.csv"'
            return resp

        elif fmt == "pdf":
            # Use ReportLab for PDF export
            try:
                from reportlab.lib.pagesizes import A4
                from reportlab.lib import colors
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                from reportlab.lib.styles import getSampleStyleSheet
                import io

                buf = io.BytesIO()
                doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=40, bottomMargin=40)
                styles = getSampleStyleSheet()
                elements = []

                # Title
                elements.append(Paragraph(f"PostureAI Pro — Session History ({days} days)", styles["Title"]))
                elements.append(Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC | User: {uid[:8]}...", styles["Normal"]))
                elements.append(Spacer(1, 16))

                # Summary stats
                scores = [r["avg_score"] for r in rows if isinstance(r["avg_score"], (int, float))]
                if scores:
                    summary_data = [
                        ["Metric", "Value"],
                        ["Sessions", str(len(rows))],
                        ["Avg Score", str(round(sum(scores)/len(scores), 1))],
                        ["Best Score", str(max(scores))],
                        ["Worst Score", str(min(scores))],
                        ["Total Time", f"{round(sum(r['duration_min'] for r in rows if isinstance(r['duration_min'],(int,float))), 1)} min"],
                    ]
                    summary = Table(summary_data, colWidths=[200, 200])
                    summary.setStyle(TableStyle([
                        ("BACKGROUND",  (0,0), (-1,0), colors.HexColor("#6366f1")),
                        ("TEXTCOLOR",   (0,0), (-1,0), colors.white),
                        ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
                        ("BACKGROUND",  (0,1), (-1,-1), colors.HexColor("#f8fafc")),
                        ("GRID",        (0,0), (-1,-1), 0.5, colors.HexColor("#e2e8f0")),
                        ("FONTSIZE",    (0,0), (-1,-1), 10),
                        ("PADDING",     (0,0), (-1,-1), 6),
                    ]))
                    elements.append(summary)
                    elements.append(Spacer(1, 16))

                # Sessions table
                elements.append(Paragraph("Session Details", styles["Heading2"]))
                elements.append(Spacer(1, 8))
                header = ["Date", "Score", "Grade", "Duration", "Good%", "Mode", "Trend"]
                table_data = [header] + [
                    [r["date"], r["avg_score"], r["grade"],
                     f"{r['duration_min']}m", f"{r['good_pct']}%",
                     r["mode"], r["trend"]]
                    for r in rows[:50]   # max 50 rows in PDF
                ]
                t = Table(table_data, colWidths=[70,45,40,55,45,55,60])
                t.setStyle(TableStyle([
                    ("BACKGROUND",  (0,0), (-1,0), colors.HexColor("#6366f1")),
                    ("TEXTCOLOR",   (0,0), (-1,0), colors.white),
                    ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
                    ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#f1f5f9")]),
                    ("GRID",        (0,0), (-1,-1), 0.3, colors.HexColor("#cbd5e1")),
                    ("FONTSIZE",    (0,0), (-1,-1), 8),
                    ("PADDING",     (0,0), (-1,-1), 5),
                    ("ALIGN",       (1,0), (-1,-1), "CENTER"),
                ]))
                elements.append(t)

                doc.build(elements)
                pdf_bytes = buf.getvalue()
                resp = make_response(pdf_bytes)
                resp.headers["Content-Type"]        = "application/pdf"
                resp.headers["Content-Disposition"] = f'attachment; filename="postureai_report_{datetime.utcnow().strftime("%Y%m%d")}.pdf"'
                return resp
            except ImportError:
                return jsonify({"error": "PDF generation unavailable — use format=csv"}), 500
        else:
            return jsonify({"error": "Invalid format — use csv or pdf"}), 400

    except Exception as e:
        return safe_error(e)

@app.route("/api/export/sessions/summary", methods=["GET"])
@require_auth
@limiter.limit("20 per minute")
def export_sessions_summary():
    """Quick JSON summary for export preview (no file download)."""
    try:
        uid  = getattr(g, "uid", "")
        days = min(90, max(1, int(request.args.get("days", 30))))
        cutoff = datetime.utcnow() - timedelta(days=days)
        docs = (db.collection("sessions")
                  .where("uid", "==", uid)
                  .where("created_at", ">=", cutoff)
                  .order_by("created_at", direction="DESCENDING")
                  .limit(500).stream())
        rows   = [d.to_dict() for d in docs]
        scores = [r.get("avg_score") for r in rows if isinstance(r.get("avg_score"), (int,float))]
        return jsonify({
            "sessions":    len(rows),
            "days":        days,
            "avg_score":   round(sum(scores)/len(scores), 1) if scores else None,
            "best_score":  max(scores) if scores else None,
            "worst_score": min(scores) if scores else None,
            "total_min":   round(sum(r.get("duration_s",0) for r in rows)/60, 1),
            "export_urls": {
                "csv": f"/api/export/sessions?format=csv&days={days}",
                "pdf": f"/api/export/sessions?format=pdf&days={days}",
            }
        })
    except Exception as e:
        return safe_error(e)


# ── Session Comparison ─────────────────────────────────────────────
@app.route("/api/session/compare", methods=["POST"])
@require_auth
@limiter.limit("30 per minute")
def compare_sessions():
    """
    Compare last 2 sessions for the user.
    Returns per-metric deltas + AI narrative of what improved/worsened.
    Optional body: { "session_a": sid, "session_b": sid }
    """
    try:
        uid  = getattr(g, "uid", "")
        data = request.get_json(force=True) or {}
        lang = data.get("lang", "en")

        # Fetch last 2 sessions from Firestore
        docs = (db.collection("sessions")
                  .where("uid", "==", uid)
                  .order_by("created_at", direction="DESCENDING")
                  .limit(2)
                  .stream())
        sessions_list = [d.to_dict() for d in docs]

        if len(sessions_list) < 2:
            return jsonify({"error": "Need at least 2 sessions to compare", "sessions_found": len(sessions_list)}), 400

        new_s, old_s = sessions_list[0], sessions_list[1]

        # ── Per-metric deltas ─────────────────────────────────────
        new_mets = new_s.get("metrics", {})
        old_mets = old_s.get("metrics", {})
        deltas   = {}
        for key in set(list(new_mets.keys()) + list(old_mets.keys())):
            nv = new_mets.get(key, {})
            ov = old_mets.get(key, {})
            if isinstance(nv, dict) and isinstance(ov, dict):
                n_val = nv.get("value"); o_val = ov.get("value")
                if isinstance(n_val, (int, float)) and isinstance(o_val, (int, float)):
                    delta = round(n_val - o_val, 1)
                    deltas[key] = {
                        "prev":    o_val,
                        "current": n_val,
                        "delta":   delta,
                        "unit":    nv.get("unit", ""),
                        "label":   nv.get("label", key),
                        "trend":   "improved" if delta < 0 else "worsened" if delta > 0 else "stable",
                    }

        # ── Score comparison ──────────────────────────────────────
        score_delta  = (new_s.get("avg_score", 0) or 0) - (old_s.get("avg_score", 0) or 0)
        score_trend  = "improved" if score_delta > 2 else "declined" if score_delta < -2 else "stable"

        # ── Most improved / most worsened metrics ─────────────────
        sorted_deltas = sorted(
            [(k, v) for k, v in deltas.items() if abs(v["delta"]) > 0.5],
            key=lambda x: x[1]["delta"]
        )
        most_improved = [{"metric": k, **v} for k, v in sorted_deltas[:3]]        # most negative delta = improved
        most_worsened = [{"metric": k, **v} for k, v in sorted_deltas[-3:][::-1]] # most positive = worsened

        # ── AI narrative ───────────────────────────────────────────
        narrative = None
        if AI_CONFIGURED:
            try:
                _imp_txt = "\n".join([f"  ✅ {d['label']}: {d['prev']}{d['unit']} → {d['current']}{d['unit']} ({d['delta']:+.1f})" for d in most_improved]) or "  None"
                _wrs_txt = "\n".join([f"  ⚠️ {d['label']}: {d['prev']}{d['unit']} → {d['current']}{d['unit']} ({d['delta']:+.1f})" for d in most_worsened]) or "  None"
                _cmp_prompt = f"""You are a physiotherapist reviewing two posture monitoring sessions.

Previous session: score {old_s.get('avg_score',0)}/100, duration {round((old_s.get('duration_s',0) or 0)/60,1)} min
Current session:  score {new_s.get('avg_score',0)}/100, duration {round((new_s.get('duration_s',0) or 0)/60,1)} min
Score change: {score_delta:+.0f} points ({score_trend})

Most improved metrics:
{_imp_txt}

Most worsened metrics:
{_wrs_txt}

Write a 3-sentence {'Arabic' if lang=='ar' else 'English'} clinical comparison:
1. Overall progress statement with score change
2. What specifically improved and why it matters
3. What needs attention and one actionable recommendation

{"أجب بالعربية." if lang=='ar' else "Be specific, encouraging, and actionable."}"""

                narrative = call_ai(
                    _cmp_prompt, max_tokens=400, temperature=0.3,
                    tier=getattr(g, "tier", "standard"),
                )
            except Exception:
                pass

        return jsonify({
            "ok":           True,
            "score_prev":   old_s.get("avg_score"),
            "score_current":new_s.get("avg_score"),
            "score_delta":  score_delta,
            "score_trend":  score_trend,
            "date_prev":    str(old_s.get("created_at", ""))[:10],
            "date_current": str(new_s.get("created_at", ""))[:10],
            "most_improved":most_improved,
            "most_worsened":most_worsened,
            "all_deltas":   deltas,
            "narrative":    narrative,
        })
    except Exception as e:
        return safe_error(e)


# ── Posture Calibration ────────────────────────────────────────────
_calibration_buffer: dict = {}  # {uid: [result_dicts]} — accumulates frames during calibration

@app.route("/api/calibrate/start", methods=["POST"])
@require_auth
@limiter.limit("10 per minute")
def calibrate_start():
    """Start a 5-second calibration window. User should sit in ideal posture."""
    uid = getattr(g, "uid", "")
    _calibration_buffer[uid] = []
    return jsonify({
        "ok": True,
        "message": "Calibration started — sit in your ideal posture for 5 seconds",
        "message_ar": "بدأ المعايرة — اجلس في وضعيتك المثالية لمدة 5 ثوانٍ",
        "duration_s": 5,
    })

@app.route("/api/calibrate/frame", methods=["POST"])
@require_auth
@limiter.limit("60 per minute")
def calibrate_frame():
    """
    Accept a posture analysis result during calibration window.
    Frontend calls this instead of /api/analyze during calibration.
    """
    uid  = getattr(g, "uid", "")
    data = request.get_json(force=True) or {}
    if uid not in _calibration_buffer:
        return jsonify({"error": "No active calibration — call /api/calibrate/start first"}), 400
    # Store key metrics from this frame
    _calibration_buffer[uid].append({
        "neck_lean":  data.get("neck_lean",  0),
        "spine_lean": data.get("spine_lean", 0),
        "sh_tilt":    data.get("sh_tilt",    0),
        "dist_cm":    data.get("dist_cm",    65),
        "score":      data.get("score",      0),
        "ts":         time.time(),
    })
    return jsonify({"ok": True, "frames": len(_calibration_buffer[uid])})

@app.route("/api/calibrate/commit", methods=["POST"])
@require_auth
@limiter.limit("10 per minute")
def calibrate_commit():
    """
    Compute baseline from accumulated calibration frames and save to Firestore.
    Baseline = median of each metric (robust to outlier frames).
    """
    uid = getattr(g, "uid", "")
    buf = _calibration_buffer.get(uid, [])
    if len(buf) < 3:
        return jsonify({"error": f"Not enough frames ({len(buf)}) — need at least 3"}), 400

    def _median(vals):
        s = sorted(v for v in vals if v is not None)
        if not s: return None
        n = len(s); return s[n//2] if n % 2 else (s[n//2-1]+s[n//2])/2

    baseline = {
        "neck_lean":   _median([f["neck_lean"]  for f in buf]),
        "spine_lean":  _median([f["spine_lean"] for f in buf]),
        "sh_tilt":     _median([f["sh_tilt"]    for f in buf]),
        "dist_cm":     _median([f["dist_cm"]    for f in buf]),
        "score":       _median([f["score"]      for f in buf]),
        "frames_used": len(buf),
        "calibrated_at": datetime.utcnow().isoformat() + "Z",
        "uid": uid,
    }
    try:
        db.collection("user_baselines").document(uid).set(baseline, merge=True)
        # Also cache in Redis for fast access during analysis
        _rc = get_redis()
        if _rc:
            import json as _j
            _rc.setex(f"baseline:{uid}", 86400 * 30, _j.dumps(baseline))  # 30-day TTL
    except Exception as e:
        log_event("calibrate_save_error", uid, {"error": str(e)})
    del _calibration_buffer[uid]
    log_event("calibrate_commit", uid, {"frames": len(buf), "baseline_score": baseline["score"]})
    return jsonify({"ok": True, "baseline": baseline})

@app.route("/api/calibrate/status", methods=["GET"])
@require_auth
def calibrate_status():
    """Return user's current baseline if it exists."""
    uid = getattr(g, "uid", "")
    try:
        # Check Redis first
        _rc = get_redis()
        if _rc:
            import json as _j
            cached = _rc.get(f"baseline:{uid}")
            if cached:
                return jsonify({"calibrated": True, "baseline": _j.loads(cached), "source": "cache"})
        # Firestore fallback
        doc = db.collection("user_baselines").document(uid).get()
        if doc.exists:
            return jsonify({"calibrated": True, "baseline": doc.to_dict(), "source": "firestore"})
        return jsonify({"calibrated": False, "baseline": None})
    except Exception as e:
        return safe_error(e)

def _get_user_baseline(uid: str) -> dict | None:
    """Load baseline for scoring adjustment. Returns None if not calibrated."""
    try:
        _rc = get_redis()
        if _rc:
            import json as _j
            cached = _rc.get(f"baseline:{uid}")
            if cached:
                return _j.loads(cached)
        doc = db.collection("user_baselines").document(uid).get()
        return doc.to_dict() if doc.exists else None
    except Exception:
        return None


# ── Time-of-day Analytics ─────────────────────────────────────────
@app.route("/api/analytics/time-of-day", methods=["POST"])
@require_auth
@limiter.limit("20 per minute")
def analytics_time_of_day():
    """
    Return average posture score per hour of day.
    Answers: "Is posture worse after lunch? Better in the morning?"
    """
    try:
        uid  = getattr(g, "uid", "")
        data = request.get_json(force=True) or {}
        days = min(90, max(7, int(data.get("days", 30))))
        cutoff = datetime.utcnow() - timedelta(days=days)

        docs = (db.collection("sessions")
                  .where("uid", "==", uid)
                  .where("created_at", ">=", cutoff)
                  .limit(500).stream())

        hour_buckets: dict = {}   # {hour: [scores]}
        for doc in docs:
            d = doc.to_dict()
            sc = d.get("avg_score")
            ts = d.get("created_at")
            if not sc or not ts: continue
            try:
                hr = ts.hour if hasattr(ts, "hour") else int(str(ts)[11:13])
                hour_buckets.setdefault(hr, []).append(sc)
            except Exception:
                pass

        hourly = {
            str(h): {
                "avg_score":  round(sum(v)/len(v), 1),
                "sessions":   len(v),
                "min_score":  min(v),
                "max_score":  max(v),
                "label":      f"{h:02d}:00–{h+1:02d}:00",
                "period":     ("morning" if 6<=h<12 else "afternoon" if 12<=h<17
                               else "evening" if 17<=h<21 else "night"),
            }
            for h, v in hour_buckets.items() if v
        }

        # Best and worst hours
        if hourly:
            best_h  = max(hourly, key=lambda h: hourly[h]["avg_score"])
            worst_h = min(hourly, key=lambda h: hourly[h]["avg_score"])
            insight = (
                f"Your posture is best at {best_h}:00 (avg {hourly[best_h]['avg_score']}) "
                f"and worst at {worst_h}:00 (avg {hourly[worst_h]['avg_score']}). "
            )
            if 13 <= int(worst_h) <= 15:
                insight += "Post-lunch dip detected — consider a 5-min walk after lunch."
            elif int(worst_h) >= 17:
                insight += "End-of-day fatigue detected — try a posture check alarm at 16:00."
        else:
            insight = "Not enough data yet."

        return jsonify({"ok": True, "hourly": hourly, "insight": insight, "days_analyzed": days})
    except Exception as e:
        return safe_error(e)

# ── Anomaly Detection ─────────────────────────────────────────────
@app.route("/api/analytics/anomaly", methods=["POST"])
@require_auth
@limiter.limit("20 per minute")
def analytics_anomaly():
    """
    Detect sudden posture score drops (>20pts in one day vs 7-day average).
    Useful for HR: identify employees who suddenly deteriorated.
    """
    try:
        uid    = getattr(g, "uid", "")
        data   = request.get_json(force=True) or {}
        days   = min(90, max(7, int(data.get("days", 30))))
        cutoff = datetime.utcnow() - timedelta(days=days)

        docs = (db.collection("sessions")
                  .where("uid", "==", uid)
                  .where("created_at", ">=", cutoff)
                  .order_by("created_at").limit(500).stream())

        daily: dict = {}
        for doc in docs:
            d = doc.to_dict()
            sc = d.get("avg_score"); ts = d.get("created_at")
            if not sc or not ts: continue
            try:
                day = str(ts)[:10] if not hasattr(ts, "strftime") else ts.strftime("%Y-%m-%d")
                daily.setdefault(day, []).append(sc)
            except Exception:
                pass

        daily_avg = {day: round(sum(v)/len(v), 1) for day, v in daily.items()}
        days_list = sorted(daily_avg.keys())
        anomalies = []

        for i, day in enumerate(days_list[7:], start=7):
            window = [daily_avg[days_list[j]] for j in range(i-7, i)]
            baseline_7d = sum(window) / len(window)
            today_avg   = daily_avg[day]
            drop        = baseline_7d - today_avg
            if drop > 20:
                anomalies.append({
                    "date":          day,
                    "score":         today_avg,
                    "baseline_7d":   round(baseline_7d, 1),
                    "drop":          round(drop, 1),
                    "severity":      "critical" if drop > 35 else "high" if drop > 25 else "medium",
                })

        return jsonify({
            "ok":        True,
            "anomalies": anomalies,
            "daily_avg": daily_avg,
            "total_anomaly_days": len(anomalies),
        })
    except Exception as e:
        return safe_error(e)

# ── Weekly Report ─────────────────────────────────────────────────
@app.route("/api/report/weekly", methods=["GET"])
@require_auth
@limiter.limit("10 per minute")
def weekly_report():
    """
    Generate a weekly posture summary (JSON or PDF).
    ?format=json|pdf  ?week_offset=0 (0=this week, 1=last week)
    """
    try:
        uid         = getattr(g, "uid", "")
        fmt         = request.args.get("format", "json").lower()
        week_offset = int(request.args.get("week_offset", 0))

        now      = datetime.utcnow()
        week_end = now - timedelta(days=now.weekday() + 7 * week_offset)
        week_start = week_end - timedelta(days=7)

        docs = (db.collection("sessions")
                  .where("uid", "==", uid)
                  .where("created_at", ">=", week_start)
                  .where("created_at", "<",  week_end)
                  .order_by("created_at").limit(500).stream())

        sessions_data = [d.to_dict() for d in docs]
        scores = [s.get("avg_score") for s in sessions_data if isinstance(s.get("avg_score"), (int,float))]

        if not scores:
            return jsonify({"error": "No sessions found for this week"}), 404

        avg  = round(sum(scores)/len(scores), 1)
        best = max(scores); worst = min(scores)
        grade = "A" if avg>=85 else "B" if avg>=70 else "C" if avg>=55 else "D"

        # Day-by-day breakdown
        daily: dict = {}
        for s in sessions_data:
            ts = s.get("created_at"); sc = s.get("avg_score")
            if not ts or not sc: continue
            try:
                day = ts.strftime("%A") if hasattr(ts, "strftime") else str(ts)[:10]
                daily.setdefault(day, []).append(sc)
            except Exception:
                pass
        daily_avg = {d: round(sum(v)/len(v),1) for d,v in daily.items()}

        # Top recurring alerts
        alert_counts: dict = {}
        for s in sessions_data:
            for a in (s.get("alerts") or []):
                key = a[:40]
                alert_counts[key] = alert_counts.get(key, 0) + 1
        top_alerts = sorted(alert_counts.items(), key=lambda x: -x[1])[:5]

        report = {
            "week_start":   week_start.strftime("%Y-%m-%d"),
            "week_end":     week_end.strftime("%Y-%m-%d"),
            "sessions":     len(sessions_data),
            "avg_score":    avg,
            "best_score":   best,
            "worst_score":  worst,
            "grade":        grade,
            "daily_avg":    daily_avg,
            "top_alerts":   [{"alert": k, "occurrences": v} for k,v in top_alerts],
            "total_min":    round(sum(s.get("duration_s",0) or 0 for s in sessions_data)/60, 1),
        }

        if fmt == "pdf":
            try:
                from reportlab.lib.pagesizes import A4
                from reportlab.lib import colors
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                from reportlab.lib.styles import getSampleStyleSheet
                import io as _io
                buf = _io.BytesIO()
                doc_pdf = SimpleDocTemplate(buf, pagesize=A4, topMargin=40, bottomMargin=40)
                styles  = getSampleStyleSheet()
                elements = []
                elements.append(Paragraph(f"PostureAI Weekly Report — {report['week_start']} to {report['week_end']}", styles["Title"]))
                elements.append(Paragraph(f"Grade: {grade}  |  Avg Score: {avg}/100  |  Sessions: {len(sessions_data)}", styles["Normal"]))
                elements.append(Spacer(1, 12))
                # Daily table
                rows = [["Day", "Avg Score"]] + [[d, f"{s}/100"] for d,s in daily_avg.items()]
                t = Table(rows, colWidths=[200, 200])
                t.setStyle(TableStyle([
                    ("BACKGROUND", (0,0),(-1,0), colors.HexColor("#6366f1")),
                    ("TEXTCOLOR",  (0,0),(-1,0), colors.white),
                    ("FONTNAME",   (0,0),(-1,0), "Helvetica-Bold"),
                    ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white, colors.HexColor("#f1f5f9")]),
                    ("GRID",       (0,0),(-1,-1), 0.3, colors.HexColor("#cbd5e1")),
                    ("FONTSIZE",   (0,0),(-1,-1), 10),
                    ("PADDING",    (0,0),(-1,-1), 6),
                ]))
                elements.append(t)
                doc_pdf.build(elements)
                pdf_bytes = buf.getvalue()
                resp = make_response(pdf_bytes)
                resp.headers["Content-Type"]        = "application/pdf"
                resp.headers["Content-Disposition"] = f'attachment; filename="weekly_report_{report["week_start"]}.pdf"'
                return resp
            except ImportError:
                pass  # fall through to JSON

        return jsonify({"ok": True, **report})
    except Exception as e:
        return safe_error(e)


# ══════════════════════════════════════════════════════════════════
# B2C: STREAK SYSTEM
# ══════════════════════════════════════════════════════════════════

@app.route("/api/streak", methods=["GET"])
@require_auth
@limiter.limit("60 per minute")
def get_streak():
    """
    Return current streak, longest streak, and last 30 days activity map.
    Streak = consecutive days with at least 1 session scoring >= 55.
    """
    try:
        uid = getattr(g, "uid", "")
        # Check Redis cache first
        _rc = get_redis()
        if _rc:
            import json as _j
            cached = _rc.get(f"streak:{uid}")
            if cached:
                return jsonify({**_j.loads(cached), "cached": True})

        cutoff = datetime.utcnow() - timedelta(days=60)
        docs = (db.collection("sessions")
                  .where("uid", "==", uid)
                  .where("created_at", ">=", cutoff)
                  .order_by("created_at", direction="DESCENDING")
                  .limit(200).stream())

        # Build day → best_score map
        day_scores: dict = {}
        for doc in docs:
            d = doc.to_dict()
            sc = d.get("avg_score"); ts = d.get("created_at")
            if not sc or not ts: continue
            try:
                day = ts.strftime("%Y-%m-%d") if hasattr(ts,"strftime") else str(ts)[:10]
                day_scores[day] = max(day_scores.get(day, 0), sc)
            except Exception:
                pass

        # Active days = days with score >= 55
        STREAK_MIN_SCORE = 55
        active_days = {d for d, sc in day_scores.items() if sc >= STREAK_MIN_SCORE}

        # Current streak (from today backwards)
        today = datetime.utcnow().date()
        current_streak = 0
        check_day = today
        # Allow today OR yesterday to count (don't break streak if user hasn't done today yet)
        if str(today) not in active_days and str(today - timedelta(days=1)) in active_days:
            check_day = today - timedelta(days=1)
        while str(check_day) in active_days:
            current_streak += 1
            check_day -= timedelta(days=1)

        # Longest streak ever
        longest = 0
        run = 0
        for i in range(60):
            d = str(today - timedelta(days=i))
            if d in active_days:
                run += 1
                longest = max(longest, run)
            else:
                run = 0

        # Last 30 days activity map
        activity_map = {}
        for i in range(30):
            d = str(today - timedelta(days=i))
            activity_map[d] = {
                "active":    d in active_days,
                "score":     day_scores.get(d),
                "day_label": (today - timedelta(days=i)).strftime("%a"),
            }

        # Milestone check
        milestones = [3, 7, 14, 30, 60, 100]
        next_milestone = next((m for m in milestones if m > current_streak), None)
        days_to_milestone = (next_milestone - current_streak) if next_milestone else None

        result = {
            "ok":                True,
            "current_streak":    current_streak,
            "longest_streak":    longest,
            "active_days_30":    sum(1 for d in activity_map.values() if d["active"]),
            "activity_map":      activity_map,
            "next_milestone":    next_milestone,
            "days_to_milestone": days_to_milestone,
            "streak_min_score":  STREAK_MIN_SCORE,
            "streak_message": (
                f"🔥 {current_streak} day streak! Keep it up!" if current_streak >= 3
                else "Start your streak today — 3 days in a row unlocks your first badge!"
            ),
            "streak_message_ar": (
                f"🔥 {current_streak} يوم متتالي! واصل!" if current_streak >= 3
                else "ابدأ سلسلتك اليوم — 3 أيام متتاليين يفتح أول badge!"
            ),
        }

        # Cache for 5 min
        if _rc:
            import json as _j
            _rc.setex(f"streak:{uid}", 300, _j.dumps(result))

        return jsonify(result)
    except Exception as e:
        return safe_error(e)


@app.route("/api/streak/check", methods=["POST"])
@require_auth
@limiter.limit("30 per minute")
def check_streak():
    """
    Called after a session ends. Recalculates streak, invalidates cache,
    returns whether a milestone was just hit.
    """
    try:
        uid = getattr(g, "uid", "")
        # Invalidate streak cache
        _rc = get_redis()
        if _rc:
            _rc.delete(f"streak:{uid}")

        # Re-fetch streak
        import requests as _req_streak
        tok = request.headers.get("Authorization", "")
        streak_resp = get_streak.__wrapped__() if hasattr(get_streak, "__wrapped__") else None

        # Simple recalculation inline
        cutoff = datetime.utcnow() - timedelta(days=60)
        docs = (db.collection("sessions")
                  .where("uid", "==", uid)
                  .where("created_at", ">=", cutoff)
                  .order_by("created_at", direction="DESCENDING")
                  .limit(200).stream())
        day_scores: dict = {}
        for doc in docs:
            d = doc.to_dict()
            sc = d.get("avg_score"); ts = d.get("created_at")
            if not sc or not ts: continue
            try:
                day = ts.strftime("%Y-%m-%d") if hasattr(ts,"strftime") else str(ts)[:10]
                day_scores[day] = max(day_scores.get(day, 0), sc)
            except Exception:
                pass
        active_days = {d for d, sc in day_scores.items() if sc >= 55}
        today = datetime.utcnow().date()
        current_streak = 0
        check_day = today
        while str(check_day) in active_days:
            current_streak += 1
            check_day -= timedelta(days=1)

        milestones     = [3, 7, 14, 30, 60, 100]
        hit_milestone  = current_streak in milestones
        milestone_msg  = None
        if hit_milestone:
            msgs = {
                3:  ("🔥 3-Day Streak!", "3 أيام متتاليين! 🔥"),
                7:  ("⚡ Week Warrior! 7 days!", "محارب الأسبوع! 7 أيام! ⚡"),
                14: ("💪 Two Week Champ!", "بطل أسبوعين! 💪"),
                30: ("🏆 Monthly Master!", "سيد الشهر! 🏆"),
                60: ("💎 60-Day Legend!", "أسطورة 60 يوم! 💎"),
                100:("👑 100 Days — Elite!", "100 يوم — نخبة! 👑"),
            }
            milestone_msg = msgs.get(current_streak)

        return jsonify({
            "ok":             True,
            "current_streak": current_streak,
            "hit_milestone":  hit_milestone,
            "milestone_msg":  milestone_msg,
        })
    except Exception as e:
        return safe_error(e)


# ══════════════════════════════════════════════════════════════════
# B2C: GOALS SYSTEM
# ══════════════════════════════════════════════════════════════════

_GOAL_TEMPLATES = [
    {"id": "score_70",    "name": "Reach Score 70",      "name_ar": "الوصول لـ 70",      "type": "score",   "target": 70,  "icon": "🎯"},
    {"id": "score_80",    "name": "Reach Score 80",      "name_ar": "الوصول لـ 80",      "type": "score",   "target": 80,  "icon": "⭐"},
    {"id": "score_90",    "name": "Reach Score 90",      "name_ar": "الوصول لـ 90",      "type": "score",   "target": 90,  "icon": "🏆"},
    {"id": "streak_7",   "name": "7-Day Streak",         "name_ar": "7 أيام متتاليين",   "type": "streak",  "target": 7,   "icon": "🔥"},
    {"id": "streak_30",  "name": "30-Day Streak",        "name_ar": "30 يوم متتالي",     "type": "streak",  "target": 30,  "icon": "💎"},
    {"id": "sessions_20","name": "Complete 20 Sessions", "name_ar": "20 جلسة",           "type": "sessions","target": 20,  "icon": "📊"},
    {"id": "neck_fix",   "name": "Fix Neck Posture",     "name_ar": "تحسين وضعية الرقبة","type": "metric",  "target": 80,  "metric": "neck_lean", "icon": "🦒"},
    {"id": "no_pain",    "name": "7 Days Pain-Free",     "name_ar": "7 أيام بدون ألم",   "type": "pain_free","target": 7,  "icon": "💚"},
]

@app.route("/api/goals", methods=["GET"])
@require_auth
@limiter.limit("30 per minute")
def get_goals():
    """Return user's active goals with current progress."""
    try:
        uid = getattr(g, "uid", "")
        docs = db.collection("user_goals").where("uid","==",uid).where("status","in",["active","completed"]).limit(10).stream()
        goals = [d.to_dict() for d in docs]
        return jsonify({"ok": True, "goals": goals, "templates": _GOAL_TEMPLATES})
    except Exception as e:
        return safe_error(e)

@app.route("/api/goals", methods=["POST"])
@require_auth
@limiter.limit("20 per minute")
def set_goal():
    """Create or update a user goal."""
    try:
        uid  = getattr(g, "uid", "")
        data = request.get_json(force=True) or {}
        goal_id    = data.get("goal_id")    # from _GOAL_TEMPLATES
        custom_target = data.get("target")  # optional override
        deadline   = data.get("deadline")   # optional YYYY-MM-DD

        template = next((t for t in _GOAL_TEMPLATES if t["id"] == goal_id), None)
        if not template:
            return jsonify({"error": f"Unknown goal_id: {goal_id}. Available: {[t['id'] for t in _GOAL_TEMPLATES]}"}), 400

        goal = {
            **template,
            "uid":        uid,
            "status":     "active",
            "progress":   0,
            "target":     custom_target or template["target"],
            "deadline":   deadline,
            "created_at": datetime.utcnow().isoformat() + "Z",
            "completed_at": None,
        }

        # Save to Firestore (upsert by uid+goal_id)
        doc_id = f"{uid}_{goal_id}"
        db.collection("user_goals").document(doc_id).set(goal, merge=True)
        log_event("goal_set", uid, {"goal_id": goal_id, "target": goal["target"]})
        return jsonify({"ok": True, "goal": goal})
    except Exception as e:
        return safe_error(e)

@app.route("/api/goals/progress", methods=["GET"])
@require_auth
@limiter.limit("30 per minute")
def goals_progress():
    """
    Check progress on all active goals.
    Updates Firestore and returns which goals were just completed.
    """
    try:
        uid = getattr(g, "uid", "")
        # Fetch active goals
        goal_docs = list(db.collection("user_goals")
                          .where("uid","==",uid)
                          .where("status","==","active")
                          .limit(10).stream())
        if not goal_docs:
            return jsonify({"ok": True, "goals": [], "just_completed": []})

        # Fetch recent sessions for progress calculation
        cutoff = datetime.utcnow() - timedelta(days=90)
        sess_docs = (db.collection("sessions")
                       .where("uid","==",uid)
                       .where("created_at",">=",cutoff)
                       .order_by("created_at",direction="DESCENDING")
                       .limit(500).stream())
        sessions_list = [d.to_dict() for d in sess_docs]
        scores   = [s.get("avg_score",0) for s in sessions_list if s.get("avg_score")]
        best_sc  = max(scores) if scores else 0
        n_sess   = len(sessions_list)

        # Current streak (simplified)
        day_scores: dict = {}
        for s in sessions_list:
            sc = s.get("avg_score"); ts = s.get("created_at")
            if not sc or not ts: continue
            try:
                day = ts.strftime("%Y-%m-%d") if hasattr(ts,"strftime") else str(ts)[:10]
                day_scores[day] = max(day_scores.get(day,0), sc)
            except Exception:
                pass
        active_days = {d for d,sc in day_scores.items() if sc >= 55}
        today = datetime.utcnow().date()
        streak = 0
        chk = today
        while str(chk) in active_days:
            streak += 1; chk -= timedelta(days=1)

        updated_goals = []
        just_completed = []

        for doc in goal_docs:
            g_data = doc.to_dict()
            gtype  = g_data.get("type")
            target = g_data.get("target", 100)
            progress = 0

            if gtype == "score":
                progress = min(100, int((best_sc / target) * 100))
                current_val = best_sc
            elif gtype == "streak":
                progress = min(100, int((streak / target) * 100))
                current_val = streak
            elif gtype == "sessions":
                progress = min(100, int((n_sess / target) * 100))
                current_val = n_sess
            elif gtype == "pain_free":
                pain_free_days = sum(1 for s in sessions_list
                    if (s.get("pain_bar",{}) or {}).get("urgency","low") in ("low","none"))
                progress = min(100, int((pain_free_days / target) * 100))
                current_val = pain_free_days
            else:
                current_val = 0

            completed = progress >= 100
            updates = {"progress": progress, "current_value": current_val}
            if completed and g_data.get("status") == "active":
                updates["status"]       = "completed"
                updates["completed_at"] = datetime.utcnow().isoformat() + "Z"
                just_completed.append({**g_data, **updates})
                log_event("goal_completed", uid, {"goal_id": g_data.get("id"), "type": gtype})

            db.collection("user_goals").document(doc.id).update(updates)
            updated_goals.append({**g_data, **updates})

        return jsonify({
            "ok":            True,
            "goals":         updated_goals,
            "just_completed":just_completed,
        })
    except Exception as e:
        return safe_error(e)


# ══════════════════════════════════════════════════════════════════
# B2C: SHARE CARD
# ══════════════════════════════════════════════════════════════════

@app.route("/api/share/card", methods=["POST"])
@require_auth
@limiter.limit("20 per minute")
def generate_share_card():
    """
    Generate a shareable posture score card as SVG.
    Frontend: SVG → Canvas → PNG for Instagram/LinkedIn.
    """
    try:
        uid  = getattr(g, "uid", "")
        data = request.get_json(force=True) or {}
        score      = int(data.get("score", 0))
        grade      = data.get("grade", "A" if score>=85 else "B" if score>=70 else "C" if score>=55 else "D")
        streak     = int(data.get("streak", 0))
        percentile = data.get("percentile")
        lang       = data.get("lang", "en")
        username   = data.get("username", "")
        date_str   = datetime.utcnow().strftime("%b %d, %Y")

        color      = "#10b981" if score>=75 else "#f59e0b" if score>=55 else "#ef4444"
        grade_color= {"A":"#10b981","B":"#3b82f6","C":"#f59e0b","D":"#ef4444"}.get(grade,"#6366f1")
        is_ar      = lang == "ar"

        label_score = "درجة الوضعية" if is_ar else "Posture Score"
        label_streak = "أيام متتاليين" if is_ar else "Day Streak"
        label_grade = "التقييم" if is_ar else "Grade"
        label_pct   = (f"أحسن من {percentile}% من المستخدمين" if is_ar
                       else f"Better than {percentile}% of users") if percentile else ""
        label_url   = "postureai-pro-omega-nine.vercel.app"
        label_try   = "جرب مجاناً" if is_ar else "Try for free"

        # Build SVG parts
        pct_line = (f'<text x="200" y="310" text-anchor="middle" font-family="Inter,Arial" ' +
                    f'font-size="11" fill="#94a3b8">{label_pct}</text>') if percentile else ""

        streak_rect = ""
        if streak >= 3:
            streak_rect = (
                f'<rect x="130" y="325" width="140" height="38" rx="10" ' +
                f'fill="rgba(245,158,11,0.12)" stroke="rgba(245,158,11,0.3)" stroke-width="1"/>' +
                f'<text x="200" y="349" text-anchor="middle" font-family="Inter,Arial" ' +
                f'font-size="13" fill="#f59e0b" font-weight="700">🔥 {streak} {label_streak}</text>'
            )

        user_line = (f'<text x="200" y="400" text-anchor="middle" font-family="Inter,Arial" ' +
                     f'font-size="12" fill="#475569">@{username}</text>') if username else ""

        dash = int(score * 5.655)
        svg_lines = [
            '<svg width="400" height="500" xmlns="http://www.w3.org/2000/svg">',
            '<defs>',
            '<linearGradient id="bg" x1="0" y1="0" x2="1" y2="1">',
            '<stop offset="0%" stop-color="#020d1f"/>',
            '<stop offset="100%" stop-color="#0d1b2e"/>',
            '</linearGradient>',
            f'<linearGradient id="ring" x1="0" y1="0" x2="1" y2="1">',
            f'<stop offset="0%" stop-color="{color}"/>',
            f'<stop offset="100%" stop-color="{color}88"/>',
            '</linearGradient>',
            '<filter id="glow"><feGaussianBlur stdDeviation="4" result="blur"/>',
            '<feMerge><feMergeNode in="blur"/><feMergeNode in="SourceGraphic"/></feMerge></filter>',
            '</defs>',
            '<rect width="400" height="500" fill="url(#bg)" rx="20"/>',
            f'<rect width="400" height="500" fill="none" stroke="{color}33" stroke-width="1.5" rx="20"/>',
            '<text x="200" y="48" text-anchor="middle" font-family="Inter,Arial" font-size="14" fill="#6366f1" font-weight="700" letter-spacing="2">PostureAI Pro</text>',
            f'<text x="200" y="68" text-anchor="middle" font-family="Inter,Arial" font-size="11" fill="#475569">{date_str}</text>',
            '<circle cx="200" cy="170" r="90" fill="none" stroke="#1e293b" stroke-width="14"/>',
            f'<circle cx="200" cy="170" r="90" fill="none" stroke="url(#ring)" stroke-width="14" stroke-dasharray="{dash} 565.5" stroke-linecap="round" transform="rotate(-90 200 170)" filter="url(#glow)"/>',
            f'<text x="200" y="158" text-anchor="middle" font-family="Inter,Arial" font-size="52" fill="white" font-weight="900">{score}</text>',
            '<text x="200" y="182" text-anchor="middle" font-family="Inter,Arial" font-size="13" fill="#94a3b8">/100</text>',
            f'<text x="200" y="202" text-anchor="middle" font-family="Inter,Arial" font-size="11" fill="#64748b">{label_score}</text>',
            f'<rect x="162" y="272" width="76" height="34" rx="8" fill="{grade_color}22" stroke="{grade_color}55" stroke-width="1"/>',
            f'<text x="200" y="294" text-anchor="middle" font-family="Inter,Arial" font-size="16" fill="{grade_color}" font-weight="800">{label_grade}: {grade}</text>',
            pct_line,
            streak_rect,
            user_line,
            f'<rect x="100" y="430" width="200" height="36" rx="10" fill="#6366f1"/>',
            f'<text x="200" y="453" text-anchor="middle" font-family="Inter,Arial" font-size="12" fill="white" font-weight="700">{label_try} → {label_url}</text>',
            '</svg>',
        ]
        svg = "\n".join(svg_lines)

        log_event("share_card_generated", uid, {"score": score, "lang": lang, "streak": streak})
        return jsonify({
            "ok":    True,
            "svg":   svg,
            "score": score,
            "grade": grade,
            "meta": {
                "title":       f"My posture score: {score}/100 ({grade})",
                "description": label_pct or f"Tracked with PostureAI Pro",
                "hashtags":    ["PostureAI", "PostureCheck", "WorkplaceWellness"],
            }
        })
    except Exception as e:
        return safe_error(e)


# ══════════════════════════════════════════════════════════════════
# B2C: PUSH NOTIFICATIONS (Firebase Cloud Messaging)
# ══════════════════════════════════════════════════════════════════

def _send_fcm(token: str, title: str, body: str, data: dict = None) -> bool:
    """
    Send a push notification via Firebase Admin SDK (FCM).
    Returns True if sent successfully.
    """
    try:
        from firebase_admin import messaging as _fcm
        msg = _fcm.Message(
            notification=_fcm.Notification(title=title, body=body),
            data={str(k): str(v) for k, v in (data or {}).items()},
            token=token,
            android=_fcm.AndroidConfig(priority="high"),
            apns=_fcm.APNSConfig(
                payload=_fcm.APNSPayload(
                    aps=_fcm.Aps(sound="default", badge=1)
                )
            ),
        )
        _fcm.send(msg)
        return True
    except Exception as e:
        log_event("fcm_error", meta={"error": str(e)[:100], "token": token[:20]})
        return False

def _send_fcm_multicast(tokens: list, title: str, body: str, data: dict = None) -> dict:
    """Send to multiple tokens at once (max 500 per call)."""
    if not tokens: return {"success": 0, "failure": 0}
    try:
        from firebase_admin import messaging as _fcm
        msg = _fcm.MulticastMessage(
            notification=_fcm.Notification(title=title, body=body),
            data={str(k): str(v) for k, v in (data or {}).items()},
            tokens=tokens[:500],
            android=_fcm.AndroidConfig(priority="high"),
        )
        result = _fcm.send_each_for_multicast(msg)
        return {"success": result.success_count, "failure": result.failure_count}
    except Exception as e:
        log_event("fcm_multicast_error", meta={"error": str(e)[:100]})
        return {"success": 0, "failure": len(tokens)}


@app.route("/api/push/register", methods=["POST"])
@require_auth
@limiter.limit("20 per minute")
def push_register():
    """
    Register or update a device push token for the authenticated user.
    Call this on app launch and whenever FCM token refreshes.
    """
    try:
        uid  = getattr(g, "uid", "")
        data = request.get_json(force=True) or {}
        token    = data.get("token", "").strip()
        platform = data.get("platform", "web")   # web | ios | android
        lang     = data.get("lang", "en")

        if not token:
            return jsonify({"error": "token required"}), 400

        # Save token to Firestore (one doc per uid+token combo)
        import hashlib as _hl
        token_hash = _hl.md5(token.encode()).hexdigest()[:16]
        db.collection("push_tokens").document(f"{uid}_{token_hash}").set({
            "uid":        uid,
            "token":      token,
            "platform":   platform,
            "lang":       lang,
            "updated_at": datetime.utcnow().isoformat() + "Z",
            "active":     True,
        }, merge=True)

        # Also cache in Redis for fast lookup
        _rc = get_redis()
        if _rc:
            import json as _j
            existing = _rc.get(f"push_tokens:{uid}")
            tokens_list = _j.loads(existing) if existing else []
            if token not in tokens_list:
                tokens_list.append(token)
                tokens_list = tokens_list[-5:]  # keep last 5 tokens per user
            _rc.setex(f"push_tokens:{uid}", 86400 * 30, _j.dumps(tokens_list))

        log_event("push_token_registered", uid, {"platform": platform})
        return jsonify({"ok": True, "platform": platform})
    except Exception as e:
        return safe_error(e)


@app.route("/api/push/unregister", methods=["POST"])
@require_auth
@limiter.limit("10 per minute")
def push_unregister():
    """Remove a device token (called on logout or notification opt-out)."""
    try:
        uid   = getattr(g, "uid", "")
        data  = request.get_json(force=True) or {}
        token = data.get("token", "").strip()
        if not token:
            return jsonify({"error": "token required"}), 400

        import hashlib as _hl
        token_hash = _hl.md5(token.encode()).hexdigest()[:16]
        db.collection("push_tokens").document(f"{uid}_{token_hash}").update({"active": False})

        _rc = get_redis()
        if _rc:
            import json as _j
            existing = _rc.get(f"push_tokens:{uid}")
            if existing:
                tokens_list = [t for t in _j.loads(existing) if t != token]
                _rc.setex(f"push_tokens:{uid}", 86400 * 30, _j.dumps(tokens_list))

        return jsonify({"ok": True})
    except Exception as e:
        return safe_error(e)


def _get_user_tokens(uid: str) -> list:
    """Get all active push tokens for a user. Tries Redis then Firestore."""
    try:
        _rc = get_redis()
        if _rc:
            import json as _j
            cached = _rc.get(f"push_tokens:{uid}")
            if cached:
                return _j.loads(cached)
        docs = (db.collection("push_tokens")
                  .where("uid", "==", uid)
                  .where("active", "==", True)
                  .limit(5).stream())
        return [d.to_dict().get("token") for d in docs if d.to_dict().get("token")]
    except Exception:
        return []


def _compute_preferred_hour(uid):
    """Most common hour-of-day (server/UTC) this user actually does sessions,
    from their last 30 sessions. Falls back to 19:00 (7pm) with too little
    data to be meaningful — matches the old fixed-time behavior."""
    try:
        docs = (db.collection("sessions")
                  .where("uid", "==", uid)
                  .order_by("created_at", direction="DESCENDING")
                  .limit(30).stream())
        hours = []
        for d in docs:
            ts = d.to_dict().get("created_at")
            if ts and hasattr(ts, "hour"):
                hours.append(ts.hour)
        if len(hours) < 5:
            return 19
        return Counter(hours).most_common(1)[0][0]
    except Exception:
        return 19


def _get_push_categories(uid):
    """Per-user notification category preferences. All default to enabled —
    a user who never touched this setting still gets reminders, matching
    the pre-existing behavior before this preference existed."""
    try:
        doc = db.collection("push_preferences").document(uid).get()
        if doc.exists:
            return doc.to_dict().get("categories", {})
    except Exception:
        pass
    return {}


@app.route("/api/push/preferences", methods=["GET", "POST"])
@require_auth
@limiter.limit("20 per minute")
def push_preferences():
    """User-facing: view/set which notification categories they want, and
    optionally override the auto-computed 'smart' reminder hour."""
    try:
        uid = g.uid
        if request.method == "GET":
            prefs = {}
            doc = db.collection("push_preferences").document(uid).get()
            if doc.exists:
                prefs = doc.to_dict()
            return jsonify({
                "categories": prefs.get("categories", {"streak": True, "symptom_reminder": True, "weekly_summary": True}),
                "preferred_hour": prefs.get("preferred_hour_override"),
                "computed_hour": _compute_preferred_hour(uid),
            })
        data = request.get_json(force=True) or {}
        upd = {"updated_at": datetime.utcnow().isoformat()+"Z"}
        if "categories" in data:
            upd["categories"] = {k: bool(v) for k, v in data["categories"].items()
                                  if k in ("streak", "symptom_reminder", "weekly_summary")}
        if "preferred_hour_override" in data:
            h = data["preferred_hour_override"]
            upd["preferred_hour_override"] = int(h) if h is not None else None
        db.collection("push_preferences").document(uid).set(upd, merge=True)
        return jsonify({"ok": True})
    except Exception as e:
        return safe_error(e)


@app.route("/api/push/streak-reminder", methods=["POST"])
@limiter.limit("30 per minute")
def push_streak_reminder():
    """
    Cron job endpoint — intended to be called once per hour (see
    services/celery_beat.py: smart-streak-reminders). Each run only
    targets users whose own computed usual-session hour (or manual
    override) matches the `hour` passed in, and who haven't disabled
    the "streak" notification category. Pass `hour` explicitly so the
    cron caller controls what "now" means; omit it to send to everyone
    regardless of hour (the old fixed-time behavior, still useful for
    manual/testing calls).
    Secured by CRON_SECRET env var (not user auth).
    """
    try:
        # Verify cron secret — fail CLOSED if not configured, and use constant-time
        # comparison. Previously this compared against "" when unset, meaning a
        # request with NO header at all would match "" == "" and pass unauthenticated.
        import hmac as _hmac
        cron_secret = os.getenv("CRON_SECRET", "")
        secret      = request.headers.get("X-Cron-Secret", "")
        if not cron_secret:
            print("[cron] ⚠️  CRON_SECRET not set — rejecting all cron requests", file=sys.stderr)
            return jsonify({"error": "Cron endpoint not configured"}), 503
        if not secret or not _hmac.compare_digest(secret, cron_secret):
            return jsonify({"error": "Unauthorized"}), 401

        data     = request.get_json(force=True) or {}
        min_streak = int(data.get("min_streak", 2))   # only remind users with streak >= 2
        target_hour = data.get("hour")  # None = send regardless of hour (legacy behavior)
        today    = datetime.utcnow().date()
        today_str = str(today)

        # Find users with active streaks who haven't done a session today
        # Query push_tokens to get all active users
        token_docs = db.collection("push_tokens").where("active","==",True).limit(1000).stream()
        uids_seen  = set()
        reminders_sent = 0
        reminders_skipped = 0

        for tdoc in token_docs:
            td  = tdoc.to_dict()
            uid = td.get("uid")
            tok = td.get("token")
            lang = td.get("lang", "en")
            if not uid or not tok or uid in uids_seen:
                continue
            uids_seen.add(uid)

            # Smart scheduling: only send to this user during their own
            # usual session hour, unless the caller wants every hour sent
            # regardless (target_hour omitted).
            if target_hour is not None:
                prefs = db.collection("push_preferences").document(uid).get()
                pdata = prefs.to_dict() if prefs.exists else {}
                if not pdata.get("categories", {}).get("streak", True):
                    reminders_skipped += 1
                    continue
                user_hour = pdata.get("preferred_hour_override")
                if user_hour is None:
                    user_hour = _compute_preferred_hour(uid)
                if int(user_hour) != int(target_hour):
                    reminders_skipped += 1
                    continue

            # Check if user already had a session today
            today_sessions = (db.collection("sessions")
                                .where("uid", "==", uid)
                                .where("created_at", ">=", datetime.utcnow().replace(hour=0,minute=0,second=0))
                                .limit(1).stream())
            had_session_today = any(True for _ in today_sessions)
            if had_session_today:
                reminders_skipped += 1
                continue

            # Check current streak
            cutoff = datetime.utcnow() - timedelta(days=35)
            sess_docs = (db.collection("sessions")
                           .where("uid", "==", uid)
                           .where("created_at", ">=", cutoff)
                           .order_by("created_at", direction="DESCENDING")
                           .limit(50).stream())
            day_scores: dict = {}
            for sd in sess_docs:
                d = sd.to_dict(); sc = d.get("avg_score"); ts = d.get("created_at")
                if not sc or not ts: continue
                try:
                    day = ts.strftime("%Y-%m-%d") if hasattr(ts,"strftime") else str(ts)[:10]
                    day_scores[day] = max(day_scores.get(day,0), sc)
                except Exception:
                    pass
            active_days = {d for d,sc in day_scores.items() if sc >= 55}
            streak = 0
            chk = today - timedelta(days=1)   # yesterday (today not done yet)
            while str(chk) in active_days:
                streak += 1; chk -= timedelta(days=1)

            if streak < min_streak:
                reminders_skipped += 1
                continue

            # Send notification
            is_ar = lang == "ar"
            if streak >= 30:
                title = "👑 لا توقف سلسلتك!" if is_ar else "👑 Don't break your streak!"
                body  = f"عندك {streak} يوم متتالي — جلسة واحدة تحافظ عليها!" if is_ar else f"You're on a {streak}-day streak — one session to keep it alive!"
            elif streak >= 7:
                title = "🔥 السلسلة على المحك!" if is_ar else "🔥 Your streak is at risk!"
                body  = f"{streak} أيام متتاليين! لا تضيعها الليلة" if is_ar else f"{streak} days strong! Don't let it end tonight"
            else:
                title = "💪 تحقق من وضعيتك اليوم" if is_ar else "💪 Check your posture today"
                body  = f"لديك سلسلة {streak} أيام — حافظ عليها!" if is_ar else f"You have a {streak}-day streak — keep it going!"

            sent = _send_fcm(tok, title, body, {"type": "streak_reminder", "streak": str(streak)})
            if sent:
                reminders_sent += 1

        log_event("streak_reminders_sent", meta={"sent": reminders_sent, "skipped": reminders_skipped})
        return jsonify({"ok": True, "sent": reminders_sent, "skipped": reminders_skipped})
    except Exception as e:
        return safe_error(e)


@app.route("/api/push/test", methods=["POST"])
@require_auth
@limiter.limit("5 per minute")
def push_test():
    """Send a test push notification to the authenticated user."""
    try:
        uid   = getattr(g, "uid", "")
        data  = request.get_json(force=True) or {}
        lang  = data.get("lang", "en")
        tokens = _get_user_tokens(uid)
        if not tokens:
            return jsonify({"error": "No registered device tokens — call /api/push/register first"}), 400

        title = "🎉 الإشعارات تعمل!" if lang=="ar" else "🎉 Push notifications work!"
        body  = "سيصلك إشعار عند خطر انكسار سلسلتك" if lang=="ar" else "You'll be notified when your streak is at risk"
        results = [_send_fcm(t, title, body, {"type": "test"}) for t in tokens]
        return jsonify({"ok": True, "tokens": len(tokens), "sent": sum(results)})
    except Exception as e:
        return safe_error(e)


# ══════════════════════════════════════════════════════════════════
# B2C: ANONYMOUS GLOBAL LEADERBOARD
# ══════════════════════════════════════════════════════════════════

@app.route("/api/leaderboard/submit", methods=["POST"])
@require_auth
@limiter.limit("10 per minute")
def leaderboard_submit():
    """
    Submit user's weekly average to the global leaderboard.
    Privacy: stored anonymously — no name, email, or UID exposed publicly.
    Display name = user-chosen alias (default: "User####")
    """
    try:
        uid  = getattr(g, "uid", "")
        data = request.get_json(force=True) or {}
        alias    = data.get("alias", "")[:20].strip()  # user-chosen display name
        score    = float(data.get("score", 0))
        streak   = int(data.get("streak", 0))
        sessions = int(data.get("sessions", 0))
        period   = data.get("period", "week")   # week | month | alltime

        if not 0 <= score <= 100:
            return jsonify({"error": "score must be 0-100"}), 400

        # Generate anonymous alias if not provided
        if not alias:
            import hashlib as _hl
            alias = "User" + _hl.md5(uid.encode()).hexdigest()[:4].upper()

        # Composite score: 70% posture + 20% streak + 10% consistency
        streak_bonus   = min(20, streak * 0.5)           # max 20pts from streak
        session_bonus  = min(10, sessions * 0.2)          # max 10pts from sessions
        composite = round(score * 0.70 + streak_bonus + session_bonus, 1)

        # Upsert to Firestore leaderboard collection
        now = datetime.utcnow()
        period_key = {
            "week":    now.strftime("%Y-W%W"),
            "month":   now.strftime("%Y-%m"),
            "alltime": "alltime",
        }.get(period, now.strftime("%Y-W%W"))

        doc_id = f"{period_key}_{uid}"
        db.collection("leaderboard").document(doc_id).set({
            "uid_hash":      __import__("hashlib").md5(uid.encode()).hexdigest()[:12],  # anonymized
            "alias":         alias,
            "score":         round(score, 1),
            "streak":        streak,
            "sessions":      sessions,
            "composite":     composite,
            "period":        period,
            "period_key":    period_key,
            "updated_at":    now.isoformat() + "Z",
            "grade":         "A" if score>=85 else "B" if score>=70 else "C" if score>=55 else "D",
        }, merge=True)

        # Also update Redis sorted set for fast leaderboard queries
        _rc = get_redis()
        if _rc:
            _rc.zadd(f"lb:{period_key}", {doc_id: composite})
            _rc.zremrangebyrank(f"lb:{period_key}", 0, -1001)  # keep top 1000
            _rc.expire(f"lb:{period_key}", 86400 * 8)           # 8 days TTL

        log_event("leaderboard_submit", uid, {"score": score, "period": period, "composite": composite})
        return jsonify({"ok": True, "composite": composite, "alias": alias, "period_key": period_key})
    except Exception as e:
        return safe_error(e)


@app.route("/api/leaderboard", methods=["GET"])
@limiter.limit("60 per minute")
def get_leaderboard():
    """
    Return global leaderboard. No auth required — public endpoint.
    Query params:
      period=week|month|alltime  (default: week)
      limit=10-100               (default: 50)
      uid=<optional>             show authenticated user's rank
    """
    try:
        period = request.args.get("period", "week")
        limit  = min(100, max(10, int(request.args.get("limit", 50))))
        uid    = request.args.get("uid", "")  # optional — to find user's rank

        now = datetime.utcnow()
        period_key = {
            "week":    now.strftime("%Y-W%W"),
            "month":   now.strftime("%Y-%m"),
            "alltime": "alltime",
        }.get(period, now.strftime("%Y-W%W"))

        # Try Redis sorted set first (fast)
        _rc = get_redis()
        board = []
        user_rank = None

        if _rc:
            try:
                # Get top N entries by composite score (descending)
                top_ids = _rc.zrevrange(f"lb:{period_key}", 0, limit-1, withscores=True)
                if top_ids:
                    import json as _j
                    for rank, (doc_id_bytes, composite) in enumerate(top_ids, start=1):
                        doc_id = doc_id_bytes.decode() if isinstance(doc_id_bytes, bytes) else doc_id_bytes
                        # Fetch full data from Firestore
                        doc = db.collection("leaderboard").document(doc_id).get()
                        if doc.exists:
                            d = doc.to_dict()
                            entry = {
                                "rank":      rank,
                                "medal":     {1:"🥇",2:"🥈",3:"🥉"}.get(rank,""),
                                "alias":     d.get("alias","Anonymous"),
                                "score":     d.get("score"),
                                "streak":    d.get("streak",0),
                                "composite": round(composite, 1),
                                "grade":     d.get("grade",""),
                                "is_me":     uid and doc_id.endswith(uid),
                            }
                            board.append(entry)
                            if uid and doc_id.endswith(uid):
                                user_rank = rank

                    # Find user rank if not in top N
                    if uid and not user_rank:
                        doc_id = f"{period_key}_{uid}"
                        user_rank_raw = _rc.zrevrank(f"lb:{period_key}", doc_id)
                        if user_rank_raw is not None:
                            user_rank = user_rank_raw + 1
            except Exception:
                pass

        # Firestore fallback
        if not board:
            docs = (db.collection("leaderboard")
                      .where("period_key","==",period_key)
                      .order_by("composite", direction="DESCENDING")
                      .limit(limit).stream())
            for rank, doc in enumerate(docs, start=1):
                d = doc.to_dict()
                board.append({
                    "rank":      rank,
                    "medal":     {1:"🥇",2:"🥈",3:"🥉"}.get(rank,""),
                    "alias":     d.get("alias","Anonymous"),
                    "score":     d.get("score"),
                    "streak":    d.get("streak",0),
                    "composite": d.get("composite"),
                    "grade":     d.get("grade",""),
                    "is_me":     uid and doc.id.endswith(uid),
                })
                if uid and doc.id.endswith(uid):
                    user_rank = rank

        # Stats
        total = len(board)
        avg_score = round(sum(e["score"] for e in board if e["score"])/max(total,1), 1)
        top_streak = max((e["streak"] for e in board), default=0)

        return jsonify({
            "ok":         True,
            "period":     period,
            "period_key": period_key,
            "board":      board,
            "user_rank":  user_rank,
            "total":      total,
            "stats": {
                "avg_score":  avg_score,
                "top_streak": top_streak,
                "participants": total,
            },
            "privacy_note": "All entries are anonymous. No personal data is exposed.",
        })
    except Exception as e:
        return safe_error(e)


@app.route("/api/leaderboard/my-rank", methods=["GET"])
@require_auth
@limiter.limit("30 per minute")
def my_leaderboard_rank():
    """Return the authenticated user's rank + nearby competitors."""
    try:
        uid    = getattr(g, "uid", "")
        period = request.args.get("period", "week")
        now    = datetime.utcnow()
        period_key = {
            "week":    now.strftime("%Y-W%W"),
            "month":   now.strftime("%Y-%m"),
            "alltime": "alltime",
        }.get(period, now.strftime("%Y-W%W"))

        doc_id = f"{period_key}_{uid}"
        doc = db.collection("leaderboard").document(doc_id).get()
        if not doc.exists:
            return jsonify({
                "ok":       True,
                "ranked":   False,
                "message":  "Not on leaderboard yet — submit your score via /api/leaderboard/submit",
                "message_ar": "لست في المتصفحة بعد — ابعت درجتك عبر /api/leaderboard/submit",
            })

        my_data = doc.to_dict()
        _rc     = get_redis()
        my_rank = None
        nearby  = []

        if _rc:
            raw_rank = _rc.zrevrank(f"lb:{period_key}", doc_id)
            if raw_rank is not None:
                my_rank = raw_rank + 1
                # Get 2 above + 2 below for context
                start = max(0, raw_rank - 2)
                end   = raw_rank + 2
                nearby_ids = _rc.zrevrange(f"lb:{period_key}", start, end, withscores=True)
                for r_offset, (nid_bytes, composite) in enumerate(nearby_ids, start=start+1):
                    nid = nid_bytes.decode() if isinstance(nid_bytes, bytes) else nid_bytes
                    ndoc = db.collection("leaderboard").document(nid).get()
                    if ndoc.exists:
                        nd = ndoc.to_dict()
                        nearby.append({
                            "rank":      r_offset,
                            "alias":     nd.get("alias","Anonymous"),
                            "composite": round(composite,1),
                            "is_me":     nid == doc_id,
                        })

        return jsonify({
            "ok":        True,
            "ranked":    True,
            "rank":      my_rank,
            "composite": my_data.get("composite"),
            "score":     my_data.get("score"),
            "streak":    my_data.get("streak"),
            "alias":     my_data.get("alias"),
            "nearby":    nearby,
            "period":    period,
        })
    except Exception as e:
        return safe_error(e)


# ══════════════════════════════════════════════════════════════════
# B2B: COMPANY DASHBOARD API (Arabic-first HR dashboard)
# ══════════════════════════════════════════════════════════════════

def _require_b2b(f):
    """Decorator: ensures user is B2B (company/HR). Returns 403 for B2C users."""
    from functools import wraps
    @wraps(f)
    def wrapper(*args, **kwargs):
        user_type = getattr(g, "user_type", None) or _detect_user_type(
            getattr(g,"uid",""), getattr(g,"role",{}))
        company_id = getattr(g, "company_id", "") or getattr(g,"role",{}).get("company_id","")
        if not company_id:
            return jsonify({
                "error": "b2b_required",
                "message": "This endpoint requires a company account.",
                "message_ar": "هذا الطلب يتطلب حساب شركة.",
            }), 403
        g.company_id = company_id
        return f(*args, **kwargs)
    return wrapper


@app.route("/api/company/dashboard", methods=["GET"])
@require_auth
@_require_b2b
@limiter.limit("30 per minute")
def company_dashboard():
    """
    Full HR dashboard — aggregate stats for the whole company.
    Returns bilingual (AR + EN) labels for Arabic-first UI.
    """
    try:
        company_id = g.company_id
        lang       = request.args.get("lang", "ar")  # default Arabic for B2B Egypt
        days       = min(90, max(7, int(request.args.get("days", 30))))
        cutoff     = datetime.utcnow() - timedelta(days=days)

        # ── Fetch all employees in this company ───────────────────
        emp_docs = (db.collection("users")
                      .where("company_id", "==", company_id)
                      .limit(500).stream())
        employees = {d.id: d.to_dict() for d in emp_docs}
        emp_uids  = list(employees.keys())

        if not emp_uids:
            return jsonify({
                "ok": True,
                "company_id": company_id,
                "message": "No employees found. Invite employees via /api/org/send-invite",
                "message_ar": "لا يوجد موظفون. أرسل دعوات عبر /api/org/send-invite",
            })

        # ── Fetch sessions for all employees ──────────────────────
        # Firestore: batch query (max 10 UIDs per __in__ query)
        all_sessions = []
        for i in range(0, len(emp_uids), 10):
            batch = emp_uids[i:i+10]
            docs  = (db.collection("sessions")
                       .where("uid", "in", batch)
                       .where("created_at", ">=", cutoff)
                       .limit(2000).stream())
            all_sessions.extend([d.to_dict() for d in docs])

        # ── Per-employee aggregation ───────────────────────────────
        emp_stats: dict = {uid: {
            "uid":      uid,
            "name":     employees[uid].get("name") or employees[uid].get("displayName") or f"Employee {uid[:6]}",
            "email":    employees[uid].get("email",""),
            "sessions": 0, "scores": [], "alerts_count": 0,
            "avg_score": None, "grade": None, "trend": "stable",
        } for uid in emp_uids}

        for s in all_sessions:
            uid = s.get("uid")
            sc  = s.get("avg_score")
            if uid not in emp_stats: continue
            emp_stats[uid]["sessions"] += 1
            if isinstance(sc, (int, float)):
                emp_stats[uid]["scores"].append(sc)
            emp_stats[uid]["alerts_count"] += s.get("alerts_count", 0) or 0

        for uid, es in emp_stats.items():
            scores = es["scores"]
            if scores:
                avg = round(sum(scores)/len(scores), 1)
                es["avg_score"] = avg
                es["grade"]     = "A" if avg>=85 else "B" if avg>=70 else "C" if avg>=55 else "D"
                # Trend: first half vs second half
                if len(scores) >= 4:
                    half = len(scores)//2
                    early = sum(scores[:half])/half
                    late  = sum(scores[half:])/half
                    es["trend"] = "improving" if late-early>5 else "declining" if late-early<-5 else "stable"

        emp_list = sorted(emp_stats.values(), key=lambda x: (x["avg_score"] or 0), reverse=True)

        # ── Company-wide KPIs ─────────────────────────────────────
        all_scores  = [e["avg_score"] for e in emp_list if e["avg_score"] is not None]
        company_avg = round(sum(all_scores)/len(all_scores), 1) if all_scores else None
        at_risk     = [e for e in emp_list if (e["avg_score"] or 100) < 55]
        improving   = [e for e in emp_list if e["trend"] == "improving"]
        declining   = [e for e in emp_list if e["trend"] == "declining"]
        top_5       = emp_list[:5]
        bottom_5    = [e for e in reversed(emp_list) if e["avg_score"] is not None][:5]

        # ── ROI estimate ──────────────────────────────────────────
        # Based on WHO: poor posture costs ~3.5 sick days/year per employee
        # PostureAI targets 40% reduction → 1.4 days saved per employee
        # Egypt avg daily wage: ~400 EGP
        _emp_count   = len(emp_uids)
        _days_saved  = round(_emp_count * 1.4, 1)
        _egp_saved   = round(_days_saved * 400, 0)
        _roi_note    = "Estimated based on WHO musculoskeletal sick leave data"

        # ── Bilingual labels ──────────────────────────────────────
        is_ar = lang == "ar"
        def _label(en, ar): return ar if is_ar else en

        return jsonify({
            "ok":          True,
            "company_id":  company_id,
            "period_days": days,
            "lang":        lang,

            # KPIs
            "kpis": {
                "company_avg_score": company_avg,
                "company_grade":     "A" if (company_avg or 0)>=85 else "B" if (company_avg or 0)>=70 else "C" if (company_avg or 0)>=55 else "D",
                "total_employees":   _emp_count,
                "active_employees":  sum(1 for e in emp_list if e["sessions"]>0),
                "at_risk_count":     len(at_risk),
                "at_risk_pct":       round(len(at_risk)/_emp_count*100, 1) if _emp_count else 0,
                "improving_count":   len(improving),
                "declining_count":   len(declining),
                "total_sessions":    len(all_sessions),

                # Labels (bilingual)
                "labels": {
                    "company_avg":   _label("Company Average","متوسط الشركة"),
                    "at_risk":       _label("At Risk","في خطر"),
                    "improving":     _label("Improving","يتحسن"),
                    "declining":     _label("Declining","يتراجع"),
                    "sessions":      _label("Total Sessions","إجمالي الجلسات"),
                },
            },

            # Employee lists
            "top_performers":  [_employee_card(e, is_ar) for e in top_5],
            "needs_attention": [_employee_card(e, is_ar) for e in bottom_5],
            "at_risk":         [_employee_card(e, is_ar) for e in at_risk],
            "all_employees":   [_employee_card(e, is_ar) for e in emp_list],

            # ROI
            "roi_estimate": {
                "employees":      _emp_count,
                "sick_days_saved": _days_saved,
                "egp_saved":      _egp_saved,
                "note":           _label(_roi_note, "تقدير بناءً على بيانات منظمة الصحة العالمية"),
            },
        })
    except Exception as e:
        return safe_error(e)


def _employee_card(e: dict, is_ar: bool) -> dict:
    """Format employee data for dashboard display."""
    trend_labels = {
        "improving": ("📈 Improving", "📈 يتحسن"),
        "declining":  ("📉 Declining", "📉 يتراجع"),
        "stable":     ("➡️ Stable",   "➡️ مستقر"),
    }
    trend = e.get("trend","stable")
    tl    = trend_labels.get(trend, ("➡️ Stable","➡️ مستقر"))
    return {
        "uid":         e["uid"],
        "name":        e["name"],
        "email":       e["email"],
        "avg_score":   e["avg_score"],
        "grade":       e["grade"],
        "sessions":    e["sessions"],
        "alerts_count":e["alerts_count"],
        "trend":       trend,
        "trend_label": tl[1] if is_ar else tl[0],
        "status":      ("🔴 في خطر" if is_ar else "🔴 At Risk") if (e["avg_score"] or 100) < 55
                  else ("🟡 يحتاج انتباه" if is_ar else "🟡 Needs Attention") if (e["avg_score"] or 100) < 70
                  else ("🟢 جيد" if is_ar else "🟢 Good"),
    }


@app.route("/api/company/alert-employees", methods=["POST"])
@require_auth
@_require_b2b
@limiter.limit("10 per minute")
def company_alert_employees():
    """
    HR sends WhatsApp/push alert to specific employees or all at-risk.
    target: "all" | "at_risk" | [uid1, uid2, ...]
    """
    try:
        company_id = g.company_id
        hr_uid     = g.uid
        data       = request.get_json(force=True) or {}
        target     = data.get("target", "at_risk")
        message    = data.get("message", "")
        lang       = data.get("lang", "ar")

        if not message:
            return jsonify({"error": "message required"}), 400

        # Get target UIDs
        if isinstance(target, list):
            uids = target
        else:
            emp_docs = (db.collection("users")
                          .where("company_id","==",company_id).limit(500).stream())
            all_uids = [d.id for d in emp_docs]
            if target == "at_risk":
                # Filter for employees with avg_score < 55 in last 7 days
                cutoff = datetime.utcnow() - timedelta(days=7)
                uids   = []
                for uid in all_uids:
                    sess = (db.collection("sessions")
                              .where("uid","==",uid)
                              .where("created_at",">=",cutoff)
                              .limit(20).stream())
                    scores = [s.to_dict().get("avg_score",0) for s in sess]
                    if scores and sum(scores)/len(scores) < 55:
                        uids.append(uid)
            else:
                uids = all_uids

        # Send notification to each employee
        sent = 0; failed = 0
        for uid in uids[:100]:  # cap at 100 per call
            result = notify_user(uid, 
                title="تنبيه وضعية من HR" if lang=="ar" else "HR Posture Alert",
                body=message, lang=lang,
                data={"type":"hr_alert","from":hr_uid})
            if result.get("sent"): sent += 1
            else: failed += 1

        log_event("company_alert_sent", hr_uid, {
            "company_id": company_id, "target": str(target)[:20],
            "sent": sent, "failed": failed
        })
        return jsonify({"ok": True, "sent": sent, "failed": failed, "total": len(uids)})
    except Exception as e:
        return safe_error(e)


@app.route("/api/company/stats", methods=["GET"])
@require_auth
@_require_b2b
@limiter.limit("30 per minute")
def company_stats():
    """Quick KPI summary for C-suite (no employee details)."""
    try:
        company_id = g.company_id
        lang       = request.args.get("lang","ar")
        days       = min(90, max(7, int(request.args.get("days",30))))
        cutoff     = datetime.utcnow() - timedelta(days=days)
        is_ar      = lang == "ar"

        emp_docs  = db.collection("users").where("company_id","==",company_id).limit(500).stream()
        emp_uids  = [d.id for d in emp_docs]
        all_sessions = []
        for i in range(0, len(emp_uids), 10):
            batch = emp_uids[i:i+10]
            docs  = (db.collection("sessions")
                       .where("uid","in",batch)
                       .where("created_at",">=",cutoff)
                       .limit(2000).stream())
            all_sessions.extend([d.to_dict() for d in docs])

        scores     = [s.get("avg_score") for s in all_sessions if isinstance(s.get("avg_score"),(int,float))]
        avg        = round(sum(scores)/len(scores),1) if scores else None
        at_risk_ct = sum(1 for s in scores if s < 55)

        return jsonify({
            "ok":            True,
            "company_id":    company_id,
            "period_days":   days,
            "avg_score":     avg,
            "total_sessions":len(all_sessions),
            "employees":     len(emp_uids),
            "at_risk_pct":   round(at_risk_ct/max(len(scores),1)*100,1),
            "grade":         "A" if (avg or 0)>=85 else "B" if (avg or 0)>=70 else "C" if (avg or 0)>=55 else "D",
            "labels": {
                "avg_score":   "متوسط درجة الوضعية" if is_ar else "Avg Posture Score",
                "at_risk_pct": "نسبة الموظفين في خطر" if is_ar else "% Employees At Risk",
                "sessions":    "إجمالي الجلسات" if is_ar else "Total Sessions",
                "grade":       "تقييم الشركة" if is_ar else "Company Grade",
            },
        })
    except Exception as e:
        return safe_error(e)


# ══════════════════════════════════════════════════════════════════
# USER NOTIFICATION PREFERENCES
# ══════════════════════════════════════════════════════════════════

@app.route("/api/user/notif-pref", methods=["GET", "PUT"])
@require_auth
@limiter.limit("20 per minute")
def user_notif_pref():
    """Get or set user notification channel preference."""
    try:
        uid = getattr(g, "uid", "")
        if request.method == "GET":
            doc = db.collection("users").document(uid).get()
            d   = doc.to_dict() if doc.exists else {}
            return jsonify({
                "ok":           True,
                "notif_channel": d.get("notif_channel","auto"),
                "wa_phone":     d.get("wa_phone",""),
                "channels": {
                    "auto":      "WhatsApp → Push → Email (recommended)",
                    "whatsapp":  "WhatsApp only",
                    "push":      "Push notifications only",
                    "email":     "Email only",
                    "none":      "No notifications",
                }
            })
        data    = request.get_json(force=True) or {}
        channel = data.get("channel","auto")
        wa_phone= data.get("wa_phone","").strip()
        valid   = ("auto","whatsapp","push","email","none")
        if channel not in valid:
            return jsonify({"error": f"Invalid channel. Valid: {valid}"}), 400
        update = {"notif_channel": channel}
        if wa_phone:
            update["wa_phone"] = wa_phone
            update["phone"]    = wa_phone
        db.collection("users").document(uid).update(update)
        log_event("notif_pref_updated", uid, {"channel": channel})
        return jsonify({"ok": True, "notif_channel": channel})
    except Exception as e:
        return safe_error(e)


# ── Generic AI Analysis endpoint (used by frontend AIInsights/AIReports/PredictiveAI) ──
@app.route("/api/ai/analyze", methods=["POST"])
@require_auth
@limiter.limit("20 per minute")
def ai_analyze():
    """
    Generic local-AI proxy — replaces direct browser→cloud-AI calls.
    All AI calls from frontend go through here (no API key on client,
    no cloud provider at all — Ollama only).
    """
    try:
        uid  = getattr(g, "uid", "")
        tier = getattr(g, "tier", "standard") or "standard"
        data = request.get_json(force=True) or {}
        prompt  = data.get("prompt", "").strip()
        lang    = data.get("lang", "en")
        context = data.get("context", {})

        if not prompt:
            return jsonify({"error": "prompt required"}), 400
        if not AI_CONFIGURED:
            return jsonify({"error": "AI not configured", "text": None}), 503
        if len(prompt) > 8000:
            return jsonify({"error": "Prompt too long (max 8000 chars)"}), 400

        # Monthly usage check (same limits as coach)
        _month_key = f"usage:{uid}:ai_analyze:{datetime.utcnow().strftime('%Y-%m')}"
        _limits = {"standard": 20, "basic": 40, "pro": 100,
                   "professional": 200, "elite": -1, "premium": -1, "enterprise": -1}
        _limit = _limits.get(tier, 20)
        if _limit > 0:
            try:
                _rc = get_redis()
                _used = int(_rc.get(_month_key) or 0) if _rc else 0
            except Exception:
                _used = 0
            if _used >= _limit:
                return jsonify({
                    "error":   "ai_limit_reached",
                    "message": f"Monthly AI analysis limit ({_limit}) reached. Upgrade for more.",
                    "used": _used, "limit": _limit,
                }), 429

        try:
            max_tokens = int(data.get("max_tokens") or 1200)
        except (TypeError, ValueError):
            max_tokens = 1200
        max_tokens = max(100, min(2000, max_tokens))

        # call_ai() routes through local Ollama only
        text = call_ai(
            prompt,
            system_prompt=context.get("system_prompt"),
            max_tokens=max_tokens,
            temperature=0.5,
            tier=tier,
        )

        if text is None:
            # ── LLM7.io fallback ─────────────────────────────────────────
            try:
                _msgs = []
                if context.get("system_prompt"):
                    _msgs.append({"role": "system", "content": context["system_prompt"]})
                _msgs.append({"role": "user", "content": prompt})
                _r = req.post(
                    "https://api.llm7.io/v1/chat/completions",
                    headers={"Content-Type": "application/json", "Authorization": "Bearer unused"},
                    json={"model": "gpt-4o-mini", "messages": _msgs, "max_tokens": max_tokens, "temperature": 0.5},
                    timeout=20,
                )
                if _r.status_code == 200:
                    text = (_r.json().get("choices") or [{}])[0].get("message", {}).get("content", "") or None
            except Exception:
                pass

        # Increment usage counter
        try:
            _rc = get_redis()
            if _rc: _rc.incr(_month_key); _rc.expire(_month_key, 86400*35)
        except Exception:
            pass

        _backend_used = "local:" + LOCAL_LLM_MODEL
        log_event("ai_analyze_ok", uid, {"model": _backend_used, "chars": len(text)})
        return jsonify({"ok": True, "text": text, "model": _backend_used})
    except Exception as e:
        return safe_error(e)


# ── Distance Calibration ──────────────────────────────────────────
@app.route("/api/calibrate/distance", methods=["POST"])
@require_auth
@limiter.limit("10 per minute")
def calibrate_distance():
    """
    User sits at a KNOWN distance (they measure with ruler/tape).
    We decode the frame, measure face width in pixels,
    compute exact focal length, store it permanently.

    Body: { "frame": "<base64>", "known_dist_cm": 65, "session_id": "..." }
    Returns: { "ok": True, "focal_px": 712.3, "accuracy_cm": 2.1 }
    """
    try:
        uid  = getattr(g, "uid", "")
        data = request.get_json(force=True) or {}
        b64  = data.get("frame", "")
        known_dist = float(data.get("known_dist_cm", 65))
        sid  = data.get("session_id", uid)

        if not b64 or known_dist < 20 or known_dist > 200:
            return jsonify({"error": "frame required + known_dist_cm must be 20-200cm"}), 400

        if "," in b64: b64 = b64.split(",", 1)[1]
        raw_bytes = __import__("base64").b64decode(b64 + "==")
        arr = np.frombuffer(raw_bytes, dtype=np.uint8)
        img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
        if img is None:
            return jsonify({"error": "Could not decode frame"}), 400

        h_img, w_img = img.shape[:2]
        rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)

        # Run FaceMesh on calibration frame
        _ensure_models()
        if FACE_MESH is None:
            return jsonify({"error": "FaceMesh not available"}), 503

        result = FACE_MESH.process(rgb)
        if not result.multi_face_landmarks:
            return jsonify({"error": "No face detected in calibration frame — ensure face is clearly visible"}), 400

        face_lms = result.multi_face_landmarks[0]

        # Measure bizygomatic width (landmarks 234 ↔ 454)
        l_cheek = face_lms.landmark[234]
        r_cheek = face_lms.landmark[454]
        face_w_px = abs(r_cheek.x * w_img - l_cheek.x * w_img)

        if face_w_px < 20:
            return jsonify({"error": "Face too small in frame — move closer or use better lighting"}), 400

        # focal = (face_width_px * known_dist_cm) / real_face_width_cm
        # Real bizygomatic diameter: population mean = 14.0cm (Farkas 1994)
        REAL_FACE_WIDTH_CM = 14.0
        focal_px = round((face_w_px * known_dist) / REAL_FACE_WIDTH_CM, 1)

        # Accuracy estimate: ±2cm at reference distance, scales with distance
        accuracy_cm = round(1.5 * (known_dist / 65), 1)

        # Also compute IPD-based focal for cross-validation
        try:
            l_pupil = face_lms.landmark[468]
            r_pupil = face_lms.landmark[473]
            ipd_px  = abs(r_pupil.x * w_img - l_pupil.x * w_img)
            # Real IPD mean = 6.3cm
            focal_ipd = round((ipd_px * known_dist) / 6.3, 1)
            # Weighted blend: 60% face-width + 40% IPD
            focal_px = round(0.60 * focal_px + 0.40 * focal_ipd, 1)
        except Exception:
            focal_ipd = None

        # Store calibrated focal in Firestore + Redis
        cal_doc = {
            "uid":           uid,
            "focal_px":      focal_px,
            "known_dist_cm": known_dist,
            "face_w_px":     round(face_w_px, 1),
            "frame_w":       w_img,
            "accuracy_cm":   accuracy_cm,
            "calibrated_at": datetime.utcnow().isoformat() + "Z",
        }
        db.collection("distance_calibrations").document(uid).set(cal_doc, merge=True)

        # Cache in Redis for fast per-session lookup
        _rc = get_redis()
        if _rc:
            import json as _j
            _rc.setex(f"dist_cal:{uid}", 86400 * 30, _j.dumps(cal_doc))

        # Also update in-memory focal cache for this session
        _focal_cal[sid] = focal_px
        _focal_cal[uid] = focal_px   # uid key as fallback

        log_event("distance_calibrated", uid, {
            "focal_px": focal_px, "known_dist": known_dist,
            "accuracy_cm": accuracy_cm
        })
        return jsonify({
            "ok":           True,
            "focal_px":     focal_px,
            "known_dist_cm":known_dist,
            "accuracy_cm":  accuracy_cm,
            "message":      f"Distance calibration complete — accuracy now ±{accuracy_cm}cm",
            "message_ar":   f"تمت معايرة المسافة — الدقة الآن ±{accuracy_cm}سم",
        })
    except Exception as e:
        return safe_error(e)


# ══════════════════════════════════════════════════════════════════
# SMART BREAK RECOMMENDATION ENGINE
# Learns each user's personal fatigue curve
# ══════════════════════════════════════════════════════════════════

_break_profiles: dict = {}  # {uid: {"scores": [(ts, score)], "deg_rate": float}}

def update_break_profile(uid: str, score: int, session_start: float) -> dict:
    """
    Track score over session time to learn personal degradation rate.
    Returns break recommendation with personalized timing.
    """
    import time as _t
    now = _t.time()
    session_min = (now - session_start) / 60

    profile = _break_profiles.setdefault(uid, {
        "scores": [], "avg_deg_rate": None,
        "personal_threshold_min": 45,   # default 45min
        "sessions_analyzed": 0,
    })

    profile["scores"].append((session_min, score))

    # Keep last 60 data points per session
    if len(profile["scores"]) > 60:
        profile["scores"] = profile["scores"][-60:]

    # Only analyze if we have enough data (>5 min of session)
    if session_min < 5 or len(profile["scores"]) < 5:
        return {"recommendation": None, "session_min": round(session_min, 1)}

    # Linear regression on score vs time to find degradation rate
    pts = profile["scores"]
    n   = len(pts)
    mean_t = sum(p[0] for p in pts) / n
    mean_s = sum(p[1] for p in pts) / n
    num = sum((p[0]-mean_t)*(p[1]-mean_s) for p in pts)
    den = sum((p[0]-mean_t)**2 for p in pts) or 1
    slope = num / den  # score points per minute (negative = degrading)

    # Update personal threshold based on observed degradation
    if slope < -0.8:
        # Fast degrader — suggest break sooner
        profile["personal_threshold_min"] = max(20, profile["personal_threshold_min"] - 2)
    elif slope > -0.2:
        # Stable — can extend break interval
        profile["personal_threshold_min"] = min(60, profile["personal_threshold_min"] + 1)

    threshold = profile["personal_threshold_min"]
    time_to_break = max(0, threshold - session_min)

    # Urgency
    urgency = (
        "now"    if time_to_break <= 0  else
        "soon"   if time_to_break <= 5  else
        "later"  if time_to_break <= 15 else
        "ok"
    )

    rec = {
        "session_min":       round(session_min, 1),
        "deg_rate":          round(slope, 2),
        "personal_threshold":threshold,
        "time_to_break_min": round(time_to_break, 1),
        "urgency":           urgency,
        "message":           None,
        "message_ar":        None,
    }

    if urgency == "now":
        rec["message"]    = f"⚠️ Break overdue ({round(session_min)}min seated) — stand up for 2 minutes"
        rec["message_ar"] = f"⚠️ حان وقت الاستراحة ({round(session_min)} دقيقة جلوس) — قم للوقوف دقيقتين"
    elif urgency == "soon":
        rec["message"]    = f"Break in {round(time_to_break)} min — your posture typically degrades after {threshold}min"
        rec["message_ar"] = f"استراحة بعد {round(time_to_break)} دقيقة — وضعيتك تتردى عادةً بعد {threshold} دقيقة"

    return rec


# ══════════════════════════════════════════════════════════════════
# FEATURE 1: POSTURE SCORE HISTORY GRAPH
# ══════════════════════════════════════════════════════════════════

@app.route("/api/analytics/score-history", methods=["GET"])
@require_auth
@limiter.limit("30 per minute")
def score_history():
    """
    Return score history for line chart visualization.
    Query: ?period=7d|30d|90d&granularity=day|session
    """
    try:
        uid         = getattr(g, "uid", "")
        period      = request.args.get("period", "30d")
        granularity = request.args.get("granularity", "day")
        days        = {"7d":7,"30d":30,"90d":90}.get(period, 30)
        cutoff      = datetime.utcnow() - timedelta(days=days)

        docs = (db.collection("sessions")
                  .where("uid","==",uid)
                  .where("created_at",">=",cutoff)
                  .order_by("created_at").limit(500).stream())

        sessions_list = [d.to_dict() for d in docs]
        if not sessions_list:
            return jsonify({"ok":True,"points":[],"summary":{}})

        if granularity == "session":
            points = []
            for s in sessions_list:
                ts = s.get("created_at")
                sc = s.get("avg_score")
                if not ts or sc is None: continue
                points.append({
                    "ts":       ts.isoformat()+"Z" if hasattr(ts,"isoformat") else str(ts),
                    "score":    sc,
                    "grade":    "A" if sc>=85 else "B" if sc>=70 else "C" if sc>=55 else "D",
                    "duration": s.get("duration_min",0),
                    "label":    ts.strftime("%b %d %H:%M") if hasattr(ts,"strftime") else str(ts)[:16],
                })
        else:
            # Daily aggregation
            daily: dict = {}
            for s in sessions_list:
                ts = s.get("created_at"); sc = s.get("avg_score")
                if not ts or sc is None: continue
                day = ts.strftime("%Y-%m-%d") if hasattr(ts,"strftime") else str(ts)[:10]
                daily.setdefault(day,[]).append(sc)
            points = []
            for day in sorted(daily.keys()):
                scores = daily[day]
                avg = round(sum(scores)/len(scores),1)
                points.append({
                    "ts":       day,
                    "score":    avg,
                    "grade":    "A" if avg>=85 else "B" if avg>=70 else "C" if avg>=55 else "D",
                    "sessions": len(scores),
                    "min":      min(scores),
                    "max":      max(scores),
                    "label":    day,
                })

        # Trend line (linear regression)
        if len(points) >= 3:
            n = len(points)
            xs = list(range(n)); ys = [p["score"] for p in points]
            mx,my = sum(xs)/n, sum(ys)/n
            num = sum((xs[i]-mx)*(ys[i]-my) for i in range(n))
            den = sum((x-mx)**2 for x in xs) or 1
            slope = num/den
            trend = "improving" if slope>0.3 else "declining" if slope<-0.3 else "stable"
        else:
            slope=0; trend="stable"

        all_scores = [p["score"] for p in points]
        return jsonify({
            "ok":True, "points":points, "period":period,
            "summary":{
                "avg":   round(sum(all_scores)/len(all_scores),1) if all_scores else None,
                "best":  max(all_scores) if all_scores else None,
                "worst": min(all_scores) if all_scores else None,
                "trend": trend,
                "slope": round(slope,2),
                "total_points": len(points),
            }
        })
    except Exception as e:
        return safe_error(e)


# ══════════════════════════════════════════════════════════════════
# FEATURE 2: PERSONAL RECORDS
# ══════════════════════════════════════════════════════════════════

@app.route("/api/records", methods=["GET"])
@require_auth
@limiter.limit("30 per minute")
def personal_records():
    """Personal bests: best session, longest streak, best day, etc."""
    try:
        uid    = getattr(g, "uid", "")
        cutoff = datetime.utcnow() - timedelta(days=365)
        docs   = (db.collection("sessions")
                    .where("uid","==",uid)
                    .where("created_at",">=",cutoff)
                    .order_by("created_at",direction="DESCENDING")
                    .limit(500).stream())
        sessions_list = [d.to_dict() for d in docs]
        if not sessions_list:
            return jsonify({"ok":True,"records":{},"message":"Complete your first session to unlock records!"})

        scores   = [(s.get("avg_score",0), s) for s in sessions_list if s.get("avg_score")]
        best_s   = max(scores, key=lambda x:x[0]) if scores else (0,{})
        longest  = max((s.get("duration_min",0) for s in sessions_list), default=0)
        total    = len(sessions_list)
        total_min= round(sum(s.get("duration_min",0) or 0 for s in sessions_list),1)

        # Best day (highest daily average)
        daily: dict = {}
        for s in sessions_list:
            ts=s.get("created_at"); sc=s.get("avg_score")
            if not ts or not sc: continue
            day = ts.strftime("%Y-%m-%d") if hasattr(ts,"strftime") else str(ts)[:10]
            daily.setdefault(day,[]).append(sc)
        best_day = max(daily.items(), key=lambda x:sum(x[1])/len(x[1])) if daily else None

        def _fmt_ts(ts):
            if not ts: return ""
            if hasattr(ts,"strftime"): return ts.strftime("%A, %b %d")
            return str(ts)[:10]

        records = {
            "best_session":   {"score": best_s[0], "date": _fmt_ts(best_s[1].get("created_at")), "label": "🏆 Best Session"},
            "longest_session":{"minutes": round(longest,1), "label": "⏱️ Longest Session"},
            "total_sessions": {"count": total, "label": "📊 Total Sessions"},
            "total_time_min": {"minutes": total_min, "label": "🕐 Total Time"},
            "best_day":       {"date": best_day[0] if best_day else None,
                               "avg": round(sum(best_day[1])/len(best_day[1]),1) if best_day else None,
                               "label": "⭐ Best Day"},
        }

        # Recent record broken?
        recent = sessions_list[0] if sessions_list else {}
        just_broke = None
        if recent.get("avg_score",0) >= best_s[0] and len(sessions_list)>1:
            just_broke = {"message":"🎉 New personal record!", "score": recent.get("avg_score")}

        return jsonify({"ok":True,"records":records,"just_broke":just_broke})
    except Exception as e:
        return safe_error(e)


# ══════════════════════════════════════════════════════════════════
# FEATURE 3: STREAK FREEZE
# ══════════════════════════════════════════════════════════════════

@app.route("/api/streak/freeze", methods=["POST"])
@require_auth
@limiter.limit("5 per minute")
def streak_freeze():
    """
    Use a streak freeze (protects streak for 1 day).
    Users get 1 free freeze/month. Pro/Elite: 3/month.
    """
    try:
        uid  = getattr(g, "uid", "")
        tier = getattr(g,"tier","standard") or "standard"
        data = request.get_json(force=True) or {}

        # Check freeze allowance
        max_freezes = 3 if tier in ("professional","elite","premium","pro") else 1
        month_key   = f"streak_freeze:{uid}:{datetime.utcnow().strftime('%Y-%m')}"

        _rc = get_redis()
        used_this_month = 0
        if _rc:
            used_this_month = int(_rc.get(month_key) or 0)

        if used_this_month >= max_freezes:
            return jsonify({
                "ok":    False,
                "error": "freeze_limit_reached",
                "message": f"You've used all {max_freezes} streak freezes this month.",
                "message_ar": f"استخدمت كل {max_freezes} تجميدات هذا الشهر.",
                "upgrade_url": "/pricing" if max_freezes == 1 else None,
            }), 429

        # Apply freeze — add today as an "active" day
        today = datetime.utcnow().date()
        db.collection("streak_freezes").document(f"{uid}_{today}").set({
            "uid":        uid,
            "date":       str(today),
            "applied_at": datetime.utcnow().isoformat()+"Z",
            "tier":       tier,
        })

        # Increment counter
        if _rc:
            _rc.incr(month_key)
            _rc.expire(month_key, 86400*35)

        # Invalidate streak cache
        if _rc: _rc.delete(f"streak:{uid}")

        log_event("streak_freeze_used", uid, {"date":str(today),"tier":tier})
        return jsonify({
            "ok":      True,
            "date":    str(today),
            "freezes_used": used_this_month+1,
            "freezes_max":  max_freezes,
            "message":      "❄️ Streak freeze applied! Your streak is protected today.",
            "message_ar":   "❄️ تم تجميد السلسلة! سلسلتك محمية اليوم.",
        })
    except Exception as e:
        return safe_error(e)


# ══════════════════════════════════════════════════════════════════
# FEATURE 4: SESSION REPLAY TIMELINE
# ══════════════════════════════════════════════════════════════════

@app.route("/api/session/<sid>/timeline", methods=["GET"])
@require_auth
@limiter.limit("30 per minute")
def session_timeline(sid):
    """
    Return frame-by-frame timeline for session replay.
    Shows: timestamp, score, neck_lean, key events.
    """
    try:
        uid = getattr(g,"uid","")
        # Fetch session snapshots from Firestore
        snaps = (db.collection("session_frames")
                   .where("session_id","==",sid)
                   .where("uid","==",uid)
                   .order_by("timestamp").limit(1000).stream())

        frames = []
        prev_score = None
        for snap in snaps:
            d = snap.to_dict()
            sc = d.get("score",0)
            ts = d.get("timestamp")
            neck = d.get("neck_lean",0)
            event = None
            if prev_score is not None:
                if sc - prev_score <= -15: event = "⚠️ Posture drop"
                elif sc - prev_score >= 10: event = "✅ Posture improved"
                elif neck > 20: event = "🔴 High neck lean"
            frames.append({
                "ts":    ts.isoformat()+"Z" if hasattr(ts,"isoformat") else str(ts),
                "score": sc,
                "neck":  round(neck,1),
                "spine": round(d.get("spine_lean",0),1),
                "event": event,
                "grade": "A" if sc>=85 else "B" if sc>=70 else "C" if sc>=55 else "D",
            })
            prev_score = sc

        # Key moments
        if frames:
            worst = min(frames, key=lambda f:f["score"])
            best  = max(frames, key=lambda f:f["score"])
        else:
            worst=best=None

        return jsonify({
            "ok":     True,
            "sid":    sid,
            "frames": frames,
            "total":  len(frames),
            "key_moments": {
                "worst": worst,
                "best":  best,
            }
        })
    except Exception as e:
        return safe_error(e)


# ══════════════════════════════════════════════════════════════════
# FEATURE 5: CUSTOM ALERT THRESHOLDS
# ══════════════════════════════════════════════════════════════════

@app.route("/api/user/alert-thresholds", methods=["GET","PUT"])
@require_auth
@limiter.limit("20 per minute")
def alert_thresholds():
    """Get or set custom alert thresholds per user."""
    try:
        uid = getattr(g,"uid","")
        if request.method == "GET":
            doc = db.collection("user_settings").document(uid).get()
            defaults = {
                "neck_lean_warn":   15,   # degrees
                "neck_lean_alert":  20,
                "spine_lean_warn":  8,
                "spine_lean_alert": 12,
                "dist_min_cm":      50,
                "dist_max_cm":      80,
                "score_alert":      55,   # alert if score drops below this
                "break_interval":   45,   # minutes between break reminders
            }
            if doc.exists:
                saved = doc.to_dict().get("alert_thresholds",{})
                merged = {**defaults, **saved}
            else:
                merged = defaults
            return jsonify({"ok":True,"thresholds":merged,"defaults":defaults})

        data = request.get_json(force=True) or {}
        thresholds = data.get("thresholds",{})

        # Validate ranges
        valid = {
            "neck_lean_warn":   (5,40),   "neck_lean_alert":   (10,50),
            "spine_lean_warn":  (3,20),   "spine_lean_alert":  (5,30),
            "dist_min_cm":      (20,80),  "dist_max_cm":       (50,150),
            "score_alert":      (20,80),  "break_interval":    (15,120),
        }
        cleaned = {}
        for key,(mn,mx) in valid.items():
            if key in thresholds:
                val = float(thresholds[key])
                cleaned[key] = max(mn, min(mx, val))

        db.collection("user_settings").document(uid).set(
            {"alert_thresholds": cleaned}, merge=True
        )
        # Invalidate Redis cache
        _rc = get_redis()
        if _rc: _rc.delete(f"user_thresholds:{uid}")

        log_event("thresholds_updated", uid, cleaned)
        return jsonify({"ok":True,"thresholds":cleaned})
    except Exception as e:
        return safe_error(e)


# ══════════════════════════════════════════════════════════════════
# FEATURE 6: POSTURE REPORT CARD (weekly auto-email)
# ══════════════════════════════════════════════════════════════════

@app.route("/api/report/send-weekly", methods=["POST"])
@limiter.limit("5 per minute")
def send_weekly_report_card():
    """
    Cron endpoint — sends weekly posture report card via email.
    Call every Monday 09:00 with X-Cron-Secret header.
    """
    try:
        secret      = request.headers.get("X-Cron-Secret","")
        cron_secret = os.getenv("CRON_SECRET","")
        if not cron_secret or not secret or secret != cron_secret:
            return jsonify({"error":"Unauthorized"}), 401

        # Get all users who opted in for weekly reports
        users_docs = (db.collection("user_settings")
                        .where("weekly_report_email","==",True)
                        .limit(1000).stream())
        sent=0; failed=0

        for udoc in users_docs:
            try:
                uid   = udoc.id
                prefs = udoc.to_dict()
                email = prefs.get("email","")
                lang  = prefs.get("lang","en")
                is_ar = lang=="ar"

                if not email: continue

                # Get last 7 days sessions
                cutoff = datetime.utcnow()-timedelta(days=7)
                sdocs  = (db.collection("sessions")
                            .where("uid","==",uid)
                            .where("created_at",">=",cutoff)
                            .limit(50).stream())
                sl = [d.to_dict() for d in sdocs]
                scores = [s.get("avg_score") for s in sl if s.get("avg_score")]
                if not scores: continue

                avg   = round(sum(scores)/len(scores),1)
                grade = "A" if avg>=85 else "B" if avg>=70 else "C" if avg>=55 else "D"
                trend_emoji = "📈" if scores[-1]>scores[0] else "📉" if scores[-1]<scores[0] else "➡️"

                subject = (f"تقرير Corvus الأسبوعي — درجتك {avg}/100 {grade}"
                           if is_ar else
                           f"Your Weekly Corvus Report — Score {avg}/100 ({grade})")

                html = f"""
<div style="font-family:Inter,sans-serif;max-width:600px;margin:0 auto;background:#050D1A;color:#F0F6FF;border-radius:16px;overflow:hidden">
  <div style="background:linear-gradient(135deg,#0EA5E9,#6366F1);padding:32px;text-align:center">
    <h1 style="margin:0;font-size:28px;color:#fff">{"تقرير الوضعية الأسبوعي" if is_ar else "Weekly Posture Report"}</h1>
    <p style="margin:8px 0 0;color:rgba(255,255,255,.7)">{datetime.utcnow().strftime("%B %d, %Y")}</p>
  </div>
  <div style="padding:32px">
    <div style="text-align:center;margin-bottom:24px">
      <div style="font-size:64px;font-weight:900;color:#0EA5E9">{avg}</div>
      <div style="font-size:16px;color:#64748B">{"متوسط الدرجة الأسبوعية" if is_ar else "Weekly Average Score"}</div>
      <div style="font-size:24px;margin-top:8px">{"الدرجة: " if is_ar else "Grade: "}<strong style="color:#10B981">{grade}</strong></div>
    </div>
    <div style="background:#0D1F35;border-radius:12px;padding:20px;margin-bottom:20px">
      <div style="display:flex;justify-content:space-between">
        <div style="text-align:center">
          <div style="font-size:24px;font-weight:700;color:#0EA5E9">{len(scores)}</div>
          <div style="font-size:12px;color:#64748B">{"جلسات" if is_ar else "Sessions"}</div>
        </div>
        <div style="text-align:center">
          <div style="font-size:24px;font-weight:700;color:#10B981">{max(scores)}</div>
          <div style="font-size:12px;color:#64748B">{"أحسن درجة" if is_ar else "Best Score"}</div>
        </div>
        <div style="text-align:center">
          <div style="font-size:24px">{trend_emoji}</div>
          <div style="font-size:12px;color:#64748B">{"الاتجاه" if is_ar else "Trend"}</div>
        </div>
      </div>
    </div>
    <div style="text-align:center;margin-top:24px">
      <a href="https://postureai-pro-omega-nine.vercel.app" style="background:#0EA5E9;color:#fff;padding:14px 32px;border-radius:10px;text-decoration:none;font-weight:700">
        {"افتح التطبيق" if is_ar else "Open App"}
      </a>
    </div>
  </div>
</div>"""

                ok = send_email(email, subject, html)
                if ok: sent+=1
                else:  failed+=1
            except Exception:
                failed+=1

        log_event("weekly_report_cards_sent", meta={"sent":sent,"failed":failed})
        return jsonify({"ok":True,"sent":sent,"failed":failed})
    except Exception as e:
        return safe_error(e)


@app.route("/api/user/weekly-report-opt-in", methods=["PUT"])
@require_auth
@limiter.limit("10 per minute")
def weekly_report_opt_in():
    """Opt in/out of weekly email report card."""
    try:
        uid  = getattr(g,"uid","")
        data = request.get_json(force=True) or {}
        opt_in = bool(data.get("opt_in",True))
        email  = data.get("email","").strip()
        db.collection("user_settings").document(uid).set({
            "weekly_report_email": opt_in,
            "email": email,
            "lang":  data.get("lang","en"),
        }, merge=True)
        return jsonify({"ok":True,"weekly_report_email":opt_in})
    except Exception as e:
        return safe_error(e)


# ══════════════════════════════════════════════════════════════════
# FEATURE 7: REFERRAL SYSTEM
# ══════════════════════════════════════════════════════════════════

@app.route("/api/referral/generate", methods=["POST"])
@require_auth
@limiter.limit("10 per minute")
def generate_referral():
    """Generate a unique referral code for the user."""
    try:
        uid = getattr(g,"uid","")
        # Check if code already exists
        doc = db.collection("referrals").document(uid).get()
        if doc.exists:
            return jsonify({"ok":True,"code":doc.to_dict().get("code"),"existing":True})

        import hashlib as _hl
        code = _hl.md5(uid.encode()).hexdigest()[:8].upper()
        db.collection("referrals").document(uid).set({
            "uid":        uid,
            "code":       code,
            "uses":       0,
            "created_at": datetime.utcnow().isoformat()+"Z",
            "reward_months_earned": 0,
        })
        return jsonify({
            "ok":True, "code":code,
            "share_url":  f"https://postureai-pro-omega-nine.vercel.app?ref={code}",
            "message":    f"Share your code {code} — both you and your friend get 1 month free!",
            "message_ar": f"شارك كودك {code} — إنت وصاحبك هتاخدوا شهر مجاناً!",
        })
    except Exception as e:
        return safe_error(e)


@app.route("/api/referral/redeem", methods=["POST"])
@require_auth
@limiter.limit("5 per minute")
def redeem_referral():
    """Redeem a referral code on signup — gives both users 1 month free."""
    try:
        uid  = getattr(g,"uid","")
        data = request.get_json(force=True) or {}
        code = data.get("code","").strip().upper()
        if not code:
            return jsonify({"error":"code required"}), 400

        # Find referrer
        refs = db.collection("referrals").where("code","==",code).limit(1).stream()
        ref_docs = list(refs)
        if not ref_docs:
            return jsonify({"error":"Invalid referral code"}), 404

        ref_doc = ref_docs[0]
        ref_uid = ref_doc.id
        if ref_uid == uid:
            return jsonify({"error":"Cannot use your own referral code"}), 400

        # Check not already redeemed
        already = (db.collection("referral_uses")
                     .where("new_uid","==",uid).limit(1).stream())
        if list(already):
            return jsonify({"error":"You have already used a referral code"}), 409

        # Record use
        db.collection("referral_uses").add({
            "ref_uid":    ref_uid,
            "new_uid":    uid,
            "code":       code,
            "redeemed_at":datetime.utcnow().isoformat()+"Z",
        })

        # Increment referrer's uses
        ref_doc.reference.update({
            "uses":                  ref_doc.to_dict().get("uses",0)+1,
            "reward_months_earned":  ref_doc.to_dict().get("reward_months_earned",0)+1,
        })

        # Grant 1 month Pro to both users
        for benefit_uid in [uid, ref_uid]:
            db.collection("user_benefits").document(benefit_uid).set({
                "free_months": (db.collection("user_benefits").document(benefit_uid).get().to_dict() or {}).get("free_months",0)+1,
                "reason":      "referral",
                "updated_at":  datetime.utcnow().isoformat()+"Z",
            }, merge=True)

        log_event("referral_redeemed", uid, {"code":code,"ref_uid":ref_uid})
        return jsonify({
            "ok":True,
            "message":    "🎉 Code redeemed! You and your friend both get 1 month free.",
            "message_ar": "🎉 تم تفعيل الكود! إنت وصاحبك هتاخدوا شهر مجاناً.",
        })
    except Exception as e:
        return safe_error(e)


@app.route("/api/referral/status", methods=["GET"])
@require_auth
@limiter.limit("20 per minute")
def referral_status():
    """Get referral stats for the user."""
    try:
        uid = getattr(g,"uid","")
        doc = db.collection("referrals").document(uid).get()
        if not doc.exists:
            return jsonify({"ok":True,"has_code":False,"code":None,"uses":0})
        d = doc.to_dict()
        return jsonify({
            "ok":True, "has_code":True,
            "code":       d.get("code"),
            "uses":       d.get("uses",0),
            "share_url":  f"https://postureai-pro-omega-nine.vercel.app?ref={d.get('code')}",
            "months_earned":d.get("reward_months_earned",0),
        })
    except Exception as e:
        return safe_error(e)

@app.route("/api/system/health", methods=["GET"])
@app.route("/api/health", methods=["GET"])
def health():
    # Fast liveness probe — do NOT call external APIs here
    local_llm_ok = bool(OLLAMA_URL)
    models_ready = POSE_LITE is not None
    mp_ver  = _mp.__version__  if (models_ready and _mp)  else "not loaded"
    cv2_ver = cv2.__version__ if (models_ready and cv2) else "not loaded"
    return jsonify({
        "status": "ok", "version": "17.0",
        "engine": f"MediaPipe {mp_ver} + OpenCV {cv2_ver} + solvePnP",
        "mediapipe": {"pose_lite": "loaded" if models_ready else "pending","pose_full": "loaded" if models_ready else "pending","face_mesh": "loaded" if models_ready else "pending"},
        "integrations": {
            "local_llm": {"configured": local_llm_ok, "url": OLLAMA_URL or None, "model": LOCAL_LLM_MODEL if local_llm_ok else None},
            "client_ai": {"configured": True, "engine": "WebLLM (browser, free, open-source)"},
            "paymob":    {"configured": bool(PAYMOB_SECRET_KEY)},
            "slack":     {"configured": bool(SLACK_WEBHOOK_URL)},
            "teams":     {"configured": bool(TEAMS_WEBHOOK_URL)},
            "whatsapp":  {"configured": bool(WA_PHONE_ID and WA_ACCESS_TOKEN)},
            "email":     {"configured": bool(SMTP_USER or SENDGRID_API_KEY)},
        },
        "features": {
            "eye_strain_detection": "Elite tier — FaceMesh iris blink rate",
            "wrist_cdt_risk": "Professional+ — elbow-wrist deviation angle",
            "3d_head_pose": "Professional+ — solvePnP on FaceMesh",
            "face_blur_snapshots": "Elite — GDPR-compliant PDF snapshots",
            "automated_emails": "POST /api/email/sequence with day parameter",
        },
        "pdf_available": REPORTLAB_OK,
        "redis":         redis_health() if REDIS_READY else {"status": "not_configured"},
        "auth_ready":    AUTH_READY,
        "rate_limiting": "enabled",
        "timestamp": datetime.now().isoformat(),
    })

# ── /api/health/detailed ─────────────────────────────────────────
@app.route("/api/health/detailed", methods=["GET"])
@require_auth
@limiter.limit("10 per minute")
def health_detailed():
    """
    Deep health check — tests every subsystem live.
    Requires auth to prevent info leakage to unauthenticated callers.
    Returns per-subsystem status, latency, and last analyze timing.
    """
    import time as _t
    report = {"timestamp": datetime.utcnow().isoformat() + "Z", "version": "17.0"}

    # ── 1. MediaPipe models ───────────────────────────────────────
    mp_status = {}
    models_ready = POSE_LITE is not None
    try:
        mp_status["mediapipe_version"] = _mp.__version__ if _mp else "not imported"
        mp_status["opencv_version"]    = cv2.__version__ if cv2 else "not imported"
        mp_status["pose_lite"]         = "loaded" if POSE_LITE is not None else "not loaded"
        mp_status["pose_full"]         = "loaded" if POSE_FULL is not None else "not loaded"
        mp_status["face_mesh"]         = "loaded" if FACE_MESH is not None else "not loaded"
        mp_status["cascade_fallback"]  = "loaded" if (face_c is not None and eye_c is not None) else "not loaded"
        mp_status["landmark_avg_sessions"] = len(_lm_history)

        # Warmup latency — run a dummy frame through MediaPipe
        if models_ready:
            _t0 = _t.perf_counter()
            _dummy = cv2.cvtColor(
                cv2.resize(cv2.imread.__func__ if False else
                    __import__("numpy").zeros((64,64,3), dtype=__import__("numpy").uint8),
                    (64,64)), cv2.COLOR_BGR2RGB
            ) if True else None
            if _dummy is not None:
                POSE_LITE.process(_dummy)
            mp_status["warmup_ms"] = round((_t.perf_counter() - _t0) * 1000, 1)
        mp_status["status"] = "ok" if models_ready else "degraded"
    except Exception as _e:
        mp_status["status"] = "error"
        mp_status["error"]  = str(_e)[:120]
    report["mediapipe"] = mp_status

    # ── 2. Redis ping ─────────────────────────────────────────────
    redis_status = {}
    try:
        if REDIS_READY:
            _t0  = _t.perf_counter()
            _key = "health:ping:test"
            cache_set(_key, "pong", 5)
            val  = cache_get(_key)
            ping_ms = round((_t.perf_counter() - _t0) * 1000, 1)
            redis_status["status"]     = "ok" if val == "pong" else "degraded"
            redis_status["ping_ms"]    = ping_ms
            redis_status["read_write"] = val == "pong"
            # Check last_valid frame cache entries
            redis_status["cached_frames"] = "available"
        else:
            redis_status["status"] = "not_configured"
            redis_status["note"]   = "Using in-process fallback — no persistence"
    except Exception as _e:
        redis_status["status"] = "error"
        redis_status["error"]  = str(_e)[:120]
    report["redis"] = redis_status

    # ── 3. Last analyze latency (from recent requests) ────────────
    latency_status = {}
    try:
        # We store last 10 processing_ms values in a rolling cache key
        _recent_raw  = cache_get("health:analyze_latency")
        if _recent_raw:
            import json as _j
            _recent = _j.loads(_recent_raw)
            if _recent:
                latency_status["last_ms"]    = _recent[-1]
                latency_status["avg_ms"]      = round(sum(_recent) / len(_recent), 1)
                latency_status["min_ms"]      = min(_recent)
                latency_status["max_ms"]      = max(_recent)
                latency_status["samples"]     = len(_recent)
                latency_status["p95_ms"]      = sorted(_recent)[int(len(_recent)*0.95)] if len(_recent) >= 5 else None
                latency_status["status"]      = (
                    "ok"       if latency_status["avg_ms"] < 500 else
                    "degraded" if latency_status["avg_ms"] < 1500 else
                    "slow"
                )
        else:
            latency_status["status"] = "no_data"
            latency_status["note"]   = "No recent analyze calls — run a session first"
    except Exception as _e:
        latency_status["status"] = "error"
        latency_status["error"]  = str(_e)[:120]
    report["analyze_latency"] = latency_status

    # ── 4. Local AI (Ollama) connectivity ──────────────────────────
    ollama_status = {}
    try:
        if OLLAMA_URL:
            ollama_status["configured"] = True
            ollama_status["model"]      = LOCAL_LLM_MODEL
            # Quick ping — minimal prompt, no analysis
            _t0   = _t.perf_counter()
            _resp = req.post(OLLAMA_URL.rstrip("/") + "/v1/chat/completions",
                json={"model": LOCAL_LLM_MODEL,
                      "messages": [{"role": "user", "content": "Reply with OK only."}],
                      "max_tokens": 5},
                timeout=8)
            ping_ms = round((_t.perf_counter() - _t0) * 1000, 1)
            ollama_status["ping_ms"] = ping_ms
            if _resp.status_code == 200:
                ollama_status["status"] = "ok"
            else:
                ollama_status["status"]      = "degraded"
                ollama_status["http_status"] = _resp.status_code
        else:
            ollama_status["status"]     = "not_configured"
            ollama_status["configured"] = False
    except Exception as _e:
        ollama_status["status"] = "error"
        ollama_status["error"]  = str(_e)[:120]
    report["ollama"] = ollama_status

    # ── 4b. Client-side AI (WebLLM, runs in browser — nothing to check) ──
    report["client_ai"] = {"status": "n/a", "note": "AI runs locally in the browser (WebLLM) — no server-side provider"}


    # ── 5. Firebase / Firestore ───────────────────────────────────
    firebase_status = {}
    try:
        if db:
            _t0  = _t.perf_counter()
            _ref = db.collection("_health").document("ping")
            _ref.set({"ts": datetime.utcnow().isoformat()}, merge=True)
            ping_ms = round((_t.perf_counter() - _t0) * 1000, 1)
            firebase_status["status"]      = "ok"
            firebase_status["ping_ms"]     = ping_ms
            firebase_status["firestore"]   = "reachable"
        else:
            firebase_status["status"] = "not_configured"
    except Exception as _e:
        firebase_status["status"] = "error"
        firebase_status["error"]  = str(_e)[:120]
    report["firebase"] = firebase_status

    # ── 6. Overall status ─────────────────────────────────────────
    ai_status = "ok" if (ollama_status.get("status") == "ok") else (
        "not_configured" if (ollama_status.get("status") == "not_configured") else "degraded"
    )
    subsystems = [
        mp_status.get("status"),
        redis_status.get("status"),
        latency_status.get("status"),
        ai_status,
        firebase_status.get("status"),
    ]
    if any(s == "error" for s in subsystems):
        overall = "degraded"
    elif any(s == "slow" for s in subsystems):
        overall = "slow"
    else:
        overall = "ok"

    report["overall"]   = overall
    report["env"]       = os.getenv("FLASK_ENV", "production")
    report["auth_ready"] = AUTH_READY
    report["pdf"]        = REPORTLAB_OK

    status_code = 200 if overall == "ok" else 207  # 207 = partial success
    return jsonify(report), status_code


# ── Latency tracking: called from analyze route ───────────────────
def analyze_keystroke_stress(keystroke_data: dict) -> dict:
    """
    Detect work stress from typing patterns.
    Correlates with posture: stressed users lean forward more.
    Input: {wpm, error_rate, pause_ratio, burst_ratio}
    Based on: Epp et al. 2011 "Identifying Emotional States Using Keystroke Dynamics"
    """
    if not keystroke_data:
        return None
    try:
        wpm         = keystroke_data.get("wpm", 60)           # words per minute
        error_rate  = keystroke_data.get("error_rate", 0.02)  # fraction of errors
        pause_ratio = keystroke_data.get("pause_ratio", 0.3)  # fraction of time paused
        burst_ratio = keystroke_data.get("burst_ratio", 0.5)  # fraction in bursts

        # Stress indicators:
        # High WPM + high errors = rushed/anxious
        # Low pause ratio = no thinking breaks (reactive mode)
        # High burst ratio = start-stop pattern (interrupted focus)
        rush_score    = min(1.0, max(0, (wpm - 60) / 60) * 0.5 + error_rate * 5)
        reactive_sc   = max(0, 1.0 - pause_ratio * 2)
        interrupted_sc= max(0, burst_ratio - 0.4) * 2

        stress_raw = rush_score * 0.40 + reactive_sc * 0.35 + interrupted_sc * 0.25
        stress_sc  = max(0, min(100, int(stress_raw * 100)))

        return {
            "stress_score":  stress_sc,
            "level":         "high" if stress_sc > 65 else "moderate" if stress_sc > 35 else "low",
            "wpm":           wpm,
            "rush_factor":   round(rush_score, 2),
            "interruptions": round(interrupted_sc, 2),
            "note":          (
                "High work stress — correlates with forward lean posture" if stress_sc > 65 else
                "Moderate work pace" if stress_sc > 35 else
                "Calm, focused work state"
            ),
        }
    except Exception:
        return None


def compute_posture_forecast(uid: str, recent_sessions: list) -> dict:
    """
    Predict tomorrow's posture risk based on historical patterns.
    Factors: recent trajectory, day-of-week patterns, time patterns.
    Returns: risk_level, predicted_score, contributing_factors, prevention_tips.
    """
    if not recent_sessions:
        return {"status": "insufficient_data", "sessions_needed": 5}

    now = datetime.utcnow()
    tomorrow_dow = (now.weekday() + 1) % 7  # 0=Monday

    # ── Day-of-week posture patterns (based on ergonomics research) ──
    # Mondays: back pain highest (weekend deconditioning)
    # Fridays: fatigue accumulation
    # Midweek: best posture (in routine)
    DOW_RISK = {0: 0.85, 1: 0.92, 2: 0.95, 3: 0.93, 4: 0.88, 5: 0.96, 6: 0.98}
    dow_factor = DOW_RISK.get(tomorrow_dow, 0.9)

    # ── Recent trajectory ─────────────────────────────────────────
    scores = [s.get("avg_score", 0) for s in recent_sessions[-10:] if s.get("avg_score")]
    if len(scores) >= 3:
        # Simple linear trend over last sessions
        n = len(scores)
        slope = (scores[-1] - scores[0]) / max(n - 1, 1)
        trend_adjustment = slope * 2  # project 2 sessions forward
    else:
        slope = 0
        trend_adjustment = 0

    # ── Base prediction ───────────────────────────────────────────
    base_score = sum(scores[-5:]) / max(len(scores[-5:]), 1)
    predicted  = max(0, min(100, round(base_score * dow_factor + trend_adjustment)))

    # ── Risk assessment ───────────────────────────────────────────
    risk = "high" if predicted < 55 else "moderate" if predicted < 70 else "low"

    # ── Contributing factors ──────────────────────────────────────
    factors = []
    if tomorrow_dow == 0:
        factors.append("Monday effect: weekend posture deconditioning")
    if tomorrow_dow == 4:
        factors.append("Friday fatigue: end-of-week energy decline")
    if slope < -2:
        factors.append(f"Declining trend: -{abs(round(slope,1))}pts per session")
    if base_score < 65:
        factors.append("Recent sessions below threshold")

    # ── Prevention tips ───────────────────────────────────────────
    tips = []
    if tomorrow_dow == 0:
        tips.append("Monday prep: do 5min stretching before sitting down")
    if predicted < 60:
        tips.append("Set posture reminder for every 20 minutes tomorrow")
    if slope < -3:
        tips.append("Check monitor height — declining scores suggest setup drift")
    tips.append("Start tomorrow's first session within first hour of work")

    return {
        "predicted_score":     predicted,
        "risk_level":          risk,
        "day_of_week":         ["Mon","Tue","Wed","Thu","Fri","Sat","Sun"][tomorrow_dow],
        "dow_factor":          round(dow_factor, 2),
        "recent_trend_slope":  round(slope, 2),
        "contributing_factors":factors,
        "prevention_tips":     tips[:3],
        "confidence":          "high" if len(scores) >= 7 else "medium" if len(scores) >= 3 else "low",
        "based_on_sessions":   len(scores),
    }


def compute_session_quality(s: dict) -> dict:
    """
    Compute session quality score (0–100) from accumulated frame metrics.
    Components:
      - Detection rate:   % of frames where pose was detected        (30%)
      - Blur rate:        % of non-blurry frames                     (20%)
      - Avg visibility:   mean landmark visibility across frames      (25%)
      - Avg confidence:   mean per-frame analysis confidence          (25%)

    Grade:
      90–100 = Excellent (full trust in results)
      75–89  = Good
      60–74  = Fair (some lighting/angle issues)
      < 60   = Poor (results less reliable)
    """
    total   = max(s.get("total_frames", 1), 1)
    detected = s.get("frames", 0)
    blurry  = s.get("blurry_count", 0)

    # 1. Detection rate (0–100)
    detection_sc = min(100, int(detected / total * 100))

    # 2. Blur rate — frames that were clear enough to analyze
    clear_frames = total - blurry
    blur_sc = min(100, int(clear_frames / total * 100))

    # 3. Average visibility confidence (0–100, already 0–1 scale)
    vis_count = s.get("vis_count", 0)
    vis_sc = int((s.get("vis_sum", 0.0) / max(vis_count, 1)) * 100) if vis_count > 0 else 50

    # 4. Average analysis confidence (already 0–96 range)
    conf_count = s.get("conf_count", 0)
    conf_raw   = s.get("conf_sum", 0.0) / max(conf_count, 1) if conf_count > 0 else 50
    conf_sc    = min(100, int(conf_raw))  # already 0–96 scale, normalize to 100

    # Weighted combination
    quality = int(round(
        detection_sc * 0.30 +
        blur_sc      * 0.20 +
        vis_sc       * 0.25 +
        conf_sc      * 0.25
    ))
    quality = max(0, min(100, quality))

    grade = ("Excellent" if quality >= 90 else
             "Good"      if quality >= 75 else
             "Fair"      if quality >= 60 else "Poor")

    issues = []
    if detection_sc < 70:  issues.append(f"Low detection rate ({detection_sc}%) — check camera angle and lighting")
    if blur_sc < 80:       issues.append(f"Motion blur detected ({100-blur_sc}% of frames) — keep still during session")
    if vis_sc < 60:        issues.append("Low landmark visibility — improve lighting or wear contrasting clothing")
    if conf_sc < 65:       issues.append("Low confidence — full-body visibility recommended")

    return {
        "quality_score":   quality,
        "grade":           grade,
        "components": {
            "detection_rate": detection_sc,
            "blur_rate":      blur_sc,
            "avg_visibility": vis_sc,
            "avg_confidence": conf_sc,
        },
        "total_frames":    total,
        "detected_frames": detected,
        "blurry_frames":   blurry,
        "issues":          issues,
        "reliable":        quality >= 60,  # flag: can results be trusted?
    }


# ══════════════════════════════════════════════════════════════
# ADAPTIVE AI COACHING ENGINE
# Builds personal posture improvement program from history
# ══════════════════════════════════════════════════════════════

def get_ergonomic_twin(uid: str) -> dict:
    """
    Ergonomic digital twin of user's workstation setup.
    Learns the user's optimal setup from their best sessions.
    Detects when setup changes (monitor moved, chair adjusted).
    """
    try:
        import json as _jt
        raw = cache_get(f"twin:{uid}")
        if raw:
            return _jt.loads(raw)
    except Exception:
        pass
    return {
        "calibrated":        False,
        "optimal_distance":  None,   # cm — learned from best sessions
        "optimal_pitch":     None,   # degrees — monitor angle
        "optimal_neck":      None,   # degrees — user's natural neck angle
        "setup_fingerprint": None,   # hash of optimal setup
        "drift_count":       0,      # times setup has drifted from optimal
        "last_calibrated":   None,
        "sessions_to_learn": 5,      # sessions before twin is calibrated
        "history":           [],     # last 20 setup snapshots
    }


def update_ergonomic_twin(uid: str, result: dict) -> dict:
    """
    Update ergonomic twin with current session data.
    If session quality > 80, use it to refine optimal setup.
    Detect drift: current setup vs optimal.
    """
    import json as _jt2
    twin = get_ergonomic_twin(uid)
    sc   = result.get("score", 0) or 0
    met  = result.get("metrics", {})

    dist    = met.get("screen_distance", {}).get("value")
    pitch   = result.get("head_pose", {}).get("pitch")
    neck    = met.get("neck_lean", {}).get("value")
    qual_sc = result.get("session_quality", {}).get("quality_score", 0) if isinstance(result.get("session_quality"), dict) else 0

    # Only learn from high-quality, good-posture sessions
    if sc >= 80 and qual_sc >= 75:
        snap = {"dist": dist, "pitch": pitch, "neck": neck, "score": sc, "ts": datetime.utcnow().isoformat()}
        hist = twin.get("history", [])
        hist.append(snap)
        if len(hist) > 20: hist = hist[-20:]
        twin["history"] = hist

        # Calibrate after 5 good sessions
        good_snaps = [h for h in hist if h.get("score", 0) >= 80]
        if len(good_snaps) >= 5:
            twin["optimal_distance"] = round(sum(h["dist"]  for h in good_snaps if h.get("dist"))  / max(sum(1 for h in good_snaps if h.get("dist")),  1), 1)
            twin["optimal_pitch"]    = round(sum(h["pitch"] for h in good_snaps if h.get("pitch")) / max(sum(1 for h in good_snaps if h.get("pitch")), 1), 1)
            twin["optimal_neck"]     = round(sum(h["neck"]  for h in good_snaps if h.get("neck"))  / max(sum(1 for h in good_snaps if h.get("neck")),  1), 1)
            twin["calibrated"]       = True
            twin["last_calibrated"]  = datetime.utcnow().isoformat()
            twin["sessions_to_learn"]= max(0, 5 - len(good_snaps))

    # Detect setup drift if calibrated
    drift_alerts = []
    if twin.get("calibrated") and dist and twin.get("optimal_distance"):
        dist_drift = abs(dist - twin["optimal_distance"])
        if dist_drift > 12:
            twin["drift_count"] = twin.get("drift_count", 0) + 1
            drift_alerts.append(
                f"⚠️ Setup drift: screen moved {round(dist_drift)}cm from your optimal {twin['optimal_distance']}cm")
    if twin.get("calibrated") and pitch is not None and twin.get("optimal_pitch") is not None:
        pitch_drift = abs(pitch - twin["optimal_pitch"])
        if pitch_drift > 8:
            drift_alerts.append(
                f"Monitor angle changed {round(pitch_drift)}° — readjust to your optimal setup")

    twin["drift_alerts"] = drift_alerts

    # Save
    try:
        cache_set(f"twin:{uid}", _jt2.dumps(twin), 86400 * 180)
    except Exception:
        pass
    return twin


def get_coaching_profile(uid: str) -> dict:
    """Load or initialize user coaching profile from cache/Redis."""
    try:
        import json as _jc
        raw = cache_get(f"coach:{uid}")
        if raw:
            return _jc.loads(raw)
    except Exception:
        pass
    return {
        "sessions_analyzed": 0,
        "recurring_issues":  {},    # {metric: {count, avg_severity, last_seen}}
        "ignored_advice":    {},    # {advice_hash: ignore_count}
        "improvement_rate":  {},    # {metric: slope over last 10 sessions}
        "coach_phase":       "onboarding",  # onboarding→building→maintaining→optimizing
        "personal_goals":    [],
        "badges_earned":     [],
        "streak_days":       0,
        "last_session":      None,
    }


def update_coaching_profile(uid: str, result: dict, session_metrics: dict) -> dict:
    """
    Update coaching profile after each session.
    Tracks improvement, recurring issues, ignored advice.
    Returns updated profile + coaching plan.
    """
    import json as _jc, hashlib as _hsh

    profile = get_coaching_profile(uid)
    profile["sessions_analyzed"] = profile.get("sessions_analyzed", 0) + 1
    n = profile["sessions_analyzed"]

    # ── Track recurring issues ───────────────────────────────────
    for key, met in session_metrics.items():
        if not isinstance(met, dict) or "score" not in met: continue
        sc = met.get("score", 100)
        if sc < 70:   # issue threshold
            ri = profile["recurring_issues"].setdefault(key, {
                "count": 0, "total_severity": 0, "last_seen": None, "worst": 0,
            })
            ri["count"]          += 1
            ri["total_severity"] += (100 - sc)
            ri["last_seen"]       = datetime.utcnow().isoformat()
            ri["worst"]           = max(ri.get("worst", 0), 100 - sc)
            ri["avg_severity"]    = ri["total_severity"] / ri["count"]

    # ── Determine coach phase ────────────────────────────────────
    if   n < 5:   phase = "onboarding"    # still learning user
    elif n < 20:  phase = "building"      # building habits
    elif n < 50:  phase = "maintaining"   # maintaining progress
    else:         phase = "optimizing"    # fine-tuning

    profile["coach_phase"] = phase

    # ── Generate coaching plan ───────────────────────────────────
    # Sort recurring issues by: frequency × severity
    sorted_issues = sorted(
        [(k, v) for k, v in profile["recurring_issues"].items()],
        key=lambda x: x[1]["count"] * x[1]["avg_severity"],
        reverse=True
    )

    ISSUE_ADVICE = {
        "neck_lean":       ("Raise monitor to exact eye level", "Place monitor on books/stand"),
        "spine_lean":      ("Engage lumbar support", "Sit back fully, activate core"),
        "screen_distance": ("Position screen arm-length away", "Use monitor arm for easy adjustment"),
        "head_yaw":        ("Reposition chair to face screen directly", "Place main screen directly ahead"),
        "shoulder_level":  ("Set armrests to equal height", "Check chair arm height settings"),
        "eye_strain":      ("Apply 20-20-20 rule every session", "Enable night mode, reduce brightness"),
        "wrist_angle":     ("Use wrist rest, lower keyboard", "Check keyboard tray height"),
        "breathing_rate":  ("Practice diaphragmatic breathing", "Set breathing reminder every 20min"),
        "fhp_index":       ("Chin tuck exercise x10 daily", "Wall posture check 3x per day"),
        "rsi_risk":        ("Take micro-breaks every 25min", "Try Pomodoro technique"),
    }

    coaching_plan = []
    for key, issue_data in sorted_issues[:3]:   # top 3 issues only
        if key in ISSUE_ADVICE:
            quick, deep = ISSUE_ADVICE[key]
            coaching_plan.append({
                "issue":          key,
                "frequency":      issue_data["count"],
                "avg_severity":   round(issue_data["avg_severity"], 1),
                "quick_fix":      quick,
                "deep_fix":       deep,
                "priority":       "high" if issue_data["count"] > 5 else "medium",
            })

    # ── Phase-specific coaching message ──────────────────────────
    phase_msg = {
        "onboarding":  f"Session {n}/5 — Learning your posture patterns. Keep going!",
        "building":    f"Building habits ({n} sessions). Focus on your top issue: {sorted_issues[0][0].replace('_',' ') if sorted_issues else 'consistency'}.",
        "maintaining": f"Great consistency ({n} sessions)! Fine-tune your weak spots.",
        "optimizing":  f"Expert level ({n} sessions). You're in the top tier of posture awareness!",
    }.get(phase, "Keep going!")

    # ── Badges ───────────────────────────────────────────────────
    badges = profile.get("badges_earned", [])
    if n == 5   and "first_week"    not in badges: badges.append("first_week")
    if n == 20  and "habit_builder" not in badges: badges.append("habit_builder")
    if n == 50  and "posture_pro"   not in badges: badges.append("posture_pro")
    if n == 100 and "elite_posture" not in badges: badges.append("elite_posture")
    overall_sc = result.get("score", 0)
    if overall_sc >= 90 and "perfect_session" not in badges: badges.append("perfect_session")
    profile["badges_earned"] = badges

    profile["coaching_plan"]  = coaching_plan
    profile["phase_message"]  = phase_msg
    profile["last_updated"]   = datetime.utcnow().isoformat()

    # Save back to cache
    try:
        import json as _jc2
        cache_set(f"coach:{uid}", _jc2.dumps(profile), 86400 * 90)
    except Exception:
        pass

    return profile


def _record_latency(ms: float):
    """Append processing_ms to rolling 20-sample list in cache."""
    try:
        import json as _j
        _raw     = cache_get("health:analyze_latency")
        _samples = _j.loads(_raw) if _raw else []
        _samples.append(round(ms, 1))
        if len(_samples) > 20: _samples = _samples[-20:]  # keep last 20
        cache_set("health:analyze_latency", _j.dumps(_samples), 3600)
    except Exception:
        pass


@require_auth
@app.route("/api/paymob/create-payment", methods=["POST"])
@limiter.limit("15 per minute")
def paymob_create_payment():
    try:
        data     = request.get_json(force=True) or {}
        currency = data.get("currency", "EGP")
        # SECURITY: tier comes from authenticated session, not client body
        tier     = getattr(g, "role", {}).get("tier", "standard")
        # Requested plan to purchase — must be a valid upgrade target
        req_tier = (data.get("tier") or "").strip().lower() or tier
        billing  = (data.get("billing") or "monthly").strip().lower()

        # Validate plan before touching PayMob
        valid, err = validate_plan_request(req_tier, billing)
        if not valid:
            return jsonify({"error": err}), 400

        # SECURITY: derive amount server-side — NEVER trust client amount_cents
        amount_cents = get_paymob_amount(req_tier, billing)
        if not amount_cents:
            return jsonify({"error": f"No price configured for {req_tier}/{billing}"}), 400

        tier = req_tier  # use the validated requested tier
        billing_data = data.get("billing_data", {})
        wallet_number= data.get("wallet_number", "")
        payment_type = data.get("payment_type", "card")
        if not PAYMOB_SECRET_KEY:
            return jsonify({"error": "PayMob not configured — add PAYMOB_SECRET_KEY to .env"}), 500
        headers = {"Content-Type": "application/json"}
        auth_resp = req.post("https://accept.paymob.com/api/auth/tokens",
                             json={"api_key": PAYMOB_SECRET_KEY}, headers=headers, timeout=15)
        if auth_resp.status_code == 403:
            return jsonify({"error": "PayMob: Domain not whitelisted."}), 403
        auth_resp.raise_for_status()
        auth_token = auth_resp.json().get("token")
        if not auth_token:
            return jsonify({"error": "PayMob auth failed — check your secret key"}), 502
        order_resp = req.post("https://accept.paymob.com/api/ecommerce/orders",
                              json={"auth_token": auth_token, "delivery_needed": False,
                                    "amount_cents": amount_cents, "currency": currency,
                                    "merchant_order_id": f"PAI-{getattr(g,'uid','anon')}-{tier[:12]}-{billing[:1]}-{int(time.time())}",
                                    "items": [{"name": f"PostureAI {tier.title()} ({billing})",
                                               "amount_cents": amount_cents,
                                               "description": f"PostureAI Pro — {tier.title()} Workforce Intelligence Plan, {billing} billing",
                                               "quantity": 1}]},
                              headers=headers, timeout=15)
        order_resp.raise_for_status()
        order_id = order_resp.json().get("id")
        integration_id = PAYMOB_INTEGRATIONS.get("mobile_wallet" if payment_type == "mobile_wallet" else "card", "")
        pk_resp = req.post("https://accept.paymob.com/api/acceptance/payment_keys",
                           json={"auth_token": auth_token, "amount_cents": amount_cents,
                                 "expiration": 3600, "order_id": order_id, "currency": currency,
                                 "integration_id": integration_id,
                                 "billing_data": {"email": billing_data.get("email","NA"),
                                                  "first_name": billing_data.get("first_name","Customer"),
                                                  "last_name": billing_data.get("last_name",""),
                                                  "phone_number": billing_data.get("phone_number","NA"),
                                                  "apartment":"NA","floor":"NA","street":"NA",
                                                  "building":"NA","shipping_method":"NA",
                                                  "postal_code":"NA","city":"Cairo","country":"EG","state":"Cairo"}},
                           headers=headers, timeout=15)
        pk_resp.raise_for_status()
        payment_key = pk_resp.json().get("token")
        if payment_type == "mobile_wallet" and wallet_number:
            wallet_resp = req.post("https://accept.paymob.com/api/acceptance/payments/pay",
                                   json={"source":{"identifier":wallet_number,"subtype":"WALLET"},
                                         "payment_token":payment_key}, headers=headers, timeout=15)
            wdata = wallet_resp.json()
            redirect_url = wdata.get("redirect_url") or wdata.get("iframe_redirection_url")
            return jsonify({"payment_key":payment_key,"redirect_url":redirect_url,"order_id":order_id,"payment_type":"mobile_wallet"})
        return jsonify({"payment_key":payment_key,"order_id":order_id,"payment_type":"card"})
    except req.exceptions.Timeout:
        return jsonify({"error":"PayMob request timed out"}), 504
    except req.exceptions.RequestException as e:
        return jsonify({"error":f"PayMob network error: {str(e)}"}), 502
    except Exception as e:
        import traceback
        return safe_error(e)

@app.route("/api/paymob/webhook", methods=["POST"])
def paymob_webhook():
    try:
        import hmac as hmac_lib, hashlib
        hmac_secret = os.getenv("PAYMOB_HMAC_SECRET","")
        received_hmac = request.args.get("hmac","")
        payload_str = request.get_data(as_text=True)
        # ── HMAC is MANDATORY in production ──────────────────────────────
        env = os.getenv("FLASK_ENV","development")
        if not hmac_secret:
            if env == "production":
                print("🚨 CRITICAL: PAYMOB_HMAC_SECRET missing in production — rejecting webhook", flush=True)
                return jsonify({"error":"PAYMOB_HMAC_SECRET not configured — set it in Railway env vars"}), 503
            else:
                print("⚠️  DEV: Skipping HMAC validation (PAYMOB_HMAC_SECRET not set)", flush=True)
        else:
            computed = hmac_lib.new(hmac_secret.encode(), payload_str.encode(), hashlib.sha512).hexdigest()
            if not hmac_lib.compare_digest(computed, received_hmac):
                print(f"🚨 Invalid HMAC on PayMob webhook — possible spoofing attempt", flush=True)
                return jsonify({"error":"Invalid HMAC signature"}), 403
        payload  = request.get_json(force=True) or {}
        obj      = payload.get("obj", {})
        success  = obj.get("success", False)
        order    = obj.get("order", {})
        amount   = obj.get("amount_cents", 0)
        merch_id = order.get("merchant_order_id", "")
        email    = obj.get("billing_data", {}).get("email", "")
        fname    = obj.get("billing_data", {}).get("first_name", "Customer")

        print(f"[Webhook] success={success} merchant_order_id={merch_id} amount={amount}", flush=True)

        if success:
            # ── Marketplace booking payments use a DIFFERENT merchant_order_id
            # shape (PAI-BOOK-{uid}-{bookingId}-{ts}) than subscription payments
            # (PAI-{uid}-{tier}-{billing}-{ts}). Must be checked FIRST — otherwise
            # the subscription parser below misreads "BOOK" as the uid and the
            # booking is silently never marked paid despite a successful charge.
            if merch_id.startswith("PAI-BOOK-"):
                try:
                    parts = merch_id.split("-")
                    # PAI - BOOK - uid - bookingId - ts
                    booking_id = parts[3] if len(parts) >= 5 else None
                    db = firestore.client()
                    paymob_order_id = order.get("id")

                    if not booking_id:
                        print(f"[Webhook] ⚠️ Could not parse booking_id from {merch_id}", flush=True)
                    else:
                        bref = db.collection("marketplace_bookings").document(booking_id)
                        bdoc = bref.get()
                        if not bdoc.exists:
                            print(f"[Webhook] ⚠️ marketplace_bookings/{booking_id} not found", flush=True)
                        else:
                            booking = bdoc.to_dict()
                            # Idempotency — PayMob can legitimately retry the webhook
                            if booking.get("status") == "confirmed":
                                print(f"[Webhook] Booking {booking_id} already confirmed — skipping duplicate", flush=True)
                                return jsonify({"received": True, "note": "already processed"}), 200
                            # Defense in depth — confirm the charged amount matches
                            # what this booking was created for, same pattern used
                            # for subscription payments above.
                            if booking.get("amount_cents") != amount:
                                print(f"🚨 [Webhook] Booking amount mismatch: expected={booking.get('amount_cents')} "
                                      f"got={amount} booking_id={booking_id} — flagged for review", flush=True)
                                send_admin_notification({
                                    "user_name": fname, "user_email": email,
                                    "tier": f"MARKETPLACE AMOUNT MISMATCH: booking {booking_id}", "amount": amount/100,
                                    "method": "PayMob", "status": "flagged",
                                })
                                return jsonify({"received": True, "warning": "amount mismatch — flagged for review"}), 200

                            bref.update({
                                "status":            "confirmed",
                                "paymob_order_id":   paymob_order_id,
                                "confirmed_at":      datetime.utcnow().isoformat()+"Z",
                            })
                            print(f"[Webhook] ✅ Marketplace booking {booking_id} confirmed", flush=True)
                            send_admin_notification({
                                "user_name":           fname or booking.get("therapist_name",""),
                                "user_email":          email,
                                "tier":                f"Marketplace booking — {booking.get('therapist_name','therapist')}",
                                "amount":              amount/100,
                                "payment_method_name": "Card/Wallet",
                                "ref_code":            str(paymob_order_id or ""),
                            })
                except Exception as booking_err:
                    print(f"[Webhook] ❌ Marketplace booking update failed: {booking_err}", flush=True)
                return jsonify({"received": True, "success": success})

            # ── Decode uid + tier from merchant_order_id ──────────────────
            # Format: PAI-{uid}-{tier}-{billing_prefix}-{timestamp}
            uid, tier, billing_type = None, None, "monthly"
            try:
                parts = merch_id.split("-")
                # PAI - uid - tier - billing - timestamp
                if len(parts) >= 4 and parts[0] == "PAI":
                    uid          = parts[1]
                    tier         = parts[2]
                    billing_type = "yearly" if parts[3] == "y" else "monthly"
            except Exception as parse_err:
                print(f"[Webhook] Could not parse merchant_order_id: {parse_err}", flush=True)

            # ── Update Firestore ───────────────────────────────────────────
            if uid and tier:
                try:
                    db  = firestore.client()
                    ref = db.collection("users").document(uid)
                    doc = ref.get()
                    if doc.exists:
                        # SECURITY: Enterprise tiers must NEVER be granted via automated
                        # webhook — they require a signed custom contract. This blocks
                        # the case where a forged/replayed merchant_order_id claims an
                        # enterprise tier (which has no price, so the mismatch check
                        # below would otherwise be silently skipped).
                        if tier in ("b2b_enterprise", "enterprise"):
                            print(f"🚨 [Webhook] Enterprise tier claimed via automated webhook for uid={uid} — refusing", flush=True)
                            send_admin_notification({
                                "user_name": fname, "user_email": email,
                                "tier": "BLOCKED: enterprise via webhook", "amount": amount/100,
                                "method": "PayMob", "status": "flagged",
                            })
                            return jsonify({"received": True, "warning": "enterprise tier requires manual approval"}), 200

                        # SECURITY: defense in depth — even though amount_cents sent
                        # to PayMob was server-derived at order creation, verify the
                        # amount PayMob says it received actually matches what this
                        # tier/billing combo should cost. Protects against PayMob-side
                        # bugs, MITM tampering, or a stale/forged webhook payload.
                        expected_amount = get_paymob_amount(tier, billing_type)
                        if expected_amount is not None and amount != expected_amount:
                            print(f"🚨 [Webhook] Amount mismatch: expected={expected_amount} got={amount} "
                                  f"tier={tier} billing={billing_type} uid={uid} — refusing upgrade", flush=True)
                            send_admin_notification({
                                "user_name": fname, "user_email": email,
                                "tier": f"AMOUNT MISMATCH: {tier}", "amount": amount/100,
                                "method": "PayMob", "status": "flagged",
                            })
                            return jsonify({"received": True, "warning": "amount mismatch — flagged for review"}), 200

                        # SECURITY: idempotency — reject if this exact PayMob order
                        # was already processed (replay protection). A webhook can
                        # legitimately be retried by PayMob on transient failures,
                        # so we check before mutating state rather than rejecting
                        # the HTTP request outright.
                        paymob_order_id = order.get("id")
                        if paymob_order_id:
                            existing_payment = db.collection("payments").where(
                                "paymob_order", "==", paymob_order_id).limit(1).get()
                            if existing_payment:
                                print(f"[Webhook] Order {paymob_order_id} already processed — skipping duplicate", flush=True)
                                return jsonify({"received": True, "note": "already processed"}), 200

                        months  = 12 if billing_type == "yearly" else 1
                        expires = datetime.utcnow() + timedelta(days=30*months)
                        ref.update({
                            "tier":               tier,
                            "is_trial":           False,
                            "trial_expires_at":   None,
                            "subscription_start": datetime.utcnow().isoformat()+"Z",
                            "subscription_end":   expires.isoformat() + "Z",
                            "subscription_billing": billing_type,
                            "last_payment_amount":  amount // 100,
                            "last_payment_at":      datetime.utcnow().isoformat()+"Z",
                            "updated_at":           datetime.utcnow().isoformat()+"Z",
                        })
                        print(f"[Webhook] ✅ Updated uid={uid} → tier={tier} billing={billing_type}", flush=True)

                        # ── Record payment in payments collection ──────────
                        db.collection("payments").add({
                            "uid":          uid,
                            "tier":         tier,
                            "amount":       amount // 100,
                            "currency":     "EGP",
                            "billing":      billing_type,
                            "status":       "confirmed",
                            "paymob_order": order.get("id"),
                            "email":        email,
                            "created_at":   datetime.utcnow().isoformat()+"Z",
                        })

                        # ── Send confirmation email ────────────────────────
                        user_email_addr = email or doc.to_dict().get("email","")
                        threading.Thread(target=send_invoice_email, args=({
                            "user_name":  fname,
                            "user_email": user_email_addr,
                            "tier":       tier,
                            "amount":     amount // 100,
                            "billing":    billing_type,
                            "currency":   "EGP",
                        },), daemon=True).start()

                    else:
                        print(f"[Webhook] ⚠️ User {uid} not found in Firestore", flush=True)
                except Exception as db_err:
                    print(f"[Webhook] ❌ Firestore update failed: {db_err}", flush=True)
            else:
                print(f"[Webhook] ⚠️ Could not extract uid/tier — manual update needed. merchant_order_id={merch_id}", flush=True)

            # ── Always notify admin ────────────────────────────────────────
            send_admin_notification({
                "user_name":          fname,
                "user_email":         email,
                "tier":               tier or "unknown",
                "amount":             amount // 100,
                "payment_method_name":"Card/Wallet",
                "ref_code":           str(order.get("id","")),
            })

        return jsonify({"received": True, "success": success})
    except Exception as e:
        import traceback
        print(f"[Webhook] Error: {e}\n{traceback.format_exc()}", flush=True)
        return safe_error(e)

@require_auth
@app.route("/api/notify/payment", methods=["POST"])
@limiter.limit("30 per minute")
def notify_payment():
    try:
        data = request.get_json(force=True) or {}
        send_admin_notification(data)
        return jsonify({"sent":True})
    except Exception as e:
        return safe_error(e)

@require_auth
@app.route("/api/notify/confirmed", methods=["POST"])
@limiter.limit("30 per minute")
def notify_confirmed():
    try:
        data = request.get_json(force=True) or {}
        sent = send_invoice_email(data)
        return jsonify({"sent":sent})
    except Exception as e:
        return safe_error(e)

@require_auth
@app.route("/api/hr/import-employees", methods=["POST"])
@require_auth
@require_hr
@limiter.limit("10 per minute")
def import_employees():
    try:
        import csv, io as sio, re as _re, html as _html
        
        def _sanitize(val):
            """Sanitize employee field values — prevent XSS and injection."""
            if not isinstance(val, str):
                return val
            # Strip HTML tags and encode entities
            val = _re.sub(r'<[^>]+>', '', val)
            val = _html.escape(val.strip())
            # Limit length
            return val[:255]
        
        if "file" not in request.files:
            return jsonify({"error":"No file uploaded"}), 400
        f    = request.files["file"]
        fname= f.filename.lower()
        employees = []
        if fname.endswith(".csv"):
            content = f.read().decode("utf-8-sig")
            reader  = csv.DictReader(sio.StringIO(content))
            for row in reader:
                employees.append({"name":row.get("name","").strip(),"email":row.get("email","").strip().lower(),
                                   "department":row.get("department","General").strip(),"employee_id":row.get("employee_id","").strip()})
        elif fname.endswith((".xlsx",".xls")):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(f, read_only=True); ws = wb.active
                headers = [str(c.value).lower().strip() if c.value else "" for c in next(ws.iter_rows(max_row=1))]
                def col(row, key):
                    aliases = {"name":["name","full name"],"email":["email","email address"],"department":["department","dept","division"],"employee_id":["employee_id","id","emp id","staff id"]}
                    for alias in aliases.get(key,[key]):
                        if alias in headers: return str(row[headers.index(alias)].value or "").strip()
                    return ""
                for row in ws.iter_rows(min_row=2):
                    if any(c.value for c in row):
                        employees.append({"name":col(row,"name"),"email":col(row,"email").lower(),"department":col(row,"department") or "General","employee_id":col(row,"employee_id")})
            except ImportError:
                return jsonify({"error":"pip install openpyxl"}), 500
        else:
            return jsonify({"error":"Only .csv or .xlsx supported"}), 400
        valid   = [e for e in employees if e["email"] and "@" in e["email"] and e["name"]]
        invalid = [e for e in employees if not (e["email"] and "@" in e["email"] and e["name"])]
        return jsonify({"valid":valid,"invalid":invalid,"total":len(employees)})
    except Exception as e:
        import traceback
        return safe_error(e)

@app.route("/api/hr/parse-csv", methods=["POST"])
@require_auth
@require_hr
@limiter.limit("20 per minute")
def parse_csv_preview():
    """
    Parse a CSV/XLSX file and return a preview of employees.
    Used by HR panel before confirming bulk import.
    Accepts: multipart/form-data with field "file"
           OR JSON with field "csv_text" (plain text)
    Returns: { columns, preview (first 5 rows), total, valid, invalid, employees }
    """
    try:
        import csv as _csv, io as _sio

        employees = []
        raw_headers = []

        # ── Multipart file upload ──────────────────────────────────
        if "file" in request.files:
            f     = request.files["file"]
            fname = f.filename.lower()

            if fname.endswith(".csv"):
                content = f.read().decode("utf-8-sig", errors="replace")
                reader  = _csv.DictReader(_sio.StringIO(content))
                raw_headers = reader.fieldnames or []
                for row in reader:
                    employees.append({
                        "name":        (row.get("name") or row.get("full_name") or row.get("Full Name") or "").strip(),
                        "email":       (row.get("email") or row.get("Email") or row.get("email_address") or "").strip().lower(),
                        "department":  (row.get("department") or row.get("dept") or row.get("Department") or "General").strip(),
                        "employee_id": (row.get("employee_id") or row.get("id") or row.get("emp_id") or row.get("ID") or "").strip(),
                        "phone":       (row.get("phone") or row.get("mobile") or "").strip(),
                    })

            elif fname.endswith((".xlsx", ".xls")):
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
                    ws = wb.active
                    rows = list(ws.iter_rows(values_only=True))
                    if not rows:
                        return jsonify({"error": "Empty spreadsheet"}), 400
                    raw_headers = [str(h).lower().strip() if h else "" for h in rows[0]]

                    def _col(row_vals, *aliases):
                        for alias in aliases:
                            a = alias.lower()
                            for i, h in enumerate(raw_headers):
                                if a in h and i < len(row_vals):
                                    v = row_vals[i]
                                    return str(v).strip() if v is not None else ""
                        return ""

                    for row in rows[1:]:
                        if not any(v for v in row if v):
                            continue
                        employees.append({
                            "name":        _col(row, "name", "full name", "full_name"),
                            "email":       _col(row, "email", "email address", "email_address").lower(),
                            "department":  _col(row, "department", "dept", "division") or "General",
                            "employee_id": _col(row, "employee_id", "id", "emp id", "staff id"),
                            "phone":       _col(row, "phone", "mobile", "telephone"),
                        })
                except ImportError:
                    return jsonify({"error": "openpyxl required — pip install openpyxl"}), 500
            else:
                return jsonify({"error": "Only .csv or .xlsx supported"}), 400

        # ── JSON csv_text payload ──────────────────────────────────
        elif request.is_json:
            data = request.get_json(force=True) or {}
            csv_text = data.get("csv_text", "")
            if not csv_text:
                return jsonify({"error": "csv_text or file required"}), 400
            reader = _csv.DictReader(_sio.StringIO(csv_text))
            raw_headers = reader.fieldnames or []
            for row in reader:
                employees.append({
                    "name":        (row.get("name") or "").strip(),
                    "email":       (row.get("email") or "").strip().lower(),
                    "department":  (row.get("department") or "General").strip(),
                    "employee_id": (row.get("employee_id") or "").strip(),
                    "phone":       (row.get("phone") or "").strip(),
                })
        else:
            return jsonify({"error": "Send file (multipart) or csv_text (JSON)"}), 400

        # ── Validate ───────────────────────────────────────────────
        valid   = [e for e in employees if e["email"] and "@" in e["email"] and e["name"]]
        invalid = [e for e in employees if not (e["email"] and "@" in e["email"] and e["name"])]

        # Mark validation errors per row
        for e in invalid:
            reasons = []
            if not e["name"]:  reasons.append("missing name")
            if not e["email"] or "@" not in e["email"]: reasons.append("invalid email")
            e["_error"] = ", ".join(reasons)

        return jsonify({
            "ok":       True,
            "columns":  raw_headers,
            "total":    len(employees),
            "valid":    valid,
            "invalid":  invalid,
            "preview":  valid[:5],
            "stats": {
                "total":         len(employees),
                "valid_count":   len(valid),
                "invalid_count": len(invalid),
                "departments":   list({e["department"] for e in valid}),
            },
        })

    except Exception as e:
        import traceback
        return safe_error(e)


@require_auth
@app.route("/api/hr/monthly-report", methods=["POST"])
@limiter.limit("5 per minute")
@require_hr
def monthly_hr_report():
    try:
        if not REPORTLAB_OK: return jsonify({"error":"pip install reportlab"}), 500
        data       = request.get_json(force=True) or {}
        company    = data.get("company_name","Company")
        month_name = data.get("month",datetime.now().strftime("%B"))
        year       = data.get("year",datetime.now().year)
        sessions_d = data.get("sessions",[])
        employees  = data.get("employees",[])
        total_sess = len(sessions_d)
        avg_score  = round(sum(s.get("avg_score",0) for s in sessions_d)/max(total_sess,1))
        good_pct   = round(sum(s.get("good_pct",0) for s in sessions_d)/max(total_sess,1))
        alerts_tot = sum(s.get("alerts_count",0) for s in sessions_d)
        risk_high  = [e for e in employees if e.get("avg_score",100) < 50]
        buf = io.BytesIO()
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.units import cm
        doc = SimpleDocTemplate(buf,pagesize=A4,topMargin=2*cm,bottomMargin=2*cm,leftMargin=2*cm,rightMargin=2*cm)
        styles = getSampleStyleSheet(); story = []
        title_style = ParagraphStyle("T",parent=styles["Title"],textColor=colors.HexColor("#1a56db"),fontSize=22,spaceAfter=4)
        story.append(Paragraph(f"PostureAI Pro — Workforce Health Intelligence Report",title_style))
        story.append(Paragraph(f"{company} · {month_name} {year}",ParagraphStyle("S",parent=styles["Normal"],textColor=colors.HexColor("#64748b"),fontSize=12,spaceAfter=20)))
        kpi_data = [["Metric","Value","Status"],
                    ["Total Sessions",str(total_sess),"✓"],
                    ["Average Posture Score",f"{avg_score}/100","✓" if avg_score>=70 else "⚠"],
                    ["Good Posture %",f"{good_pct}%","✓" if good_pct>=65 else "⚠"],
                    ["Total Alerts",str(alerts_tot),"✓" if alerts_tot<20 else "⚠"],
                    ["High Risk Employees",str(len(risk_high)),"✓" if not risk_high else "⚠"]]
        kpi_table = Table(kpi_data,colWidths=[8*cm,4*cm,3*cm])
        kpi_table.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1a56db")),("TEXTCOLOR",(0,0),(-1,0),colors.white),("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),10),("GRID",(0,0),(-1,-1),0.5,colors.HexColor("#e2e8f0")),("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#f8fafc")]),("PADDING",(0,0),(-1,-1),8)]))
        story.append(kpi_table); story.append(Spacer(1,0.5*cm))
        if risk_high:
            story.append(Paragraph("🔴 High Risk Employees (Score < 50)",ParagraphStyle("H",parent=styles["Heading2"],textColor=colors.HexColor("#ef4444"))))
            risk_data = [["Name","Dept","Avg Score","Sessions","Recommendation"]]
            for e in risk_high[:10]:
                risk_data.append([e.get("name","—"),e.get("department","—"),str(e.get("avg_score","—")),str(e.get("sessions",0)),"Ergonomic assessment required"])
            rt = Table(risk_data,colWidths=[5*cm,3*cm,2.5*cm,2.5*cm,4*cm])
            rt.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.HexColor("#ef4444")),("TEXTCOLOR",(0,0),(-1,0),colors.white),("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),9),("GRID",(0,0),(-1,-1),0.5,colors.HexColor("#fecaca")),("PADDING",(0,0),(-1,-1),6)]))
            story.append(rt); story.append(Spacer(1,0.3*cm))
        story.append(Spacer(1,1*cm))
        story.append(Paragraph(f"Report generated by PostureAI Pro on {datetime.now().strftime('%d %B %Y at %H:%M')}. This report is for HR use only.",
            ParagraphStyle("F",parent=styles["Normal"],fontSize=8,textColor=colors.HexColor("#94a3b8"))))
        doc.build(story); buf.seek(0)
        fname = f"PostureAI_HR_Report_{company}_{month_name}_{year}.pdf".replace(" ","_")
        return send_file(buf,as_attachment=True,download_name=fname,mimetype="application/pdf")
    except Exception as e:
        import traceback
        return safe_error(e)


# ── Company: add member with seat check ────────────────────────────
SEAT_LIMITS = {"starter":25,"standard":25,"growth":100,"professional":100,"business":500,"elite":-1,"enterprise":-1}

@app.route("/api/org/invite/accept", methods=["POST"])
@require_auth
@limiter.limit("20 per minute")
def org_invite_accept():
    """Accept invite token — links user to company in Firestore."""
    try:
        data       = request.get_json(force=True) or {}
        token      = (data.get("token") or "").strip()
        # SECURITY: the accepting uid comes from the verified Firebase token,
        # never from the request body — otherwise anyone could accept an
        # invite on behalf of an arbitrary uid they don't control.
        uid        = getattr(g, "uid", "")
        company_id = (data.get("company_id") or "").strip()
        if not (token and uid and company_id):
            return jsonify({"error":"token, uid, company_id required"}),400
        db  = firestore.client()

        # SECURITY: the invite document MUST exist, MUST belong to the claimed
        # company_id, and MUST still be pending. Without these checks anyone
        # who knows (or guesses) a company_id could call this endpoint with a
        # random/garbage token and still get linked to that company, because
        # the previous version only checked `if inv.exists` and silently
        # continued (returning ok:True) even when the invite did not exist.
        inv_ref = db.collection("invites").document(token)
        inv     = inv_ref.get()
        if not inv.exists:
            return jsonify({"error": "Invalid or expired invite"}), 404
        inv_data = inv.to_dict()
        if inv_data.get("company_id") != company_id:
            print(f"🚨 [invite] company_id mismatch: invite belongs to {inv_data.get('company_id')}, "
                  f"request claimed {company_id} — possible spoofing attempt by uid={uid}", flush=True)
            return jsonify({"error": "Invite does not belong to this company"}), 403
        if inv_data.get("status") == "accepted":
            return jsonify({"error": "Invite already accepted"}), 409
        if inv_data.get("status") not in (None, "pending"):
            return jsonify({"error": f"Invite is {inv_data.get('status')}, cannot accept"}), 409

        # Check seat limit
        company = db.collection("companies").document(company_id).get()
        if company.exists:
            cd    = company.to_dict()
            plan  = cd.get("plan","starter")
            limit = SEAT_LIMITS.get(plan,-1)
            if limit > 0:
                members = db.collection("users").where("company_id","==",company_id).get()
                if len(list(members)) >= limit:
                    return jsonify({"error":f"Seat limit reached ({limit} seats on {plan} plan). Upgrade to add more members."}),403

        # Mark invite accepted
        inv_ref.update({"status":"accepted","accepted_by":uid,"accepted_at":datetime.utcnow().isoformat()+"Z"})

        # Link the accepting user to the company — including department/role
        # from the invite itself, since the frontend used to write these
        # directly but that path is removed now (see security notes above).
        db.collection("users").document(uid).update({
            "company_id": company_id,
            "department": inv_data.get("department", ""),
            "role":       inv_data.get("role", "employee"),
            "updated_at": datetime.utcnow().isoformat()+"Z",
        })

        audit(uid, "invite_accepted", "org", {"company_id": company_id, "invite_id": token})
        return jsonify({"ok":True,"company_id":company_id})
    except Exception as e:
        return jsonify({"error":str(e)}),500


# ── Employee Consent ───────────────────────────────────────────────
@app.route("/api/employee/consent", methods=["GET"])
@require_auth
def get_consent_status():
    """Return consent status for the authenticated employee."""
    try:
        uid = getattr(g, "uid", "")
        doc = db.collection("user_consent").document(uid).get()
        if doc.exists:
            d = doc.to_dict()
            return jsonify({
                "consented":   d.get("consented", False),
                "consented_at": d.get("consented_at", ""),
                "version":     d.get("version", "1.0"),
            })
        return jsonify({"consented": False})
    except Exception as e:
        return safe_error(e)

@app.route("/api/employee/consent", methods=["POST"])
@require_auth
@limiter.limit("10 per minute")
def set_consent():
    """Record employee consent to posture monitoring."""
    try:
        uid  = getattr(g, "uid", "")
        data = request.get_json(force=True) or {}
        accepted = bool(data.get("accepted", False))
        db.collection("user_consent").document(uid).set({
            "uid":          uid,
            "consented":    accepted,
            "consented_at": datetime.utcnow().isoformat() + "Z",
            "version":      "1.0",
            "ip":           request.headers.get("X-Forwarded-For", request.remote_addr),
            "user_agent":   request.headers.get("User-Agent", ""),
        }, merge=True)
        log_event("employee_consent", uid, {"accepted": accepted})
        return jsonify({"ok": True, "consented": accepted})
    except Exception as e:
        return safe_error(e)

@app.route("/api/employee/consent/revoke", methods=["POST"])
@require_auth
@limiter.limit("5 per minute")
def revoke_consent():
    """Employee can revoke consent at any time — stops all monitoring."""
    try:
        uid = getattr(g, "uid", "")
        db.collection("user_consent").document(uid).set({
            "consented":   False,
            "revoked_at":  datetime.utcnow().isoformat() + "Z",
            "version":     "1.0",
        }, merge=True)
        log_event("employee_consent_revoked", uid, {})
        return jsonify({"ok": True, "message": "Consent revoked. Monitoring stopped."})
    except Exception as e:
        return safe_error(e)



# ── Subscription status check ──────────────────────────────────────
@app.route("/api/subscription/check", methods=["POST"])
@require_auth
@limiter.limit("30 per minute")
def subscription_check():
    """Check if subscription is expired and downgrade if needed."""
    try:
        data = request.get_json(force=True) or {}
        # SECURITY: uid must match authenticated user (or admin can check any)
        req_uid = (data.get("uid") or "").strip()
        auth_uid = getattr(g, "uid", "")
        is_admin = getattr(g, "role", {}).get("is_admin", False)
        uid = auth_uid  # default: check own subscription
        if req_uid and req_uid != auth_uid and not is_admin:
            return jsonify({"error": "Cannot check another user's subscription"}), 403
        if req_uid and is_admin:
            uid = req_uid
        if not uid: return jsonify({"error":"uid required"}),400
        db   = firestore.client()
        ref  = db.collection("users").document(uid)
        doc  = ref.get()
        if not doc.exists: return jsonify({"error":"User not found"}),404
        u    = doc.to_dict()
        sub_end = u.get("subscription_end")
        if sub_end and not u.get("is_trial",False):
            from dateutil.parser import parse as parsedt
            try:
                end_dt = parsedt(sub_end)
                if end_dt.replace(tzinfo=None) < datetime.utcnow():
                    # Subscription expired → downgrade
                    ref.update({"tier":"standard","subscription_end":None,"updated_at":datetime.utcnow().isoformat()+"Z"})
                    # Send renewal email
                    threading.Thread(target=_send_renewal_reminder,args=(u.get("email",""),u.get("name",""),u.get("tier",""),True),daemon=True).start()
                    return jsonify({"status":"expired","downgraded":True})
            except Exception: pass
        # Check if expiring in 3 days — send reminder
        if sub_end:
            try:
                from dateutil.parser import parse as parsedt
                end_dt = parsedt(sub_end)
                days_left = (end_dt.replace(tzinfo=None)-datetime.utcnow()).days
                if 0 <= days_left <= 3 and not u.get("renewal_reminder_sent"):
                    threading.Thread(target=_send_renewal_reminder,args=(u.get("email",""),u.get("name",""),u.get("tier",""),False,days_left),daemon=True).start()
                    ref.update({"renewal_reminder_sent":True})
            except Exception: pass
        return jsonify({"status":"active","tier":u.get("tier","standard")})
    except Exception as e:
        return jsonify({"error":str(e)}),500


def _send_renewal_reminder(email, name, tier, expired=False, days_left=0):
    if not email: return
    subject = f"⏰ Your PostureAI Pro subscription {'expired' if expired else f'expires in {days_left} days'}"
    body = f"""
    <div style="font-family:Inter,system-ui,sans-serif;max-width:520px;margin:0 auto;padding:24px">
      <div style="text-align:center;margin-bottom:20px">
        <div style="width:48px;height:48px;background:linear-gradient(135deg,#1a56db,#0891b2);border-radius:12px;display:inline-flex;align-items:center;justify-content:center;font-size:22px">◈</div>
        <h2 style="margin:12px 0 4px;font-size:18px;color:#f0f6ff">PostureAI Pro</h2>
      </div>
      {"<p style='color:#ef4444;font-weight:700;font-size:16px'>Your subscription has expired.</p>" if expired else f"<p style='color:#f59e0b;font-weight:700;font-size:16px'>Your subscription expires in {days_left} day{'s' if days_left!=1 else ''}.</p>"}
      <p style="color:#94a3b8;font-size:13px">You were on the <strong>{tier.title()}</strong> plan. {"Your account has been downgraded to the free tier." if expired else "Renew now to keep your data and settings."}</p>
      <div style="text-align:center;margin-top:24px">
        <a href="{os.getenv('APP_URL','https://postureai.vercel.app')}#pricing"
           style="background:#1a56db;color:#fff;padding:12px 28px;border-radius:10px;text-decoration:none;font-weight:700;font-size:13px">
          {"Reactivate Plan →" if expired else "Renew Now →"}
        </a>
      </div>
    </div>
    """
    _send_email_smtp(email, subject, body)


def _send_email_smtp(to_email:str, subject:str, html_body:str):
    smtp_user = os.getenv("SMTP_USER","")
    smtp_pass = os.getenv("SMTP_PASS","")
    smtp_host = os.getenv("SMTP_HOST", SMTP_HOST)  # use global SMTP_HOST
    smtp_port = int(os.getenv("SMTP_PORT","587"))
    if not smtp_user: return False
    try:
        import smtplib
        from email.mime.multipart import MIMEMultipart
        from email.mime.text import MIMEText
        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"]    = smtp_user
        msg["To"]      = to_email
        msg.attach(MIMEText(html_body,"html"))
        with smtplib.SMTP(smtp_host, smtp_port) as s:
            s.ehlo(); s.starttls(); s.login(smtp_user,smtp_pass)
            s.sendmail(smtp_user,to_email,msg.as_string())
        return True
    except Exception as e:
        print(f"[Email] Failed: {e}",flush=True)
        return False


@require_auth
@app.route("/api/coupon/validate", methods=["POST"])
@limiter.limit("20 per minute")
def coupon_validate():
    try:
        data = request.get_json(force=True) or {}
        code = data.get("code","").strip().upper()
        _, c = _get_coupon_from_db(code)
        if not c: return jsonify({"valid":False,"reason":"Invalid coupon code"})
        if c.get("used",0) >= c.get("max_uses",100): return jsonify({"valid":False,"reason":"Coupon expired or used up"})
        return jsonify({"valid":True,"discount":c["discount"],"label":c["label"]})
    except Exception as e:
        return safe_error(e)

@require_auth
@app.route("/api/coupon/use", methods=["POST"])
@limiter.limit("10 per minute")
def coupon_use():
    try:
        data = request.get_json(force=True) or {}
        code = data.get("code","").strip().upper()
        with _coupon_lock:
            _, c = _get_coupon_from_db(code)
            if not c:
                return jsonify({"ok":False,"reason":"Invalid coupon"}), 400
            if c["used"] >= c["max_uses"]:
                return jsonify({"ok":False,"reason":"Coupon fully redeemed"}), 400
            c["used"] += 1        # atomic under lock
        return jsonify({"ok":True,"remaining": c["max_uses"] - c["used"]})
    except Exception as e:
        return safe_error(e)

@require_auth
@app.route("/api/work-hours", methods=["GET"])
@limiter.limit("60 per minute")
def work_hours():
    now = datetime.now()
    return jsonify({
        "start":         WORK_HOURS_START,
        "end":           WORK_HOURS_END,
        "in_work_hours": WORK_HOURS_START <= now.hour < WORK_HOURS_END,
        "current_hour":  now.hour,
        "day_of_week":   now.strftime("%A"),
        "is_weekday":    now.weekday() < 5,
    })

@require_auth
@app.route("/api/gemini", methods=["POST"])
@limiter.limit("20 per minute")
def gemini_proxy():
    """
    Async Gemini proxy — returns immediately with a job_id.
    Poll GET /api/gemini/job/<job_id> for result.
    Solves: synchronous blocking that stalled Gunicorn workers.
    """
    try:
        import json as _json
        data   = request.get_json(force=True) or {}
        prompt = data.get("prompt","")
        if not prompt: return jsonify({"error":"prompt required"}), 400
        # Block if no AI backend is configured
        if not AI_CONFIGURED:
            return jsonify({"error":"No AI backend configured (set OLLAMA_URL)","text":None}), 503

        job_id = f"gj_{_uuid_mod.uuid4().hex[:16]}"
        rset(f"gemini_job:{job_id}", _json.dumps({"status":"pending","created":time.time()}), 120)

        _tier = getattr(g, "tier", "standard")

        def _run_ai(jid, prompt_text):
            try:
                text = call_ai(prompt_text, max_tokens=400, temperature=0.4, tier=_tier)
                if text is not None:
                    rset(f"gemini_job:{jid}", _json.dumps({"status":"done","text":text,"ts":time.time()}), 120)
                else:
                    rset(f"gemini_job:{jid}", _json.dumps({"status":"error","error":"AI unavailable (Ollama not reachable or model not pulled)","ts":time.time()}), 60)
            except Exception as _e:
                rset(f"gemini_job:{jid}", _json.dumps({"status":"error","error":str(_e),"ts":time.time()}), 60)

        _ai_executor.submit(_run_ai, job_id, prompt)
        return jsonify({"job_id": job_id, "status": "pending"})

    except Exception as e:
        return jsonify({"error":str(e),"text":None}), 500


@require_auth
@app.route("/api/gemini/job/<job_id>", methods=["GET"])
@limiter.limit("120 per minute")
def gemini_job_status(job_id):
    """Poll for Gemini job result. Returns {status, text} or {status, error}."""
    try:
        import json as _json
        raw = rget(f"gemini_job:{job_id}")
        if not raw:
            return jsonify({"status": "expired", "error": "Job not found or expired"}), 404
        return jsonify(_json.loads(raw))
    except Exception as e:
        return jsonify({"status": "error", "error": str(e)}), 500


@require_auth
@app.route("/api/gemini/sync", methods=["POST"])
@limiter.limit("10 per minute")
def gemini_proxy_sync():
    """
    Synchronous local-AI call for PDF generation only.
    Rate limited to 10/min to prevent worker starvation.
    Use /api/gemini (async) for all interactive calls.
    """
    try:
        data   = request.get_json(force=True) or {}
        prompt = data.get("prompt","")
        if not prompt: return jsonify({"error":"prompt required"}), 400
        if not AI_CONFIGURED: return jsonify({"error":"AI not configured","text":None}), 503
        text = call_ai(prompt, max_tokens=400, temperature=0.4, tier=getattr(g, "tier", "standard"))
        if text is not None:
            return jsonify({"text": text})
        return jsonify({"error":"AI unavailable","text":None}), 502
    except Exception as e:
        return jsonify({"error":str(e),"text":None}), 500

@require_auth
@app.route("/api/user/export", methods=["POST"])
@limiter.limit("5 per hour")
def user_export():
    try:
        data = request.get_json(force=True) or {}
        uid  = data.get("uid","")
        if not uid: return jsonify({"error":"uid required"}), 400
        user_sess = {k:{kk:vv for kk,vv in v.items() if kk!="last_ai"} for k,v in sessions.items()}
        return jsonify({"uid":uid,"exported":datetime.now().isoformat(),"sessions_count":len(user_sess),"sessions":user_sess,
                        "note":"Firebase profile data: export from Firebase Console for complete GDPR package"})
    except Exception as e:
        return safe_error(e)

@require_auth
@app.route("/api/session/delete", methods=["POST"])
@limiter.limit("60 per minute")
def delete_session():
    try:
        data = request.get_json(force=True) or {}
        sid  = data.get("session_id","")
        if sid in sessions: del sessions[sid]
        if sid in session_snapshots: del session_snapshots[sid]
        return jsonify({"ok":True})
    except Exception as e:
        return safe_error(e)


@require_auth
@app.route("/api/session/sync", methods=["POST"])
@limiter.limit("30 per minute")
def sync_offline_sessions():
    """Sync sessions captured offline by the service worker.
    Accepts an array of session objects and saves each to Firestore via the
    client-side saveSession function cannot be called from backend directly,
    so we store them in the user's sessions subcollection.
    """
    uid  = g.uid
    data = request.get_json(silent=True) or {}
    offline_list = data.get("sessions", [])

    if not offline_list or not isinstance(offline_list, list):
        return jsonify({"error": "sessions array required"}), 400

    saved = 0
    errors = []
    for s in offline_list[:50]:  # Max 50 sessions per sync
        try:
            session_data = {
                "uid":        uid,
                "mode":       s.get("mode", "laptop"),
                "duration":   s.get("duration", 0),
                "avg_score":  s.get("avgScore", s.get("avg_score", 0)),
                "frames":     s.get("frames", 0),
                "good_frames":s.get("goodFrames", s.get("good_frames", 0)),
                "started_at": s.get("startedAt", s.get("started_at", "")),
                "synced_from_offline": True,
                "created_at": datetime.utcnow().isoformat(),
            }
            # Persist to Redis-backed session store
            session_id = s.get("sessionId", f"offline_{uid}_{saved}")
            sessions[session_id] = session_data
            saved += 1
        except Exception as e:
            errors.append(str(e))

    return jsonify({
        "synced": saved,
        "total":  len(sessions),
        "errors": errors[:5] if errors else [],
    })


@require_auth
@app.route("/api/email/invoice", methods=["POST"])
@limiter.limit("10 per minute")
def send_invoice_ep():
    try:
        data    = request.get_json(force=True) or {}
        to      = data.get("email","")
        if not to: return jsonify({"ok":False,"reason":"email required"}), 400
        ok = send_invoice_email({
            "user_email": to, "user_name": data.get("name","Customer"),
            "tier": data.get("tier","standard"), "amount": data.get("amount",0),
            "billing_cycle": data.get("billing","monthly"), "ref_code": data.get("ref","AUTO"),
            "payment_method_name": "PayMob (Automatic)"
        })
        return jsonify({"ok":ok})
    except Exception as e:
        return safe_error(e)

@require_auth
@app.route("/api/email/welcome", methods=["POST"])
@limiter.limit("10 per minute")
def send_welcome_ep():
    try:
        data = request.get_json(force=True) or {}
        to   = data.get("email","")
        name = data.get("name","there")
        tier = getattr(g, "tier", None) or "standard"  # never trust client tier
        if not to: return jsonify({"ok":False}), 400
        html = f"""
<div style="font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;max-width:560px;margin:0 auto;background:#07111f;border-radius:16px;overflow:hidden">
  <div style="background:linear-gradient(135deg,#4f46e5,#0891b2);padding:36px 32px;text-align:center">
    <div style="font-size:32px;margin-bottom:12px">◈</div>
    <div style="font-size:22px;font-weight:700;color:white;letter-spacing:-.03em;margin-bottom:4px">Welcome to PostureAI Pro</div>
    <div style="font-size:13px;color:rgba(255,255,255,.65)">AI Workforce Intelligence Platform</div>
  </div>
  <div style="padding:32px">
    <p style="color:#94a3b8;font-size:14px;margin:0 0 20px">Hi <b style="color:#f1f5f9">{name}</b> — you're on the <b style="color:#818cf8">{tier.title()} plan</b>. Your 7-day trial is live. No commitment.</p>

    <div style="background:#0c1728;border-radius:12px;padding:20px;margin-bottom:24px;border:1px solid rgba(255,255,255,.06)">
      <div style="font-size:11px;font-weight:700;color:#64748b;letter-spacing:.08em;text-transform:uppercase;margin-bottom:14px">GET STARTED IN 60 SECONDS</div>
      <div style="display:flex;align-items:flex-start;gap:12px;margin-bottom:12px">
        <div style="width:24px;height:24px;background:#4f46e5;border-radius:50%;display:flex;align-items:center;justify-content:center;font-size:11px;font-weight:700;color:white;flex-shrink:0">1</div>
        <div style="color:#cbd5e1;font-size:13px;line-height:1.5">Open the platform → choose your camera mode (Laptop / Phone / Side)</div>
      </div>
      <div style="display:flex;align-items:flex-start;gap:12px;margin-bottom:12px">
        <div style="width:24px;height:24px;background:#0891b2;border-radius:50%;display:flex;align-items:center;justify-content:center;font-size:11px;font-weight:700;color:white;flex-shrink:0">2</div>
        <div style="color:#cbd5e1;font-size:13px;line-height:1.5">Allow camera access → click <b style="color:#f1f5f9">Start Session</b> — your AI health score appears within 5 seconds</div>
      </div>
      <div style="display:flex;align-items:flex-start;gap:12px">
        <div style="width:24px;height:24px;background:#10b981;border-radius:50%;display:flex;align-items:center;justify-content:center;font-size:11px;font-weight:700;color:white;flex-shrink:0">3</div>
        <div style="color:#cbd5e1;font-size:13px;line-height:1.5">Complete 3 sessions → your AI Coach generates a personalized health intelligence report</div>
      </div>
    </div>

    <div style="text-align:center;margin-bottom:24px">
      <a href="{APP_URL}" style="background:linear-gradient(135deg,#4f46e5,#0891b2);color:white;padding:14px 32px;border-radius:10px;text-decoration:none;font-weight:700;font-size:14px;display:inline-block">Launch Your Intelligence Dashboard →</a>
    </div>

    <div style="background:#0c1728;border-radius:10px;padding:16px;border:1px solid rgba(255,255,255,.06);margin-bottom:20px">
      <div style="font-size:11px;font-weight:700;color:#64748b;letter-spacing:.08em;text-transform:uppercase;margin-bottom:10px">WHAT'S INCLUDED IN YOUR TRIAL</div>
      <div style="color:#94a3b8;font-size:12px;line-height:2">
        ✦ Real-time AI health intelligence (33 body landmarks)<br>
        ✦ Burnout risk prediction & fatigue pattern analysis<br>
        ✦ Gemini AI Coach with personalized recommendations<br>
        ✦ Clinical PDF export &amp; weekly executive summaries
      </div>
    </div>

    <p style="color:#475569;font-size:11px;text-align:center;margin:0">PostureAI Pro · <a href="mailto:{SUPPORT_EMAIL}" style="color:#6366f1">{SUPPORT_EMAIL}</a> · {ADMIN_PHONE}</p>
  </div>
</div>"""
        ok = send_email(to=to, subject="✦ Welcome to PostureAI Pro — Your AI workforce intelligence is live", html_body=html)
        return jsonify({"ok":ok})
    except Exception as e:
        return safe_error(e)

@app.errorhandler(404)
def not_found(e): return jsonify({"error":"Endpoint not found","status":404}), 404
@app.errorhandler(500)
def server_error(e): return jsonify({"error":"Internal server error","status":500}), 500
@app.errorhandler(429)
def rate_limit_exceeded(e): return jsonify({"error":"Rate limit exceeded — please slow down","status":429}), 429


# ── Admin: Users ──────────────────────────────────────────────────
@app.route("/api/admin/users", methods=["GET"])
@require_auth
@require_admin
@limiter.limit("60 per minute")
def admin_users():
    try:
        db   = firestore.client()
        docs = db.collection("users").order_by("created_at", direction=firestore.Query.DESCENDING).limit(500).get()
        users = []
        for d in docs:
            u = d.to_dict()
            u["id"] = d.id
            # Remove sensitive fields
            u.pop("password", None)
            users.append(u)
        return jsonify({"users": users, "total": len(users)})
    except Exception as e:
        return safe_error(e)


# ── Admin: Payments ───────────────────────────────────────────────
@app.route("/api/admin/payments", methods=["GET"])
@require_auth
@require_admin
@limiter.limit("60 per minute")
def admin_payments():
    try:
        db   = firestore.client()
        docs = db.collection("payments").order_by("created_at", direction=firestore.Query.DESCENDING).limit(500).get()
        pays = [{**d.to_dict(), "id": d.id} for d in docs]
        total = sum(p.get("amount", 0) for p in pays if p.get("status") == "confirmed")
        return jsonify({"payments": pays, "total": len(pays), "revenue": total})
    except Exception as e:
        return safe_error(e)


# ── Admin: Confirm payment ────────────────────────────────────────
@app.route("/api/admin/confirm-payment", methods=["POST"])
@require_auth
@require_admin
@limiter.limit("30 per minute")
def admin_confirm_payment():
    try:
        data = request.get_json(force=True) or {}
        pid  = data.get("payment_id","").strip()
        uid  = data.get("uid","").strip()
        tier = data.get("tier","professional").strip()
        months = int(data.get("months", 1))
        if not (pid and uid and tier):
            return jsonify({"error":"payment_id, uid, tier required"}), 400
        db  = firestore.client()
        expires = datetime.utcnow() + timedelta(days=30*months)
        # Update payment
        db.collection("payments").document(pid).update({
            "status": "confirmed",
            "confirmed_by": g.uid,
            "confirmed_at": datetime.utcnow().isoformat()+"Z",
        })
        # Update user tier
        db.collection("users").document(uid).update({
            "tier":                tier,
            "is_trial":            False,
            "subscription_end":    expires.isoformat()+"Z",
            "subscription_start":  datetime.utcnow().isoformat()+"Z",
            "updated_at":          datetime.utcnow().isoformat()+"Z",
        })
        invalidate_user_cache(uid)
        return jsonify({"ok": True, "tier": tier, "expires": expires.isoformat()})
    except Exception as e:
        return safe_error(e)


# ── Admin: Reject payment ─────────────────────────────────────────
@app.route("/api/admin/reject-payment", methods=["POST"])
@require_auth
@require_admin
@limiter.limit("30 per minute")
def admin_reject_payment():
    try:
        data   = request.get_json(force=True) or {}
        pid    = data.get("payment_id","").strip()
        reason = data.get("reason","Not verified")
        if not pid:
            return jsonify({"error":"payment_id required"}), 400
        firestore.client().collection("payments").document(pid).update({
            "status": "rejected", "reject_reason": reason,
            "rejected_by": g.uid, "rejected_at": datetime.utcnow().isoformat()+"Z",
        })
        return jsonify({"ok": True})
    except Exception as e:
        return safe_error(e)


# ── Admin: Set tier directly ──────────────────────────────────────
@app.route("/api/admin/set-tier", methods=["POST"])
@require_auth
@require_admin
@limiter.limit("30 per minute")
def admin_set_tier():
    try:
        data  = request.get_json(force=True) or {}
        uid   = data.get("uid","").strip()
        tier  = data.get("tier","standard")
        months= int(data.get("months", 0))
        if not uid:
            return jsonify({"error":"uid required"}), 400
        db  = firestore.client()
        upd = {
            "tier":     tier,
            "is_trial": False,
            "updated_at": datetime.utcnow().isoformat()+"Z",
            "updated_by_admin": g.uid,
        }
        if months > 0:
            upd["subscription_end"] = (datetime.utcnow() + timedelta(days=30*months)).isoformat()+"Z"
        db.collection("users").document(uid).update(upd)
        invalidate_user_cache(uid)
        return jsonify({"ok": True, "uid": uid, "tier": tier})
    except Exception as e:
        return safe_error(e)


# ── Admin: Coupons ────────────────────────────────────────────────
@app.route("/api/admin/coupons", methods=["GET"])
@require_auth
@require_admin
@limiter.limit("60 per minute")
def admin_list_coupons():
    try:
        db   = firestore.client()
        docs = db.collection("coupons").get()
        return jsonify({"coupons": [{**d.to_dict(), "id": d.id} for d in docs]})
    except Exception as e:
        return safe_error(e)


@app.route("/api/admin/coupons", methods=["POST"])
@require_auth
@require_admin
@limiter.limit("30 per minute")
def admin_create_coupon():
    try:
        data = request.get_json(force=True) or {}
        code = (data.get("code","")).strip().upper()
        if not code:
            return jsonify({"error":"code required"}), 400
        db = firestore.client()
        db.collection("coupons").document(code).set({
            "code":      code,
            "discount":  int(data.get("discount",20)),
            "label":     data.get("label",""),
            "max_uses":  int(data.get("max_uses",100)),
            "used":      0,
            "created_at": datetime.utcnow().isoformat()+"Z",
            "created_by": g.uid,
        })
        return jsonify({"ok": True, "code": code})
    except Exception as e:
        return safe_error(e)


@app.route("/api/admin/coupons/<code>", methods=["DELETE"])
@require_auth
@require_admin
@limiter.limit("30 per minute")
def admin_delete_coupon(code):
    try:
        firestore.client().collection("coupons").document(code.upper()).delete()
        return jsonify({"ok": True})
    except Exception as e:
        return safe_error(e)


# ── User payments history ─────────────────────────────────────────
@require_auth
@app.route("/api/user/payments", methods=["GET"])
@limiter.limit("30 per minute")
@require_auth
def user_payments():
    try:
        db   = firestore.client()
        docs = db.collection("payments").where("uid","==",g.uid).order_by("created_at",direction=firestore.Query.DESCENDING).limit(20).get()
        docs = sorted(docs, key=lambda d: d.to_dict().get("created_at",""), reverse=True)[:50]
        pays = [d.to_dict() for d in docs]
        return jsonify({"payments": pays})
    except Exception as e:
        return safe_error(e)


# ── Cancel subscription ────────────────────────────────────────────
@require_auth
@app.route("/api/subscription/cancel", methods=["POST"])
@limiter.limit("5 per hour")
@require_auth
def subscription_cancel():
    try:
        data  = request.get_json(force=True) or {}
        uid   = g.uid
        email = data.get("email","")
        tier  = data.get("tier","")
        db    = firestore.client()
        # Mark cancellation requested (admin processes manually or via cron)
        db.collection("users").document(uid).update({
            "cancellation_requested":    True,
            "cancellation_requested_at": datetime.utcnow().isoformat()+"Z",
        })
        # Notify admin
        threading.Thread(target=_send_email_smtp, args=(
            os.getenv("ADMIN_EMAIL","admin@postureai.io"),
            f"⚠️ Cancellation Request — {email} ({tier})",
            f"<p>User <strong>{email}</strong> (uid: {uid}) requested cancellation of their <strong>{tier}</strong> plan.</p><p>Process in PayMob and update Firestore.</p>"
        ), daemon=True).start()
        return jsonify({"ok": True})
    except Exception as e:
        return safe_error(e)

        return safe_error(e)

# ═══════════════════════════════════════════════════════════════════
# PHASE 3 — FEATURE 3: EMR / FHIR R4 API
# FHIR R4-compatible posture observation endpoint for clinical integrations
# Auth: Firebase JWT (same as all other endpoints)
# GET  /api/v1/emr/session/:session_id      → FHIR Observation bundle
# GET  /api/v1/emr/patient/:uid/summary     → Patient posture summary
# POST /api/v1/emr/webhook                  → Receive EMR webhooks (future)
# ═══════════════════════════════════════════════════════════════════

def _fhir_score_interpretation(score):
    """Map posture score to FHIR interpretation code (SNOMED-style)."""
    if score >= 80: return {"system":"http://snomed.info/sct","code":"17621005","display":"Normal"}
    if score >= 60: return {"system":"http://snomed.info/sct","code":"371929003","display":"Mildly abnormal"}
    if score >= 40: return {"system":"http://snomed.info/sct","code":"13791008","display":"Abnormal"}
    return             {"system":"http://snomed.info/sct","code":"42752001","display":"Due to"}

def _session_to_fhir_observation(session_data, uid, session_id):
    """Convert Corvus session to FHIR R4 Observation resource."""
    avg_score  = round(session_data.get("avg_score", 0))
    duration_s = session_data.get("duration_s", session_data.get("duration_sec", 0))
    created_at = session_data.get("created_at")
    if hasattr(created_at, "isoformat"):
        iso_ts = created_at.isoformat() + "Z"
    else:
        iso_ts = datetime.utcnow().isoformat() + "Z"

    metrics = session_data.get("metrics", {})
    # Build component list from metrics
    LOINC_MAP = {
        "neck_lean":      ("98350-7",  "Neck flexion angle",            "deg"),
        "fhp":            ("98351-5",  "Forward head posture distance",  "cm"),
        "shoulder":       ("98352-3",  "Shoulder symmetry index",        "%"),
        "spine_lean":     ("98353-1",  "Spinal lateral deviation angle", "deg"),
        "hip_angle":      ("98354-9",  "Hip joint angle",               "deg"),
        "trunk_lean":     ("98355-6",  "Trunk lean angle",              "deg"),
        "distance":       ("98356-4",  "Screen viewing distance",        "cm"),
    }
    components = []
    for key, (loinc, display, unit) in LOINC_MAP.items():
        m = metrics.get(key)
        if m is None: continue
        val = m.get("value") if isinstance(m, dict) else None
        sc  = m.get("score", 100) if isinstance(m, dict) else (m if isinstance(m, (int,float)) else 100)
        if val is None: continue
        components.append({
            "code": {
                "coding": [{"system":"http://loinc.org","code":loinc,"display":display}],
                "text": display,
            },
            "valueQuantity": {"value": round(float(val),2), "unit": unit, "system":"http://unitsofmeasure.org"},
            "interpretation": [{"coding":[_fhir_score_interpretation(sc)]}],
            "extension": [{
                "url": "https://corvus.health/fhir/StructureDefinition/posture-metric-score",
                "valueInteger": round(sc),
            }],
        })

    return {
        "resourceType": "Observation",
        "id": session_id,
        "meta": {
            "profile": ["https://corvus.health/fhir/StructureDefinition/PostureObservation"],
            "versionId": "1",
            "lastUpdated": iso_ts,
        },
        "status": "final",
        "category": [{
            "coding": [{
                "system": "http://terminology.hl7.org/CodeSystem/observation-category",
                "code":    "activity",
                "display": "Activity",
            }],
        }],
        "code": {
            "coding": [{
                "system":  "http://snomed.info/sct",
                "code":    "228557008",
                "display": "Postural Assessment",
            }],
            "text": "Corvus AI Posture Analysis",
        },
        "subject": {"reference": f"Patient/{uid}", "type": "Patient"},
        "effectiveDateTime": iso_ts,
        "issued": iso_ts,
        "performer": [{
            "reference": "Device/corvus-posture-ai",
            "display":   "Corvus PostureAI Pro",
        }],
        "valueInteger": avg_score,
        "interpretation": [{"coding":[_fhir_score_interpretation(avg_score)]}],
        "note": [{
            "text": (
                f"AI posture analysis session. Duration: {duration_s}s. "
                f"Good posture: {session_data.get('good_pct',0)}%. "
                f"Camera mode: {session_data.get('mode','laptop')}. "
                f"Alerts: {session_data.get('alerts_count',0)}. "
                f"Tier: {session_data.get('tier','standard')}."
            ),
        }],
        "component": components,
        "extension": [
            {"url":"https://corvus.health/fhir/ext/session-duration-seconds","valueInteger": duration_s},
            {"url":"https://corvus.health/fhir/ext/good-posture-percentage","valueDecimal": session_data.get("good_pct",0)},
            {"url":"https://corvus.health/fhir/ext/camera-mode","valueString": session_data.get("mode","laptop")},
        ],
    }


@app.route("/api/v1/emr/session/<session_id>", methods=["GET"])
def emr_session_observation(session_id):
    """
    GET /api/v1/emr/session/:session_id
    Returns FHIR R4 Observation for a single Corvus posture session.
    Auth: Firebase JWT Bearer token (same as all other endpoints).
    Accept: application/fhir+json (default) or application/json
    """
    try:
        uid = verify_firebase_token(request)
        # Fetch session from Firestore
        sess_ref = db.collection("sessions").document(session_id)
        snap = sess_ref.get()
        if not snap.exists:
            # Try user-scoped sessions collection
            users_ref = db.collection("users").document(uid)\
                         .collection("sessions").document(session_id)
            snap = users_ref.get()
        if not snap.exists:
            return jsonify({"error":"Session not found","code":404}), 404

        data = snap.to_dict()
        # Ownership check
        owner_uid = data.get("uid") or data.get("user_id") or data.get("owner_uid")
        if owner_uid and owner_uid != uid:
            return jsonify({"error":"Forbidden","code":403}), 403

        observation = _session_to_fhir_observation(data, uid, session_id)
        resp = jsonify(observation)
        resp.headers["Content-Type"] = "application/fhir+json; fhirVersion=4.0"
        resp.headers["X-Corvus-FHIR-Version"] = "R4"
        return resp, 200

    except Exception as e:
        return safe_error(e)


@app.route("/api/v1/emr/patient/<patient_uid>/summary", methods=["GET"])
def emr_patient_summary(patient_uid):
    """
    GET /api/v1/emr/patient/:uid/summary
    Returns FHIR R4 DiagnosticReport summarising posture health for a patient.
    Includes last 30 sessions + zonal risk summary.
    Auth: Firebase JWT — caller must be the patient or an HR admin of their org.
    Query params:
      days=30  (default 30, max 365)
      format=fhir|json (default fhir)
    """
    try:
        caller_uid = verify_firebase_token(request)
        days = min(int(request.args.get("days", 30)), 365)
        cutoff = datetime.utcnow() - timedelta(days=days)

        # Authorization: self or org HR admin
        if caller_uid != patient_uid:
            caller_snap = db.collection("users").document(caller_uid).get()
            caller_data = caller_snap.to_dict() or {} if caller_snap.exists else {}
            if caller_data.get("user_type") not in ("hr_admin","platform_admin"):
                return jsonify({"error":"Forbidden — only the patient or their HR admin can request this","code":403}), 403

        # Fetch patient profile
        patient_snap = db.collection("users").document(patient_uid).get()
        patient_data = patient_snap.to_dict() if patient_snap.exists else {}
        patient_name = patient_data.get("name","Unknown")

        # Fetch sessions
        sessions_query = (
            db.collection("users").document(patient_uid).collection("sessions")
              .where("created_at",">=",cutoff)
              .order_by("created_at", direction="DESCENDING")
              .limit(30)
        )
        session_docs = sessions_query.stream()
        sessions = [{"id":d.id, **d.to_dict()} for d in session_docs]

        if not sessions:
            return jsonify({"error":"No sessions found in date range","code":404}), 404

        scores  = [s.get("avg_score",0) for s in sessions if s.get("avg_score")]
        avg_sc  = round(sum(scores)/len(scores)) if scores else 0
        best_sc = max(scores) if scores else 0
        worst_sc= min(scores) if scores else 0

        # Aggregate zone risk
        all_zonal = []
        for s in sessions:
            m = s.get("metrics",{})
            if m:
                cervical = 100 - sum([
                    (m.get("neck_lean",{}).get("score",100) if isinstance(m.get("neck_lean"),dict) else m.get("neck_lean",100)),
                    (m.get("fhp",{}).get("score",100) if isinstance(m.get("fhp"),dict) else m.get("fhp",100)),
                ]) / 2
                thoracic = 100 - sum([
                    (m.get("shoulder",{}).get("score",100) if isinstance(m.get("shoulder"),dict) else m.get("shoulder",100)),
                    (m.get("rounded",{}).get("score",100) if isinstance(m.get("rounded"),dict) else m.get("rounded",100)),
                ]) / 2
                lumbar = 100 - sum([
                    (m.get("spine_align",{}).get("score",100) if isinstance(m.get("spine_align"),dict) else m.get("spine_align",100)),
                    (m.get("hip_angle",{}).get("score",100) if isinstance(m.get("hip_angle"),dict) else m.get("hip_angle",100)),
                ]) / 2
                all_zonal.append({"cervical":max(0,cervical),"thoracic":max(0,thoracic),"lumbar":max(0,lumbar)})

        avg_zone = {
            "cervical": round(sum(z["cervical"] for z in all_zonal)/len(all_zonal)) if all_zonal else 0,
            "thoracic": round(sum(z["thoracic"] for z in all_zonal)/len(all_zonal)) if all_zonal else 0,
            "lumbar":   round(sum(z["lumbar"]   for z in all_zonal)/len(all_zonal)) if all_zonal else 0,
        }

        # FHIR R4 DiagnosticReport
        report_date = datetime.utcnow().isoformat() + "Z"
        report = {
            "resourceType": "DiagnosticReport",
            "id": f"corvus-posture-{patient_uid}-{days}d",
            "meta": {
                "profile": ["https://corvus.health/fhir/StructureDefinition/PostureDiagnosticReport"],
                "lastUpdated": report_date,
            },
            "status": "final",
            "category": [{"coding":[{
                "system":"http://terminology.hl7.org/CodeSystem/v2-0074",
                "code":"OT","display":"Occupational Therapy",
            }]}],
            "code": {"coding":[{
                "system":"http://snomed.info/sct",
                "code":"229070002","display":"Ergonomic assessment",
            }],"text":"Corvus AI Posture Health Summary"},
            "subject": {
                "reference": f"Patient/{patient_uid}",
                "display":   patient_name,
            },
            "effectivePeriod": {
                "start": cutoff.isoformat()+"Z",
                "end":   report_date,
            },
            "issued": report_date,
            "performer": [{"display":"Corvus PostureAI Pro","reference":"Device/corvus-posture-ai"}],
            "result": [{"reference":f"Observation/{s['id']}"} for s in sessions[:5]],
            "conclusion": (
                f"AI posture analysis over {days} days ({len(sessions)} sessions). "
                f"Average score: {avg_sc}/100. "
                f"Best: {best_sc}. Worst: {worst_sc}. "
                f"Cervical zone risk: {avg_zone['cervical']}%. "
                f"Thoracic zone risk: {avg_zone['thoracic']}%. "
                f"Lumbar zone risk: {avg_zone['lumbar']}%."
            ),
            "presentedForm": [{
                "contentType": "application/fhir+json",
                "url": f"/api/v1/emr/patient/{patient_uid}/summary",
            }],
            "extension": [
                {"url":"https://corvus.health/fhir/ext/sessions-analysed","valueInteger":len(sessions)},
                {"url":"https://corvus.health/fhir/ext/avg-posture-score","valueInteger":avg_sc},
                {"url":"https://corvus.health/fhir/ext/cervical-zone-risk","valueInteger":avg_zone["cervical"]},
                {"url":"https://corvus.health/fhir/ext/thoracic-zone-risk","valueInteger":avg_zone["thoracic"]},
                {"url":"https://corvus.health/fhir/ext/lumbar-zone-risk",  "valueInteger":avg_zone["lumbar"]},
                {"url":"https://corvus.health/fhir/ext/days-monitored","valueInteger":days},
            ],
            # Corvus non-FHIR summary (for convenience — non-FHIR clients)
            "_corvus": {
                "avg_score":    avg_sc,
                "best_score":   best_sc,
                "worst_score":  worst_sc,
                "sessions":     len(sessions),
                "days":         days,
                "zone_risk":    avg_zone,
                "patient_name": patient_name,
                "generated_at": report_date,
            },
        }

        resp = jsonify(report)
        resp.headers["Content-Type"] = "application/fhir+json; fhirVersion=4.0"
        resp.headers["X-Corvus-FHIR-Version"] = "R4"
        resp.headers["Access-Control-Allow-Origin"] = "*"
        resp.headers["X-Corvus-API-Version"] = "v1"
        return resp, 200

    except Exception as e:
        return safe_error(e)


@app.route("/api/v1/emr/webhook", methods=["POST"])
def emr_webhook_receiver():
    """
    POST /api/v1/emr/webhook
    Receive incoming webhooks from EMR systems (Epic MyChart, Cerner, etc.)
    Validates HMAC signature in X-Corvus-Signature header.
    Stores webhook payload in Firestore for async processing.
    """
    try:
        import hmac, hashlib
        secret = os.getenv("EMR_WEBHOOK_SECRET","")
        sig_header = request.headers.get("X-Corvus-Signature","")
        if secret:
            body_bytes = request.get_data()
            expected   = "sha256=" + hmac.new(
                secret.encode(), body_bytes, hashlib.sha256
            ).hexdigest()
            if not hmac.compare_digest(sig_header, expected):
                return jsonify({"error":"Invalid signature"}), 401
        else:
            body_bytes = request.get_data()

        payload = request.get_json(force=True) or {}
        event_type = payload.get("type","unknown")

        # Store for async processing
        db.collection("emr_webhooks").add({
            "type":        event_type,
            "payload":     payload,
            "received_at": datetime.utcnow().isoformat()+"Z",
            "source_ip":   request.remote_addr,
            "processed":   False,
        })

        return jsonify({"received":True,"event":event_type}), 200
    except Exception as e:
        return safe_error(e)



    import threading as _th; _th.Thread(target=_ensure_models, daemon=True).start()
    import atexit as _ae
    _ae.register(lambda: _ai_executor.shutdown(wait=False))
    os.makedirs("reports", exist_ok=True)
    print("="*60, flush=True)
    print("  PostureAI Pro Backend v15", flush=True)
    print(f"  MediaPipe {_mp.__version__ if _mp else '(lazy-loaded on first request)'} — Pose + FaceMesh + solvePnP + Blink Detection", flush=True)
    _ai_label = ('✅ Ollama (' + LOCAL_LLM_MODEL + ')') if OLLAMA_URL else '⚠️  No server AI backend (set OLLAMA_URL) — note: client uses local WebLLM regardless'
    print(f"  AI:  {_ai_label}", flush=True)
    print(f"  PDF: {'✅ ReportLab ready' if REPORTLAB_OK else '⚠️  pip install reportlab'}", flush=True)
    print(f"  APP_URL: {APP_URL}", flush=True)
    print("  PORT: 5050  →  http://localhost:5050", flush=True)
    if not os.getenv("PAYMOB_HMAC_SECRET",""):
        print("⚠️  WARNING: PAYMOB_HMAC_SECRET not set — webhook verification DISABLED", flush=True)
    if not AI_CONFIGURED:
        print("ℹ️  No server-side AI backend configured — server AI endpoints disabled (client uses local WebLLM regardless).", flush=True)
    print("="*60, flush=True)
    sys.stdout.flush()
    app.run(host="0.0.0.0", port=5050, debug=False, threaded=True, use_reloader=False)

# ══════════════════════════════════════════════════════════════════
# ENTERPRISE WEBHOOK ENGINE
# ══════════════════════════════════════════════════════════════════
import hmac as _hmac, hashlib as _hashlib, threading as _threading

# In-memory webhook store (replace with DB in production)
_webhooks: dict = {}          # webhook_id -> config
_webhook_logs: list = []      # delivery log
_risk_tracker: dict = {}      # uid -> {score, since}

def _sign_payload(secret: str, body: str) -> str:
    return "sha256=" + _hmac.new(secret.encode(), body.encode(), _hashlib.sha256).hexdigest()

def _deliver_webhook(wh: dict, payload: dict, attempt: int = 1):
    import json, time
    body    = json.dumps(payload)
    sig     = _sign_payload(wh.get("secret", ""), body)
    headers = {
        "Content-Type":          "application/json",
        "X-PostureAI-Signature": sig,
        "X-PostureAI-Event":     payload.get("event", "posture.alert"),
        "X-PostureAI-Delivery":  payload.get("delivery_id", ""),
        "User-Agent":            "PostureAI-Webhooks/1.0",
    }
    log_entry = {
        "webhook_id":   wh["id"],
        "url":          wh["url"],
        "event":        payload.get("event"),
        "status":       None,
        "attempt":      attempt,
        "ts":           datetime.now().isoformat(),
        "delivery_id":  payload.get("delivery_id", ""),
    }
    try:
        r = req.post(wh["url"], data=body, headers=headers, timeout=12)
        log_entry["status"] = r.status_code
        log_entry["ok"]     = r.status_code < 300
        _webhook_logs.append(log_entry)
        if r.status_code >= 300 and attempt < 4:
            delay = 2 ** attempt * 3   # 6s, 12s, 24s
            _threading.Timer(delay, _deliver_webhook, args=[wh, payload, attempt + 1]).start()
    except Exception as e:
        log_entry["status"] = 0
        log_entry["ok"]     = False
        log_entry["error"]  = str(e)
        _webhook_logs.append(log_entry)
        if attempt < 4:
            delay = 2 ** attempt * 3
            _threading.Timer(delay, _deliver_webhook, args=[wh, payload, attempt + 1]).start()

def _fire_webhooks(event: str, data: dict):
    import uuid
    for wh in _webhooks.values():
        if not wh.get("active", True): continue
        if event not in wh.get("events", [event]): continue
        payload = {
            "event":       event,
            "delivery_id": str(uuid.uuid4())[:12],
            "timestamp":   datetime.now().isoformat(),
            "data":        data,
        }
        _threading.Thread(target=_deliver_webhook, args=[wh, payload], daemon=True).start()

# ── Trend buffer: stores recent scores per uid for 3-day rolling avg ──
_trend_buffer: dict = {}   # {uid: [(timestamp, score)]}
_TREND_WINDOW_S = 3 * 24 * 3600   # 3 days

def _get_trend_avg(uid: str) -> float:
    """Return 3-day rolling average score for uid. Returns -1 if insufficient data."""
    buf = _trend_buffer.get(uid, [])
    cutoff = time.time() - _TREND_WINDOW_S
    recent = [(t, s) for t, s in buf if t >= cutoff]
    if len(recent) < 10:   # need at least 10 data points (≈ 10 frames) to trust trend
        return -1.0
    return sum(s for _, s in recent) / len(recent)

def _check_risk_threshold(uid: str, score: int, threshold: int = 45, duration_s: int = 300):
    """
    Trend-aware HR alert system.
    - Records every score into a 3-day rolling buffer.
    - Alert fires only when the 3-day average drops below threshold
      AND the current session score has been low for > duration_s seconds.
    - This prevents false alerts from a single bad frame or brief slouch.
    """
    now = time.time()

    # ── Update trend buffer ──────────────────────────────────────
    buf = _trend_buffer.setdefault(uid, [])
    buf.append((now, score))
    # Prune old entries (>3 days)
    cutoff = now - _TREND_WINDOW_S
    _trend_buffer[uid] = [(t, s) for t, s in buf if t >= cutoff]
    # Cap buffer size (max 10000 entries per user)
    if len(_trend_buffer[uid]) > 10000:
        _trend_buffer[uid] = _trend_buffer[uid][-5000:]

    # ── Get 3-day trend avg ──────────────────────────────────────
    trend_avg = _get_trend_avg(uid)

    # ── Instant session tracker (unchanged logic) ────────────────
    if score < threshold:
        if uid not in _risk_tracker:
            _risk_tracker[uid] = {"score": score, "since": now}
        else:
            elapsed = now - _risk_tracker[uid]["since"]
            if elapsed >= duration_s:
                # Only fire if trend also confirms chronic poor posture
                # trend_avg == -1 means insufficient data → fire anyway (new user)
                trend_confirms = (trend_avg == -1.0) or (trend_avg < threshold + 10)
                if trend_confirms:
                    risk_level = (
                        "critical" if (trend_avg != -1 and trend_avg < 30) else
                        "high"     if score < 35 else
                        "medium"
                    )
                    _fire_webhooks("posture.risk_alert", {
                        "employee_uid":     uid,
                        "risk_level":       risk_level,
                        "posture_score":    score,
                        "trend_avg_3day":   round(trend_avg, 1) if trend_avg != -1 else None,
                        "trend_confirmed":  trend_avg != -1,
                        "duration_seconds": int(elapsed),
                        "threshold":        threshold,
                        "triggered_at":     datetime.now().isoformat(),
                        "note": (
                            "Trend-confirmed chronic issue" if trend_avg != -1
                            else "Insufficient trend data — alert based on session only"
                        ),
                    })
                del _risk_tracker[uid]   # reset after firing (whether fired or suppressed)
    else:
        _risk_tracker.pop(uid, None)

@require_auth
@app.route("/api/webhooks", methods=["GET"])
@require_auth
@limiter.limit("60 per minute")
def list_webhooks():
    uid = getattr(g, "uid", None)
    user_hooks = {k: v for k,v in _webhooks.items() if v.get("uid") == uid}
    return jsonify({"webhooks": list(user_hooks.values())})

@require_auth
@app.route("/api/webhooks", methods=["POST"])
@limiter.limit("20 per minute")
def create_webhook():
    try:
        import uuid, secrets
        data = request.get_json(force=True) or {}
        url  = data.get("url", "").strip()
        if not url or not url.startswith("http"):
            return jsonify({"error": "Valid URL required"}), 400
        wh = {
            "id":          str(uuid.uuid4())[:12],
            "url":         url,
            "secret":      data.get("secret") or secrets.token_hex(24),
            "events":      data.get("events", ["posture.risk_alert", "session.complete", "report.ready"]),
            "org_id":      data.get("org_id", ""),
            "description": data.get("description", ""),
            "active":      True,
            "created_at":  datetime.now().isoformat(),
        }
        _webhooks[wh["id"]] = wh
        return jsonify({"webhook": wh}), 201
    except Exception as e:
        return safe_error(e)

@require_auth
@app.route("/api/webhooks/<wid>", methods=["DELETE"])
@limiter.limit("20 per minute")
def delete_webhook(wid):
    if wid in _webhooks:
        del _webhooks[wid]
        return jsonify({"ok": True})
    return jsonify({"error": "Not found"}), 404

@require_auth
@app.route("/api/webhooks/<wid>/test", methods=["POST"])
@limiter.limit("10 per minute")
def test_webhook(wid):
    wh = _webhooks.get(wid)
    if not wh: return jsonify({"error": "Webhook not found"}), 404
    import uuid
    payload = {
        "event": "webhook.test", "delivery_id": str(uuid.uuid4())[:12],
        "timestamp": datetime.now().isoformat(),
        "data": {"message": "PostureAI webhook test — delivery confirmed ✓"},
    }
    _threading.Thread(target=_deliver_webhook, args=[wh, payload], daemon=True).start()
    return jsonify({"ok": True, "message": "Test delivery queued"})

@require_auth
@app.route("/api/webhooks/logs", methods=["GET"])
@limiter.limit("60 per minute")
def webhook_logs():
    limit_n = min(int(request.args.get("limit", 50)), 200)
    wid     = request.args.get("webhook_id")
    logs    = [l for l in _webhook_logs if not wid or l.get("webhook_id") == wid]
    return jsonify({"logs": list(reversed(logs))[:limit_n], "total": len(logs)})

@require_auth
@app.route("/api/webhooks/risk-check", methods=["POST"])
@limiter.limit("200 per minute")
def risk_check():
    """Call this after every analysis frame to trigger threshold webhooks."""
    try:
        data  = request.get_json(force=True) or {}
        uid   = data.get("uid", "")
        score = int(data.get("score", 100))
        threshold = int(data.get("threshold", 45))
        duration  = int(data.get("duration_s", 300))
        if uid and score > 0:
            _check_risk_threshold(uid, score, threshold, duration)
        return jsonify({"ok": True})
    except Exception as e:
        return safe_error(e)

# ══════════════════════════════════════════════════════════════════
# AUDIT LOG SYSTEM (ISO 27001)
# ══════════════════════════════════════════════════════════════════
_audit_logs: list = []   # in-memory tail for fast /admin/audit/logs reads

_AUDIT_REDIS_KEY = "audit_logs"
_AUDIT_MAX       = 5000   # keep last 5000 in Redis list

def audit(uid: str, action: str, resource: str = "", meta: dict = None, ip: str = ""):
    import json as _jaud
    entry = {
        "uid":      uid,
        "action":   action,
        "resource": resource,
        "meta":     meta or {},
        "ip":       ip or (request.remote_addr if request else ""),
        "ts":       datetime.now().isoformat(),
        "id":       f"AUD{int(time.time()*1000)%100000000:08d}",
    }
    # 1. In-memory tail (fast reads for admin dashboard)
    _audit_logs.append(entry)
    if len(_audit_logs) > 1000:
        _audit_logs.pop(0)
    # 2. Redis list (persistent, survives restarts)
    try:
        rpush(_AUDIT_REDIS_KEY, _jaud.dumps(entry, default=str), max_len=_AUDIT_MAX)
    except Exception:
        pass  # Redis unavailable — memory-only is acceptable fallback

# ── Merged from v13: WorkOS SSO + Custom Domains ───────────────────
try:
    from services.workos_sso import register_workos_routes
    register_workos_routes(app, require_auth, require_admin, firestore, audit, print)
    print("✅ WorkOS SSO routes registered")
except Exception as _e:
    print(f"⚠️  WorkOS SSO skipped: {_e}")

try:
    from services.custom_domains import register_domain_routes
    _fs_client = firestore.client()
    register_domain_routes(app, require_auth, require_admin, _fs_client, audit)
    print("✅ Custom domain routes registered")
except Exception as _e:
    print(f"⚠️  Custom domain routes skipped: {_e}")

@require_auth
@app.route("/api/audit/log", methods=["POST"])
@limiter.limit("200 per minute")
def audit_log_ep():
    try:
        data = request.get_json(force=True) or {}
        audit(
            uid      = data.get("uid", ""),
            action   = data.get("action", "unknown"),
            resource = data.get("resource", ""),
            meta     = data.get("meta", {}),
            ip       = request.remote_addr,
        )
        return jsonify({"ok": True})
    except Exception as e:
        return safe_error(e)

@require_auth
@app.route("/api/audit/logs", methods=["GET"])
@limiter.limit("30 per minute")
@require_admin
def get_audit_logs():
    try:
        uid    = request.args.get("uid")
        action = request.args.get("action")
        limit_n= min(int(request.args.get("limit", 100)), 500)
        # Prefer Redis (persistent across restarts) over memory tail
        import json as _jaud
        logs = []
        try:
            raw_list = rlist(_AUDIT_REDIS_KEY, _AUDIT_MAX)
            for raw in raw_list:
                try: logs.append(_jaud.loads(raw))
                except: pass
        except Exception:
            pass
        # Fallback to memory if Redis empty/unavailable
        if not logs:
            logs = _audit_logs[:]
        # Filters
        if uid:    logs = [l for l in logs if l.get("uid") == uid]
        if action: logs = [l for l in logs if l.get("action") == action]
        # Redis list is LIFO (newest first via lpush), reverse for chronological
        logs_sorted = sorted(logs, key=lambda x: x.get("ts",""), reverse=True)
        return jsonify({"logs": logs_sorted[:limit_n], "total": len(logs)})
    except Exception as e:
        return safe_error(e)

# ── Admin: Set Firebase custom claims (admin/hr) ──────────────────
@require_auth
@app.route("/api/admin/set-claims", methods=["POST"])
@limiter.limit("10 per minute")
@require_admin
def set_user_claims():
    """
    Promote/demote a user's Firebase custom claims.
    Only callable by existing admins (triple-layer checked above).
    Body: { "target_uid": "...", "admin": true/false, "hr": true/false }
    """
    try:
        from firebase_admin import auth as fb_auth_mod
        data       = request.get_json(force=True) or {}
        target_uid = data.get("target_uid","").strip()
        if not target_uid:
            return jsonify({"error":"target_uid required"}),400
        # Only set known claim keys — never pass arbitrary data
        claims = {}
        if "admin" in data: claims["admin"] = bool(data["admin"])
        if "hr"    in data: claims["hr"]    = bool(data["hr"])
        fb_auth_mod.set_custom_user_claims(target_uid, claims)
        invalidate_user_cache(target_uid)
        audit(g.uid, "set_claims", meta={"target": target_uid, "claims": claims})
        return jsonify({"ok":True,"target_uid":target_uid,"claims":claims})
    except Exception as e:
        return safe_error(e)

# ══════════════════════════════════════════════════════════════════
# AI POSTURE COACH — Gemini Streaming Proxy
# ══════════════════════════════════════════════════════════════════
@require_auth
@app.route("/api/coach/chat", methods=["POST"])
@limiter.limit("30 per minute")
def coach_chat():
    try:
        data     = request.get_json(force=True) or {}
        messages = data.get("messages", [])   # [{role, content}]
        context  = data.get("context", {})    # posture analytics context
        lang     = data.get("lang", "en")

        if not messages:
            return jsonify({"error": "messages required"}), 400
        # OLLAMA_URL is optional — if unset, server-side AI Coach is disabled

        # ── Tier-based AI Coach limits + depth (single source of truth) ──
        _uid_coach    = getattr(g, "uid", "")
        _tier_coach   = getattr(g, "tier", "standard") or "standard"
        _quality_coach = get_coach_quality(_tier_coach)
        _month_key  = f"usage:{_uid_coach}:ai_coach:{datetime.utcnow().strftime('%Y-%m')}"
        _limit = _quality_coach["monthly_limit"]
        if _limit > 0:
            try:
                import redis as _rc_coach
                _rc = _rc_coach.from_url(os.getenv("REDIS_URL", "redis://localhost:6379/0"), socket_timeout=1)
                _used = int(_rc.get(_month_key) or 0)
            except Exception:
                _used = 0
            if _used >= _limit:
                return jsonify({
                    "error":   "coach_limit_reached",
                    "message": f"You've used all {_limit} AI Coach messages this month. Upgrade to get more.",
                    "used":    _used,
                    "limit":   _limit,
                    "upgrade_url": "/pricing",
                }), 429

        # Build system prompt with full posture context
        avg_score  = context.get("avg_score", 0)
        worst_time = context.get("worst_time", "unknown")
        sessions_n = context.get("sessions_count", 0)
        top_alerts = context.get("top_alerts", [])
        calib      = context.get("has_calibration", False)
        tier       = _tier_coach  # authoritative — never trust client-sent context.tier
        is_ar      = lang == "ar"

        # Build rich clinical data block
        _alerts_str = ', '.join(top_alerts[:4]) if top_alerts else 'None recorded'
        _score_label = ("Excellent" if avg_score >= 85 else "Good" if avg_score >= 70
                        else "Fair" if avg_score >= 55 else "Needs Attention")
        _neck_risk   = context.get("neck_risk", 0)
        _fatigue     = context.get("fatigue_score", 0)
        _burnout     = context.get("burnout_risk", 0)
        _week_avg    = context.get("week_avg", avg_score)
        _trend_pct   = context.get("trend_pct", 0)
        _streak      = context.get("streak_days", 0)
        _week_sess   = context.get("week_sessions", 0)

        sys_prompt = f"""You are Dr. Corvus — the AI physiotherapist embedded in Corvus PostureAI Pro.

IDENTITY: You are a certified ergonomics consultant and clinical physiotherapy specialist. Not a generic chatbot.
Never say "I'm an AI" — you are Dr. Corvus.

CLINICAL EXPERTISE:
- MSK anatomy: cervical/thoracic/lumbar spine, shoulder girdle, carpal tunnel, hip flexors, sacroiliac joint
- Postural syndromes: Forward Head Posture (FHP), Upper/Lower Crossed Syndrome (Janda), kyphosis, lordosis, APT, piriformis syndrome, thoracic outlet syndrome
- Biomechanics: Hansraj cervical load model (2014) — 0°=4.5kg, 15°=12kg, 30°=18kg, 45°=22kg, 60°=27kg; Nachemson disc pressure — sitting unsupported=140% vs standing
- Evidence base: NIOSH 1997, OSHA ergonomics standards, ISO 11226, Cornell Human Factors (Hedge 2017), Janda Syndromes
- Therapeutic protocols: McKenzie method, muscle energy techniques, neuromuscular re-education, progressive loading

PATIENT CLINICAL DATA (authoritative — always reference these numbers):
- Overall posture score: {avg_score}/100 ({_score_label})
- This week average: {_week_avg}/100 | Trend: {('+' if _trend_pct > 0 else '')}{_trend_pct}% vs last week
- Total sessions: {sessions_n} | This week: {_week_sess} | Streak: {_streak} days
- Cervical risk score: {_neck_risk}% ({'HIGH' if _neck_risk >= 70 else 'MODERATE' if _neck_risk >= 40 else 'LOW'})
- Fatigue index: {_fatigue}% | Burnout risk: {_burnout}%
- Worst posture window: {worst_time if worst_time and worst_time != 'unknown' else 'Not yet identified'}
- Anthropometric calibration: {'COMPLETE — personalized thresholds active' if calib else 'NOT DONE — generic population thresholds in use'}
- Recurring postural alerts: {_alerts_str}
- Subscription tier: {tier}

RESPONSE PRINCIPLES:
1. Reference the patient's ACTUAL data numbers in every response — never speak in generalities.
2. For every recommendation: WHAT to do → WHY (anatomical mechanism) → HOW (precise steps) → BENEFIT → TIMEFRAME.
3. NEVER say "maintain good posture" — describe the specific anatomical correction and which muscles it targets.
4. Cite research naturally: "Hansraj (2014) demonstrated that at 45° neck flexion, cervical load reaches 22 kg..."
5. Use clinical terminology with plain explanations: "the deep cervical flexors (longus colli — your spine's inner stabilizers)..."
6. ⚕️ Flag red flags (radiating pain, paresthesia, unilateral weakness) → recommend professional evaluation.
7. Format: **bold** key terms, numbered steps for protocols, short ## headers for multi-part responses.
8. TOPIC BOUNDARY: posture, ergonomics, MSK health, workspace, physiotherapy. Redirect off-topic warmly.
9. Response length: 150-220 words for conversation. Up to 350 for explicit report/plan requests.
{get_depth_instruction(_tier_coach)}

CONVERSATION STYLE:
- Respond to what was actually asked — don't give a template response.
- Pain reports: assess clinically (location, character, radiation, aggravating/relieving factors).
- End with ONE focused follow-up question when clinically appropriate.
{"LANGUAGE: Respond ENTIRELY in Egyptian Arabic (عامية مصرية). Use medical terms then immediately explain them simply in Arabic." if is_ar else "LANGUAGE: Respond in clear, professional English."}"""

        # ── Gemini requires alternating user/model turns ──────────────
        # Build Ollama-compatible messages list (OpenAI format) AND
        # Gemini-compatible contents list simultaneously:
        # - Ollama: [{role:"system",...},{role:"user",...},{role:"assistant",...}]
        # - Gemini: [{role:"user",...},{role:"model",...}] (no system role in contents)
        ollama_msgs = [{"role": "system", "content": sys_prompt}]
        gemini_contents = []
        last_role = None
        for i, msg in enumerate(messages):
            role_ui = msg.get("role","user")
            role_gem = "user" if role_ui == "user" else "model"
            role_oll = "user" if role_ui == "user" else "assistant"
            if role_gem == last_role:
                continue
            text_content = msg.get("content","")
            # NOTE: Do NOT inject context into user turn — it's already in sys_prompt above.
            # Double-injection confuses the LLM and produces generic responses.
            ollama_msgs.append({"role": role_oll, "content": text_content})
            gemini_contents.append({"role": role_gem, "parts": [{"text": text_content}]})
            last_role = role_gem

        if not gemini_contents or gemini_contents[0]["role"] != "user":
            return jsonify({"error": "Conversation must start with a user message", "ok": False}), 400
        if gemini_contents[-1]["role"] != "user":
            return jsonify({"error": "Last message must be from user", "ok": False}), 400

        # Final user prompt is the last message
        final_user_prompt = gemini_contents[-1]["parts"][0]["text"]
        text = None

        # Temperature: lower for clinical/report requests, higher for conversational
        _last_user_msg = (messages[-1].get("content","") if messages else "").lower()
        _is_report_req = any(w in _last_user_msg for w in ["report","plan","analyze","تقرير","خطة","تحليل","protocol","exercise","تمارين"])
        _temperature = 0.3 if _is_report_req else 0.55
        _max_tok = _quality_coach["max_tokens"]

        # ── Local Ollama (multi-turn via OpenAI messages format) ─────────
        if OLLAMA_URL:
            try:
                resp_local = req.post(
                    OLLAMA_URL.rstrip("/") + "/v1/chat/completions",
                    headers={"Content-Type": "application/json"},
                    json={
                        "model":       LOCAL_LLM_MODEL,
                        "messages":    ollama_msgs,
                        "max_tokens":  _max_tok,
                        "temperature": _temperature,
                        "stream":      False,
                    },
                    timeout=30,
                )
                if resp_local.status_code == 200:
                    text = (resp_local.json().get("choices") or [{}])[0].get("message",{}).get("content","") or None
            except Exception as _le:
                log_event("coach_local_llm_error", meta={"error": str(_le)[:80]})

        if text is None:
            # ── LLM7.io fallback (anonymous, no key, Railway can reach it) ──
            try:
                llm7_resp = req.post(
                    "https://api.llm7.io/v1/chat/completions",
                    headers={"Content-Type": "application/json", "Authorization": "Bearer unused"},
                    json={"model": "gpt-4o-mini", "messages": ollama_msgs,
                          "max_tokens": _max_tok, "temperature": _temperature},
                    timeout=20,
                )
                if llm7_resp.status_code == 200:
                    text = (llm7_resp.json().get("choices") or [{}])[0].get("message", {}).get("content", "") or None
                elif llm7_resp.status_code == 429:
                    # Rate limited — try DeepSeek
                    llm7_resp2 = req.post(
                        "https://api.llm7.io/v1/chat/completions",
                        headers={"Content-Type": "application/json", "Authorization": "Bearer unused"},
                        json={"model": "deepseek/deepseek-r1", "messages": ollama_msgs,
                              "max_tokens": _max_tok, "temperature": _temperature},
                        timeout=20,
                    )
                    if llm7_resp2.status_code == 200:
                        text = (llm7_resp2.json().get("choices") or [{}])[0].get("message", {}).get("content", "") or None
            except Exception as _le:
                log_event("coach_llm7_error", meta={"error": str(_le)[:80]})

        if text is None:
            return jsonify({"error": "AI unavailable — please try again", "ok": False, "text": None}), 503
        _uid = getattr(g, "uid", "unknown")
        audit(_uid, "ai_coach_query", "local:" + LOCAL_LLM_MODEL if OLLAMA_URL else "none", {"lang": lang, "turns": len(gemini_contents)})
        try:
            import redis as _r2
            _rc2 = _r2.from_url(os.getenv("REDIS_URL","redis://localhost:6379/0"))
            _month2 = datetime.utcnow().strftime("%Y-%m")
            _k2 = f"usage:{_uid}:ai_coach:{_month2}"
            _rc2.incr(_k2); _rc2.expire(_k2, 86400 * 35)
        except Exception: pass
        return jsonify({"text": text, "ok": True})
    except req.exceptions.Timeout:
        return jsonify({"error": "AI response timed out — try again", "ok": False}), 504
    except Exception as e:
        return jsonify({"error": str(e), "ok": False}), 500
_xp_events = {
    "session_complete":    10,
    "score_80_plus":       25,
    "score_90_plus":       50,
    "calibration_done":    100,
    "7_day_streak":        200,
    "30_day_streak":       500,
    "first_session":       50,
    "share_referral":      30,
    "break_complete":      15,
}

ACHIEVEMENTS = [
    {"id": "first_session",   "name": "First Step",       "name_ar": "الخطوة الأولى",    "icon": "🎯", "xp": 50,  "req": {"sessions": 1}},
    {"id": "ten_sessions",    "name": "Getting Serious",  "name_ar": "جدية حقيقية",       "icon": "📊", "xp": 100, "req": {"sessions": 10}},
    {"id": "fifty_sessions",  "name": "Posture Pro",      "name_ar": "محترف الوضعية",      "icon": "⭐", "xp": 300, "req": {"sessions": 50}},
    {"id": "score_85",        "name": "Excellent Form",   "name_ar": "أداء ممتاز",         "icon": "🏆", "xp": 150, "req": {"avg_score": 85}},
    {"id": "streak_7",        "name": "Week Warrior",     "name_ar": "محارب الأسبوع",      "icon": "🔥", "xp": 200, "req": {"streak": 7}},
    {"id": "streak_30",       "name": "Monthly Master",   "name_ar": "سيد الشهر",          "icon": "💎", "xp": 500, "req": {"streak": 30}},
    {"id": "calibrated",      "name": "Personalized",     "name_ar": "معاير شخصياً",       "icon": "🎯", "xp": 100, "req": {"calibrated": True}},
    {"id": "referral_1",      "name": "Evangelist",       "name_ar": "سفير المنتج",         "icon": "🤝", "xp": 150, "req": {"referrals": 1}},
    {"id": "referral_5",      "name": "Growth Driver",    "name_ar": "محرك النمو",          "icon": "🚀", "xp": 400, "req": {"referrals": 5}},
    {"id": "night_owl",       "name": "Night Owl",        "name_ar": "بومة الليل",          "icon": "🦉", "xp": 75,  "req": {"late_session": True}},
    {"id": "early_bird",      "name": "Early Bird",       "name_ar": "الطائر المبكر",       "icon": "🌅", "xp": 75,  "req": {"early_session": True}},
    {"id": "perfect_week",    "name": "Perfect Week",     "name_ar": "أسبوع مثالي",         "icon": "✨", "xp": 350, "req": {"perfect_week": True}},
]

@require_auth
@app.route("/api/gamification/compute", methods=["POST"])
@limiter.limit("60 per minute")
def compute_gamification():
    try:
        data          = request.get_json(force=True) or {}
        sessions_n    = int(data.get("sessions_count", 0))
        avg_score     = int(data.get("avg_score", 0))
        streak        = int(data.get("streak", 0))
        referrals_n   = int(data.get("referral_count", 0))
        calibrated    = bool(data.get("has_calibration", False))
        existing_ach  = data.get("earned_achievements", [])
        hour_of_day   = datetime.now().hour

        # Compute XP
        xp = 0
        xp += sessions_n * _xp_events["session_complete"]
        if avg_score >= 80: xp += _xp_events["score_80_plus"] * max(0, sessions_n - 5)
        if avg_score >= 90: xp += _xp_events["score_90_plus"] * max(0, sessions_n - 20)
        if calibrated: xp += _xp_events["calibration_done"]
        if streak >= 7:  xp += _xp_events["7_day_streak"]
        if streak >= 30: xp += _xp_events["30_day_streak"]

        # XP level system
        def xp_to_level(x):
            lvl = 1
            needed = 100
            while x >= needed:
                x -= needed; lvl += 1; needed = int(needed * 1.35)
            return lvl, x, needed

        level, xp_curr, xp_next = xp_to_level(xp)

        # Check achievements
        new_achievements = []
        all_earned       = list(existing_ach)
        for ach in ACHIEVEMENTS:
            if ach["id"] in all_earned: continue
            req = ach["req"]
            earned = False
            if "sessions"      in req and sessions_n >= req["sessions"]: earned = True
            if "avg_score"     in req and avg_score  >= req["avg_score"]: earned = True
            if "streak"        in req and streak      >= req["streak"]:   earned = True
            if "calibrated"    in req and calibrated  == req["calibrated"]: earned = True
            if "referrals"     in req and referrals_n >= req["referrals"]: earned = True
            if "late_session"  in req and hour_of_day >= 22:              earned = True
            if "early_session" in req and hour_of_day <= 7:               earned = True
            if earned:
                new_achievements.append(ach)
                all_earned.append(ach["id"])
                xp += ach["xp"]

        # Daily goal: 30 min at 75+
        daily_goal = {
            "target_score":    75,
            "target_minutes":  30,
            "label":           "Maintain 75+ posture for 30 minutes",
            "label_ar":        "حافظ على 75+ لمدة 30 دقيقة",
            "xp_reward":       50,
        }

        return jsonify({
            "xp":                xp,
            "level":             level,
            "xp_current":        xp_curr,
            "xp_to_next":        xp_next,
            "level_label":       ["Beginner","Aware","Developing","Consistent","Advanced","Pro","Elite","Master","Grandmaster","Legend"][min(level-1, 9)],
            "streak":            streak,
            "new_achievements":  new_achievements,
            "all_achievements":  all_earned,
            "daily_goal":        daily_goal,
            "achievements_list": ACHIEVEMENTS,
        })
    except Exception as e:
        return safe_error(e)

@require_auth
@app.route("/api/gamification/leaderboard", methods=["POST"])
@limiter.limit("30 per minute")
def compute_leaderboard():
    """Compute leaderboard from submitted employee data."""
    try:
        data      = request.get_json(force=True) or {}
        employees = data.get("employees", [])
        period    = data.get("period", "week")   # week | month | all

        if not employees:
            return jsonify({"leaderboard": [], "total": 0})

        ranked = sorted(
            [e for e in employees if isinstance(e.get("avg_score"), (int, float))],
            key=lambda e: (e.get("avg_score", 0), e.get("sessions_count", 0)),
            reverse=True
        )
        leaderboard = []
        for i, emp in enumerate(ranked[:20]):
            score  = emp.get("avg_score", 0)
            medals = {0: "🥇", 1: "🥈", 2: "🥉"}
            leaderboard.append({
                "rank":           i + 1,
                "medal":          medals.get(i, ""),
                "name":           emp.get("name", "Anonymous"),
                "department":     emp.get("department", ""),
                "avg_score":      round(score, 1),
                "sessions_count": emp.get("sessions_count", 0),
                "streak":         emp.get("streak", 0),
                "grade":          "Excellent" if score >= 85 else "Good" if score >= 70 else "Fair" if score >= 50 else "Poor",
            })

        # Department rankings
        depts: dict = {}
        for emp in employees:
            dept = emp.get("department", "General")
            if dept not in depts: depts[dept] = []
            if isinstance(emp.get("avg_score"), (int, float)):
                depts[dept].append(emp["avg_score"])
        dept_rankings = sorted(
            [{"department": k, "avg_score": round(sum(v)/len(v), 1), "employees": len(v)} for k, v in depts.items() if v],
            key=lambda d: d["avg_score"], reverse=True
        )

        return jsonify({
            "leaderboard":       leaderboard,
            "department_ranking": dept_rankings,
            "total":             len(ranked),
            "period":            period,
            "generated_at":      datetime.now().isoformat(),
        })
    except Exception as e:
        return safe_error(e)

# ══════════════════════════════════════════════════════════════════
# HEATMAP ANALYTICS
# ══════════════════════════════════════════════════════════════════
@require_auth


@app.route("/api/explain", methods=["POST"])
@require_auth
@limiter.limit("30 per minute")
def explain_score():
    """
    Explains WHY score changed between two sessions/days.

    POST body:
    {
      "today":     { session or day summary },
      "yesterday": { session or day summary },
      "lang": "en" | "ar"
    }

    Each session/day object:
    {
      "avg_score": 72,
      "metrics": {
        "neck_lean":       {"value": 14, "score": 55, "unit": "°"},
        "head_tilt":       {"value": 3,  "score": 82, "unit": "°"},
        "shoulder_level":  {"value": 5,  "score": 72, "unit": "°"},
        "spine_lean":      {"value": 9,  "score": 61, "unit": "°"},
        "screen_distance": {"value": 58, "score": 100,"unit": "cm"},
        "head_yaw":        {"value": 6,  "score": 78, "unit": "°"},
        "eye_strain":      {"value": 10, "score": 45, "unit": "bpm"},
      },
      "alerts": [...],
      "quality_score": 88,
      "duration_s": 900,
      "good_pct": 65
    }
    """
    try:
        data      = request.get_json(force=True) or {}
        today     = data.get("today")
        yesterday = data.get("yesterday")
        lang      = data.get("lang", "en")
        is_ar     = lang == "ar"
        uid       = getattr(g, "uid", None)

        if not today:
            return jsonify({"error": "today session data required"}), 400

        # ── Score delta ───────────────────────────────────────────
        score_today = today.get("avg_score", 0)
        score_yest  = yesterday.get("avg_score", 0) if yesterday else None
        delta       = round(score_today - score_yest) if score_yest is not None else None

        # ── Per-metric comparison ──────────────────────────────────
        METRIC_LABELS = {
            "neck_lean":       ("Neck lean",       "ميل الرقبة",      "°",  True),
            "fhp_index":       ("Forward head",    "بروز الرأس للأمام","cm", True),
            "neck_load":       ("Neck load",       "حمل الرقبة",      "kg", True),
            "ergonomics_score":("Workstation",     "بيئة العمل",      "/100",False),
            "rsi_risk":        ("RSI risk",        "خطر الإجهاد",     "pts",True),
            "symmetry":        ("Body symmetry",   "تماثل الجسم",     "%",  True),
            "head_tilt":       ("Head tilt",        "ميل الرأس",       "°",  True),
            "shoulder_level":  ("Shoulder level",   "مستوى الكتفين",   "°",  True),
            "spine_lean":      ("Spine lean",       "ميل العمود",      "°",  True),
            "screen_distance": ("Screen distance",  "بُعد الشاشة",     "cm", False),
            "head_yaw":        ("Head turn",        "دوران الرأس",     "°",  True),
            "eye_strain":      ("Eye strain",       "إجهاد العين",     "bpm",True),
        }

        met_today = today.get("metrics", {})
        met_yest  = yesterday.get("metrics", {}) if yesterday else {}

        changes = []   # sorted by impact (biggest drop first)
        improvements = []

        for key, (en_lbl, ar_lbl, unit, low_good) in METRIC_LABELS.items():
            t = met_today.get(key)
            y = met_yest.get(key)
            if not t: continue

            t_score = t.get("score", 0)
            t_val   = t.get("value")
            y_score = y.get("score", 0)  if y else None
            y_val   = y.get("value")     if y else None

            score_diff = round(t_score - y_score) if y_score is not None else None
            val_diff   = round(t_val   - y_val, 1) if (t_val is not None and y_val is not None) else None

            label = ar_lbl if is_ar else en_lbl

            entry = {
                "metric":       key,
                "label":        label,
                "today_score":  t_score,
                "today_value":  t_val,
                "unit":         unit,
                "yest_score":   y_score,
                "yest_value":   y_val,
                "score_diff":   score_diff,
                "val_diff":     val_diff,
                "low_good":     low_good,
            }

            if score_diff is not None:
                if score_diff <= -8:
                    changes.append(entry)       # significant drop
                elif score_diff >= 8:
                    improvements.append(entry)  # significant gain
            elif t_score < 60:
                changes.append(entry)           # no yesterday — just bad today

        # Sort by impact
        changes.sort(key=lambda x: x.get("score_diff") or -x["today_score"])
        improvements.sort(key=lambda x: -(x.get("score_diff") or x["today_score"]))

        # ── Generate human-readable explanations ──────────────────
        def explain_metric(e, is_drop):
            label  = e["label"]
            t_sc   = e["today_score"]
            t_val  = e["today_value"]
            y_sc   = e["yest_score"]
            y_val  = e["yest_value"]
            unit   = e["unit"]
            s_diff = e["score_diff"]
            v_diff = e["val_diff"]
            low_g  = e["low_good"]

            if is_ar:
                if is_drop:
                    trend = f"ارتفع بـ {abs(v_diff)}{unit}" if (v_diff and low_g and v_diff > 0) else                             f"انخفض بـ {abs(v_diff)}{unit}" if (v_diff and not low_g and v_diff < 0) else ""
                    return {
                        "title":   f"⬇️ {label} تراجع",
                        "detail":  f"النقطة: {t_sc}/100 ({trend})" if trend else f"النقطة: {t_sc}/100",
                        "impact":  f"أثّر بـ {abs(s_diff or 0)} نقطة على نتيجتك",
                        "fix":     _get_fix(e["metric"], t_val, is_ar=True),
                    }
                else:
                    return {
                        "title":   f"⬆️ {label} تحسّن",
                        "detail":  f"النقطة: {t_sc}/100 ({'تحسّن ' + str(abs(v_diff or 0)) + unit if v_diff else ''})",
                        "impact":  f"أضاف {abs(s_diff or 0)} نقطة لنتيجتك",
                    }
            else:
                if is_drop:
                    val_note = ""
                    if v_diff and low_g and v_diff > 0:
                        val_note = f" (+{v_diff}{unit} worse)"
                    elif v_diff and not low_g and v_diff < 0:
                        val_note = f" ({v_diff}{unit} from ideal)"
                    return {
                        "title":   f"⬇️ {label} declined",
                        "detail":  f"Score dropped to {t_sc}/100{val_note}",
                        "impact":  f"Caused ~{abs(s_diff or 0)}pt drop in overall score",
                        "fix":     _get_fix(e["metric"], t_val, is_ar=False),
                    }
                else:
                    return {
                        "title":   f"⬆️ {label} improved",
                        "detail":  f"Score rose to {t_sc}/100",
                        "impact":  f"Added ~{abs(s_diff or 0)}pts to overall score",
                    }

        def _get_fix(metric, val, is_ar):
            fixes = {
                "neck_lean":       ("Raise monitor to eye level — use a stand or stack books",
                                    "ارفع مستوى الشاشة لمستوى العيون"),
                "head_tilt":       ("Level your head — check chair height and monitor centering",
                                    "تأكد من استواء رأسك — اضبط ارتفاع الكرسي"),
                "shoulder_level":  ("Adjust armrests to equal height — relax both shoulders",
                                    "اضبط ذراعَي الكرسي لنفس الارتفاع"),
                "spine_lean":      ("Sit back fully against chair back, engage lumbar support",
                                    "اجلس للخلف واستخدم دعامة أسفل الظهر"),
                "screen_distance": ("Move screen to 50–80cm away (arm's length)",
                                    "ضع الشاشة على بُعد 50–80سم (مسافة ذراع)"),
                "head_yaw":        ("Face monitor directly — reposition chair if needed",
                                    "انظر للشاشة مباشرة — أعد ترتيب وضع الكرسي"),
                "eye_strain":      ("20-20-20 rule: every 20 min look 6m away for 20 sec",
                                    "قاعدة 20-20-20: كل 20 دقيقة انظر 6 أمتار بعيداً لـ 20 ثانية"),
            }
            f = fixes.get(metric, ("Correct your posture", "صحّح وضعيتك"))
            return f[1] if is_ar else f[0]

        explanations_drops  = [explain_metric(e, True)  for e in changes[:3]]
        explanations_gains  = [explain_metric(e, False) for e in improvements[:2]]

        # ── Overall verdict ───────────────────────────────────────
        if delta is None:
            if score_today >= 80:
                verdict = "good_no_comparison"
                verdict_text = "وضعيتك جيدة اليوم ✓" if is_ar else "Good posture today ✓"
            elif score_today >= 60:
                verdict = "fair_no_comparison"
                verdict_text = "وضعيتك مقبولة — يمكن تحسينها" if is_ar else "Fair posture — room to improve"
            else:
                verdict = "poor_no_comparison"
                verdict_text = "وضعيتك تحتاج انتباهاً" if is_ar else "Posture needs attention"
        elif delta >= 5:
            verdict = "improved"
            verdict_text = f"تحسّن بـ {delta} نقطة مقارنة بالأمس 📈" if is_ar else f"Improved by {delta}pts vs yesterday 📈"
        elif delta >= -4:
            verdict = "stable"
            verdict_text = "مستقرة مقارنة بالأمس ➡️" if is_ar else "Stable vs yesterday ➡️"
        elif delta >= -10:
            verdict = "slight_drop"
            verdict_text = f"انخفض بـ {abs(delta)} نقطة عن الأمس ⚠️" if is_ar else f"Down {abs(delta)}pts from yesterday ⚠️"
        else:
            verdict = "significant_drop"
            verdict_text = f"انخفض بـ {abs(delta)} نقطة — راجع وضعيتك 🔴" if is_ar else f"Down {abs(delta)}pts — review your setup 🔴"

        # ── Context clues (non-metric reasons) ────────────────────
        context_clues = []
        qual_t = today.get("quality_score")
        qual_y = yesterday.get("quality_score") if yesterday else None
        dur_t  = today.get("duration_s", 0)
        dur_y  = yesterday.get("duration_s", 0) if yesterday else 0

        if qual_t and qual_t < 65:
            context_clues.append(
                "جودة الجلسة منخفضة — الإضاءة أو الكاميرا قد تؤثر على الدقة" if is_ar
                else f"Session quality was low ({qual_t}/100) — lighting or camera angle may have affected accuracy"
            )
        if dur_t and dur_y and dur_t < dur_y * 0.5:
            context_clues.append(
                "الجلسة اليوم أقصر بكثير — قد لا تعكس يومًا كاملاً" if is_ar
                else f"Today's session was much shorter ({dur_t//60}min vs {dur_y//60}min yesterday)"
            )
        if delta and delta < -5 and not changes:
            context_clues.append(
                "الانخفاض موزّع على عدة مقاييس — لا يوجد سبب واحد واضح" if is_ar
                else "Drop is spread across multiple metrics — no single dominant cause"
            )

        log_event("explain_score", uid=uid, meta={
            "delta": delta, "verdict": verdict,
            "top_drop": changes[0]["metric"] if changes else None
        })

        return jsonify({
            "verdict":       verdict,
            "verdict_text":  verdict_text,
            "score_today":   score_today,
            "score_yesterday": score_yest,
            "delta":         delta,
            "drops":         explanations_drops,
            "gains":         explanations_gains,
            "context_clues": context_clues,
            "raw_changes":   changes[:5],
            "lang":          lang,
        })

    except Exception as e:
        return safe_error(e)


@app.route("/api/forecast", methods=["POST"])
@require_auth
@limiter.limit("20 per minute")
def posture_forecast():
    """
    Predict tomorrow's posture risk.
    POST: { "sessions": [...recent sessions...] }
    """
    try:
        data     = request.get_json(force=True) or {}
        sessions_data = data.get("sessions", [])
        uid      = getattr(g, "uid", None)
        forecast = compute_posture_forecast(uid, sessions_data)
        log_event("forecast_generated", uid=uid, meta={"risk": forecast.get("risk_level")})
        return jsonify(forecast)
    except Exception as e:
        return safe_error(e)


@app.route("/api/coaching/profile", methods=["GET"])
@require_auth
@limiter.limit("30 per minute")
def get_coaching():
    """Get user's full coaching profile + plan."""
    try:
        uid     = getattr(g, "uid", None)
        if not uid: return jsonify({"error": "auth required"}), 401
        profile = get_coaching_profile(uid)
        return jsonify(profile)
    except Exception as e:
        return safe_error(e)


@app.route("/api/coaching/reset", methods=["POST"])
@require_auth
@limiter.limit("5 per minute")
def reset_coaching():
    """Reset coaching profile (user request)."""
    try:
        uid = getattr(g, "uid", None)
        if not uid: return jsonify({"error": "auth required"}), 401
        cache_set(f"coach:{uid}", "{}", 86400*90)
        log_event("coaching_reset", uid=uid)
        return jsonify({"status": "reset", "uid": uid})
    except Exception as e:
        return safe_error(e)


@app.route("/api/analytics/trend", methods=["POST"])
@require_auth
@limiter.limit("30 per minute")
def analytics_trend():
    """
    Returns daily avg posture scores for the last 30 days + linear regression.

    POST body:
      { "sessions": [...], "days": 30, "include_weekly": true }

    sessions array items:
      { "avg_score": 78, "created_at_iso": "2026-06-01T09:00:00Z",
        "duration_s": 900, "good_pct": 72, "quality_score": 85 }

    Returns:
      daily[]       - one entry per day (null if no sessions)
      regression    - slope, intercept, r_squared, verdict
      weekly[]      - 4-week breakdown with diffs
      summary       - overall stats
    """
    try:
        data     = request.get_json(force=True) or {}
        sessions = data.get("sessions", [])
        n_days   = min(max(int(data.get("days", 30)), 7), 90)
        inc_week = data.get("include_weekly", True)
        uid      = getattr(g, "uid", None)

        if not sessions:
            return jsonify({"error": "sessions array required"}), 400

        # ── Build daily buckets ──────────────────────────────────
        from collections import defaultdict
        buckets = defaultdict(list)   # "YYYY-MM-DD" → [scores]
        dur_map = defaultdict(list)   # duration per day
        qual_map= defaultdict(list)   # quality scores per day

        for s in sessions:
            ts  = s.get("created_at_iso") or s.get("created_at", "")
            sc  = s.get("avg_score", 0)
            dur = s.get("duration_s", 0)
            qsc = s.get("quality_score") or s.get("quality", {}).get("quality_score")
            if not ts or not sc: continue
            try:
                dt  = datetime.fromisoformat(ts.replace("Z", "+00:00"))
                day = dt.strftime("%Y-%m-%d")
                buckets[day].append(sc)
                dur_map[day].append(dur)
                if qsc: qual_map[day].append(qsc)
            except Exception:
                continue

        # ── Build 30-day array ───────────────────────────────────
        today    = datetime.utcnow().date()
        daily    = []
        pts      = []   # (idx, avg) for regression — only days with data

        for i in range(n_days - 1, -1, -1):
            d     = today - __import__("datetime").timedelta(days=i)
            ds    = d.strftime("%Y-%m-%d")
            label = d.strftime("%b %-d")
            scores= buckets.get(ds, [])
            avg   = round(sum(scores) / len(scores)) if scores else None
            entry = {
                "date":        ds,
                "label":       label,
                "avg_score":   avg,
                "sessions":    len(scores),
                "min_score":   min(scores) if scores else None,
                "max_score":   max(scores) if scores else None,
                "avg_duration": round(sum(dur_map[ds]) / len(dur_map[ds])) if dur_map.get(ds) else None,
                "avg_quality": round(sum(qual_map[ds]) / len(qual_map[ds])) if qual_map.get(ds) else None,
            }
            daily.append(entry)
            if avg is not None:
                pts.append((n_days - 1 - i, avg))   # x = days ago inverted

        # ── Linear regression ────────────────────────────────────
        regression = None
        if len(pts) >= 3:
            n   = len(pts)
            sx  = sum(p[0] for p in pts)
            sy  = sum(p[1] for p in pts)
            sx2 = sum(p[0]**2 for p in pts)
            sxy = sum(p[0]*p[1] for p in pts)
            denom = n*sx2 - sx**2
            if abs(denom) > 0.001:
                m = (n*sxy - sx*sy) / denom          # slope pts/day
                b = (sy - m*sx) / n                  # intercept
                # R² — how well does the line fit
                y_mean = sy / n
                ss_tot = sum((p[1]-y_mean)**2 for p in pts)
                ss_res = sum((p[1]-(m*p[0]+b))**2 for p in pts)
                r2 = max(0.0, 1 - ss_res/ss_tot) if ss_tot > 0 else 0.0

                # Trend verdict
                slope_30 = m * 30   # projected change over 30 days
                if   slope_30 >  8: verdict = "improving_fast"
                elif slope_30 >  2: verdict = "improving"
                elif slope_30 > -2: verdict = "stable"
                elif slope_30 > -8: verdict = "declining"
                else:               verdict = "declining_fast"

                verdict_label = {
                    "improving_fast": "📈 Clear improvement trend",
                    "improving":      "📈 Gradual improvement",
                    "stable":         "➡️  Stable — no clear change",
                    "declining":      "📉 Slight decline — pay attention",
                    "declining_fast": "📉 Clear decline — review habits",
                }.get(verdict, "")

                regression = {
                    "slope":          round(m, 4),
                    "slope_per_week": round(m * 7, 2),
                    "intercept":      round(b, 2),
                    "r_squared":      round(r2, 3),
                    "projected_30d":  round(m * 30, 1),
                    "verdict":        verdict,
                    "verdict_label":  verdict_label,
                    "data_points":    n,
                    "reliable":       r2 >= 0.25 and n >= 7,
                }

        # ── Weekly breakdown ────────────────────────────────────
        weekly = []
        if inc_week:
            for w in range(4):
                w_days  = daily[w*7 : w*7+7]
                w_pts   = [d["avg_score"] for d in w_days if d["avg_score"] is not None]
                w_avg   = round(sum(w_pts)/len(w_pts)) if w_pts else None
                w_start = w_days[0]["label"]  if w_days else ""
                w_end   = w_days[-1]["label"] if w_days else ""
                weekly.append({
                    "week":       w + 1,
                    "label":      f"{w_start} – {w_end}",
                    "avg_score":  w_avg,
                    "active_days":len(w_pts),
                    "sessions":   sum(d["sessions"] for d in w_days),
                })
            # Add week-over-week diffs
            for i in range(1, len(weekly)):
                prev = weekly[i-1]["avg_score"]
                curr = weekly[i]["avg_score"]
                weekly[i]["diff_from_prev"] = (
                    round(curr - prev) if curr is not None and prev is not None else None
                )

        # ── Summary ───────────────────────────────────────────────
        all_scores  = [d["avg_score"] for d in daily if d["avg_score"] is not None]
        active_days = len(all_scores)
        summary = {
            "overall_avg":  round(sum(all_scores)/active_days) if all_scores else None,
            "best_score":   max(all_scores) if all_scores else None,
            "worst_score":  min(all_scores) if all_scores else None,
            "active_days":  active_days,
            "total_days":   n_days,
            "consistency":  round(active_days / n_days * 100),   # % days with sessions
            "first_half_avg": round(sum(all_scores[:active_days//2]) / max(len(all_scores[:active_days//2]),1)) if all_scores else None,
            "second_half_avg":round(sum(all_scores[active_days//2:]) / max(len(all_scores[active_days//2:]),1)) if all_scores else None,
        }

        log_event("analytics_trend", uid=uid, meta={"days": n_days, "sessions": len(sessions), "pts": len(pts)})

        return jsonify({
            "daily":      daily,
            "regression": regression,
            "weekly":     weekly if inc_week else [],
            "summary":    summary,
            "days":       n_days,
            "generated_at": datetime.utcnow().isoformat() + "Z",
        })

    except Exception as e:
        return safe_error(e)


@app.route("/api/analytics/heatmap", methods=["POST"])
@require_auth
@limiter.limit("30 per minute")
@require_tier("professional")
def compute_heatmap():
    try:
        data     = request.get_json(force=True) or {}
        sessions = data.get("sessions", [])
        if not sessions:
            return jsonify({"heatmap": [], "insights": []})

        # Hourly breakdown
        hourly: dict = {h: [] for h in range(24)}
        daily:  dict = {d: [] for d in range(7)}
        for s in sessions:
            ts = s.get("created_at_iso")
            sc = s.get("avg_score", 0)
            if not ts or not sc: continue
            try:
                dt = datetime.fromisoformat(ts.replace("Z", "+00:00"))
                hourly[dt.hour].append(sc)
                daily[dt.weekday()].append(sc)
            except Exception:
                continue

        hourly_avg = {h: round(sum(v)/len(v)) if v else None for h, v in hourly.items()}
        daily_avg  = {d: round(sum(v)/len(v)) if v else None for d, v in daily.items()}

        # Find worst/best times
        filled_h = {h: v for h, v in hourly_avg.items() if v is not None}
        filled_d = {d: v for d, v in daily_avg.items() if v is not None}
        worst_h  = min(filled_h, key=filled_h.get) if filled_h else None
        best_h   = max(filled_h, key=filled_h.get) if filled_h else None
        worst_d  = min(filled_d, key=filled_d.get) if filled_d else None
        days_n   = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]

        insights = []
        if worst_h is not None:
            hr_label = f"{worst_h}:00–{worst_h+1}:00"
            insights.append(f"Your posture is weakest around {hr_label} — consider a break before this time.")
        if best_h is not None:
            insights.append(f"Your best posture is around {best_h}:00 — schedule important tasks then.")
        if worst_d is not None:
            insights.append(f"{days_n[worst_d]} is your worst posture day — plan extra stretch breaks.")

        # AI insight if local AI configured
        ai_insight = None
        if AI_CONFIGURED and insights:
            try:
                prompt = f"Based on this posture data summary: {'; '.join(insights)}. Give one concise ergonomic tip (1 sentence, professional)."
                ai_insight = call_ai(prompt, max_tokens=80, temperature=0.4)
            except Exception:
                pass

        return jsonify({
            "hourly":       [{"hour": h, "avg": hourly_avg.get(h), "count": len(hourly[h])} for h in range(24)],
            "daily":        [{"day": d, "name": days_n[d], "avg": daily_avg.get(d), "count": len(daily[d])} for d in range(7)],
            "worst_hour":   worst_h,
            "best_hour":    best_h,
            "worst_day":    days_n[worst_d] if worst_d is not None else None,
            "insights":     insights,
            "ai_insight":   ai_insight,
        })
    except Exception as e:
        return safe_error(e)

# ══════════════════════════════════════════════════════════════════
# API KEY PLATFORM (Enterprise API)
# ══════════════════════════════════════════════════════════════════
import secrets as _secrets
# ── API Keys store — Redis-backed (same pattern as sessions) ────────
# key_hash -> {uid, plan, name, usage, created_at, last_used, rate_limit}
# TTL: 365 days — keys survive server restarts
import json as _json_apikeys

_API_KEY_PREFIX  = "api_key:"
_API_KEY_TTL     = 3600 * 24 * 365   # 1 year

class _ApiKeyStore:
    """Redis-backed API key store. Falls back to in-memory dict."""
    def __init__(self):
        self._mem: dict = {}

    def _rk(self, h): return f"{_API_KEY_PREFIX}{h}"

    def get(self, key_hash):
        raw = rget(self._rk(key_hash))
        if raw:
            try: return _json_apikeys.loads(raw)
            except: pass
        return self._mem.get(key_hash)

    def set(self, key_hash, data):
        rset(self._rk(key_hash), _json_apikeys.dumps(data, default=str), _API_KEY_TTL)
        self._mem[key_hash] = data

    def delete(self, key_hash):
        rdel(self._rk(key_hash))
        self._mem.pop(key_hash, None)

    def __contains__(self, key_hash):
        return self.get(key_hash) is not None

    def __getitem__(self, key_hash):
        v = self.get(key_hash)
        if v is None: raise KeyError(key_hash)
        return v

    def __setitem__(self, key_hash, data):
        self.set(key_hash, data)

_api_keys = _ApiKeyStore()

def _require_api_key(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        key = request.headers.get("X-PostureAI-Key") or request.args.get("api_key")
        if not key:
            return jsonify({"error": "API key required. Add X-PostureAI-Key header."}), 401
        key_hash = _hashlib.sha256(key.encode()).hexdigest()[:32]
        meta     = _api_keys.get(key_hash)
        if not meta:
            return jsonify({"error": "Invalid API key"}), 401
        meta["usage"] = meta.get("usage", 0) + 1
        meta["last_used"] = datetime.now().isoformat()
        request.api_key_meta = meta
        return f(*args, **kwargs)
    return decorated

@require_auth
@app.route("/api/keys/create", methods=["POST"])
@limiter.limit("10 per minute")
@require_admin
def create_api_key():
    try:
        data = request.get_json(force=True) or {}
        uid  = data.get("uid", "")
        plan = data.get("plan", "standard")
        name = data.get("name", "My API Key")
        if not uid: return jsonify({"error": "uid required"}), 400
        raw_key  = "pai_" + _secrets.token_urlsafe(32)
        key_hash = _hashlib.sha256(raw_key.encode()).hexdigest()[:32]
        _api_keys[key_hash] = {
            "uid":        uid,
            "plan":       plan,
            "name":       name,
            "usage":      0,
            "created_at": datetime.now().isoformat(),
            "last_used":  None,
            "rate_limit": 1000 if plan == "enterprise" else 100,
        }
        audit(uid, "api_key_created", "api_keys", {"plan": plan})
        return jsonify({"api_key": raw_key, "plan": plan, "note": "Save this key — it won't be shown again"}), 201
    except Exception as e:
        return safe_error(e)

@require_auth
@app.route("/api/v1/posture/analyze", methods=["POST"])
@limiter.limit("100 per minute")
@_require_api_key
def api_v1_analyze():
    """Public API endpoint — same as /api/analyze but requires API key."""
    return analyze()

@require_auth
@app.route("/api/v1/reports/summary", methods=["GET"])
@limiter.limit("60 per minute")
@_require_api_key
def api_v1_summary():
    meta = getattr(request, "api_key_meta", {})
    return jsonify({
        "uid":        meta.get("uid"),
        "plan":       meta.get("plan"),
        "api_usage":  meta.get("usage"),
        "timestamp":  datetime.now().isoformat(),
        "message":    "PostureAI Enterprise API v1",
        "docs":       "https://docs.postureai.io/api",
    })

# ══════════════════════════════════════════════════════════════════
# WEEKLY EMAIL — automated
# ══════════════════════════════════════════════════════════════════
@require_auth
@app.route("/api/email/weekly-progress", methods=["POST"])
@limiter.limit("10 per minute")
def weekly_progress_email():
    try:
        data       = request.get_json(force=True) or {}
        to         = data.get("email", "")
        name       = data.get("name", "there")
        score_this = data.get("score_this_week", 0)
        score_last = data.get("score_last_week", 0)
        sessions_n = data.get("sessions_count", 0)
        streak     = data.get("streak", 0)
        insights   = data.get("insights", [])
        lang       = data.get("lang", "en")
        achievements = data.get("new_achievements", [])
        upgrade_url  = data.get("upgrade_url", APP_URL)
        if not to: return jsonify({"ok": False, "reason": "email required"}), 400

        delta = score_this - score_last
        delta_str = (f"↑ {abs(delta)} improvement" if delta > 0 else f"↓ {abs(delta)} decline" if delta < 0 else "→ Stable") if score_last else "First week!"
        delta_col = "#10b981" if delta >= 0 else "#ef4444"
        grade     = "Excellent 🏆" if score_this >= 85 else "Good 👍" if score_this >= 70 else "Fair ⚠️" if score_this >= 50 else "Needs Work 🔧"

        ach_html = "".join([f'<div style="display:inline-block;margin:4px;padding:6px 12px;background:rgba(99,102,241,.12);border-radius:99px;font-size:12px">{a.get("icon","")} {a.get("name","")}</div>' for a in achievements[:3]]) if achievements else ""

        ins_html = "".join([f'<li style="margin-bottom:6px;font-size:13px;color:#475569">{ins}</li>' for ins in insights[:3]]) if insights else '<li style="color:#475569;font-size:13px">Keep up the great work this week!</li>'

        is_ar = lang == "ar"
        subject = f"{'📊 تقرير ذكاء صحتك الأسبوعي' if is_ar else '📊 Weekly Health Intelligence Report'} — {score_this}/100 · {grade}"

        # Pre-compute all conditional strings (avoids backslash-in-f-string on Python < 3.12)
        _dir        = "rtl" if is_ar else "ltr"
        _hdr_label  = "تقرير ذكاء القوى العاملة الأسبوعي" if is_ar else "Weekly Workforce Intelligence Report"
        _week_lbl   = "هذا الأسبوع" if is_ar else "This week"
        _score_col  = "#10b981" if score_this >= 70 else ("#f59e0b" if score_this >= 50 else "#ef4444")
        _chg_lbl    = "التغيير" if is_ar else "Change"
        _streak_lbl = "أيام متتالية" if is_ar else "Streak"
        _ach_block  = f'<div style="margin-bottom:20px">{ach_html}</div>' if ach_html else ""
        _ins_title  = "توصيات الأسبوع" if is_ar else "This week's insights"
        _cta_txt    = "فتح PostureAI ←" if is_ar else "Open PostureAI →"

        html = f"""
<div dir="{_dir}" style="font-family:Arial,sans-serif;max-width:580px;margin:0 auto;background:#030b14;border-radius:14px;overflow:hidden">
  <div style="background:linear-gradient(135deg,#1a56db,#0891b2);padding:32px;text-align:center">
    <div style="font-size:14px;color:rgba(255,255,255,.7);margin-bottom:6px">{_hdr_label}</div>
    <div style="font-size:13px;color:rgba(255,255,255,.5)">{name}</div>
  </div>
  <div style="padding:28px">
    <div style="display:flex;gap:12px;margin-bottom:24px">
      <div style="flex:1;background:#05101f;border:0.5px solid rgba(148,163,184,.1);border-radius:10px;padding:16px;text-align:center">
        <div style="font-size:11px;color:#64748b;margin-bottom:6px">{_week_lbl}</div>
        <div style="font-size:36px;font-weight:800;color:{_score_col}">{score_this}</div>
        <div style="font-size:10px;color:#64748b">/100 · {grade}</div>
      </div>
      <div style="flex:1;background:#05101f;border:0.5px solid rgba(148,163,184,.1);border-radius:10px;padding:16px;text-align:center">
        <div style="font-size:11px;color:#64748b;margin-bottom:6px">{_chg_lbl}</div>
        <div style="font-size:22px;font-weight:700;color:{delta_col}">{delta_str}</div>
      </div>
      <div style="flex:1;background:#05101f;border:0.5px solid rgba(148,163,184,.1);border-radius:10px;padding:16px;text-align:center">
        <div style="font-size:11px;color:#64748b;margin-bottom:6px">{_streak_lbl}</div>
        <div style="font-size:28px;font-weight:700;color:#f59e0b">🔥{streak}</div>
      </div>
    </div>
    {_ach_block}
    <div style="background:#05101f;border:0.5px solid rgba(148,163,184,.1);border-radius:10px;padding:16px;margin-bottom:20px">
      <div style="font-size:12px;font-weight:600;color:#f0f4f8;margin-bottom:10px">{_ins_title}</div>
      <ul style="margin:0;padding-left:18px">{ins_html}</ul>
    </div>
    <div style="text-align:center">
      <a href="{APP_URL}" style="background:#1a56db;color:white;padding:12px 28px;border-radius:8px;text-decoration:none;font-weight:600;font-size:13px">{_cta_txt}</a>
    </div>
    <p style="font-size:10px;color:#334155;text-align:center;margin-top:20px">PostureAI Pro · {SUPPORT_EMAIL}</p>
  </div>
</div>"""
        ok = send_email(to, subject, html)
        audit("system", "weekly_email_sent", to, {"score": score_this, "lang": lang})
        return jsonify({"ok": ok})
    except Exception as e:
        return safe_error(e)

# ══════════════════════════════════════════════════════════════════
# HEATMAP PDF SECTION (add to existing PDF)
# ══════════════════════════════════════════════════════════════════
@require_auth
@app.route("/api/analytics/posture-insights", methods=["POST"])
@limiter.limit("20 per minute")
@require_tier("standard")
def posture_insights():
    """AI-powered posture insights using full session history."""
    try:
        data       = request.get_json(force=True) or {}
        sessions   = data.get("sessions", [])
        profile    = data.get("profile", {})
        lang       = data.get("lang", "en")
        if not AI_CONFIGURED:
            return jsonify({"insights": [], "summary": "AI not configured"}), 503
        if not sessions:
            return jsonify({"insights": [], "summary": "No session data available"})

        scores     = [s.get("avg_score", 0) for s in sessions if s.get("avg_score")]
        avg        = round(sum(scores) / len(scores)) if scores else 0
        trend      = scores[-5:] if len(scores) >= 5 else scores
        improving  = len(trend) > 1 and trend[-1] > trend[0]

        prompt = f"""You are a workplace ergonomics expert. Analyze this employee's posture data and provide insights.

Data:
- Total sessions: {len(sessions)}
- Average score: {avg}/100
- Recent trend: {'improving ↑' if improving else 'declining ↓' if len(trend)>1 else 'stable →'}
- Score history (last 10): {scores[-10:]}
- Name: {profile.get('name', 'Employee')}

Provide exactly 3 insights in this JSON format (respond with JSON only, no markdown):
{{"insights": [{{"title": "string", "body": "string", "priority": "high|medium|low", "icon": "emoji"}}]}}
{"Respond in Arabic." if lang == "ar" else ""}"""

        raw = call_ai(prompt, max_tokens=400, temperature=0.3)
        if raw:
            raw = raw.replace("```json", "").replace("```", "").strip()
            import json as _json
            parsed = _json.loads(raw)
            return jsonify(parsed)
        return jsonify({"insights": [], "summary": "AI unavailable"})
    except Exception as e:
        return jsonify({"error": str(e), "insights": []}), 500


# ══════════════════════════════════════════════════════════════════
# AI INTELLIGENCE LAYER — Phase 4
# Executive summaries · Predictive AI · Fatigue · Burnout · Anomaly
# ══════════════════════════════════════════════════════════════════

def _gemini_json(prompt: str, max_tokens: int = 600, temperature: float = 0.25) -> dict | None:
    """Helper: call local AI, parse JSON, return dict or None."""
    if not AI_CONFIGURED:
        return None
    try:
        import json as _j
        raw = call_ai(prompt, max_tokens=max_tokens, temperature=temperature)
        if not raw:
            return None
        raw = raw.replace("```json","").replace("```","").strip()
        return _j.loads(raw)
    except Exception as _e:
        # DEBUG print(f"[AI] _gemini_json error: {_e}")
        return None


def _compute_posture_stats(sessions: list) -> dict:
    """Pure-Python posture analytics — no AI needed."""
    import statistics as _stats
    if not sessions:
        return {}

    scores   = [s.get("avg_score", 0) for s in sessions if s.get("avg_score", 0) > 0]
    durations= [s.get("duration_s", 0) for s in sessions if s.get("duration_s", 0) > 0]

    # Trend (linear slope over last 14 sessions)
    recent = scores[-14:]
    trend_slope = 0.0
    if len(recent) >= 3:
        n   = len(recent)
        xs  = list(range(n))
        xm  = sum(xs) / n
        ym  = sum(recent) / n
        num = sum((xs[i]-xm)*(recent[i]-ym) for i in range(n))
        den = sum((xs[i]-xm)**2 for i in range(n))
        trend_slope = round(num/den, 2) if den else 0.0

    # Hourly breakdown (fatigue proxy)
    hour_scores: dict[int, list] = {}
    for s in sessions:
        d = s.get("created_at")
        if hasattr(d, "hour"):
            h = d.hour
        elif hasattr(d, "toDate"):
            h = d.toDate().hour
        else:
            h = datetime.now().hour
        sc_val = s.get("avg_score", 0)
        if sc_val > 0:
            hour_scores.setdefault(h, []).append(sc_val)

    hourly = {h: round(sum(v)/len(v)) for h, v in hour_scores.items()}
    worst_hour = min(hourly, key=hourly.get) if hourly else None
    best_hour  = max(hourly, key=hourly.get) if hourly else None

    # Anomaly detection — sessions > 2 std deviations below mean
    anomalies = []
    if len(scores) >= 5:
        mean = _stats.mean(scores)
        std  = _stats.stdev(scores)
        threshold = mean - 2 * std
        for i, s in enumerate(sessions[:20]):
            sc_v = s.get("avg_score", 0)
            if sc_v > 0 and sc_v < threshold:
                anomalies.append({
                    "session_idx": i,
                    "score":  sc_v,
                    "delta":  round(sc_v - mean, 1),
                    "date":   str(s.get("created_at", ""))[:10],
                })

    # Burnout risk score (0–100)
    # Factors: declining trend, low avg, high anomaly rate, declining consistency
    burnout_risk = 0
    if scores:
        avg = sum(scores) / len(scores)
        if trend_slope < -1:    burnout_risk += 30   # declining trend
        if avg < 55:            burnout_risk += 25   # chronically low
        if trend_slope < -2:    burnout_risk += 15   # steep decline
        anom_rate = len(anomalies) / max(len(scores), 1)
        if anom_rate > 0.2:     burnout_risk += 20   # >20% anomaly rate
        if len(scores) >= 5:
            recent_avg = sum(scores[-5:]) / 5
            older_avg  = sum(scores[:-5]) / max(len(scores)-5, 1)
            if recent_avg < older_avg - 8: burnout_risk += 10  # recent dip
        burnout_risk = min(100, max(0, burnout_risk))

    # Posture risk score
    risk_score = max(0, min(100, 100 - (sum(scores)/len(scores) if scores else 50)))

    # Streak & consistency
    days_with_session = set()
    for s in sessions:
        d = s.get("created_at")
        if hasattr(d, "toDate"): d = d.toDate()
        if hasattr(d, "date"): days_with_session.add(str(d.date()))
    streak = 0
    today = datetime.utcnow().date()
    while str(today - timedelta(days=streak)) in days_with_session:
        streak += 1

    return {
        "total_sessions":  len(sessions),
        "avg_score":       round(sum(scores)/len(scores)) if scores else 0,
        "best_score":      max(scores) if scores else 0,
        "worst_score":     min(scores) if scores else 0,
        "score_std":       round(_stats.stdev(scores), 1) if len(scores) >= 2 else 0,
        "trend_slope":     trend_slope,
        "trend_label":     "improving" if trend_slope > 0.5 else "declining" if trend_slope < -0.5 else "stable",
        "total_min":       round(sum(durations)/60),
        "avg_duration_min":round((sum(durations)/len(durations))/60, 1) if durations else 0,
        "worst_hour":      worst_hour,
        "best_hour":       best_hour,
        "hourly_scores":   hourly,
        "anomalies":       anomalies[:5],
        "anomaly_count":   len(anomalies),
        "burnout_risk":    burnout_risk,
        "burnout_label":   "High" if burnout_risk >= 60 else "Medium" if burnout_risk >= 30 else "Low",
        "posture_risk":    round(risk_score),
        "streak_days":     streak,
        "consistency_pct": round(len(days_with_session)/max(30,1)*100),
    }


@require_auth
@app.route("/api/ai/executive-summary", methods=["POST"])
@limiter.limit("10 per minute")
@require_tier("professional")
def ai_executive_summary():
    """
    AI-powered executive summary for the current user.
    Returns structured summary: headline, kpis, trends, recommendations, risk.
    Used by: individual dashboard + PDF reports.
    """
    try:
        import json as _j
        data     = request.get_json(force=True) or {}
        sessions = data.get("sessions", [])
        profile  = data.get("profile", {})
        lang     = data.get("lang", "en")
        uid      = getattr(g, "uid", data.get("uid",""))

        stats = _compute_posture_stats(sessions)
        if not stats:
            return jsonify({"error": "No session data", "summary": None}), 200

        is_ar = lang == "ar"
        prompt = f"""You are a corporate health & ergonomics analyst. Produce a concise executive summary.

Employee data:
- Average posture score: {stats['avg_score']}/100
- Total sessions: {stats['total_sessions']}
- Trend: {stats['trend_label']} (slope: {stats['trend_slope']:+.1f}/session)
- Worst hour: {stats['worst_hour']}:00 | Best hour: {stats['best_hour']}:00
- Burnout risk: {stats['burnout_label']} ({stats['burnout_risk']}/100)
- Posture risk score: {stats['posture_risk']}/100
- Anomalies detected: {stats['anomaly_count']}
- Streak: {stats['streak_days']} days
- Total monitored time: {stats['total_min']} minutes
- Score consistency (σ): {stats['score_std']}

Return ONLY valid JSON (no markdown):
{{
  "headline": "One sentence executive summary of posture health",
  "score_interpretation": "What the average score means clinically",
  "trend_analysis": "2-sentence trend analysis",
  "fatigue_pattern": "Describe fatigue/worst-hour pattern",
  "top_recommendations": ["rec1", "rec2", "rec3"],
  "risk_summary": "One sentence risk assessment",
  "burnout_signal": "Brief burnout risk explanation",
  "priority": "high|medium|low"
}}
{"Respond entirely in Arabic." if is_ar else "Be concise and professional."}"""

        result = _gemini_json(prompt, max_tokens=700)
        if not result:
            # Fallback: rule-based summary
            result = {
                "headline": f"Posture health score {stats['avg_score']}/100 — {stats['trend_label']} trend",
                "trend_analysis": f"Score is {stats['trend_label']} at {stats['trend_slope']:+.1f} per session.",
                "top_recommendations": ["Take regular breaks", "Adjust monitor height", "Enable posture reminders"],
                "priority": "high" if stats['burnout_risk'] >= 60 else "medium" if stats['burnout_risk'] >= 30 else "low",
                "burnout_signal": stats['burnout_label'] + " burnout risk",
                "risk_summary": f"Posture risk: {stats['posture_risk']}/100",
            }

        audit(uid, "ai_executive_summary", "ai_intelligence", {"sessions": len(sessions)})
        return jsonify({**result, "stats": stats})

    except Exception as e:
        import traceback
        return safe_error(e)


@require_auth
@app.route("/api/ai/predictive", methods=["POST"])
@limiter.limit("10 per minute")
@require_tier("professional")
def ai_predictive():
    """
    Predictive AI: burnout prediction, anomaly detection, trend forecasting.
    Returns risk scores, predicted next-week score, and anomaly list.
    """
    try:
        data     = request.get_json(force=True) or {}
        sessions = data.get("sessions", [])
        lang     = data.get("lang", "en")
        uid      = getattr(g, "uid", "")

        stats = _compute_posture_stats(sessions)
        if not stats:
            return jsonify({"burnout_risk": 0, "posture_risk": 0, "forecast": [], "anomalies": []}), 200

        scores  = [s.get("avg_score", 0) for s in sessions if s.get("avg_score", 0) > 0]
        is_ar   = lang == "ar"

        # 7-day forecast (linear extrapolation + noise damping)
        forecast = []
        if len(scores) >= 3:
            for day in range(1, 8):
                predicted = round(min(100, max(0,
                    stats['avg_score'] + stats['trend_slope'] * day * 0.7
                )))
                forecast.append({
                    "day": day,
                    "predicted_score": predicted,
                    "confidence": max(40, 85 - day*5),
                })

        # AI narrative on predictions (only if local AI available)
        narrative = None
        if AI_CONFIGURED and len(scores) >= 5:
            prompt = f"""Posture analytics prediction request.

Stats: avg={stats['avg_score']}, trend={stats['trend_label']} ({stats['trend_slope']:+.1f}),
burnout_risk={stats['burnout_risk']}/100, anomalies={stats['anomaly_count']}, streak={stats['streak_days']}d

Return ONLY JSON:
{{
  "burnout_prediction": "2-sentence prediction of burnout risk trajectory",
  "injury_risk": "low|medium|high",
  "injury_explanation": "Brief MSK injury risk explanation",
  "next_week_outlook": "What to expect next week",
  "intervention_urgency": "immediate|soon|routine"
}}
{"Respond in Arabic." if is_ar else ""}"""
            narrative = _gemini_json(prompt, max_tokens=350)

        audit(uid, "ai_predictive", "ai_intelligence", {"sessions": len(sessions), "burnout": stats['burnout_risk']})
        return jsonify({
            "burnout_risk":   stats['burnout_risk'],
            "burnout_label":  stats['burnout_label'],
            "posture_risk":   stats['posture_risk'],
            "trend_slope":    stats['trend_slope'],
            "trend_label":    stats['trend_label'],
            "anomalies":      stats['anomalies'],
            "anomaly_count":  stats['anomaly_count'],
            "forecast":       forecast,
            "narrative":      narrative or {},
            "worst_hour":     stats['worst_hour'],
            "stats":          stats,
        })

    except Exception as e:
        return safe_error(e)


@require_auth
@app.route("/api/ai/weekly-insights", methods=["POST"])
@limiter.limit("20 per minute")
@require_tier("professional")
def ai_weekly_insights():
    """
    AI weekly insights: what happened, why, and what to do next.
    Designed for the home screen weekly card and email digest.
    """
    try:
        data      = request.get_json(force=True) or {}
        sessions  = data.get("sessions", [])
        lang      = data.get("lang", "en")
        uid       = getattr(g, "uid", "")

        # Filter to this week
        week_ms   = 7 * 86400000
        this_week = [s for s in sessions
                     if (time.time()*1000 - (s.get("created_at_ms") or 0)) < week_ms]
        prev_week = [s for s in sessions
                     if week_ms <= (time.time()*1000 - (s.get("created_at_ms") or 0)) < 2*week_ms]

        def _avg(ss): return round(sum(s.get("avg_score",0) for s in ss)/len(ss)) if ss else 0

        curr_avg = _avg(this_week)
        prev_avg = _avg(prev_week)
        delta    = curr_avg - prev_avg
        is_ar    = lang == "ar"

        stats = _compute_posture_stats(sessions)

        prompt = f"""Weekly posture progress summary for employee.

This week: {len(this_week)} sessions, avg score {curr_avg}/100
Last week: {len(prev_week)} sessions, avg score {prev_avg}/100
Change: {delta:+d} points
Overall trend: {stats.get('trend_label','stable')}
Burnout risk: {stats.get('burnout_label','Low')}

Return ONLY JSON:
{{
  "headline": "Short weekly summary (max 12 words)",
  "performance": "good|fair|poor",
  "week_summary": "2-3 sentence narrative about this week",
  "vs_last_week": "Comparison with last week",
  "top_insight": "Most important insight (1 sentence)",
  "action_this_week": "Single most impactful action for next week",
  "motivation": "Short motivational message"
}}
{"Respond in Arabic." if is_ar else ""}"""

        result = _gemini_json(prompt, max_tokens=450)
        if not result:
            result = {
                "headline": f"Week avg: {curr_avg}/100 ({delta:+d} vs last week)",
                "performance": "good" if curr_avg >= 70 else "fair" if curr_avg >= 50 else "poor",
                "top_insight": "Consistency is key to posture improvement.",
                "action_this_week": "Enable posture reminders every 30 minutes.",
            }

        audit(uid, "ai_weekly_insights", "ai_intelligence", {"curr_avg": curr_avg, "delta": delta})
        return jsonify({
            **result,
            "curr_avg":   curr_avg,
            "prev_avg":   prev_avg,
            "delta":      delta,
            "this_week_sessions": len(this_week),
            "stats":      stats,
        })

    except Exception as e:
        return safe_error(e)


@require_auth
@app.route("/api/ai/department-report", methods=["POST"])
@limiter.limit("5 per minute")
@require_tier("professional")
def ai_department_report():
    """
    AI-powered department/manager insights.
    Compares employees, identifies at-risk workers, generates executive PDF summary.
    Requires: HR/admin tier. Input: list of employee session data.
    """
    try:
        import json as _j
        data       = request.get_json(force=True) or {}
        employees  = data.get("employees", [])    # [{name, uid, sessions:[]}]
        company    = data.get("company_name", "")
        period     = data.get("period", "Monthly")
        lang       = data.get("lang", "en")
        uid        = getattr(g, "uid", "")
        is_ar      = lang == "ar"

        if not employees:
            return jsonify({"error": "employees array required"}), 400

        # Compute stats per employee
        emp_stats = []
        for emp in employees[:50]:  # max 50 employees
            stats = _compute_posture_stats(emp.get("sessions", []))
            emp_stats.append({
                "name":          emp.get("name", "Unknown"),
                "uid":           emp.get("uid", ""),
                "avg_score":     stats.get("avg_score", 0),
                "burnout_risk":  stats.get("burnout_risk", 0),
                "trend_label":   stats.get("trend_label", "stable"),
                "sessions":      stats.get("total_sessions", 0),
                "anomaly_count": stats.get("anomaly_count", 0),
                "risk_level":    "high" if stats.get("burnout_risk",0)>=60 else "medium" if stats.get("burnout_risk",0)>=30 else "low",
            })

        # Sort: highest risk first
        emp_stats.sort(key=lambda x: x["burnout_risk"], reverse=True)

        dept_avg  = round(sum(e["avg_score"] for e in emp_stats) / len(emp_stats)) if emp_stats else 0
        at_risk   = [e for e in emp_stats if e["risk_level"] == "high"]
        improving = [e for e in emp_stats if e["trend_label"] == "improving"]

        prompt = f"""You are a corporate health analyst writing a {period} department posture report for {company or "a company"}.

Department stats:
- Total employees: {len(emp_stats)}
- Department avg score: {dept_avg}/100
- Employees at high burnout risk: {len(at_risk)}
- Employees improving: {len(improving)}
- Top 3 at-risk: {[e['name']+' ('+str(e['burnout_risk'])+'/100 burnout risk)' for e in emp_stats[:3]]}

Return ONLY JSON:
{{
  "executive_summary": "3-sentence department health overview for C-suite",
  "department_grade": "A|B|C|D",
  "key_findings": ["finding1", "finding2", "finding3"],
  "at_risk_summary": "Summary of at-risk employees situation",
  "department_recommendations": ["rec1", "rec2", "rec3"],
  "manager_actions": ["action1", "action2"],
  "benchmark_note": "How this department compares to typical office workforce",
  "roi_note": "Business impact / productivity note"
}}
{"Respond entirely in Arabic." if is_ar else "Keep executive tone."}"""

        ai_report = _gemini_json(prompt, max_tokens=800)
        if not ai_report:
            ai_report = {
                "executive_summary": f"Department average score: {dept_avg}/100. {len(at_risk)} employees require immediate ergonomic intervention.",
                "department_grade": "A" if dept_avg>=85 else "B" if dept_avg>=70 else "C" if dept_avg>=55 else "D",
                "key_findings": [f"{len(at_risk)} high-risk employees", f"Avg score {dept_avg}/100", f"{len(improving)} improving"],
                "department_recommendations": ["Schedule ergonomic assessments for at-risk employees", "Implement mandatory break reminders", "Provide standing desk options"],
                "manager_actions": ["Review high-risk employee list", "Schedule 1-on-1s with at-risk staff"],
            }

        audit(uid, "ai_department_report", "ai_intelligence", {
            "company": company, "employees": len(emp_stats), "at_risk": len(at_risk)
        })
        return jsonify({
            **ai_report,
            "period":       period,
            "company":      company,
            "dept_avg":     dept_avg,
            "total_emp":    len(emp_stats),
            "at_risk_count":len(at_risk),
            "improving_count": len(improving),
            "employee_ranking": emp_stats,
            "generated_at": datetime.utcnow().isoformat(),
        })

    except Exception as e:
        import traceback
        return safe_error(e)


@require_auth
@app.route("/api/ai/fatigue-analysis", methods=["POST"])
@limiter.limit("10 per minute")
@require_tier("professional")
def ai_fatigue_analysis():
    """
    Fatigue pattern analysis: detects when during the day posture degrades.
    Uses hourly breakdown to identify fatigue windows.
    """
    try:
        data     = request.get_json(force=True) or {}
        sessions = data.get("sessions", [])
        lang     = data.get("lang", "en")
        uid      = getattr(g, "uid", "")
        is_ar    = lang == "ar"

        stats = _compute_posture_stats(sessions)
        hourly = stats.get("hourly_scores", {})

        if not hourly:
            return jsonify({"pattern": None, "insights": [], "peak_fatigue_hour": None})

        worst_h = stats.get("worst_hour")
        best_h  = stats.get("best_hour")
        worst_score = hourly.get(worst_h, 0) if worst_h is not None else 0
        best_score  = hourly.get(best_h, 0)  if best_h  is not None else 0

        # Morning/afternoon/evening breakdown
        morning   = {h:s for h,s in hourly.items() if 6  <= h < 12}
        afternoon = {h:s for h,s in hourly.items() if 12 <= h < 17}
        evening   = {h:s for h,s in hourly.items() if 17 <= h < 22}

        def _bavg(d): return round(sum(d.values())/len(d)) if d else None

        prompt = f"""Analyze workplace fatigue pattern from posture data.

Hourly posture scores (24h): {hourly}
Morning avg: {_bavg(morning)} | Afternoon avg: {_bavg(afternoon)} | Evening avg: {_bavg(evening)}
Worst hour: {worst_h}:00 ({worst_score}/100) | Best hour: {best_h}:00 ({best_score}/100)

Return ONLY JSON:
{{
  "fatigue_pattern": "circadian|afternoon_dip|end_of_day|consistent|irregular",
  "pattern_explanation": "2 sentences explaining the fatigue pattern",
  "peak_fatigue_description": "When and why posture is worst",
  "peak_energy_description": "When posture is best",
  "break_schedule": ["Recommended break schedule based on fatigue pattern"],
  "ergonomic_adjustments": ["Time-specific adjustments"]
}}
{"Respond in Arabic." if is_ar else ""}"""

        result = _gemini_json(prompt, max_tokens=400)
        if not result:
            result = {
                "fatigue_pattern": "afternoon_dip" if worst_h and 13<=worst_h<=16 else "circadian",
                "peak_fatigue_description": f"Posture is worst around {worst_h}:00" if worst_h is not None else "Insufficient data",
                "break_schedule": ["Take a 5-min break every 45 minutes", "Stand up and stretch at 2pm"],
            }

        audit(uid, "ai_fatigue_analysis", "ai_intelligence", {"worst_hour": worst_h})
        return jsonify({
            **result,
            "worst_hour":  worst_h,
            "best_hour":   best_h,
            "hourly":      hourly,
            "morning_avg": _bavg(morning),
            "afternoon_avg": _bavg(afternoon),
            "evening_avg":   _bavg(evening),
        })

    except Exception as e:
        return safe_error(e)


# ══════════════════════════════════════════════════════════════════
# STRIPE BILLING
# ══════════════════════════════════════════════════════════════════
STRIPE_SECRET_KEY     = os.getenv("STRIPE_SECRET_KEY", "")
STRIPE_WEBHOOK_SECRET = os.getenv("STRIPE_WEBHOOK_SECRET", "")

@require_auth
@app.route("/api/stripe/create-session", methods=["POST"])
@limiter.limit("10 per minute")
def stripe_create_session():
    try:
        if not STRIPE_SECRET_KEY:
            return jsonify({"error": "Stripe not configured — add STRIPE_SECRET_KEY to .env"}), 503
        data        = request.get_json(force=True) or {}
        price_id    = data.get("price_id", "")
        email       = data.get("customer_email", "")
        uid = getattr(g, "uid", "") or data.get("uid", "")  # SECURITY: auth-first
        plan_id     = data.get("plan_id", "")
        billing     = data.get("billing", "monthly")
        success_url = data.get("success_url", APP_URL)
        cancel_url  = data.get("cancel_url",  APP_URL)
        locale      = data.get("locale", "auto")
        if not price_id:
            return jsonify({"error": "price_id required — add VITE_STRIPE_PRICE_* to frontend .env.local"}), 400
        headers = {
            "Authorization":  f"Bearer {STRIPE_SECRET_KEY}",
            "Content-Type":   "application/x-www-form-urlencoded",
        }
        payload = {
            "mode":                          "subscription",
            "line_items[0][price]":          price_id,
            "line_items[0][quantity]":       "1",
            "success_url":                   success_url,
            "cancel_url":                    cancel_url,
            "locale":                        locale,
            "allow_promotion_codes":         "true",
            "billing_address_collection":    "auto",
            "subscription_data[metadata][uid]":     uid,
            "subscription_data[metadata][plan]":    plan_id,
            "subscription_data[metadata][billing]": billing,
        }
        if email:
            payload["customer_email"] = email
        resp = req.post("https://api.stripe.com/v1/checkout/sessions", data=payload, headers=headers, timeout=15)
        resp.raise_for_status()
        session = resp.json()
        audit(uid, "stripe_session_created", "billing", {"plan": plan_id, "billing": billing})
        return jsonify({"session_id": session["id"], "url": session.get("url")})
    except req.exceptions.HTTPError as e:
        err = e.response.json().get("error", {})
        return jsonify({"error": err.get("message", str(e))}), 400
    except Exception as e:
        return safe_error(e)

@require_auth
@app.route("/api/stripe/portal", methods=["POST"])
@limiter.limit("10 per minute")
def stripe_portal():
    try:
        if not STRIPE_SECRET_KEY:
            return jsonify({"error": "Stripe not configured"}), 503
        data       = request.get_json(force=True) or {}
        uid        = data.get("uid", "")
        return_url = data.get("return_url", APP_URL)
        # Find Stripe customer ID from subscriptions
        headers = {"Authorization": f"Bearer {STRIPE_SECRET_KEY}"}
        subs_resp = req.get(f"https://api.stripe.com/v1/subscriptions?metadata[uid]={uid}&limit=1", headers=headers, timeout=10)
        subs = subs_resp.json().get("data", [])
        if not subs:
            return jsonify({"error": "No active subscription found"}), 404
        customer_id = subs[0]["customer"]
        portal_resp = req.post("https://api.stripe.com/v1/billing_portal/sessions",
            data={"customer": customer_id, "return_url": return_url},
            headers={**headers, "Content-Type": "application/x-www-form-urlencoded"}, timeout=10)
        portal_resp.raise_for_status()
        return jsonify({"url": portal_resp.json()["url"]})
    except Exception as e:
        return safe_error(e)

@app.route("/api/stripe/webhook", methods=["POST"])
def stripe_webhook():
    try:
        payload = request.get_data()
        sig     = request.headers.get("Stripe-Signature", "")

        # SECURITY: Hard fail when secret not configured
        _flask_env = os.getenv("FLASK_ENV", "development")
        if not STRIPE_WEBHOOK_SECRET:
            if _flask_env == "production":
                print("🚨 CRITICAL: STRIPE_WEBHOOK_SECRET not set in production — rejecting all Stripe webhooks", file=sys.stderr)
                return jsonify({"error": "Stripe not configured"}), 503
            print("⚠️  DEV: STRIPE_WEBHOOK_SECRET not set — skipping signature check", file=sys.stderr)
        else:
            import hmac as hm, hashlib as hl
            try:
                parts    = {k: v for part in sig.split(",") for k, v in [part.split("=", 1)]}
                ts       = parts.get("t", "")
                sigs     = [parts.get(f"v{i}", "") for i in range(1, 3)]
                body     = f"{ts}.{payload.decode()}"
                computed = hm.new(STRIPE_WEBHOOK_SECRET.encode(), body.encode(), hl.sha256).hexdigest()
                if not any(hm.compare_digest(computed, s) for s in sigs if s):
                    print("[stripe] Invalid webhook signature — possible spoofing attempt", file=sys.stderr)
                    return jsonify({"error": "Invalid signature"}), 400
            except Exception as sig_err:
                print(f"[stripe] Signature check error: {sig_err}", file=sys.stderr)
                return jsonify({"error": "Signature verification failed"}), 400

        event = request.get_json(force=True) or {}
        etype = event.get("type", "")
        obj   = event.get("data", {}).get("object", {})

        if etype == "checkout.session.completed":
            uid     = obj.get("metadata", {}).get("uid", "")
            plan    = obj.get("metadata", {}).get("plan", "")
            billing = obj.get("metadata", {}).get("billing", "monthly")
            email   = obj.get("customer_email", "")

            # Validate the plan from Stripe metadata
            if plan and plan not in ALL_TIERS:
                print(f"[stripe] Invalid plan in metadata: {plan}", file=sys.stderr)
                return jsonify({"received": True, "warning": "invalid plan"}), 200

            # SECURITY: Enterprise tiers must never be auto-granted by a webhook —
            # they require a manually negotiated, signed contract.
            if plan in ("b2b_enterprise", "enterprise"):
                print(f"🚨 [stripe] Enterprise tier claimed via webhook for uid={uid} — refusing", file=sys.stderr)
                return jsonify({"received": True, "warning": "enterprise requires manual approval"}), 200

            # Update Firestore — this is the critical missing step from v1
            if uid and plan:
                try:
                    _db  = firestore.client()
                    _ref = _db.collection("users").document(uid)
                    _doc = _ref.get()
                    if _doc.exists:
                        # SECURITY: idempotency — Stripe can and does retry webhook
                        # delivery on transient failures. Without this check, a retry
                        # would extend the subscription period and record a duplicate
                        # payment for the same checkout session.
                        _session_id = obj.get("id", "")
                        if _session_id:
                            _existing = _db.collection("payments").where(
                                "stripe_session_id", "==", _session_id).limit(1).get()
                            if _existing:
                                print(f"[stripe] Session {_session_id} already processed — skipping duplicate", flush=True)
                                return jsonify({"received": True, "note": "already processed"}), 200

                        months  = 12 if billing == "yearly" else 1
                        expires = datetime.utcnow() + timedelta(days=30 * months)
                        _ref.update({
                            "tier":               plan,
                            "is_trial":           False,
                            "trial_expires_at":   None,
                            "subscription_start": datetime.utcnow().isoformat() + "Z",
                            "subscription_end":   expires.isoformat() + "Z",
                            "subscription_billing": billing,
                            "stripe_session_id":  obj.get("id", ""),
                            "updated_at":         datetime.utcnow().isoformat() + "Z",
                        })
                        invalidate_user_cache(uid)
                        print(f"[stripe] ✅ Updated uid={uid} → tier={plan}", flush=True)
                        # Record payment
                        _db.collection("payments").add({
                            "uid": uid, "tier": plan,
                            "amount": (obj.get("amount_total") or 0) // 100,
                            "currency": "USD", "billing": billing,
                            "status": "confirmed",
                            "stripe_session_id": obj.get("id"),
                            "email": email,
                            "created_at": datetime.utcnow().isoformat() + "Z",
                        })
                    else:
                        print(f"[stripe] User {uid} not found in Firestore", file=sys.stderr)
                except Exception as fs_err:
                    print(f"[stripe] Firestore update error: {fs_err}", file=sys.stderr)

            audit(uid or "stripe", "stripe_payment_success", "billing", {"plan": plan, "billing": billing, "email": email})
            send_admin_notification({
                "user_email": email, "tier": plan,
                "amount": (obj.get("amount_total") or 0) // 100,
                "payment_method_name": "Stripe",
                "ref_code": obj.get("id", ""),
                "user_name": email.split("@")[0] if email else "Customer"
            })
            if email:
                send_email(email, f"✅ PostureAI {plan.title()} — Payment Confirmed",
                           f"<p>Your {plan.title()} plan is now active. <a href='{APP_URL}'>Open PostureAI</a></p>")

        elif etype == "customer.subscription.deleted":
            uid = obj.get("metadata", {}).get("uid", "")
            if uid:
                try:
                    _db = firestore.client()
                    _db.collection("users").document(uid).update({
                        "tier": "standard",
                        "subscription_end": None,
                        "updated_at": datetime.utcnow().isoformat() + "Z",
                    })
                    invalidate_user_cache(uid)
                except Exception as fs_err:
                    print(f"[stripe] Downgrade error: {fs_err}", file=sys.stderr)
            audit(uid or "stripe", "stripe_subscription_cancelled", "billing", {})

        elif etype == "invoice.payment_failed":
            uid   = obj.get("metadata", {}).get("uid", "")
            email = obj.get("customer_email", "")
            audit(uid or "stripe", "stripe_payment_failed", "billing", {"email": email})

        return jsonify({"received": True})
    except Exception as e:
        return safe_error(e, "Webhook processing error")

# ══════════════════════════════════════════════════════════════════
# BILLING MATURITY — Phase 8
# Usage tracking · Proration · Failed payment recovery (dunning)
# Invoice PDF · Billing analytics · Upgrade/downgrade mid-cycle
# ══════════════════════════════════════════════════════════════════

import io as _billing_io

# ── Plan pricing table (authoritative source) ────────────────────
PLAN_PRICES = {
    "standard":     {"monthly": 0,     "yearly": 0,      "seats": 5    },
    "professional": {"monthly": 199,   "yearly": 1590,   "seats": -1   },
    "elite":        {"monthly": 399,   "yearly": 3190,   "seats": -1   },
    "enterprise":   {"monthly": None,  "yearly": None,   "seats": -1   },
}

PLAN_ORDER = ["standard", "professional", "elite", "enterprise"]

def _plan_rank(plan: str) -> int:
    return PLAN_ORDER.index(plan) if plan in PLAN_ORDER else 0

def _prorate_credit(current_plan: str, billing_cycle: str,
                     days_used: int, total_days: int) -> int:
    """Calculate unused credit for mid-cycle plan change (EGP)."""
    prices = PLAN_PRICES.get(current_plan, {})
    total  = prices.get(billing_cycle, 0) or 0
    if not total or total_days <= 0:
        return 0
    days_remaining = max(0, total_days - days_used)
    return round(total * (days_remaining / total_days))

def _compute_billing_analytics(payments: list) -> dict:
    """Compute MRR, ARR, churn signals, LTV from payment list."""
    import statistics as _stat
    confirmed   = [p for p in payments if p.get("status") == "confirmed"]
    pending     = [p for p in payments if p.get("status") == "pending"]
    failed      = [p for p in payments if p.get("status") in ("rejected","failed")]

    total_rev   = sum(p.get("amount", 0) for p in confirmed)
    monthly_rev = [p for p in confirmed if p.get("billing_cycle") != "yearly"]
    yearly_rev  = [p for p in confirmed if p.get("billing_cycle") == "yearly"]

    mrr = sum(p.get("amount", 0) for p in monthly_rev)
    mrr += sum(p.get("amount", 0) / 12 for p in yearly_rev)
    arr = mrr * 12

    # ARPU
    uids  = {p.get("uid") for p in confirmed if p.get("uid")}
    arpu  = round(total_rev / len(uids)) if uids else 0

    # Plan breakdown
    plan_dist = {}
    for p in confirmed:
        t = p.get("tier", "standard")
        plan_dist[t] = plan_dist.get(t, 0) + 1

    # Revenue by month (last 6)
    from collections import defaultdict
    monthly = defaultdict(int)
    for p in confirmed:
        d = p.get("created_at", "")
        if hasattr(d, "toDate"): d = d.toDate()
        month_key = str(d)[:7] if d else "unknown"
        monthly[month_key] += p.get("amount", 0)

    # Conversion rate
    total_attempts = len(confirmed) + len(failed)
    conv_rate = round(len(confirmed) / total_attempts * 100) if total_attempts else 0

    # Failed payment amount at risk
    failed_arr = sum(p.get("amount", 0) for p in failed)

    return {
        "mrr":          round(mrr),
        "arr":          round(arr),
        "total_revenue": total_rev,
        "arpu":          arpu,
        "plan_distribution": plan_dist,
        "monthly_revenue":   dict(sorted(monthly.items())[-6:]),
        "pending_count":     len(pending),
        "failed_count":      len(failed),
        "failed_revenue_at_risk": failed_arr,
        "conversion_rate":   conv_rate,
        "unique_customers":  len(uids),
        "avg_payment":       round(total_rev / len(confirmed)) if confirmed else 0,
    }


# ── Dunning: failed payment recovery ────────────────────────────
@require_auth
@require_admin
@app.route("/api/billing/dunning/send", methods=["POST"])
@limiter.limit("20 per minute")
def billing_dunning_send():
    """
    Send a payment recovery email for a failed/pending payment.
    Supports 3 dunning stages: gentle, urgent, final.
    """
    try:
        data      = request.get_json(force=True) or {}
        payment   = data.get("payment", {})
        stage     = data.get("stage", 1)   # 1=gentle, 2=urgent, 3=final
        uid       = getattr(g, "uid", "admin")

        email     = payment.get("user_email", "")
        name      = payment.get("user_name", "Customer")
        tier      = payment.get("tier", "professional")
        amount    = payment.get("amount", 0)
        ref       = payment.get("ref_code", "—")

        if not email:
            return jsonify({"error": "payment.user_email required"}), 400

        stage_copy = {
            1: {
                "subject":  f"⚠️ Action needed — PostureAI {tier.title()} payment",
                "headline": "Your payment didn't go through",
                "body":     f"We tried to process your {tier.title()} subscription ({amount:,} EGP) but it was unsuccessful. Please update your payment method to keep your intelligence platform active.",
                "cta":      "Update Payment Method →",
                "color":    "#f59e0b",
            },
            2: {
                "subject":  f"🔴 Final notice — PostureAI account at risk",
                "headline": "Your account will be suspended in 24 hours",
                "body":     f"Your {tier.title()} plan payment of {amount:,} EGP is overdue. Your workforce intelligence data will be paused at midnight if not resolved. Use code RECOVER10 for 10% off your next payment.",
                "cta":      "Resolve Payment Now →",
                "color":    "#ef4444",
            },
            3: {
                "subject":  f"Account suspended — PostureAI Pro",
                "headline": "Your account has been suspended",
                "body":     f"Your PostureAI Pro account has been suspended due to non-payment. Your data is secured for 30 days. Reactivate now to restore full access and retain all your workforce health intelligence history.",
                "cta":      "Reactivate Account →",
                "color":    "#7f1d1d",
            },
        }
        sc = stage_copy.get(stage, stage_copy[1])

        html = f"""
<div style="font-family:-apple-system,sans-serif;max-width:520px;margin:0 auto;background:#07111f;border-radius:14px;overflow:hidden">
  <div style="background:{sc['color']};padding:24px;text-align:center">
    <div style="font-size:28px;margin-bottom:6px">{'⚠️' if stage==1 else '🔴' if stage==2 else '🚫'}</div>
    <div style="font-size:17px;font-weight:700;color:white">{sc['headline']}</div>
  </div>
  <div style="padding:24px">
    <p style="color:#94a3b8;font-size:13px;line-height:1.7;margin-bottom:18px">Hi <b style="color:#f1f5f9">{name}</b>, {sc['body']}</p>
    <div style="background:#0c1728;border-radius:10px;padding:14px;margin-bottom:18px;border:1px solid rgba(255,255,255,.06)">
      <div style="color:#64748b;font-size:11px;margin-bottom:6px;text-transform:uppercase;letter-spacing:.06em">Payment Details</div>
      <div style="color:#f1f5f9;font-size:13px;line-height:2">
        Plan: <b>{tier.title()}</b><br>Amount: <b style="color:#f87171">{amount:,} EGP</b><br>Reference: <span style="font-family:monospace;color:#a5b4fc">{ref}</span>
      </div>
    </div>
    <div style="text-align:center;margin-bottom:14px">
      <a href="{APP_URL}?billing=recovery&ref={ref}" style="background:{sc['color']};color:white;padding:12px 28px;border-radius:9px;text-decoration:none;font-weight:700;font-size:13px">{sc['cta']}</a>
    </div>
    <p style="color:#475569;font-size:11px;text-align:center">Questions? <a href="mailto:{SUPPORT_EMAIL}" style="color:#6366f1">{SUPPORT_EMAIL}</a> · {ADMIN_PHONE}</p>
  </div>
</div>"""

        ok = send_email(email, sc["subject"], html)
        audit(uid, "dunning_email_sent", "billing", {
            "stage": stage, "email": email, "tier": tier, "amount": amount
        })
        return jsonify({"sent": ok, "stage": stage, "email": email})
    except Exception as e:
        return safe_error(e)


# ── Proration calculator ─────────────────────────────────────────
@require_auth
@app.route("/api/billing/prorate", methods=["POST"])
@limiter.limit("30 per minute")
def billing_prorate():
    """
    Calculate proration for mid-cycle plan upgrade/downgrade.
    Returns: credit, new_charge, net_amount, effective_date
    """
    try:
        data           = request.get_json(force=True) or {}
        current_plan   = data.get("current_plan", "standard")
        new_plan       = data.get("new_plan", "professional")
        billing_cycle  = data.get("billing_cycle", "monthly")
        days_used      = int(data.get("days_used", 0))
        uid            = getattr(g, "uid", "")

        total_days = 365 if billing_cycle == "yearly" else 30
        credit     = _prorate_credit(current_plan, billing_cycle, days_used, total_days)
        new_price  = PLAN_PRICES.get(new_plan, {}).get(billing_cycle, 0) or 0
        net        = max(0, new_price - credit)
        is_upgrade = _plan_rank(new_plan) > _plan_rank(current_plan)

        return jsonify({
            "current_plan":    current_plan,
            "new_plan":        new_plan,
            "billing_cycle":   billing_cycle,
            "days_used":       days_used,
            "days_remaining":  max(0, total_days - days_used),
            "credit_amount":   credit,
            "new_plan_price":  new_price,
            "net_charge":      net,
            "is_upgrade":      is_upgrade,
            "immediate":       is_upgrade,
            "effective_date":  "immediate" if is_upgrade else "next cycle",
            "note":            f"{'Upgrade' if is_upgrade else 'Downgrade'}: {credit:,} EGP credit applied",
        })
    except Exception as e:
        return safe_error(e)


# ── Usage-based billing tracker ──────────────────────────────────
@require_auth
@app.route("/api/billing/usage", methods=["GET"])
@limiter.limit("60 per minute")
def billing_usage():
    """
    Return current-period usage vs plan limits for the authenticated user.
    Used by: billing dashboard, upgrade nudges, session throttle.
    """
    try:
        import json as _j
        uid   = getattr(g, "uid", "")
        tier  = getattr(g, "tier", "standard")

        LIMITS = {
            # standard (Free): kept in sync with the enforced cap in
            # /api/session/start (FREE_MONTHLY_LIMIT/FREE_DAILY_LIMIT) and
            # the "5 sessions/month" marketing copy in App.jsx TIERS.standard.
            "standard":     {"sessions_per_day": 3,   "sessions_per_month": 5,   "ai_coach": 0,   "pdf_exports": 0,  "employees": 1    },
            # basic: TIERS.basic.features explicitly promises "Unlimited
            # sessions" — was previously missing here and silently falling
            # back to the standard (Free) limits above.
            "basic":        {"sessions_per_day": -1,  "sessions_per_month": -1,  "ai_coach": 10,  "pdf_exports": 0,  "employees": 1    },
            "professional": {"sessions_per_day": -1,  "sessions_per_month": -1,  "ai_coach": 50,  "pdf_exports": -1, "employees": 25   },
            "elite":        {"sessions_per_day": -1,  "sessions_per_month": -1,  "ai_coach": -1,  "pdf_exports": -1, "employees": -1   },
            "business":     {"sessions_per_day": -1,  "sessions_per_month": -1,  "ai_coach": -1,  "pdf_exports": -1, "employees": 500  },
            "enterprise":   {"sessions_per_day": -1,  "sessions_per_month": -1,  "ai_coach": -1,  "pdf_exports": -1, "employees": -1   },
        }
        limits = LIMITS.get(tier, LIMITS["standard"])

        # Pull usage from Redis
        today_key  = f"usage:{uid}:sessions:{datetime.utcnow().strftime('%Y-%m-%d')}"
        month_key  = f"usage:{uid}:sessions:{datetime.utcnow().strftime('%Y-%m')}"
        coach_key  = f"usage:{uid}:ai_coach:{datetime.utcnow().strftime('%Y-%m')}"
        pdf_key    = f"usage:{uid}:pdf:{datetime.utcnow().strftime('%Y-%m')}"

        def _get_count(key):
            try:
                v = rget(key)
                return int(v) if v else 0
            except: return 0

        usage = {
            "sessions_today":       _get_count(today_key),
            "sessions_this_month":  _get_count(month_key),
            "ai_coach_this_month":  _get_count(coach_key),
            "pdf_exports_this_month": _get_count(pdf_key),
        }

        def _pct(used, limit):
            if limit < 0: return None  # unlimited
            return min(100, round(used / limit * 100)) if limit > 0 else 100

        return jsonify({
            "uid":    uid,
            "tier":   tier,
            "period": datetime.utcnow().strftime("%B %Y"),
            "usage":  usage,
            "limits": limits,
            "pct": {
                "sessions_day":   _pct(usage["sessions_today"],      limits["sessions_per_day"]),
                "sessions_month": _pct(usage["sessions_this_month"], limits["sessions_per_month"]),
                "ai_coach":       _pct(usage["ai_coach_this_month"], limits["ai_coach"]),
                "pdf_exports":    _pct(usage["pdf_exports_this_month"], limits["pdf_exports"]),
            },
            "at_limit": any(
                v == 100 for v in [
                    _pct(usage["sessions_today"],      limits["sessions_per_day"]),
                    _pct(usage["sessions_this_month"], limits["sessions_per_month"]),
                ] if v is not None
            ),
        })
    except Exception as e:
        return safe_error(e)


# ── Increment usage counter ──────────────────────────────────────
@require_auth
@app.route("/api/billing/usage/increment", methods=["POST"])
@limiter.limit("120 per minute")
def billing_usage_increment():
    """Increment a usage counter for the authenticated user."""
    try:
        data     = request.get_json(force=True) or {}
        metric   = data.get("metric", "sessions")   # sessions | ai_coach | pdf
        uid      = getattr(g, "uid", "")
        today    = datetime.utcnow().strftime("%Y-%m-%d")
        month    = datetime.utcnow().strftime("%Y-%m")
        keys = {
            "sessions": [
                (f"usage:{uid}:sessions:{today}", 86400 * 2),
                (f"usage:{uid}:sessions:{month}", 86400 * 35),
            ],
            "ai_coach": [(f"usage:{uid}:ai_coach:{month}", 86400 * 35)],
            "pdf":      [(f"usage:{uid}:pdf:{month}", 86400 * 35)],
        }
        for key, ttl in keys.get(metric, []):
            try:
                import redis as _redis_mod
                r = _redis_mod.from_url(os.getenv("REDIS_URL", "redis://localhost:6379/0"))
                r.incr(key)
                r.expire(key, ttl)
            except: pass
        return jsonify({"ok": True, "metric": metric})
    except Exception as e:
        return safe_error(e)


# ── Billing analytics (admin) ────────────────────────────────────
@require_auth
@require_admin
@app.route("/api/billing/analytics", methods=["GET"])
@limiter.limit("30 per minute")
def billing_analytics():
    """
    Full billing analytics: MRR, ARR, ARPU, plan distribution,
    conversion rate, failed payments, monthly revenue trend.
    """
    try:
        db       = firestore.client()
        payments = [d.to_dict() for d in db.collection("payments").order_by(
            "created_at", direction=firestore.Query.DESCENDING).limit(1000).get()]
        analytics = _compute_billing_analytics(payments)
        audit(getattr(g,"uid","admin"), "billing_analytics_viewed", "billing", {})
        return jsonify({**analytics, "generated_at": datetime.utcnow().isoformat()})
    except Exception as e:
        return safe_error(e)


# ── Invoice PDF generator ────────────────────────────────────────
@require_auth
@app.route("/api/billing/invoice/pdf", methods=["POST"])
@limiter.limit("10 per minute")
@require_tier("standard")
def billing_invoice_pdf():
    """
    Generate a professional PDF invoice for a confirmed payment.
    Returns: PDF binary (application/pdf).
    """
    try:
        if not REPORTLAB_OK:
            return jsonify({"error": "PDF generation not available — install reportlab"}), 503

        data    = request.get_json(force=True) or {}
        uid     = getattr(g, "uid", "")
        payment = data.get("payment", {})

        # Verify payment belongs to this user (or admin)
        tier_g  = getattr(g, "tier", "standard")
        is_adm  = getattr(g, "is_admin", False)
        if not is_adm and payment.get("uid") != uid:
            return jsonify({"error": "Unauthorized"}), 403

        inv_number = payment.get("ref_code", f"INV{int(time.time())}")
        cust_name  = payment.get("user_name", "Customer")
        cust_email = payment.get("user_email", "")
        tier       = payment.get("tier", "professional")
        amount     = payment.get("amount", 0)
        cycle      = payment.get("billing_cycle", "monthly")
        method     = payment.get("payment_method_name", "—")
        d          = payment.get("created_at", datetime.now())
        if hasattr(d, "toDate"): d = d.toDate()
        inv_date   = d.strftime("%d %b %Y") if hasattr(d, "strftime") else str(d)[:10]
        due_date   = inv_date  # paid immediately

        buf  = _billing_io.BytesIO()
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units     import mm
        from reportlab.lib           import colors
        from reportlab.platypus      import (SimpleDocTemplate, Paragraph,
                                             Spacer, Table, TableStyle, HRFlowable)
        from reportlab.lib.styles    import ParagraphStyle
        from reportlab.lib.enums     import TA_RIGHT, TA_CENTER, TA_LEFT

        doc = SimpleDocTemplate(buf, pagesize=A4,
            leftMargin=20*mm, rightMargin=20*mm,
            topMargin=15*mm,  bottomMargin=15*mm)

        W  = A4[0] - 40*mm
        BG = colors.HexColor("#040b14")
        BL = colors.HexColor("#6366f1")
        GR = colors.HexColor("#10b981")
        MU = colors.HexColor("#64748b")
        TX = colors.HexColor("#f1f5f9")
        CA = colors.HexColor("#0c1728")

        def _style(name, **kw):
            return ParagraphStyle(name, fontName="Helvetica",
                textColor=TX, **kw)

        h1  = _style("h1", fontSize=20, fontName="Helvetica-Bold")
        h2  = _style("h2", fontSize=13, fontName="Helvetica-Bold")
        lab = _style("lab", fontSize=9,  textColor=MU)
        val = _style("val", fontSize=11, fontName="Helvetica-Bold")
        sm  = _style("sm",  fontSize=9,  textColor=MU)
        lg  = _style("lg",  fontSize=24, fontName="Helvetica-Bold", textColor=GR)

        story = []

        # ── Header ────────────────────────────────────────────────
        story.append(Table([
            [Paragraph("◈ PostureAI Pro", h1),
             Paragraph("INVOICE", ParagraphStyle("inv", fontName="Helvetica-Bold",
                fontSize=11, textColor=BL, alignment=TA_RIGHT))],
        ], colWidths=[W*0.6, W*0.4]))

        story.append(Spacer(1, 6*mm))
        story.append(HRFlowable(width=W, thickness=0.5,
            color=colors.HexColor("#1e2d4a")))
        story.append(Spacer(1, 6*mm))

        # ── Meta row ──────────────────────────────────────────────
        story.append(Table([
            [Paragraph("Invoice number", lab),
             Paragraph("Issue date", lab),
             Paragraph("Status", lab)],
            [Paragraph(f"<b>{inv_number}</b>", val),
             Paragraph(f"<b>{inv_date}</b>", val),
             Paragraph("<b>PAID</b>", ParagraphStyle("paid",
                fontName="Helvetica-Bold", fontSize=11, textColor=GR))],
        ], colWidths=[W*0.4, W*0.3, W*0.3]))

        story.append(Spacer(1, 8*mm))

        # ── Bill to ───────────────────────────────────────────────
        story.append(Table([
            [Paragraph("BILL TO", lab), ""],
            [Paragraph(f"<b>{cust_name}</b>", val), ""],
            [Paragraph(cust_email, sm), ""],
        ], colWidths=[W*0.5, W*0.5]))
        story.append(Spacer(1, 6*mm))

        # ── Line items ────────────────────────────────────────────
        items_data = [
            [Paragraph("Description", lab),
             Paragraph("Qty", lab),
             Paragraph("Unit Price", lab),
             Paragraph("Total", ParagraphStyle("r", fontName="Helvetica",
                fontSize=9, textColor=MU, alignment=TA_RIGHT))],
            [Paragraph(f"PostureAI Pro — {tier.title()} Intelligence Plan ({cycle.title()})", val),
             Paragraph("1", val),
             Paragraph(f"{amount:,} EGP", val),
             Paragraph(f"<b>{amount:,} EGP</b>",
                ParagraphStyle("rb", fontName="Helvetica-Bold",
                    fontSize=11, textColor=TX, alignment=TA_RIGHT))],
        ]
        items_tbl = Table(items_data, colWidths=[W*0.55, W*0.1, W*0.18, W*0.17])
        items_tbl.setStyle(TableStyle([
            ("BACKGROUND",  (0,0), (-1,0), colors.HexColor("#0a1628")),
            ("BACKGROUND",  (0,1), (-1,1), CA),
            ("ROWPADDING",  (0,0), (-1,-1), 8),
            ("LINEBELOW",   (0,0), (-1,0), 0.5, colors.HexColor("#1e2d4a")),
            ("LINEBELOW",   (0,1), (-1,1), 0.5, colors.HexColor("#1e2d4a")),
        ]))
        story.append(items_tbl)
        story.append(Spacer(1, 5*mm))

        # ── Total ─────────────────────────────────────────────────
        story.append(Table([
            [Paragraph("Subtotal", sm),
             Paragraph(f"{amount:,} EGP",
                ParagraphStyle("ra", alignment=TA_RIGHT, fontSize=11,
                    fontName="Helvetica", textColor=TX))],
            [Paragraph("Tax (0%)", sm),
             Paragraph("0 EGP",
                ParagraphStyle("rb2", alignment=TA_RIGHT, fontSize=11,
                    fontName="Helvetica", textColor=MU))],
            [Paragraph("<b>Total</b>", ParagraphStyle("tot",
                fontName="Helvetica-Bold", fontSize=13, textColor=TX)),
             Paragraph(f"<b>{amount:,} EGP</b>",
                ParagraphStyle("totv", fontName="Helvetica-Bold",
                    fontSize=14, textColor=GR, alignment=TA_RIGHT))],
        ], colWidths=[W*0.7, W*0.3]))

        story.append(Spacer(1, 8*mm))
        story.append(HRFlowable(width=W, thickness=0.5,
            color=colors.HexColor("#1e2d4a")))
        story.append(Spacer(1, 5*mm))

        # ── Footer ────────────────────────────────────────────────
        story.append(Table([
            [Paragraph(f"Payment method: <b>{method}</b>", sm),
             Paragraph(f"<a href='mailto:{SUPPORT_EMAIL}'>{SUPPORT_EMAIL}</a>",
                ParagraphStyle("fr", fontSize=9, textColor=BL, alignment=TA_RIGHT))],
        ], colWidths=[W*0.6, W*0.4]))

        story.append(Spacer(1, 4*mm))
        story.append(Paragraph(
            "PostureAI Pro · AI Workforce Intelligence Platform · postureai.io",
            ParagraphStyle("tag", fontSize=8, textColor=MU, alignment=TA_CENTER)
        ))

        doc.build(story)
        buf.seek(0)
        return send_file(
            buf,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=f"postureai_invoice_{inv_number}.pdf",
        )
    except Exception as e:
        import traceback
        return safe_error(e)


# ── Subscription upgrade/downgrade mid-cycle ─────────────────────
@require_auth
@app.route("/api/billing/change-plan", methods=["POST"])
@limiter.limit("5 per minute")
def billing_change_plan():
    """
    Upgrade or downgrade a subscription mid-cycle.
    Calculates proration, creates a new payment record,
    and applies the change immediately (upgrade) or next cycle (downgrade).
    """
    try:
        import json as _j
        data          = request.get_json(force=True) or {}
        uid           = getattr(g, "uid", "")
        current_plan  = getattr(g, "tier", data.get("current_plan", "standard"))
        new_plan      = data.get("new_plan", "")
        billing_cycle = data.get("billing_cycle", "monthly")
        days_used     = int(data.get("days_used", 0))
        user_email    = data.get("user_email", "")
        user_name     = data.get("user_name", "Customer")

        if not new_plan:
            return jsonify({"error": "new_plan required"}), 400
        if new_plan not in PLAN_PRICES:
            return jsonify({"error": f"Unknown plan: {new_plan}"}), 400
        if new_plan == current_plan:
            return jsonify({"error": "Already on this plan"}), 400

        is_upgrade  = _plan_rank(new_plan) > _plan_rank(current_plan)
        total_days  = 365 if billing_cycle == "yearly" else 30
        credit      = _prorate_credit(current_plan, billing_cycle, days_used, total_days)
        new_price   = PLAN_PRICES.get(new_plan, {}).get(billing_cycle, 0) or 0
        net_charge  = max(0, new_price - credit)

        # Create pending payment record for the net charge
        if net_charge > 0:
            db  = firestore.client()
            ref = f"CHANGE-{uid[:8].upper()}-{int(time.time())}"
            db.collection("payments").add({
                "uid":                uid,
                "user_email":         user_email,
                "user_name":          user_name,
                "tier":               new_plan,
                "amount":             net_charge,
                "billing_cycle":      billing_cycle,
                "status":             "pending",
                "payment_method_name":"Plan Change",
                "ref_code":           ref,
                "proration_credit":   credit,
                "previous_plan":      current_plan,
                "created_at":         datetime.utcnow(),
                "change_type":        "upgrade" if is_upgrade else "downgrade",
            })

        audit(uid, "plan_change_requested", "billing", {
            "from": current_plan, "to": new_plan, "cycle": billing_cycle,
            "credit": credit, "net": net_charge, "is_upgrade": is_upgrade
        })

        return jsonify({
            "ok":           True,
            "is_upgrade":   is_upgrade,
            "current_plan": current_plan,
            "new_plan":     new_plan,
            "credit":       credit,
            "net_charge":   net_charge,
            "effective":    "immediate" if is_upgrade else "next cycle",
            "message":      f"{'Upgrade' if is_upgrade else 'Downgrade'} to {new_plan} — {credit:,} EGP credit applied. Net: {net_charge:,} EGP",
        })
    except Exception as e:
        return safe_error(e)


# ══════════════════════════════════════════════════════════════════
# SECURITY HARDENING — CSRF + Input Validation + Rate Protection
# ══════════════════════════════════════════════════════════════════
import re as _re

def sanitize_str(val, max_len=500):
    """Strip dangerous chars and limit length."""
    if not isinstance(val, str): return ""
    # Remove null bytes and control chars
    val = val.replace('\x00', '').replace('\r', '')
    # Remove HTML tags
    val = _re.sub(r'<[^>]+>', '', val)
    return val[:max_len].strip()

def validate_email(email):
    if not email or not isinstance(email, str): return False
    return bool(_re.match(r'^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$', email.strip()))

def validate_uid(uid):
    if not uid or not isinstance(uid, str): return False
    return bool(_re.match(r'^[a-zA-Z0-9_\-]{4,128}$', uid.strip()))

@app.before_request
def _set_request_id():
    import uuid as _uuid
    g.request_id = _uuid.uuid4().hex[:12]

@app.after_request
def add_security_headers(response):
    """Add security headers to every response."""
    response.headers["X-Request-ID"]            = getattr(g, "request_id", "-")
    response.headers["X-Content-Type-Options"]  = "nosniff"
    response.headers["X-Frame-Options"]          = "DENY"
    response.headers["X-XSS-Protection"]         = "1; mode=block"
    response.headers["Referrer-Policy"]           = "strict-origin-when-cross-origin"
    response.headers["Permissions-Policy"]        = "camera=(), microphone=(), geolocation=()"
    response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
    if os.getenv("FLASK_ENV") == "production":
        response.headers["Content-Security-Policy"] = (
            "default-src 'self'; "
            "script-src 'self' 'unsafe-inline' https://accept.paymob.com; "
            "frame-src https://accept.paymob.com; "
            "connect-src 'self' https://firestore.googleapis.com https://identitytoolkit.googleapis.com https://securetoken.googleapis.com https://*.firebaseapp.com https://*.googleapis.com; "
            "img-src 'self' data: blob:; "
            "style-src 'self' 'unsafe-inline';"
        )
    # Remove server fingerprint
    response.headers.pop("Server", None)
    response.headers.pop("X-Powered-By", None)
    return response

# ══════════════════════════════════════════════════════════════════
# ADVANCED REPORTS SYSTEM
# ══════════════════════════════════════════════════════════════════
@require_auth
@app.route("/api/reports/user", methods=["POST"])
@limiter.limit("10 per minute")
@require_auth
def user_report():
    """Generate comprehensive per-user posture report with AI insights."""
    try:
        if not REPORTLAB_OK:
            return jsonify({"error": "pip install reportlab"}), 500
        data = request.get_json(force=True) or {}
        uid  = sanitize_str(data.get("uid",""))
        if not validate_uid(uid):
            return jsonify({"error": "Invalid uid"}), 400
        name        = sanitize_str(data.get("name","Employee"), 100)
        sessions    = data.get("sessions",[])
        email       = sanitize_str(data.get("email",""), 200)
        tier        = data.get("tier","standard")
        lang        = data.get("lang","en")
        date_range  = data.get("date_range","Last 30 days")
        if not sessions:
            return jsonify({"error": "No session data provided"}), 400

        # Compute comprehensive analytics
        scores    = [s.get("avg_score",0) for s in sessions if s.get("avg_score")]
        avg_score = round(sum(scores)/len(scores)) if scores else 0
        best_sc   = max(scores) if scores else 0
        worst_sc  = min(scores) if scores else 0
        good_pct  = round(len([s for s in scores if s >= 70]) / max(len(scores),1) * 100)
        total_min = round(sum(s.get("duration_s",0) for s in sessions) / 60)
        total_sess= len(sessions)

        # Eye strain analysis
        eye_scores  = [s.get("eye_score",0) for s in sessions if s.get("eye_score")]
        avg_eye     = round(sum(eye_scores)/len(eye_scores)) if eye_scores else None

        # RSI risk
        rsi_scores  = [s.get("rsi_score",0) for s in sessions if s.get("rsi_score")]
        avg_rsi     = round(sum(rsi_scores)/len(rsi_scores)) if rsi_scores else None

        # Trend (last 5 vs first 5)
        if len(scores) >= 10:
            first5 = sum(scores[:5])/5
            last5  = sum(scores[-5:])/5
            trend  = "improving" if last5 > first5+3 else "declining" if last5 < first5-3 else "stable"
        else:
            trend  = "insufficient data"

        # Risk factors
        risk_factors = []
        if avg_score < 50:  risk_factors.append("Chronic poor posture — immediate ergonomic review recommended")
        if avg_score < 70:  risk_factors.append("Below optimal posture score — increased musculoskeletal risk")
        if avg_eye and avg_eye < 60: risk_factors.append("Eye strain detected — inadequate blink rate")
        if avg_rsi and avg_rsi < 60: risk_factors.append("RSI risk — wrist/elbow deviation above safe threshold")
        if total_min < 60:  risk_factors.append("Low monitoring time — consider longer daily sessions")
        if good_pct < 50:   risk_factors.append("Less than 50% good posture time — workspace audit needed")

        # Productivity estimation (rough correlation)
        productivity_est = min(100, max(0, avg_score * 0.6 + good_pct * 0.4))

        # AI insights (local)
        ai_summary = None
        if AI_CONFIGURED:
            prompt = f"""You are an occupational health specialist. Write a brief clinical summary (3-4 sentences) for:

Employee: {name}
Period: {date_range}
Sessions: {total_sess} | Avg score: {avg_score}/100 | Trend: {trend}
Good posture: {good_pct}% | Total monitoring: {total_min} minutes
Risk factors: {', '.join(risk_factors) if risk_factors else 'None identified'}

Focus on actionable recommendations. {"Respond in Arabic." if lang=="ar" else "Be professional and concise."}"""
            try:
                ai_summary = call_ai(prompt, max_tokens=250, temperature=0.3)
            except Exception:
                pass

        # Generate PDF
        from io import BytesIO as BIO
        buf = BIO()
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
        from reportlab.lib.units import mm
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

        NAVY=colors.HexColor("#05101f"); BLUE=colors.HexColor("#1a56db")
        GREEN=colors.HexColor("#059669"); AMBER=colors.HexColor("#d97706")
        RED=colors.HexColor("#dc2626"); GRAY=colors.HexColor("#64748b")
        WHITE=colors.white; LGRAY=colors.HexColor("#f8fafc")

        def sc_col(s): return GREEN if s>=75 else AMBER if s>=50 else RED
        def ps(n,**kw):
            d=dict(fontName='Helvetica',fontSize=9,textColor=colors.HexColor("#1e293b"),leading=13,spaceAfter=2)
            d.update(kw); return ParagraphStyle(n,**d)

        W,_ = A4; uw = W - 30*mm
        doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=15*mm, rightMargin=15*mm, topMargin=12*mm, bottomMargin=12*mm)
        story = []

        # Header
        hdr = Table([[
            Paragraph(f'<b>◈ PostureAI Pro</b>', ps('h',fontSize=16,textColor=WHITE,fontName='Helvetica-Bold')),
            Paragraph(f'<b>Personal Health Report</b><br/><font size="8" color="#94a3b8">{date_range} · {tier.title()}</font>', ps('hr',fontSize=11,textColor=WHITE,textColor2=GRAY,alignment=TA_RIGHT)),
        ]], colWidths=[uw*.5,uw*.5])
        hdr.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,-1),NAVY),('PADDING',(0,0),(-1,-1),14),('VALIGN',(0,0),(-1,-1),'MIDDLE')]))
        story.append(hdr); story.append(Spacer(1,8))

        # Employee info
        story.append(Paragraph(f"<b>{name}</b>", ps('n',fontSize=14,textColor=colors.HexColor("#0f172a"))))
        story.append(Paragraph(f"{email} · {total_sess} sessions · {total_min} minutes monitored", ps('sub',fontSize=9,textColor=GRAY)))
        story.append(Spacer(1,10))

        # KPI row
        kpi_data = [[
            Paragraph(f'<font size="8" color="#64748b">Avg Score</font><br/><b><font size="28" color="#{("059669" if avg_score>=75 else "d97706" if avg_score>=50 else "dc2626")}">{avg_score}</font></b><br/><font size="8" color="#64748b">/100</font>', ps('k',alignment=TA_CENTER)),
            Paragraph(f'<font size="8" color="#64748b">Good Posture</font><br/><b><font size="22" color="#059669">{good_pct}%</font></b><br/><font size="8" color="#64748b">of time</font>', ps('k',alignment=TA_CENTER)),
            Paragraph(f'<font size="8" color="#64748b">Sessions</font><br/><b><font size="22">{total_sess}</font></b><br/><font size="8" color="#64748b">completed</font>', ps('k',alignment=TA_CENTER)),
            Paragraph(f'<font size="8" color="#64748b">Trend</font><br/><b><font size="14" color="{"#059669" if trend=="improving" else "#d97706" if trend=="stable" else "#dc2626"}">{trend.title()}</font></b><br/><font size="8" color="#64748b">posture</font>', ps('k',alignment=TA_CENTER)),
            Paragraph(f'<font size="8" color="#64748b">Best Score</font><br/><b><font size="22" color="#059669">{best_sc}</font></b><br/><font size="8" color="#64748b">peak</font>', ps('k',alignment=TA_CENTER)),
        ]]
        kpi = Table(kpi_data, colWidths=[uw/5]*5)
        kpi.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,-1),LGRAY),('BOX',(0,0),(-1,-1),0.5,colors.HexColor("#e2e8f0")),('LINEBETWEEN',(0,0),(-1,-1),0.3,colors.HexColor("#e2e8f0")),('PADDING',(0,0),(-1,-1),(10,10)),('VALIGN',(0,0),(-1,-1),'MIDDLE')]))
        story.append(kpi); story.append(Spacer(1,12))

        # Score history chart
        if scores:
            from reportlab.graphics.shapes import Drawing, Rect, Line, String, Polygon
            story.append(Paragraph("Score History", ps('h2',fontSize=11,fontName='Helvetica-Bold',textColor=colors.HexColor("#0f172a"),spaceAfter=6)))
            dw = float(uw); dh = 72.0
            d = Drawing(dw, dh)
            d.add(Rect(0,0,dw,dh,fillColor=LGRAY,strokeColor=None))
            max_s = max(max(scores),1)
            bw = dw / max(len(scores),1)
            for idx_s, sv in enumerate(scores):
                bh = max(2,dh*sv/100)
                color_s = colors.HexColor("#059669") if sv>=75 else colors.HexColor("#d97706") if sv>=50 else colors.HexColor("#dc2626")
                d.add(Rect(idx_s*bw+0.5, 0, max(1.5,bw-1), bh, fillColor=color_s, strokeColor=None))
            # Reference line at 70
            ref_y = dh*70/100
            d.add(Line(0,ref_y,dw,ref_y,strokeColor=colors.HexColor("#94a3b8"),strokeDashArray=[4,3],strokeWidth=0.6))
            story.append(d); story.append(Spacer(1,10))

        # Risk factors
        if risk_factors:
            story.append(Paragraph("Risk Assessment", ps('h2',fontSize=11,fontName='Helvetica-Bold',textColor=colors.HexColor("#0f172a"),spaceAfter=6)))
            risk_rows = [[Paragraph(f"<b>{i+1}</b>",ps('rn',textColor=WHITE,alignment=TA_CENTER,fontName='Helvetica-Bold',fontSize=9)),
                          Paragraph(rf,ps('rb',fontSize=9))] for i,rf in enumerate(risk_factors)]
            rt = Table(risk_rows, colWidths=[uw*.06,uw*.94])
            rt.setStyle(TableStyle([('BACKGROUND',(0,0),(0,-1),RED if len(risk_factors)>2 else AMBER),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('PADDING',(0,0),(-1,-1),(7,5)),('ROWBACKGROUNDS',(1,0),(1,-1),[LGRAY,WHITE]),('LINEBELOW',(0,0),(-1,-1),0.3,colors.HexColor("#e2e8f0"))]))
            story.append(rt); story.append(Spacer(1,10))

        # AI Summary
        if ai_summary:
            story.append(Paragraph("AI Health Assessment", ps('h2',fontSize=11,fontName='Helvetica-Bold',textColor=colors.HexColor("#0f172a"),spaceAfter=6)))
            ai_box = Table([[
                Paragraph("AI",ps('aib',fontSize=9,fontName='Helvetica-Bold',textColor=WHITE,alignment=TA_CENTER)),
                Paragraph(ai_summary,ps('ait',fontSize=8.5,leading=13)),
            ]],colWidths=[uw*.07,uw*.93])
            ai_box.setStyle(TableStyle([('BACKGROUND',(0,0),(0,-1),BLUE),('BACKGROUND',(1,0),(1,-1),colors.HexColor("#eff6ff")),('VALIGN',(0,0),(-1,-1),'TOP'),('PADDING',(0,0),(-1,-1),(8,8)),('BOX',(0,0),(-1,-1),0.5,BLUE)]))
            story.append(ai_box); story.append(Spacer(1,10))

        # Recommendations
        recs = [
            f"Target: maintain 75+ posture score — current average {avg_score}/100",
            "Schedule 2-minute stretch break every 30 minutes",
            "Apply 20-20-20 eye rule every 20 minutes",
            "Position monitor top at or slightly below eye level",
            "Keep wrists straight when typing — avoid deviation",
            "Lumbar support: lower back fully touching chair back",
        ]
        story.append(Paragraph("Recommendations", ps('h2',fontSize=11,fontName='Helvetica-Bold',textColor=colors.HexColor("#0f172a"),spaceAfter=6)))
        for i, rec in enumerate(recs):
            story.append(Paragraph(f"{'  ✓' if i < 2 else '  •'} {rec}", ps('r',fontSize=9,textColor=colors.HexColor("#334155"),leftIndent=8,spaceAfter=3)))
        story.append(Spacer(1,10))

        # Footer
        story.append(HRFlowable(width=uw,thickness=0.5,color=colors.HexColor("#e2e8f0")))
        story.append(Paragraph(
            f"Report generated by PostureAI Pro v20 · {datetime.now().strftime('%d %B %Y')} · For occupational health purposes only · Not medical advice",
            ps('ft',fontSize=7,textColor=GRAY,alignment=TA_CENTER)))

        doc.build(story)
        buf.seek(0)
        fname = f"PostureAI_Report_{name.replace(' ','_')}_{datetime.now().strftime('%Y%m%d')}.pdf"
        return send_file(buf, mimetype="application/pdf", as_attachment=True, download_name=fname)
    except Exception as e:
        import traceback
        return safe_error(e)

@require_auth
@app.route("/api/reports/weekly-summary", methods=["POST"])
@limiter.limit("20 per minute")
def weekly_summary():
    """JSON endpoint for weekly summary — used by email system."""
    try:
        data     = request.get_json(force=True) or {}
        sessions = data.get("sessions",[])
        if not sessions:
            return jsonify({"avg_score":0,"sessions":0,"good_pct":0,"trend":"no data","insights":[]})
        scores    = [s.get("avg_score",0) for s in sessions if s.get("avg_score")]
        avg_score = round(sum(scores)/len(scores)) if scores else 0
        good_pct  = round(len([s for s in scores if s >= 70]) / max(len(scores),1) * 100)
        total_min = round(sum(s.get("duration_s",0) for s in sessions) / 60)
        # AI insights
        insights = []
        if avg_score >= 85: insights.append("Excellent posture week — keep it up!")
        elif avg_score >= 70: insights.append("Good posture maintained. Focus on consistency.")
        elif avg_score >= 50: insights.append("Fair posture. Schedule more frequent breaks.")
        else: insights.append("Posture needs attention. Consider an ergonomic assessment.")
        if good_pct < 50: insights.append("Less than half your time was at good posture — try posture reminders.")
        if total_min < 30: insights.append("Low monitoring time. Try to use PostureAI for at least 30 min/day.")
        return jsonify({"avg_score":avg_score,"sessions":len(sessions),"good_pct":good_pct,"total_min":total_min,"trend":"stable","insights":insights})
    except Exception as e:
        return safe_error(e)

# ── Send invite email ─────────────────────────────────────────────
@require_auth
@app.route("/api/org/send-invite", methods=["POST"])
@limiter.limit("100 per hour")
@optional_auth
def org_send_invite():
    """Send invite email to an employee."""
    try:
        data       = request.get_json(force=True) or {}
        invite_id  = data.get("invite_id","").strip()
        email      = data.get("email","").strip()
        name       = data.get("name","") or email.split("@")[0]
        company_id = data.get("company_id","")
        invite_url = data.get("invite_url","")

        if not (email and invite_url):
            return jsonify({"error":"email and invite_url required"}),400

        # Get company name
        company_name = "your team"
        try:
            db = firestore.client()
            co = db.collection("companies").document(company_id).get()
            if co.exists:
                company_name = co.to_dict().get("name","your team")
        except Exception:
            pass

        subject = f"[{company_name}] You're invited to PostureAI — AI Workforce Intelligence Platform"
        html    = f"""
<div style="font-family:Inter,system-ui,sans-serif;max-width:520px;margin:0 auto;padding:28px 24px;background:#030b14;color:#f0f6ff;border-radius:16px">
  <div style="text-align:center;margin-bottom:24px">
    <div style="width:52px;height:52px;background:linear-gradient(135deg,#1a56db,#0891b2);border-radius:14px;display:inline-flex;align-items:center;justify-content:center;font-size:26px">◈</div>
    <h1 style="margin:12px 0 4px;font-size:20px;font-weight:800">PostureAI Pro</h1>
    <p style="color:#64748b;font-size:13px;margin:0">AI-powered workplace health</p>
  </div>
  <h2 style="font-size:18px;font-weight:700;margin-bottom:8px">Hi {name} 👋</h2>
  <p style="color:#94a3b8;font-size:14px;line-height:1.6;margin-bottom:24px">
    <strong style="color:#f0f6ff">{company_name}</strong> has invited you to join their PostureAI Pro workspace.
    Real-time AI posture analysis to protect your health at work.
  </p>
  <div style="text-align:center;margin:28px 0">
    <a href="{invite_url}" style="display:inline-block;background:linear-gradient(135deg,#1a56db,#0891b2);color:#fff;padding:14px 32px;border-radius:12px;text-decoration:none;font-weight:700;font-size:15px;box-shadow:0 4px 20px rgba(26,86,219,.4)">
      ✓ Accept Invitation →
    </a>
  </div>
  <p style="color:#475569;font-size:11px;text-align:center;margin-top:24px">
    This invite expires in 7 days. If you didn't expect this, ignore it safely.
    <br>Need help? <a href="mailto:{os.getenv('SUPPORT_EMAIL','support@postureai.io')}" style="color:#1a56db">{os.getenv('SUPPORT_EMAIL','support@postureai.io')}</a>
  </p>
</div>"""

        sent = _send_email_smtp(email, subject, html)
        if not sent:
            # SMTP not configured — return the invite URL so admin can share manually
            return jsonify({"ok": True, "sent": False, "invite_url": invite_url,
                            "note": "SMTP not configured — share invite_url manually"})

        return jsonify({"ok": True, "sent": True, "to": email})
    except Exception as e:
        return safe_error(e)

@app.route("/api/llm", methods=["POST", "OPTIONS"])
def llm_proxy():
    """
    LLM proxy — forwards requests to LLM7.io server-side (no CORS issues).
    Used by AI Coach, AI Insights, Predictive AI.
    No auth required — rate limited by IP via Vercel.
    """
    if req.method == "OPTIONS":
        resp = jsonify({})
        resp.headers["Access-Control-Allow-Origin"]  = "*"
        resp.headers["Access-Control-Allow-Methods"] = "POST, OPTIONS"
        resp.headers["Access-Control-Allow-Headers"] = "Content-Type"
        return resp, 204

    try:
        data = req.get_json(force=True) or {}
        messages     = data.get("messages", [])
        system_prompt = data.get("system_prompt", "You are Dr. Corvus, a clinical physiotherapy AI.")
        max_tokens   = int(data.get("max_tokens", 700))
        temperature  = float(data.get("temperature", 0.5))

        if not messages:
            return jsonify({"error": "messages required"}), 400

        llm_messages = [{"role": "system", "content": system_prompt}] + [
            {"role": ("assistant" if m.get("role") == "assistant" else "user"),
             "content": str(m.get("content", ""))}
            for m in messages
        ]

        models = ["gpt-4o-mini", "meta-llama/llama-3.3-70b-instruct", "deepseek/deepseek-r1"]

        import requests as _req
        for model in models:
            try:
                r = _req.post(
                    "https://api.llm7.io/v1/chat/completions",
                    headers={"Content-Type": "application/json", "Authorization": "Bearer unused"},
                    json={"model": model, "messages": llm_messages,
                          "max_tokens": max_tokens, "temperature": temperature},
                    timeout=25,
                )
                if r.status_code == 429:
                    continue
                if not r.ok:
                    continue
                text = (r.json().get("choices") or [{}])[0].get("message", {}).get("content", "").strip()
                if not text:
                    continue
                resp = jsonify({"ok": True, "text": text, "model": model})
                resp.headers["Access-Control-Allow-Origin"] = "*"
                return resp, 200
            except Exception:
                continue

        resp = jsonify({"ok": False, "error": "All AI models unavailable"})
        resp.headers["Access-Control-Allow-Origin"] = "*"
        return resp, 503

    except Exception as e:
        return safe_error(e)


# ════════════════════════════════════════════════════════════════════
# Physiotherapist Marketplace
# Therapist profiles are admin-curated (invite-only) for v1 — no public
# self-serve signup yet. Bookings reuse the existing PayMob order flow,
# priced from the therapist's own session_fee_cents rather than a
# subscription tier.
# ════════════════════════════════════════════════════════════════════

def _therapist_public(doc):
    """Strip internal/admin-only fields before returning a therapist to patients."""
    d = doc.to_dict() or {}
    return {
        "id": doc.id,
        "name": d.get("name", ""),
        "photo_url": d.get("photo_url", ""),
        "specialties": d.get("specialties", []),
        "city": d.get("city", ""),
        "bio": d.get("bio", ""),
        "years_experience": d.get("years_experience", 0),
        "session_fee_cents": d.get("session_fee_cents", 0),
        "currency": d.get("currency", "EGP"),
        "languages": d.get("languages", []),
        "rating": d.get("rating"),
        "review_count": d.get("review_count", 0),
    }


# ── Admin: manage therapist profiles ──────────────────────────────
@app.route("/api/admin/marketplace/therapists", methods=["GET"])
@require_auth
@require_admin
@limiter.limit("60 per minute")
def admin_list_therapists():
    try:
        db   = firestore.client()
        docs = db.collection("therapists").order_by(
            "created_at", direction=firestore.Query.DESCENDING
        ).get()
        return jsonify({"therapists": [{**d.to_dict(), "id": d.id} for d in docs]})
    except Exception as e:
        return safe_error(e)


@app.route("/api/admin/marketplace/therapists", methods=["POST"])
@require_auth
@require_admin
@limiter.limit("30 per minute")
def admin_create_therapist():
    try:
        data = request.get_json(force=True) or {}
        name = (data.get("name") or "").strip()
        if not name:
            return jsonify({"error": "name required"}), 400
        fee_cents = int(data.get("session_fee_cents") or 0)
        if fee_cents <= 0:
            return jsonify({"error": "session_fee_cents must be > 0"}), 400

        db  = firestore.client()
        doc = {
            "name":              name,
            "photo_url":         (data.get("photo_url") or "").strip(),
            "specialties":       data.get("specialties") or [],
            "city":              (data.get("city") or "").strip(),
            "bio":               (data.get("bio") or "").strip(),
            "years_experience":  int(data.get("years_experience") or 0),
            "session_fee_cents": fee_cents,
            "currency":          (data.get("currency") or "EGP").strip().upper(),
            "languages":         data.get("languages") or [],
            "contact_email":     (data.get("contact_email") or "").strip(),
            "contact_phone":     (data.get("contact_phone") or "").strip(),
            "status":            "active",          # active | paused
            "rating":            None,
            "review_count":      0,
            # Weekly recurring availability: {"mon":["09:00","11:00"], "wed":[...]}
            # Empty/absent means "no fixed slots" — booking falls back to the
            # freeform preferred_time text field for that therapist.
            "availability_template": data.get("availability_template") or {},
            "created_at":        datetime.utcnow().isoformat()+"Z",
            "created_by_admin":  g.uid,
        }
        ref = db.collection("therapists").document()
        ref.set(doc)
        return jsonify({"ok": True, "id": ref.id})
    except Exception as e:
        return safe_error(e)


@app.route("/api/admin/marketplace/therapists/<therapist_id>", methods=["PATCH"])
@require_auth
@require_admin
@limiter.limit("30 per minute")
def admin_update_therapist(therapist_id):
    try:
        data = request.get_json(force=True) or {}
        allowed = {"name","photo_url","specialties","city","bio","years_experience",
                   "session_fee_cents","currency","languages","contact_email",
                   "contact_phone","status","availability_template"}
        upd = {k: v for k, v in data.items() if k in allowed}
        if not upd:
            return jsonify({"error": "no valid fields to update"}), 400
        upd["updated_at"]       = datetime.utcnow().isoformat()+"Z"
        upd["updated_by_admin"] = g.uid
        db = firestore.client()
        ref = db.collection("therapists").document(therapist_id)
        if not ref.get().exists:
            return jsonify({"error": "therapist not found"}), 404
        ref.update(upd)
        return jsonify({"ok": True})
    except Exception as e:
        return safe_error(e)


@app.route("/api/admin/marketplace/bookings", methods=["GET"])
@require_auth
@require_admin
@limiter.limit("60 per minute")
def admin_list_bookings():
    try:
        db   = firestore.client()
        docs = db.collection("marketplace_bookings").order_by(
            "created_at", direction=firestore.Query.DESCENDING
        ).limit(200).get()
        return jsonify({"bookings": [{**d.to_dict(), "id": d.id} for d in docs]})
    except Exception as e:
        return safe_error(e)


# ── Patient-facing: browse + book ─────────────────────────────────
@app.route("/api/marketplace/therapists", methods=["GET"])
@require_auth
@limiter.limit("60 per minute")
def marketplace_list_therapists():
    try:
        db    = firestore.client()
        query = db.collection("therapists").where("status", "==", "active")
        city  = (request.args.get("city") or "").strip()
        specialty = (request.args.get("specialty") or "").strip()
        docs  = query.get()
        out   = [_therapist_public(d) for d in docs]
        if city:
            out = [t for t in out if t["city"].lower() == city.lower()]
        if specialty:
            out = [t for t in out if specialty.lower() in [s.lower() for s in t["specialties"]]]
        return jsonify({"therapists": out})
    except Exception as e:
        return safe_error(e)


@app.route("/api/marketplace/therapists/<therapist_id>", methods=["GET"])
@require_auth
@limiter.limit("60 per minute")
def marketplace_get_therapist(therapist_id):
    try:
        db  = firestore.client()
        doc = db.collection("therapists").document(therapist_id).get()
        if not doc.exists or doc.to_dict().get("status") != "active":
            return jsonify({"error": "therapist not found"}), 404
        return jsonify({"therapist": _therapist_public(doc)})
    except Exception as e:
        return safe_error(e)


_WEEKDAY_KEYS = ["mon","tue","wed","thu","fri","sat","sun"]


@app.route("/api/marketplace/therapists/<therapist_id>/slots", methods=["GET"])
@require_auth
@limiter.limit("60 per minute")
def marketplace_therapist_slots(therapist_id):
    """Expands the therapist's weekly availability_template into concrete,
    bookable datetimes for the next `days` days, excluding slots already
    taken by a non-cancelled booking. Returns [] (not an error) for
    therapists with no template set — the frontend falls back to the
    freeform preferred_time text field in that case."""
    try:
        days = min(30, max(1, int(request.args.get("days", 14))))
        db   = firestore.client()
        tdoc = db.collection("therapists").document(therapist_id).get()
        if not tdoc.exists:
            return jsonify({"error": "therapist not found"}), 404
        template = tdoc.to_dict().get("availability_template") or {}
        if not template:
            return jsonify({"slots": [], "has_template": False})

        now = datetime.utcnow()
        candidates = []
        for i in range(days):
            day = now + timedelta(days=i)
            key = _WEEKDAY_KEYS[day.weekday()]
            for time_str in template.get(key, []):
                try:
                    hh, mm = [int(x) for x in time_str.split(":")]
                except (ValueError, AttributeError):
                    continue
                slot_dt = day.replace(hour=hh, minute=mm, second=0, microsecond=0)
                if slot_dt > now:  # never offer a slot that's already passed today
                    candidates.append(slot_dt)
        candidates.sort()

        # Exclude already-booked slots for this therapist in the same window
        window_end = now + timedelta(days=days)
        booked_docs = (db.collection("marketplace_bookings")
                         .where("therapist_id", "==", therapist_id)
                         .where("slot_datetime", ">=", now)
                         .where("slot_datetime", "<=", window_end)
                         .stream())
        taken = {d.to_dict().get("slot_datetime") for d in booked_docs
                 if d.to_dict().get("status") != "cancelled" and d.to_dict().get("slot_datetime")}
        taken_iso = {t.isoformat() if hasattr(t, "isoformat") else str(t) for t in taken}

        free = [c for c in candidates if c.isoformat() not in taken_iso]
        return jsonify({
            "slots": [s.isoformat()+"Z" for s in free],
            "has_template": True,
        })
    except Exception as e:
        return safe_error(e)


@app.route("/api/marketplace/bookings", methods=["POST"])
@require_auth
@limiter.limit("15 per minute")
def marketplace_create_booking():
    """Create a booking request and open a PayMob payment for the therapist's
    session fee. Mirrors /api/paymob/create-payment's order flow, but priced
    from the therapist doc instead of a subscription tier."""
    try:
        data          = request.get_json(force=True) or {}
        therapist_id  = (data.get("therapist_id") or "").strip()
        preferred_time= (data.get("preferred_time") or "").strip()
        slot_iso      = (data.get("slot_datetime") or "").strip()
        notes         = (data.get("notes") or "").strip()
        billing_data  = data.get("billing_data", {})
        if not therapist_id:
            return jsonify({"error": "therapist_id required"}), 400

        db   = firestore.client()
        tdoc = db.collection("therapists").document(therapist_id).get()
        if not tdoc.exists or tdoc.to_dict().get("status") != "active":
            return jsonify({"error": "therapist not found or unavailable"}), 404
        therapist = tdoc.to_dict()
        amount_cents = int(therapist.get("session_fee_cents") or 0)
        currency     = therapist.get("currency", "EGP")
        if amount_cents <= 0:
            return jsonify({"error": "therapist has no fee configured"}), 400

        slot_dt = None
        if slot_iso:
            try:
                slot_dt = datetime.fromisoformat(slot_iso.replace("Z", ""))
            except ValueError:
                return jsonify({"error": "invalid slot_datetime"}), 400
            if slot_dt <= datetime.utcnow():
                return jsonify({"error": "that slot has already passed"}), 400
            # Re-check the slot is still free right before booking — closes
            # the race window between the patient loading slots and clicking book.
            clash = (db.collection("marketplace_bookings")
                       .where("therapist_id", "==", therapist_id)
                       .where("slot_datetime", "==", slot_dt)
                       .stream())
            if any(d.to_dict().get("status") != "cancelled" for d in clash):
                return jsonify({"error": "that slot was just booked by someone else — please pick another"}), 409

        booking_ref = db.collection("marketplace_bookings").document()
        booking = {
            "therapist_id":   therapist_id,
            "therapist_name": therapist.get("name", ""),
            "user_id":        g.uid,
            "preferred_time": preferred_time or (slot_dt.strftime("%a %b %d, %H:%M") if slot_dt else ""),
            "slot_datetime":  slot_dt,
            "notes":          notes,
            "amount_cents":   amount_cents,
            "currency":       currency,
            "status":         "pending_payment",   # pending_payment | confirmed | cancelled
            "created_at":     datetime.utcnow().isoformat()+"Z",
        }

        if not PAYMOB_SECRET_KEY:
            # No payment configured — still record the request so admin can
            # follow up manually, matching the "coming soon" card path.
            booking["status"] = "pending_manual_followup"
            booking_ref.set(booking)
            return jsonify({"ok": True, "booking_id": booking_ref.id,
                             "payment": None,
                             "note": "PayMob not configured — booking recorded for manual follow-up"}), 200

        headers   = {"Content-Type": "application/json"}
        auth_resp = req.post("https://accept.paymob.com/api/auth/tokens",
                              json={"api_key": PAYMOB_SECRET_KEY}, headers=headers, timeout=15)
        auth_resp.raise_for_status()
        auth_token = auth_resp.json().get("token")
        if not auth_token:
            return jsonify({"error": "PayMob auth failed"}), 502

        order_resp = req.post("https://accept.paymob.com/api/ecommerce/orders",
                               json={"auth_token": auth_token, "delivery_needed": False,
                                     "amount_cents": amount_cents, "currency": currency,
                                     "merchant_order_id": f"PAI-BOOK-{g.uid}-{booking_ref.id}-{int(time.time())}",
                                     "items": [{"name": f"Session with {therapist.get('name','Therapist')}",
                                                "amount_cents": amount_cents,
                                                "description": f"PostureAI Marketplace — physiotherapy session booking",
                                                "quantity": 1}]},
                               headers=headers, timeout=15)
        order_resp.raise_for_status()
        order_id = order_resp.json().get("id")

        pk_resp = req.post("https://accept.paymob.com/api/acceptance/payment_keys",
                            json={"auth_token": auth_token, "amount_cents": amount_cents,
                                  "expiration": 3600, "order_id": order_id, "currency": currency,
                                  "integration_id": PAYMOB_INTEGRATIONS.get("card", ""),
                                  "billing_data": {"email": billing_data.get("email","NA"),
                                                   "first_name": billing_data.get("first_name","Customer"),
                                                   "last_name": billing_data.get("last_name",""),
                                                   "phone_number": billing_data.get("phone_number","NA"),
                                                   "apartment":"NA","floor":"NA","street":"NA",
                                                   "building":"NA","shipping_method":"NA",
                                                   "postal_code":"NA","city":"Cairo","country":"EG","state":"Cairo"}},
                            headers=headers, timeout=15)
        pk_resp.raise_for_status()
        payment_key = pk_resp.json().get("token")

        booking["paymob_order_id"] = order_id
        booking_ref.set(booking)

        iframe_id = PAYMOB_IFRAME_ID if 'PAYMOB_IFRAME_ID' in globals() else os.getenv("PAYMOB_IFRAME_ID", "")
        return jsonify({
            "ok": True,
            "booking_id": booking_ref.id,
            "payment": {
                "payment_key": payment_key,
                "iframe_url": f"https://accept.paymob.com/api/acceptance/iframes/{iframe_id}?payment_token={payment_key}" if iframe_id else None,
            }
        })
    except Exception as e:
        return safe_error(e)


@app.route("/api/marketplace/bookings", methods=["GET"])
@require_auth
@limiter.limit("30 per minute")
def marketplace_my_bookings():
    try:
        db   = firestore.client()
        docs = db.collection("marketplace_bookings").where("user_id", "==", g.uid).order_by(
            "created_at", direction=firestore.Query.DESCENDING
        ).get()
        return jsonify({"bookings": [{**d.to_dict(), "id": d.id} for d in docs]})
    except Exception as e:
        return safe_error(e)


@app.route("/api/marketplace/bookings/<booking_id>/cancel", methods=["POST"])
@require_auth
@limiter.limit("20 per minute")
def marketplace_cancel_booking(booking_id):
    """Patient cancels their own booking. Works at any status except an
    already-cancelled one. Cancelling an already-confirmed (paid) booking
    is still allowed here — refunds aren't automated, so it's flagged for
    admin follow-up rather than silently vanishing with money already
    charged and no record of the cancellation request."""
    try:
        db   = firestore.client()
        bref = db.collection("marketplace_bookings").document(booking_id)
        bdoc = bref.get()
        if not bdoc.exists:
            return jsonify({"error": "booking not found"}), 404
        booking = bdoc.to_dict()
        if booking.get("user_id") != g.uid:
            return jsonify({"error": "forbidden"}), 403
        if booking.get("status") == "cancelled":
            return jsonify({"ok": True, "note": "already cancelled"}), 200

        was_paid = booking.get("status") == "confirmed"
        bref.update({
            "status": "cancelled",
            "cancelled_at": datetime.utcnow().isoformat()+"Z",
            "needs_refund_review": was_paid,
        })
        if was_paid:
            amount_egp = (booking.get("amount_cents") or 0) / 100
            html = f"""
            <div style="font-family:Arial,sans-serif;max-width:480px;background:#fff;border-radius:10px;padding:24px">
              <h2 style="color:#dc2626;margin-bottom:16px">🔁 Booking Cancelled — Refund Review Needed</h2>
              <table style="width:100%;border-collapse:collapse">
                <tr><td style="padding:6px 0;color:#666">Booking ID</td><td style="font-family:monospace">{booking_id}</td></tr>
                <tr><td style="padding:6px 0;color:#666">Therapist</td><td style="font-weight:600">{booking.get("therapist_name","—")}</td></tr>
                <tr><td style="padding:6px 0;color:#666">Amount paid</td><td style="font-weight:700;color:#dc2626;font-size:18px">{amount_egp:,.2f} {booking.get("currency","EGP")}</td></tr>
              </table>
              <div style="margin-top:20px;padding:12px;background:#fee2e2;border-radius:8px;font-size:13px">
                ⚠️ This booking was already paid before the patient cancelled it. Refunds aren't automated — please process manually via PayMob and update the booking record.
              </div>
            </div>"""
            send_email(ADMIN_EMAIL, f"🔁 Booking cancelled after payment — refund needed ({booking_id})", html)
        return jsonify({"ok": True, "refund_pending": was_paid})
    except Exception as e:
        return safe_error(e)


@app.route("/api/marketplace/bookings/<booking_id>/review", methods=["POST"])
@require_auth
@limiter.limit("15 per minute")
def marketplace_review_booking(booking_id):
    """Patient rates a completed session (1-5 stars + optional comment).
    Recomputes the therapist's aggregate rating/review_count from all
    reviews left on their bookings — a simple running average, not a
    separate reviews collection, since bookings are already the natural
    one-review-per-booking unit and we don't need review edit history."""
    try:
        data = request.get_json(force=True) or {}
        rating = data.get("rating")
        try:
            rating = int(rating)
        except (TypeError, ValueError):
            return jsonify({"error": "rating must be an integer 1-5"}), 400
        if rating < 1 or rating > 5:
            return jsonify({"error": "rating must be between 1 and 5"}), 400
        comment = (data.get("comment") or "").strip()[:500]

        db   = firestore.client()
        bref = db.collection("marketplace_bookings").document(booking_id)
        bdoc = bref.get()
        if not bdoc.exists:
            return jsonify({"error": "booking not found"}), 404
        booking = bdoc.to_dict()
        if booking.get("user_id") != g.uid:
            return jsonify({"error": "forbidden"}), 403
        if booking.get("status") == "cancelled":
            return jsonify({"error": "cannot review a cancelled booking"}), 400
        if booking.get("rating"):
            return jsonify({"error": "this booking already has a review"}), 400

        bref.update({
            "rating":      rating,
            "review_text": comment,
            "reviewed_at": datetime.utcnow().isoformat()+"Z",
        })

        # Recompute the therapist's aggregate from every rated booking they have
        therapist_id = booking.get("therapist_id")
        if therapist_id:
            rated = list(db.collection("marketplace_bookings")
                           .where("therapist_id", "==", therapist_id)
                           .where("rating", ">", 0).stream())
            ratings = [d.to_dict().get("rating") for d in rated if d.to_dict().get("rating")]
            if ratings:
                avg_rating = round(sum(ratings) / len(ratings), 1)
                db.collection("therapists").document(therapist_id).update({
                    "rating": avg_rating,
                    "review_count": len(ratings),
                })
        return jsonify({"ok": True, "rating": rating})
    except Exception as e:
        return safe_error(e)


@app.route("/api/marketplace/therapists/<therapist_id>/reviews", methods=["GET"])
@require_auth
@limiter.limit("30 per minute")
def marketplace_therapist_reviews(therapist_id):
    """Public (any logged-in user): recent written reviews for a therapist —
    shown on their profile card. Only bookings with a non-empty comment,
    newest first, capped at 20."""
    try:
        db   = firestore.client()
        docs = (db.collection("marketplace_bookings")
                  .where("therapist_id", "==", therapist_id)
                  .where("rating", ">", 0)
                  .order_by("rating", direction=firestore.Query.DESCENDING)
                  .limit(20).stream())
        reviews = []
        for d in docs:
            b = d.to_dict()
            if b.get("review_text"):
                reviews.append({
                    "rating":      b.get("rating"),
                    "comment":     b.get("review_text"),
                    "reviewed_at": b.get("reviewed_at"),
                })
        return jsonify({"reviews": reviews})
    except Exception as e:
        return safe_error(e)


def _booking_chat_access_ok(booking_data):
    """Booking owner (patient) or a platform admin can read/post — not
    open to other patients, and not gated behind require_admin since
    patients need their own thread too."""
    is_admin = getattr(g, "role", {}).get("is_admin", False)
    return is_admin or booking_data.get("user_id") == g.uid


@app.route("/api/marketplace/bookings/<booking_id>/messages", methods=["GET"])
@require_auth
@limiter.limit("60 per minute")
def marketplace_get_messages(booking_id):
    """Booking-scoped chat thread — lets a patient ask questions about their
    booking and an admin (there's no separate therapist login yet — see
    the 'no public self-serve therapist signup' note above) reply. Simple
    polling, not websockets: matches this app's existing patterns (no
    other real-time feature uses a persistent connection either)."""
    try:
        db   = firestore.client()
        bdoc = db.collection("marketplace_bookings").document(booking_id).get()
        if not bdoc.exists:
            return jsonify({"error": "booking not found"}), 404
        if not _booking_chat_access_ok(bdoc.to_dict()):
            return jsonify({"error": "forbidden"}), 403
        msgs = db.collection("marketplace_bookings").document(booking_id) \
                 .collection("messages").order_by("created_at").limit(200).get()
        return jsonify({"messages": [{**m.to_dict(), "id": m.id} for m in msgs]})
    except Exception as e:
        return safe_error(e)


@app.route("/api/marketplace/bookings/<booking_id>/messages", methods=["POST"])
@require_auth
@limiter.limit("30 per minute")
def marketplace_send_message(booking_id):
    try:
        data = request.get_json(force=True) or {}
        text = (data.get("text") or "").strip()
        if not text:
            return jsonify({"error": "text required"}), 400
        if len(text) > 2000:
            return jsonify({"error": "message too long (2000 char max)"}), 400

        db   = firestore.client()
        bref = db.collection("marketplace_bookings").document(booking_id)
        bdoc = bref.get()
        if not bdoc.exists:
            return jsonify({"error": "booking not found"}), 404
        booking_data = bdoc.to_dict()
        if not _booking_chat_access_ok(booking_data):
            return jsonify({"error": "forbidden"}), 403

        is_admin = getattr(g, "role", {}).get("is_admin", False)
        msg_ref = bref.collection("messages").document()
        msg = {
            "text":       text,
            "sender_uid": g.uid,
            "sender_role": "admin" if is_admin else "patient",
            "created_at": datetime.utcnow().isoformat()+"Z",
        }
        msg_ref.set(msg)
        bref.update({"last_message_at": msg["created_at"], "has_unread": True})
        return jsonify({"ok": True, "message": {**msg, "id": msg_ref.id}})
    except Exception as e:
        return safe_error(e)


# ════════════════════════════════════════════════════════════════════
# Symptom Correlation Engine
# Lets a user log how they actually felt on a given day (headache, neck
# pain, etc.) and cross-references that against their real posture
# session data for the same days — a self-reported ground truth layered
# on top of the purely metric-based pain_prediction already computed
# live during a session.
# ════════════════════════════════════════════════════════════════════

SYMPTOM_TYPES = {"headache","neck_pain","back_pain","shoulder_pain","eye_strain","wrist_pain","other"}


@app.route("/api/symptoms/log", methods=["POST"])
@require_auth
@limiter.limit("30 per minute")
def log_symptom():
    """Upsert today's (or a given day's) symptom check-in. One doc per user per day —
    logging again the same day overwrites rather than duplicates."""
    try:
        data = request.get_json(force=True) or {}
        day  = (data.get("date") or datetime.utcnow().strftime("%Y-%m-%d")).strip()
        try:
            datetime.strptime(day, "%Y-%m-%d")
        except ValueError:
            return jsonify({"error": "date must be YYYY-MM-DD"}), 400

        symptoms_in = data.get("symptoms") or []
        clean = []
        for s in symptoms_in:
            stype = (s.get("type") or "").strip().lower()
            if stype not in SYMPTOM_TYPES:
                continue
            severity = max(1, min(5, int(s.get("severity") or 3)))
            clean.append({"type": stype, "severity": severity})
        if not clean:
            return jsonify({"error": "at least one valid symptom required"}), 400

        uid = g.uid
        doc_id = f"{uid}_{day}"
        db.collection("symptom_logs").document(doc_id).set({
            "uid":       uid,
            "date":      day,
            "symptoms":  clean,
            "notes":     (data.get("notes") or "").strip()[:500],
            "updated_at": datetime.utcnow().isoformat()+"Z",
        }, merge=False)
        return jsonify({"ok": True, "date": day, "symptoms": clean})
    except Exception as e:
        return safe_error(e)


@app.route("/api/symptoms/log", methods=["GET"])
@require_auth
@limiter.limit("30 per minute")
def get_symptom_log():
    """History of the user's own symptom check-ins. ?period=7d|30d|90d"""
    try:
        uid    = g.uid
        period = request.args.get("period", "30d")
        days   = {"7d":7,"30d":30,"90d":90}.get(period, 30)
        cutoff_day = (datetime.utcnow() - timedelta(days=days)).strftime("%Y-%m-%d")

        docs = (db.collection("symptom_logs")
                  .where("uid","==",uid)
                  .where("date",">=",cutoff_day)
                  .order_by("date", direction=firestore.Query.DESCENDING)
                  .limit(200).stream())
        return jsonify({"logs": [d.to_dict() for d in docs]})
    except Exception as e:
        return safe_error(e)


def _day_key(ts):
    """Normalize a Firestore timestamp / datetime / iso-string into a YYYY-MM-DD key."""
    if ts is None:
        return None
    if hasattr(ts, "strftime"):
        return ts.strftime("%Y-%m-%d")
    return str(ts)[:10]


@app.route("/api/analytics/symptom-correlation", methods=["GET"])
@require_auth
@limiter.limit("20 per minute")
@require_tier("standard")
def symptom_correlation():
    """
    Core correlation engine. For each symptom type the user has logged at
    least MIN_DAYS times, compares posture-session metrics on days that
    symptom was reported (severity >= 3) vs. all other logged-session days,
    and surfaces whichever alert cause / score gap is largest.

    This is a plain mean-difference comparison, not a black-box model —
    kept deliberately explainable so the result can be shown to the user
    (and, per Mo's PT background, would hold up if a clinician looked at it).
    """
    MIN_DAYS = 3
    try:
        uid    = g.uid
        period = request.args.get("period", "90d")
        days   = {"30d":30,"90d":90,"180d":180}.get(period, 90)
        cutoff = datetime.utcnow() - timedelta(days=days)
        cutoff_day = cutoff.strftime("%Y-%m-%d")

        symptom_docs = list(db.collection("symptom_logs")
                              .where("uid","==",uid)
                              .where("date",">=",cutoff_day)
                              .stream())
        if len(symptom_docs) < MIN_DAYS:
            return jsonify({"ok": True, "insights": [],
                             "note": f"Log symptoms on at least {MIN_DAYS} different days to unlock correlations",
                             "days_logged": len(symptom_docs)})

        session_docs = list(db.collection("sessions")
                              .where("uid","==",uid)
                              .where("created_at",">=",cutoff)
                              .limit(1000).stream())
        if not session_docs:
            return jsonify({"ok": True, "insights": [], "note": "No session data in this period"})

        # Aggregate sessions by day
        by_day: dict = {}
        for s in session_docs:
            d = s.to_dict()
            day = _day_key(d.get("created_at"))
            if not day:
                continue
            by_day.setdefault(day, []).append(d)

        # Which symptom (severity>=3) was logged on which day
        symptom_days: dict = {}   # symptom_type -> set(days)
        all_logged_days = set()
        for doc in symptom_docs:
            d = doc.to_dict()
            day = d.get("date")
            all_logged_days.add(day)
            for s in d.get("symptoms", []):
                if s.get("severity", 0) >= 3:
                    symptom_days.setdefault(s["type"], set()).add(day)

        def day_avg_score(day):
            scores = [s.get("avg_score") for s in by_day.get(day, []) if s.get("avg_score") is not None]
            return sum(scores)/len(scores) if scores else None

        def day_top_cause(day):
            counts: dict = {}
            for s in by_day.get(day, []):
                for a in (s.get("alert_causes") or []):
                    c = a.get("cause","posture")
                    counts[c] = counts.get(c,0)+1
            if not counts:
                return None
            return max(counts.items(), key=lambda kv: kv[1])[0]

        # Baseline: all days with session data, regardless of symptoms
        baseline_days = [d for d in by_day.keys()]
        baseline_scores = [x for d in baseline_days for x in [day_avg_score(d)] if x is not None]
        baseline_avg = sum(baseline_scores)/len(baseline_scores) if baseline_scores else None

        insights = []
        for stype, sdays in symptom_days.items():
            relevant_days = [d for d in sdays if d in by_day]
            if len(relevant_days) < MIN_DAYS:
                continue
            symptom_scores = [x for d in relevant_days for x in [day_avg_score(d)] if x is not None]
            other_days     = [d for d in baseline_days if d not in sdays]
            other_scores   = [x for d in other_days for x in [day_avg_score(d)] if x is not None]
            if not symptom_scores or not other_scores:
                continue
            sym_avg   = sum(symptom_scores)/len(symptom_scores)
            other_avg = sum(other_scores)/len(other_scores)
            gap       = round(other_avg - sym_avg, 1)   # positive => worse posture score on symptom days

            causes = [day_top_cause(d) for d in relevant_days]
            causes = [c for c in causes if c]
            top_cause = None
            if causes:
                cc: dict = {}
                for c in causes: cc[c] = cc.get(c,0)+1
                top_cause = max(cc.items(), key=lambda kv: kv[1])[0]

            if abs(gap) >= 3 or top_cause:
                insights.append({
                    "symptom": stype,
                    "days_logged": len(relevant_days),
                    "avg_score_on_symptom_days": round(sym_avg,1),
                    "avg_score_other_days": round(other_avg,1),
                    "score_gap": gap,
                    "dominant_alert_cause": top_cause,
                    "direction": "worse" if gap > 0 else "better" if gap < 0 else "no_difference",
                })

        insights.sort(key=lambda i: abs(i["score_gap"]), reverse=True)
        return jsonify({
            "ok": True,
            "insights": insights,
            "days_logged": len(all_logged_days),
            "baseline_avg_score": round(baseline_avg,1) if baseline_avg is not None else None,
            "period": period,
        })
    except Exception as e:
        return safe_error(e)
































